from datetime import date
from django.contrib import admin, messages
from django.shortcuts import render, redirect
from django.urls import path
from import_export import resources, fields
from import_export.admin import ImportExportModelAdmin
from import_export.widgets import ForeignKeyWidget
from openpyxl import load_workbook

from .models import Torneo, Organizador, Categoria, Documento, Equipo, Jugador, Partido, Gol, Tarjeta, AlineacionPartido, SustitucionPartido, ReglaEdadCategoria
from django.contrib import admin

admin.site.register(Torneo)
admin.site.register(Organizador)

def limpiar_texto(valor):
    return '' if valor is None else str(valor).strip()


def limpiar_cedula(valor):
    if valor is None:
        return ''
    valor = str(valor).strip().replace('.', '').replace(',', '').replace(' ', '')
    return valor[:-2] if valor.endswith('.0') else valor


def limpiar_entero(valor):
    if valor in [None, '']:
        return None
    try:
        return int(float(valor))
    except Exception:
        return None


def normalizar_anio(anio):
    anio = limpiar_entero(anio)
    if anio is None:
        return None
    if anio < 100:
        return 2000 + anio if anio <= 30 else 1900 + anio
    return anio


def construir_fecha_nacimiento(dia, mes, anio):
    dia = limpiar_entero(dia)
    mes = limpiar_entero(mes)
    anio = normalizar_anio(anio)
    if not dia or not mes or not anio:
        return None
    try:
        return date(anio, mes, dia)
    except Exception:
        return None


def calcular_edad(fecha_nacimiento):
    if not fecha_nacimiento:
        return None
    hoy = date.today()
    return hoy.year - fecha_nacimiento.year - ((hoy.month, hoy.day) < (fecha_nacimiento.month, fecha_nacimiento.day))


def rango_edad(fecha_nacimiento):
    edad = calcular_edad(fecha_nacimiento)
    if edad is None:
        return 'SIN FECHA'
    if edad >= 50:
        return '50+'
    if 45 <= edad <= 49:
        return '45-49'
    if 40 <= edad <= 44:
        return '40-44'
    return 'FUERA DE RANGO'


def obtener_hoja_planilla(workbook):
    for nombre in ['Planilla inscripcion', 'Planilla inscripción', 'PLANILLA INSCRIPCION', 'PLANILLA INSCRIPCIÓN', 'Inscripcion', 'Inscripción']:
        if nombre in workbook.sheetnames:
            return workbook[nombre]
    return workbook.active


def importar_planilla_inscripcion(request):
    if request.method == 'POST':
        archivo = request.FILES.get('archivo_excel')
        if not archivo:
            messages.error(request, 'Debes seleccionar un archivo Excel.')
            return redirect('/admin/importar-planilla-inscripcion/')
        try:
            wb = load_workbook(archivo, data_only=True)
            ws = obtener_hoja_planilla(wb)
            categoria_nombre = limpiar_texto(ws['D3'].value)
            equipo_nombre = limpiar_texto(ws['I3'].value)
            delegado = limpiar_texto(ws['D4'].value)
            telefono_delegado = limpiar_cedula(ws['I4'].value)
            director_tecnico = limpiar_texto(ws['C39'].value)
            telefono_dt = limpiar_cedula(ws['G39'].value)
            asistente_tecnico = limpiar_texto(ws['C40'].value)
            telefono_at = limpiar_cedula(ws['G40'].value)
            if not categoria_nombre:
                messages.error(request, 'No se encontró la categoría en la celda D3.')
                return redirect('/admin/importar-planilla-inscripcion/')
            if not equipo_nombre:
                messages.error(request, 'No se encontró el equipo en la celda I3.')
                return redirect('/admin/importar-planilla-inscripcion/')
            categoria = Categoria.objects.filter(nombre__iexact=categoria_nombre).first()
            if not categoria:
                messages.error(request, f'No existe la categoría: {categoria_nombre}. Créala primero en el admin.')
                return redirect('/admin/importar-planilla-inscripcion/')
            equipo = Equipo.objects.filter(nombre__iexact=equipo_nombre, categoria=categoria).first()
            if not equipo:
                equipo = Equipo.objects.create(nombre=equipo_nombre.upper(), categoria=categoria, activo=True)
            equipo.delegado = delegado.upper() if delegado else equipo.delegado
            equipo.telefono = telefono_delegado or equipo.telefono
            equipo.director_tecnico = director_tecnico.upper() if director_tecnico else equipo.director_tecnico
            equipo.telefono_dt = telefono_dt or equipo.telefono_dt
            equipo.asistente_tecnico = asistente_tecnico.upper() if asistente_tecnico else equipo.asistente_tecnico
            equipo.telefono_at = telefono_at or equipo.telefono_at
            equipo.activo = True
            equipo.save()
            creados = actualizados = omitidos = 0
            errores = []
            for fila in range(8, 38):
                nombre = limpiar_texto(ws[f'C{fila}'].value)
                dorsal = limpiar_entero(ws[f'D{fila}'].value)
                dia = ws[f'E{fila}'].value
                mes = ws[f'F{fila}'].value
                anio = ws[f'G{fila}'].value
                cedula = limpiar_cedula(ws[f'H{fila}'].value)
                if not nombre and not cedula:
                    continue
                if not nombre:
                    omitidos += 1
                    errores.append(f'Fila {fila}: falta el nombre del jugador.')
                    continue
                if not cedula:
                    omitidos += 1
                    errores.append(f'Fila {fila}: falta la cédula de {nombre}.')
                    continue
                fecha_nacimiento = construir_fecha_nacimiento(dia, mes, anio)
                if not fecha_nacimiento:
                    omitidos += 1
                    errores.append(f'Fila {fila}: fecha de nacimiento inválida para {nombre}.')
                    continue
                jugador_misma_categoria = Jugador.objects.filter(
                    cedula=cedula,
                    equipo__categoria=categoria,
                ).exclude(equipo=equipo).select_related('equipo').first()
                if jugador_misma_categoria:
                    omitidos += 1
                    errores.append(
                        f'Fila {fila}: {nombre} ya esta inscrito en {jugador_misma_categoria.equipo.nombre} '
                        f'para {categoria.nombre}.'
                    )
                    continue
                _, creado = Jugador.objects.update_or_create(
                    equipo=equipo,
                    cedula=cedula,
                    defaults={'dorsal': dorsal, 'nombres': nombre.upper(), 'fecha_nacimiento': fecha_nacimiento, 'estado': 'ACTIVO'},
                )
                if creado:
                    creados += 1
                else:
                    actualizados += 1
            messages.success(request, f'Planilla importada: {equipo.nombre} / {categoria.nombre}. Nuevos: {creados}. Actualizados: {actualizados}. Omitidos: {omitidos}.')
            for error in errores[:15]:
                messages.warning(request, error)
            if len(errores) > 15:
                messages.warning(request, f'Hay {len(errores) - 15} advertencias adicionales no mostradas.')
            return redirect('admin:torneos_jugador_changelist')
        except Exception as exc:
            messages.error(request, f'Error importando la planilla: {exc}')
            return redirect('/admin/importar-planilla-inscripcion/')
    return render(request, 'admin/importar_planilla_inscripcion.html')


_original_get_urls = admin.site.get_urls

def ir_a_generar_fixture(request):
    return redirect('gestion_generar_fixture')


def get_admin_urls():
    return [
        path('importar-planilla-inscripcion/', admin.site.admin_view(importar_planilla_inscripcion), name='importar_planilla_inscripcion'),
        path('generar-fixture/', admin.site.admin_view(ir_a_generar_fixture), name='generar_fixture'),
    ] + _original_get_urls()

admin.site.get_urls = get_admin_urls


class CategoriaWidget(ForeignKeyWidget):
    def clean(self, value, row=None, *args, **kwargs):
        if not value:
            return None
        categoria = Categoria.objects.filter(nombre__iexact=str(value).strip()).first()
        if not categoria:
            raise ValueError(f'No existe la categoría: {value}')
        return categoria


class EquipoWidget(ForeignKeyWidget):
    def clean(self, value, row=None, *args, **kwargs):
        if not value:
            return None
        nombre_equipo = str(value).strip()
        qs = Equipo.objects.filter(nombre__iexact=nombre_equipo)
        categoria_nombre = row.get('categoria') or row.get('Categoría') or row.get('CATEGORIA') if row else None
        if categoria_nombre:
            qs = qs.filter(categoria__nombre__iexact=str(categoria_nombre).strip())
        equipo = qs.first()
        if not equipo:
            raise ValueError(f'No existe el equipo: {nombre_equipo}')
        return equipo


class JugadorWidget(ForeignKeyWidget):
    def clean(self, value, row=None, *args, **kwargs):
        if not value:
            return None
        valor = str(value).strip()
        jugadores = Jugador.objects.all()
        equipo_nombre = row.get('equipo') or row.get('Equipo') or row.get('EQUIPO') if row else None
        categoria_nombre = row.get('categoria') or row.get('Categoría') or row.get('CategorÃ­a') or row.get('CATEGORIA') if row else None
        if equipo_nombre:
            jugadores = jugadores.filter(equipo__nombre__iexact=str(equipo_nombre).strip())
        if categoria_nombre:
            jugadores = jugadores.filter(equipo__categoria__nombre__iexact=str(categoria_nombre).strip())
        jugador = jugadores.filter(cedula=valor).first() or jugadores.filter(nombres__iexact=valor).first()
        if not jugador:
            raise ValueError(f'No existe el jugador: {valor}')
        return jugador


class CategoriaResource(resources.ModelResource):
    class Meta:
        model = Categoria
        import_id_fields = ('nombre',)
        fields = ('id', 'nombre', 'descripcion', 'edad_minima', 'edad_maxima', 'torneo', 'controlar_foraneos', 'porcentaje_minimo_foraneos')


class EquipoResource(resources.ModelResource):
    categoria = fields.Field(column_name='categoria', attribute='categoria', widget=CategoriaWidget(Categoria, 'nombre'))
    class Meta:
        model = Equipo
        import_id_fields = ('nombre', 'categoria')
        fields = ('id', 'nombre', 'categoria', 'delegado', 'telefono', 'director_tecnico', 'telefono_dt', 'asistente_tecnico', 'telefono_at', 'activo')
        skip_unchanged = True
        report_skipped = True


class JugadorResource(resources.ModelResource):
    equipo = fields.Field(column_name='equipo', attribute='equipo', widget=EquipoWidget(Equipo, 'nombre'))
    class Meta:
        model = Jugador
        import_id_fields = ('equipo', 'cedula')
        fields = ('id', 'equipo', 'dorsal', 'nombres', 'cedula', 'fecha_nacimiento', 'telefono', 'es_foraneo', 'estado')
        skip_unchanged = True
        report_skipped = True


class PartidoResource(resources.ModelResource):
    categoria = fields.Field(column_name='categoria', attribute='categoria', widget=CategoriaWidget(Categoria, 'nombre'))
    equipo_local = fields.Field(column_name='equipo_local', attribute='equipo_local', widget=EquipoWidget(Equipo, 'nombre'))
    equipo_visitante = fields.Field(column_name='equipo_visitante', attribute='equipo_visitante', widget=EquipoWidget(Equipo, 'nombre'))
    class Meta:
        model = Partido
        import_id_fields = ('categoria', 'fase', 'numero_fecha', 'equipo_local', 'equipo_visitante')
        fields = ('id', 'categoria', 'equipo_local', 'equipo_visitante', 'fecha', 'hora', 'goles_local', 'goles_visitante', 'estado', 'observaciones', 'numero_fecha', 'grupo', 'cancha', 'fase', 'ajuste_puntos_local', 'ajuste_puntos_visitante', 'observacion_comite', 'goles_local_penales', 'goles_visitante_penales')
        skip_unchanged = True
        report_skipped = True


class GolResource(resources.ModelResource):
    jugador = fields.Field(column_name='jugador', attribute='jugador', widget=JugadorWidget(Jugador, 'cedula'))
    equipo = fields.Field(column_name='equipo', attribute='equipo', widget=EquipoWidget(Equipo, 'nombre'))
    class Meta:
        model = Gol
        fields = ('id', 'partido', 'jugador', 'equipo', 'cantidad', 'es_autogol', 'es_penal')
        skip_unchanged = True
        report_skipped = True


class TarjetaResource(resources.ModelResource):
    jugador = fields.Field(column_name='jugador', attribute='jugador', widget=JugadorWidget(Jugador, 'cedula'))
    equipo = fields.Field(column_name='equipo', attribute='equipo', widget=EquipoWidget(Equipo, 'nombre'))
    class Meta:
        model = Tarjeta
        fields = ('id', 'partido', 'jugador', 'equipo', 'tipo')
        skip_unchanged = True
        report_skipped = True


class ReglaEdadCategoriaInline(admin.TabularInline):
    model = ReglaEdadCategoria
    extra = 1
    fields = ('etiqueta', 'edad_minima', 'edad_maxima', 'minimo_titulares', 'orden', 'activa')
    ordering = ('orden', 'edad_minima')


@admin.register(Categoria)
class CategoriaAdmin(ImportExportModelAdmin):
    resource_class = CategoriaResource
    list_display = ('nombre', 'torneo', 'edad_minima', 'edad_maxima', 'controlar_foraneos', 'porcentaje_minimo_foraneos')
    list_filter = ('torneo',)
    search_fields = ('nombre', 'torneo__nombre')
    inlines = [ReglaEdadCategoriaInline]


@admin.register(Documento)
class DocumentoAdmin(admin.ModelAdmin):
    list_display = ('tipo', 'torneo', 'titulo', 'activo', 'creado_en')
    list_filter = ('torneo', 'tipo', 'activo')
    search_fields = ('titulo', 'descripcion', 'torneo__nombre')
    ordering = ('tipo', '-creado_en', 'titulo')


class JugadorInline(admin.TabularInline):
    model = Jugador
    extra = 0
    fields = ('dorsal', 'nombres', 'cedula', 'fecha_nacimiento', 'es_foraneo', 'estado')
    ordering = ('dorsal', 'nombres')


@admin.register(Equipo)
class EquipoAdmin(ImportExportModelAdmin):
    resource_class = EquipoResource
    list_display = ('nombre', 'categoria', 'delegado', 'telefono', 'director_tecnico', 'telefono_dt', 'asistente_tecnico', 'telefono_at', 'activo')
    list_filter = ('categoria__torneo', 'categoria', 'activo')
    search_fields = ('nombre', 'delegado', 'telefono', 'director_tecnico', 'telefono_dt', 'asistente_tecnico', 'telefono_at')
    inlines = [JugadorInline]


@admin.register(Jugador)
class JugadorAdmin(ImportExportModelAdmin):
    resource_class = JugadorResource
    change_list_template = "admin/jugador_changelist.html"
    list_display = ('dorsal', 'nombres', 'equipo', 'cedula', 'fecha_nacimiento', 'edad_actual', 'rango', 'es_foraneo', 'estado')
    list_filter = ('equipo', 'equipo__categoria', 'es_foraneo', 'estado')
    search_fields = ('nombres', 'cedula', 'equipo__nombre')
    ordering = ('equipo__nombre', 'dorsal', 'nombres')

    @admin.display(description='Edad')
    def edad_actual(self, obj):
        edad = calcular_edad(obj.fecha_nacimiento)
        return edad if edad is not None else '-'

    @admin.display(description='Rango')
    def rango(self, obj):
        return rango_edad(obj.fecha_nacimiento)


class GolInline(admin.TabularInline):
    model = Gol
    extra = 0


class TarjetaInline(admin.TabularInline):
    model = Tarjeta
    extra = 0


class AlineacionInline(admin.TabularInline):
    model = AlineacionPartido
    extra = 0


class SustitucionInline(admin.TabularInline):
    model = SustitucionPartido
    extra = 0


@admin.register(Partido)
class PartidoAdmin(ImportExportModelAdmin):
    resource_class = PartidoResource
    list_display = ('categoria', 'grupo', 'numero_fecha', 'fase', 'equipo_local', 'equipo_visitante', 'goles_local', 'goles_visitante', 'estado', 'fecha', 'hora', 'cancha', 'ajuste_puntos_local', 'ajuste_puntos_visitante', 'goles_local_penales', 'goles_visitante_penales')
    list_filter = ('categoria__torneo', 'categoria', 'grupo', 'numero_fecha', 'fase', 'estado')
    search_fields = ('equipo_local__nombre', 'equipo_visitante__nombre', 'cancha')
    filter_horizontal = ('planilleros',)
    inlines = [GolInline, TarjetaInline, AlineacionInline, SustitucionInline]
    ordering = ('categoria__nombre', 'grupo', 'numero_fecha', 'fase', 'fecha', 'hora')


@admin.register(Gol)
class GolAdmin(ImportExportModelAdmin):
    resource_class = GolResource
    list_display = ('jugador', 'equipo', 'cantidad', 'es_autogol', 'es_penal', 'partido')
    list_filter = ('es_autogol', 'es_penal', 'equipo', 'partido__categoria', 'partido__grupo', 'partido__fase')
    search_fields = ('jugador__nombres', 'jugador__cedula', 'equipo__nombre')


@admin.register(Tarjeta)
class TarjetaAdmin(ImportExportModelAdmin):
    resource_class = TarjetaResource
    list_display = ('jugador', 'equipo', 'tipo', 'partido')
    list_filter = ('tipo', 'equipo', 'partido__categoria', 'partido__grupo', 'partido__fase')
    search_fields = ('jugador__nombres', 'jugador__cedula', 'equipo__nombre')


@admin.register(AlineacionPartido)
class AlineacionPartidoAdmin(admin.ModelAdmin):
    list_display = ('partido', 'equipo', 'jugador', 'rol')
    list_filter = ('equipo', 'rol', 'partido__categoria', 'partido__fase')
    search_fields = ('jugador__nombres', 'jugador__cedula', 'equipo__nombre')


@admin.register(ReglaEdadCategoria)
class ReglaEdadCategoriaAdmin(admin.ModelAdmin):
    list_display = ('categoria', 'etiqueta', 'edad_minima', 'edad_maxima', 'minimo_titulares', 'orden', 'activa')
    list_filter = ('categoria__torneo', 'categoria', 'activa')
    search_fields = ('categoria__nombre', 'etiqueta')
    ordering = ('categoria__nombre', 'orden', 'edad_minima')


@admin.register(SustitucionPartido)
class SustitucionPartidoAdmin(admin.ModelAdmin):
    list_display = ('partido', 'equipo', 'jugador_sale', 'jugador_entra', 'minuto')
    list_filter = ('equipo', 'partido__categoria', 'partido__fase')
    search_fields = ('jugador_sale__nombres', 'jugador_entra__nombres', 'equipo__nombre')

