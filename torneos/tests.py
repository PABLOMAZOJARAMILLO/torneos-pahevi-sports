from datetime import date, time, timedelta
from io import BytesIO
import os
from unittest.mock import patch

from types import SimpleNamespace

from django.core.exceptions import ValidationError
from django.contrib.auth.models import User
from django.contrib.auth.models import AnonymousUser
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db.models import Q
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.test import RequestFactory, TestCase, TransactionTestCase, override_settings
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from PIL import Image

from .forms import EquipoDelegadoForm, EquipoForm, JugadorForm, PartidoForm, PartidoProgramacionForm, TorneoForm
from .models import AlineacionPartido, EntregaAlineacionPartido, AdminOrganizador, AdminTorneo, Categoria, CobroPenal, Documento, Equipo, Gol, IncidenciaReglaEdad, Jugador, Organizador, Partido, ReglaEdadCategoria, RegistroActividad, VisitaPublicaDiaria, SolicitudValidacion, SustitucionPartido, Tarjeta, Torneo, ruta_escudo_equipo
from .middleware import AuditoriaModificacionesMiddleware
from .media_cleanup import eliminar_imagenes_sin_referencia, nombres_imagenes_instancias
from .planillas_pdf import _dorsal, _edad, _header_image_sources, _jugadores, _team_shield_source, _draw_team_watermark, _titulo_planilla, _nombre_jugador_planilla
from .storage_backends import CloudinaryMediaStorage
from .views import DocumentoStorageError, buscar_planilleros_excel, construir_estructura, construir_estadisticas_foraneos, construir_partidos_portada, construir_partidos_programacion, enriquecer_registros_actividad_legacy, fechas_presentes_en_programacion, foraneos_no_habilitados_fase_final, _clave_orden_evento_resumen, _equipo_turno_tanda, _minuto_evento_en_vivo, _sincronizar_no_disponibles_por_tarjetas, etiqueta_columna_planilla, etiqueta_edad_jugador, jugadores_actuales_en_cancha, nombre_corto_jugador, nombre_resumen_jugador, puede_descargar_programacion, podios_torneo, politica_reemplazo_jugador, reglas_edad_para_frontend, subir_documento_supabase, subir_documento_torneo, tercera_fecha_iniciada, texto_edad_jugador, tabla_general_mata_mata_ida_vuelta, url_imagen_cloudinary, validar_reglas_edad_titulares


class VisibilidadPublicaTorneoTests(TestCase):
    def setUp(self):
        self.visible = Torneo.objects.create(
            nombre="Copa pública", fecha_inicio=date(2026, 1, 1), visible_publico=True,
        )
        self.oculto = Torneo.objects.create(
            nombre="Copa privada", fecha_inicio=date(2026, 2, 1), visible_publico=False,
        )
        categoria = Categoria.objects.create(
            nombre="Única", torneo=self.oculto, edad_minima=18, edad_maxima=80,
        )
        local = Equipo.objects.create(nombre="Local privado", categoria=categoria)
        visitante = Equipo.objects.create(nombre="Visitante privado", categoria=categoria)
        self.partido_oculto = Partido.objects.create(
            categoria=categoria, equipo_local=local, equipo_visitante=visitante,
            fecha=date(2026, 2, 2), hora=time(16),
        )

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_portal_publico_no_muestra_torneo_oculto(self):
        respuesta = self.client.get("/?portal=1")

        self.assertContains(respuesta, "Copa pública")
        self.assertNotContains(respuesta, "Copa privada")

    def test_portal_publico_abre_con_el_almacenamiento_estatico_de_produccion(self):
        respuesta = self.client.get("/?portal=1")

        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, '<link rel="manifest" href="/static/manifest.')

    def test_enlace_directo_de_partido_oculto_no_es_publico(self):
        self.assertEqual(self.client.get(f"/partido/{self.partido_oculto.id}/live/").status_code, 404)
        self.assertEqual(self.client.get(f"/partido/{self.partido_oculto.id}/live/revision/").status_code, 404)

    def test_administrador_puede_cambiar_visibilidad(self):
        admin = User.objects.create_superuser("admin-visibilidad", password="clave")
        self.client.force_login(admin)

        respuesta = self.client.post(f"/gestion/torneos/{self.visible.id}/visibilidad/")

        self.assertEqual(respuesta.status_code, 302)
        self.visible.refresh_from_db()
        self.assertFalse(self.visible.visible_publico)
        self.assertEqual(self.client.get(f"/partido/{self.partido_oculto.id}/live/").status_code, 200)

    def test_formulario_incluye_control_de_visibilidad(self):
        self.assertIn("visible_publico", TorneoForm().fields)


class EquipoCuerpoTecnicoFormTests(TestCase):
    def test_datos_de_cada_miembro_aparecen_consecutivos(self):
        consecutivos = [
            "delegado", "telefono", "foto_delegado",
            "administrador_app", "telefono_administrador_app", "foto_administrador_app",
            "director_tecnico", "telefono_dt", "foto_director_tecnico",
            "asistente_tecnico", "telefono_at", "foto_asistente_tecnico",
        ]

        campos_admin = list(EquipoForm().fields)
        campos_delegado = list(EquipoDelegadoForm().fields)
        self.assertEqual([campo for campo in campos_admin if campo in consecutivos], consecutivos)
        self.assertEqual([campo for campo in campos_delegado if campo in consecutivos], consecutivos)


class CloudinaryStorageTests(TestCase):
    def test_tamanos_segun_tipo_de_imagen(self):
        storage = CloudinaryMediaStorage()

        self.assertEqual(storage._image_width("equipos/senior/escudo.png"), 160)
        self.assertEqual(storage._image_width("jugadores/senior/jugador.jpg"), 320)
        self.assertEqual(storage._image_width("equipos/senior/cuerpo_tecnico_1.jpg"), 320)
        self.assertEqual(storage._image_width("torneos/veranero/imagen_central.jpg"), 900)

    @override_settings(CLOUDINARY_URL="cloudinary://key:secret@test-cloud")
    def test_url_de_imagen_usa_transformacion_liviana(self):
        try:
            import cloudinary.utils  # noqa: F401
        except ModuleNotFoundError:
            self.skipTest("cloudinary no esta instalado en este entorno local")

        url = CloudinaryMediaStorage().url("equipos/senior/escudo.png")

        self.assertIn("f_auto", url)
        self.assertIn("q_auto", url)
        self.assertIn("c_limit", url)
        self.assertIn("w_160", url)

        foto = CloudinaryMediaStorage().url("jugadores/senior/equipo/jugador.jpg")
        portada = CloudinaryMediaStorage().url("torneos/veranero/imagen_central.jpg")

        self.assertIn("w_320", foto)
        self.assertIn("w_900", portada)

    @override_settings(CLOUDINARY_URL="cloudinary://key:secret@test-cloud")
    def test_miniatura_cloudinary_usa_transformacion_liviana(self):
        try:
            import cloudinary.utils  # noqa: F401
        except ModuleNotFoundError:
            self.skipTest("cloudinary no esta instalado en este entorno local")

        url = url_imagen_cloudinary("jugadores/senior/jugador", ancho=320)

        self.assertIn("f_auto", url)
        self.assertIn("q_auto", url)
        self.assertIn("c_limit", url)
        self.assertIn("w_320", url)


class LimpiezaImagenesTests(TestCase):
    class StorageFalso:
        def __init__(self):
            self.eliminadas = []

        def _public_id(self, name):
            nombre = str(name).replace("\\", "/").lstrip("/")
            return os.path.splitext(nombre)[0]

        def delete(self, name):
            self.eliminadas.append(name)

    def setUp(self):
        self.torneo = Torneo.objects.create(
            nombre="Torneo limpieza",
            fecha_inicio=date(2026, 1, 1),
        )
        self.categoria = Categoria.objects.create(
            nombre="Senior",
            edad_minima=18,
            edad_maxima=80,
            torneo=self.torneo,
        )

    def test_conserva_imagen_compartida_por_reinscripcion(self):
        origen = Equipo.objects.create(
            nombre="Equipo origen",
            categoria=self.categoria,
            escudo="equipos/senior/equipo/escudo.png",
        )
        destino = Equipo.objects.create(
            nombre="Equipo destino",
            categoria=self.categoria,
            escudo="equipos/senior/equipo/escudo",
        )
        storage = self.StorageFalso()
        imagenes = nombres_imagenes_instancias([origen])

        origen.delete()
        eliminadas = eliminar_imagenes_sin_referencia(imagenes, storage=storage)

        self.assertEqual(eliminadas, [])
        self.assertEqual(storage.eliminadas, [])
        self.assertTrue(Equipo.objects.filter(pk=destino.pk).exists())

    def test_elimina_imagen_cuando_desaparece_la_ultima_referencia(self):
        equipo = Equipo.objects.create(
            nombre="Equipo unico",
            categoria=self.categoria,
            escudo="equipos/senior/unico/escudo.png",
        )
        jugador = Jugador.objects.create(
            equipo=equipo,
            nombres="Jugador unico",
            cedula="123456",
            fecha_nacimiento=date(1990, 1, 1),
            foto="jugadores/senior/unico/123456.jpg",
        )
        storage = self.StorageFalso()
        imagenes = nombres_imagenes_instancias([equipo, jugador])

        equipo.delete()
        eliminadas = eliminar_imagenes_sin_referencia(imagenes, storage=storage)

        self.assertCountEqual(eliminadas, imagenes)
        self.assertCountEqual(storage.eliminadas, imagenes)


@override_settings(STORAGES={
    "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
    "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
})
class ArchivoTorneoYCanchasTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser(
            username="admin-archivo",
            email="admin@example.com",
            password="clave",
        )
        self.torneo = Torneo.objects.create(
            nombre="Copa Municipal",
            fecha_inicio=date(2026, 1, 1),
            canchas_habilitadas="Estadio Municipal\nCancha Auxiliar\nestadio municipal",
        )
        self.categoria = Categoria.objects.create(
            nombre="Senior",
            edad_minima=18,
            edad_maxima=80,
            torneo=self.torneo,
        )
        self.campeon = Equipo.objects.create(nombre="Equipo Campeon", categoria=self.categoria)
        self.subcampeon = Equipo.objects.create(nombre="Equipo Subcampeon", categoria=self.categoria)

    def crear_final(self):
        return Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.campeon,
            equipo_visitante=self.subcampeon,
            fecha=date(2026, 12, 20),
            hora=time(16, 0),
            fase="FINAL",
            estado="FINALIZADO",
            goles_local=2,
            goles_visitante=1,
        )

    def test_torneo_normaliza_canchas_y_formulario_usa_selector(self):
        self.assertEqual(
            self.torneo.lista_canchas(),
            ["Estadio Municipal", "Cancha Auxiliar"],
        )

        form = PartidoProgramacionForm(torneo=self.torneo)

        self.assertEqual(
            list(form.fields["cancha"].choices),
            [
                ("", "Selecciona una cancha"),
                ("Estadio Municipal", "Estadio Municipal"),
                ("Cancha Auxiliar", "Cancha Auxiliar"),
            ],
        )

    def test_finalizar_y_archivar_muestra_solo_podio_al_publico(self):
        self.crear_final()
        self.client.force_login(self.admin)

        respuesta = self.client.post(f"/gestion/torneos/{self.torneo.id}/finalizar/")
        self.assertRedirects(respuesta, "/gestion/torneos/")
        self.torneo.refresh_from_db()
        self.assertEqual(self.torneo.estado, "FINALIZADO")
        self.assertIsNotNone(self.torneo.fecha_fin)

        respuesta = self.client.post(f"/gestion/torneos/{self.torneo.id}/archivar/")
        self.assertRedirects(respuesta, "/gestion/torneos/")
        self.torneo.refresh_from_db()
        self.assertEqual(self.torneo.estado, "ARCHIVADO")

        self.client.logout()
        respuesta = self.client.get(f"/?torneo={self.torneo.id}")
        self.assertContains(respuesta, "Torneo finalizado y archivado")
        self.assertContains(respuesta, "Equipo Campeon")
        self.assertContains(respuesta, "Equipo Subcampeon")
        self.assertNotContains(respuesta, "PROGRAMACIÓN DE PARTIDOS")

        self.client.force_login(self.admin)
        respuesta = self.client.get(f"/?torneo={self.torneo.id}")
        self.assertContains(respuesta, "Torneo finalizado y archivado")
        self.assertNotContains(respuesta, "PROGRAMACIÓN DE PARTIDOS")

    def test_no_archiva_sin_final_cerrada(self):
        self.torneo.estado = "FINALIZADO"
        self.torneo.save(update_fields=["estado"])
        self.client.force_login(self.admin)

        self.client.post(f"/gestion/torneos/{self.torneo.id}/archivar/")

        self.torneo.refresh_from_db()
        self.assertEqual(self.torneo.estado, "FINALIZADO")
        self.assertEqual(podios_torneo(self.torneo), [])


class AuditoriaUsuariosTests(TestCase):
    def setUp(self):
        self.torneo = Torneo.objects.create(nombre="Auditoría", fecha_inicio=date(2026, 1, 1))
        self.categoria = Categoria.objects.create(
            nombre="Senior",
            edad_minima=18,
            edad_maxima=80,
            torneo=self.torneo,
        )
        self.delegado = User.objects.create_user("delegado-auditado", password="clave-segura")
        Equipo.objects.create(nombre="Equipo auditado", categoria=self.categoria, responsable=self.delegado)
        self.admin = User.objects.create_user("admin-auditoria", password="clave", is_staff=True)
        AdminTorneo.objects.create(usuario=self.admin, torneo=self.torneo)

    def test_registra_inicio_y_cierre_de_sesion_del_delegado(self):
        respuesta = self.client.post("/ingresar/", {
            "username": self.delegado.username,
            "password": "clave-segura",
        })

        self.assertEqual(respuesta.status_code, 200)
        ingreso = RegistroActividad.objects.get(usuario=self.delegado, accion="INICIAR_SESION")
        self.assertEqual(ingreso.datos["tipo_usuario"], "Delegado")
        self.assertEqual(ingreso.torneo, self.torneo)

        self.client.get("/salir/")

        self.assertTrue(RegistroActividad.objects.filter(
            usuario=self.delegado,
            accion="CERRAR_SESION",
        ).exists())

    def test_middleware_registra_una_modificacion_sin_guardar_formulario(self):
        request = RequestFactory().post("/delegado/accion/", {"dato_sensible": "no guardar"})
        request.user = self.delegado
        request.session = {"torneo_id": self.torneo.id}
        request.resolver_match = SimpleNamespace(url_name="accion_delegado")
        middleware = AuditoriaModificacionesMiddleware(lambda _request: HttpResponse(status=302))

        middleware(request)

        registro = RegistroActividad.objects.get(usuario=self.delegado, accion="ACCION_DELEGADO")
        self.assertEqual(registro.torneo, self.torneo)
        self.assertEqual(registro.datos["tipo_usuario"], "Delegado")
        self.assertEqual(registro.datos["ruta"], "/delegado/accion/")
        self.assertNotIn("dato_sensible", registro.datos)

    def test_middleware_asocia_torneo_desde_equipo_a_movimiento_delegado(self):
        equipo = Equipo.objects.get(responsable=self.delegado)
        request = RequestFactory().post(f"/equipos/delegado/{equipo.id}/editar/", {})
        request.user = self.delegado
        request.session = {}
        request.resolver_match = SimpleNamespace(
            url_name="delegado_equipo_editar",
            kwargs={"equipo_id": equipo.id},
        )
        middleware = AuditoriaModificacionesMiddleware(lambda _request: HttpResponse(status=302))

        middleware(request)

        registro = RegistroActividad.objects.get(usuario=self.delegado, accion="DELEGADO_EQUIPO_EDITAR")
        self.assertEqual(registro.torneo, self.torneo)
        self.assertEqual(registro.datos["tipo_usuario"], "Delegado")
        self.assertIn("Equipo auditado", registro.descripcion)

    def test_middleware_detalla_partido_equipo_jugador_e_infraccion(self):
        equipo = Equipo.objects.get(responsable=self.delegado)
        rival = Equipo.objects.create(nombre="Equipo rival", categoria=self.categoria)
        jugador = Jugador.objects.create(
            equipo=equipo,
            nombres="Jugador Infractor",
            cedula="AUD-1",
            fecha_nacimiento=date(1990, 1, 1),
        )
        partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=equipo,
            equipo_visitante=rival,
            fecha=date(2026, 8, 9),
            hora=time(14, 0),
            estado="EN_JUEGO",
        )
        request = RequestFactory().post(
            f"/partido/{partido.id}/agregar-tarjeta-movil/",
            {"equipo": equipo.id, "jugador": jugador.id, "tipo": "ROJA", "minuto_manual": "63"},
        )
        request.user = self.delegado
        request.session = {"torneo_id": self.torneo.id}
        request.resolver_match = SimpleNamespace(
            url_name="agregar_tarjeta_movil",
            kwargs={"partido_id": partido.id},
        )
        middleware = AuditoriaModificacionesMiddleware(lambda _request: HttpResponse(status=302))

        middleware(request)

        registro = RegistroActividad.objects.get(usuario=self.delegado, accion="REGISTRAR_INFRACCION")
        self.assertIn("Equipo auditado vs Equipo rival", registro.descripcion)
        self.assertIn("Equipo infractor: Equipo auditado", registro.descripcion)
        self.assertIn("Jugador Infractor", registro.descripcion)
        self.assertIn("ROJA", registro.descripcion)
        self.assertIn("63", registro.descripcion)
        self.assertEqual(registro.datos["partido_id"], partido.id)

    def test_middleware_detalla_equipo_y_jugadores_de_sustitucion(self):
        equipo = Equipo.objects.get(responsable=self.delegado)
        rival = Equipo.objects.create(nombre="Rival sustitución", categoria=self.categoria)
        sale = Jugador.objects.create(equipo=equipo, nombres="Carlos Sale", cedula="AUD-S", fecha_nacimiento=date(1990, 1, 1))
        entra = Jugador.objects.create(equipo=equipo, nombres="Pedro Entra", cedula="AUD-E", fecha_nacimiento=date(1991, 1, 1))
        partido = Partido.objects.create(
            categoria=self.categoria, equipo_local=equipo, equipo_visitante=rival,
            fecha=date(2026, 8, 9), hora=time(16, 0), estado="EN_JUEGO",
        )
        request = RequestFactory().post(
            f"/partido/{partido.id}/agregar-sustitucion-movil/",
            {"equipo": equipo.id, "jugador_sale": sale.id, "jugador_entra": entra.id, "minuto": "54"},
        )
        request.user = self.delegado
        request.session = {"torneo_id": self.torneo.id}
        request.resolver_match = SimpleNamespace(url_name="agregar_sustitucion_movil", kwargs={"partido_id": partido.id})
        middleware = AuditoriaModificacionesMiddleware(lambda _request: HttpResponse(status=302))

        middleware(request)

        registro = RegistroActividad.objects.get(usuario=self.delegado, accion="REGISTRAR_SUSTITUCION")
        self.assertIn("Equipo auditado vs Rival sustitución", registro.descripcion)
        self.assertIn("Equipo: Equipo auditado", registro.descripcion)
        self.assertIn("Salió: Carlos Sale", registro.descripcion)
        self.assertIn("Entró: Pedro Entra", registro.descripcion)
        self.assertIn("Minuto: 54", registro.descripcion)
        self.assertEqual(registro.datos["jugador_sale"], "Carlos Sale")
        self.assertEqual(registro.datos["jugador_entra"], "Pedro Entra")

    def test_auditoria_recupera_detalle_de_sustitucion_antigua(self):
        equipo = Equipo.objects.get(responsable=self.delegado)
        rival = Equipo.objects.create(nombre="Rival histórico", categoria=self.categoria)
        sale = Jugador.objects.create(equipo=equipo, nombres="Sale Histórico", cedula="H-S", fecha_nacimiento=date(1990, 1, 1))
        entra = Jugador.objects.create(equipo=equipo, nombres="Entra Histórico", cedula="H-E", fecha_nacimiento=date(1991, 1, 1))
        partido = Partido.objects.create(
            categoria=self.categoria, equipo_local=equipo, equipo_visitante=rival,
            fecha=date(2026, 8, 8), hora=time(17, 0), estado="EN_JUEGO",
        )
        SustitucionPartido.objects.create(
            partido=partido, equipo=equipo, jugador_sale=sale, jugador_entra=entra, minuto=54,
        )
        registro = RegistroActividad.objects.create(
            usuario=self.delegado, torneo=self.torneo, accion="MODIFICAR",
            descripcion=f"Operación POST en /partido/{partido.id}/agregar-sustitucion-movil/.",
            datos={"ruta": f"/partido/{partido.id}/agregar-sustitucion-movil/", "metodo": "POST"},
        )

        enriquecido = enriquecer_registros_actividad_legacy([registro])[0]

        self.assertEqual(enriquecido.accion, "REGISTRAR_SUSTITUCION")
        self.assertIn("Equipo auditado vs Rival histórico", enriquecido.descripcion)
        self.assertIn("Sale Histórico", enriquecido.descripcion)
        self.assertIn("Entra Histórico", enriquecido.descripcion)
        self.assertEqual(enriquecido.datos["equipo"], "Equipo auditado")

    def test_auditoria_recupera_detalle_de_gol_y_tarjeta_antiguos(self):
        equipo = Equipo.objects.get(responsable=self.delegado)
        rival = Equipo.objects.create(nombre="Rival eventos", categoria=self.categoria)
        jugador = Jugador.objects.create(equipo=equipo, nombres="Autor Evento", cedula="EV-1", fecha_nacimiento=date(1990, 1, 1))
        partido = Partido.objects.create(
            categoria=self.categoria, equipo_local=equipo, equipo_visitante=rival,
            fecha=date(2026, 8, 8), hora=time(18, 0), estado="EN_JUEGO",
        )
        Gol.objects.create(partido=partido, equipo=equipo, jugador=jugador, cantidad=1, minuto=35)
        Tarjeta.objects.create(partido=partido, equipo=equipo, jugador=jugador, tipo="AMARILLA", minuto=42)
        registro_gol = RegistroActividad.objects.create(
            usuario=self.delegado, torneo=self.torneo, accion="MODIFICAR",
            descripcion="Operación POST.",
            datos={"ruta": f"/partido/{partido.id}/agregar-gol-movil/", "metodo": "POST"},
        )
        registro_tarjeta = RegistroActividad.objects.create(
            usuario=self.delegado, torneo=self.torneo, accion="MODIFICAR",
            descripcion="Operación POST.",
            datos={"ruta": f"/partido/{partido.id}/agregar-tarjeta-movil/", "metodo": "POST"},
        )

        gol_enriquecido, tarjeta_enriquecida = enriquecer_registros_actividad_legacy([registro_gol, registro_tarjeta])

        self.assertEqual(gol_enriquecido.accion, "REGISTRAR_GOL")
        self.assertIn("Equipo auditado vs Rival eventos", gol_enriquecido.descripcion)
        self.assertIn("Autor Evento", gol_enriquecido.descripcion)
        self.assertIn("35", gol_enriquecido.descripcion)
        self.assertEqual(tarjeta_enriquecida.accion, "REGISTRAR_INFRACCION")
        self.assertIn("AMARILLA", tarjeta_enriquecida.descripcion.upper())
        self.assertIn("Autor Evento", tarjeta_enriquecida.descripcion)
        self.assertIn("42", tarjeta_enriquecida.descripcion)

    def test_auditoria_recupera_info_cronometro_penales_y_eliminaciones_antiguas(self):
        equipo = Equipo.objects.get(responsable=self.delegado)
        rival = Equipo.objects.create(nombre="Rival penales", categoria=self.categoria)
        jugador = Jugador.objects.create(equipo=equipo, nombres="Cobrador Histórico", cedula="P-H", fecha_nacimiento=date(1990, 1, 1))
        partido = Partido.objects.create(
            categoria=self.categoria, equipo_local=equipo, equipo_visitante=rival,
            fecha=date(2026, 8, 8), hora=time(19, 0), estado="EN_JUEGO",
            fase="SEMIFINAL", goles_local=1, goles_visitante=1, equipo_inicia_penales=equipo,
        )
        CobroPenal.objects.create(partido=partido, equipo=equipo, jugador=jugador, orden=1, convertido=True)

        def antiguo(ruta):
            return RegistroActividad.objects.create(
                usuario=self.delegado, torneo=self.torneo, accion="MODIFICAR",
                descripcion="Operación POST.", datos={"ruta": ruta, "metodo": "POST"},
            )

        info = antiguo(f"/partido/{partido.id}/guardar-info-movil/")
        pausa = antiguo(f"/partido/{partido.id}/cronometro/pausar/")
        inicio_penales = antiguo(f"/partido/{partido.id}/cronometro/penales/iniciar/")
        cobro = antiguo(f"/partido/{partido.id}/cronometro/penales/cobro/")
        eliminacion = antiguo("/gol/98765/eliminar-movil/")

        enriquecidos = enriquecer_registros_actividad_legacy([info, pausa, inicio_penales, cobro, eliminacion])

        self.assertEqual(enriquecidos[0].accion, "ACTUALIZAR_PARTIDO")
        self.assertIn("Equipo auditado vs Rival penales", enriquecidos[0].descripcion)
        self.assertEqual(enriquecidos[1].accion, "PAUSAR_CRONOMETRO")
        self.assertEqual(enriquecidos[2].accion, "INICIAR_PENALES")
        self.assertIn("Equipo auditado", enriquecidos[2].descripcion)
        self.assertEqual(enriquecidos[3].accion, "REGISTRAR_COBRO_PENAL")
        self.assertIn("Cobrador Histórico", enriquecidos[3].descripcion)
        self.assertIn("anotó", enriquecidos[3].descripcion)
        self.assertEqual(enriquecidos[4].accion, "ELIMINAR_GOL")
        self.assertIn("registro antiguo", enriquecidos[4].descripcion)

    def test_middleware_detalla_el_cobro_que_se_va_a_deshacer(self):
        equipo = Equipo.objects.get(responsable=self.delegado)
        rival = Equipo.objects.create(nombre="Rival deshacer", categoria=self.categoria)
        jugador = Jugador.objects.create(equipo=equipo, nombres="Último Cobrador", cedula="P-D", fecha_nacimiento=date(1990, 1, 1))
        partido = Partido.objects.create(
            categoria=self.categoria, equipo_local=equipo, equipo_visitante=rival,
            fecha=date(2026, 8, 9), hora=time(20, 0), estado="EN_JUEGO", fase="FINAL",
        )
        CobroPenal.objects.create(partido=partido, equipo=equipo, jugador=jugador, orden=3, convertido=False)
        request = RequestFactory().post(f"/partido/{partido.id}/cronometro/penales/deshacer/", {})
        request.user = self.delegado
        request.session = {"torneo_id": self.torneo.id}
        request.resolver_match = SimpleNamespace(url_name="deshacer_cobro_penal", kwargs={"partido_id": partido.id})
        middleware = AuditoriaModificacionesMiddleware(lambda _request: HttpResponse(status=302))

        middleware(request)

        registro = RegistroActividad.objects.get(usuario=self.delegado, accion="DESHACER_COBRO_PENAL")
        self.assertIn("Equipo auditado vs Rival deshacer", registro.descripcion)
        self.assertIn("Último Cobrador", registro.descripcion)
        self.assertIn("orden #3", registro.descripcion)

    def test_admin_puede_descargar_auditoria_csv(self):
        RegistroActividad.objects.create(
            usuario=self.delegado,
            torneo=self.torneo,
            accion="EDITAR",
            descripcion="Actualizó el equipo.",
            datos={"tipo_usuario": "Delegado", "ruta": "/delegado/equipo/"},
        )
        superusuario = User.objects.create_superuser("super-csv", password="clave")
        self.client.force_login(superusuario)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

        respuesta = self.client.get("/gestion/actividad/?formato=csv")

        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(respuesta["Content-Type"], "text/csv; charset=utf-8")
        contenido = respuesta.content.decode("utf-8-sig")
        self.assertIn("delegado-auditado", contenido)
        self.assertIn("Actualizó el equipo.", contenido)

    def test_visita_publica_cuenta_una_vez_por_dispositivo_torneo_y_dia(self):
        middleware = AuditoriaModificacionesMiddleware(lambda _request: HttpResponse(status=200))
        request = RequestFactory().get("/")
        request.user = AnonymousUser()
        request.session = {"torneo_id": self.torneo.id}
        request.resolver_match = SimpleNamespace(url_name="panel", kwargs={})

        respuesta = middleware(request)

        self.assertEqual(VisitaPublicaDiaria.objects.count(), 1)
        visita = VisitaPublicaDiaria.objects.get()
        self.assertEqual(visita.torneo, self.torneo)
        self.assertEqual(visita.canal, "ESCRITORIO")
        self.assertNotEqual(visita.visitante_hash, respuesta.cookies["pahevi_visitante"].value)

        segundo_request = RequestFactory().get("/")
        segundo_request.user = AnonymousUser()
        segundo_request.session = {"torneo_id": self.torneo.id}
        segundo_request.resolver_match = SimpleNamespace(url_name="panel", kwargs={})
        segundo_request.COOKIES["pahevi_visitante"] = respuesta.cookies["pahevi_visitante"].value
        segundo_request.COOKIES["pahevi_visita_contada"] = respuesta.cookies["pahevi_visita_contada"].value

        middleware(segundo_request)

        self.assertEqual(VisitaPublicaDiaria.objects.count(), 1)

    def test_metricas_publicas_solo_son_visibles_para_superusuario(self):
        VisitaPublicaDiaria.objects.create(
            fecha=date.today(),
            torneo=self.torneo,
            visitante_hash="hash-anonimo",
            canal="APK",
        )
        self.client.force_login(self.admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

        respuesta_admin = self.client.get("/gestion/actividad/")

        self.assertEqual(respuesta_admin.status_code, 403)
        self.assertContains(
            respuesta_admin,
            "únicamente para superusuarios",
            status_code=403,
        )

        superusuario = User.objects.create_superuser("super-auditoria", password="clave")
        self.client.force_login(superusuario)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

        respuesta_superusuario = self.client.get("/gestion/actividad/")

        self.assertContains(respuesta_superusuario, "Visitas públicas anónimas")
        self.assertContains(respuesta_superusuario, "Visible solo para superusuarios")

    def test_admin_de_comite_no_ve_enlace_de_auditoria(self):
        self.client.force_login(self.admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

        respuesta = self.client.get("/gestion/")

        self.assertEqual(respuesta.status_code, 200)
        self.assertNotContains(respuesta, "Auditoría de usuarios")
        self.assertNotContains(respuesta, "/gestion/actividad/")

    def test_delegado_ve_movimientos_de_todos_los_usuarios_de_su_torneo_sin_descarga(self):
        otro_torneo = Torneo.objects.create(nombre="Otro torneo", fecha_inicio=date(2026, 2, 1))
        RegistroActividad.objects.create(
            usuario=self.admin,
            torneo=self.torneo,
            accion="EDITAR",
            descripcion="Movimiento transparente del administrador.",
            ip="10.0.0.8",
            user_agent="Dispositivo privado",
        )
        RegistroActividad.objects.create(
            usuario=self.admin,
            torneo=otro_torneo,
            accion="EDITAR",
            descripcion="Movimiento de otro torneo oculto.",
        )
        self.client.force_login(self.delegado)

        respuesta = self.client.get("/gestion/actividad/")

        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, "Movimiento transparente del administrador.")
        self.assertNotContains(respuesta, "Movimiento de otro torneo oculto.")
        self.assertNotContains(respuesta, "Descargar CSV")
        self.assertNotContains(respuesta, "10.0.0.8")
        self.assertNotContains(respuesta, "Dispositivo privado")

        respuesta_csv = self.client.get("/gestion/actividad/?formato=csv")
        self.assertEqual(respuesta_csv.status_code, 403)


class CambioContrasenaTests(TestCase):
    def setUp(self):
        self.torneo = Torneo.objects.create(nombre="Contraseñas", fecha_inicio=date(2026, 1, 1))
        self.categoria = Categoria.objects.create(
            nombre="Senior",
            edad_minima=18,
            edad_maxima=80,
            torneo=self.torneo,
        )
        self.usuario = User.objects.create_user("delegado-clave", password="ClaveActual2026!")
        Equipo.objects.create(nombre="Equipo clave", categoria=self.categoria, responsable=self.usuario)

    def test_usuario_asignado_puede_cambiar_su_propia_contrasena_sin_cerrar_sesion(self):
        self.client.force_login(self.usuario)

        respuesta = self.client.post("/mi-cuenta/cambiar-contrasena/", {
            "old_password": "ClaveActual2026!",
            "new_password1": "NuevaClaveSegura2026!",
            "new_password2": "NuevaClaveSegura2026!",
        })

        self.assertRedirects(respuesta, "/mi-cuenta/cambiar-contrasena/", fetch_redirect_response=False)
        self.usuario.refresh_from_db()
        self.assertTrue(self.usuario.check_password("NuevaClaveSegura2026!"))
        self.assertIn("_auth_user_id", self.client.session)
        registro = RegistroActividad.objects.get(usuario=self.usuario, accion="CAMBIAR_CONTRASENA")
        self.assertEqual(registro.torneo, self.torneo)
        self.assertNotIn("password", str(registro.datos).lower())

    def test_visitante_no_puede_abrir_cambio_de_contrasena(self):
        respuesta = self.client.get("/mi-cuenta/cambiar-contrasena/")

        self.assertEqual(respuesta.status_code, 302)
        self.assertIn("/ingresar/", respuesta.url)


class EscudoEquipoTests(TestCase):
    def test_ruta_escudo_equipo_genera_nombre_unico_para_evitar_cache(self):
        torneo = Torneo.objects.create(nombre="Veranero", fecha_inicio=date(2026, 1, 1))
        categoria = Categoria.objects.create(nombre="Senior Master", edad_minima=40, edad_maxima=80, torneo=torneo)
        equipo = Equipo(nombre="Paraiso", categoria=categoria)

        primera_ruta = ruta_escudo_equipo(equipo, "escudo.png")
        segunda_ruta = ruta_escudo_equipo(equipo, "escudo.png")

        self.assertNotEqual(primera_ruta, segunda_ruta)
        self.assertTrue(primera_ruta.startswith("equipos/SENIOR_MASTER/PARAISO/escudo_"))
        self.assertTrue(primera_ruta.endswith(".png"))


class DocumentosTorneoTests(TestCase):
    def setUp(self):
        self.torneo = Torneo.objects.create(nombre="Veranero", fecha_inicio=date(2026, 1, 1))
        self.otro_torneo = Torneo.objects.create(nombre="Copa Antigua", fecha_inicio=date(2025, 1, 1))
        self.usuario = User.objects.create_user("admin-docs", password="clave")
        AdminTorneo.objects.create(usuario=self.usuario, torneo=self.torneo, puede_editar=True, puede_validar=True, puede_programar=True)
        self.documento_actual = Documento.objects.create(
            torneo=self.torneo,
            tipo="REGLAMENTO",
            titulo="Reglamento Veranero",
            archivo="https://example.com/veranero.pdf",
            activo=True,
        )
        self.documento_otro = Documento.objects.create(
            torneo=self.otro_torneo,
            tipo="REGLAMENTO",
            titulo="Reglamento Copa Antigua",
            archivo="https://example.com/antigua.pdf",
            activo=True,
        )
        self.documento_global = Documento.objects.create(
            tipo="REGLAMENTO",
            titulo="Reglamento Sin Torneo",
            archivo="https://example.com/global.pdf",
            activo=True,
        )

    def seleccionar_torneo(self):
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_panel_publico_muestra_solo_documentos_del_torneo_actual(self):
        self.seleccionar_torneo()
        response = self.client.get("/")

        self.assertContains(response, "Reglamento Veranero")
        self.assertNotContains(response, "Reglamento Copa Antigua")
        self.assertNotContains(response, "Reglamento Sin Torneo")

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_panel_publico_muestra_planillas_activas_del_torneo(self):
        categoria = Categoria.objects.create(
            torneo=self.torneo,
            nombre="Senior Master",
            edad_minima=40,
            edad_maxima=80,
        )
        local = Equipo.objects.create(nombre="Templarios", categoria=categoria)
        visitante = Equipo.objects.create(nombre="Villamatoso", categoria=categoria)
        Documento.objects.create(
            torneo=self.torneo,
            tipo="PLANILLA_JUEGO",
            titulo="Planilla Mama Mata Fecha 1",
            archivo="https://example.com/planilla-mama-mata.jpg",
            categoria=categoria,
            equipo_local=local,
            equipo_visitante=visitante,
            numero_fecha="1",
            activo=True,
        )
        Documento.objects.create(
            torneo=self.otro_torneo,
            tipo="PLANILLA_JUEGO",
            titulo="Planilla Copa Antigua",
            archivo="https://example.com/planilla-antigua.jpg",
            activo=True,
        )
        self.seleccionar_torneo()

        response = self.client.get("/")

        self.assertContains(response, 'data-documento-seccion="planillas"')
        self.assertContains(response, "Planilla Mama Mata Fecha 1")
        self.assertContains(response, '<details class="planillas-categoria"')
        self.assertContains(response, "Senior Master")
        self.assertContains(response, '<details class="planillas-fecha"')
        self.assertContains(response, "<summary>FECHA 1</summary>", html=True)
        self.assertNotContains(response, "Planilla Copa Antigua")

    def test_planilla_publica_abre_archivo_original_sin_visor_de_google(self):
        documento = Documento.objects.create(
            torneo=self.torneo,
            tipo="PLANILLA_JUEGO",
            titulo="Planilla en imagen",
            archivo="https://example.com/planilla.jpg",
            activo=True,
        )
        self.seleccionar_torneo()

        response = self.client.get(f"/documentos/{documento.id}/")

        self.assertRedirects(
            response,
            "https://example.com/planilla.jpg",
            fetch_redirect_response=False,
        )

    def test_documento_publico_envia_al_visor_con_url_externa(self):
        self.seleccionar_torneo()

        response = self.client.get(f"/documentos/{self.documento_actual.id}/")

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response.url,
            "https://docs.google.com/gview?embedded=1&url=https%3A%2F%2Fexample.com%2Fveranero.pdf",
        )
        self.assertNotIn("/documentos/", response.url)

    def test_ruta_antigua_de_archivo_redirige_sin_retransmitir(self):
        self.seleccionar_torneo()

        response = self.client.get(
            f"/documentos/{self.documento_actual.id}/archivo.pdf",
        )

        self.assertRedirects(
            response,
            "https://example.com/veranero.pdf",
            fetch_redirect_response=False,
        )

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_panel_publico_muestra_otros_documentos_del_torneo(self):
        Documento.objects.create(
            torneo=self.torneo,
            tipo="OTRO",
            titulo="Circular informativa",
            archivo="https://example.com/circular.pdf",
            activo=True,
        )
        self.seleccionar_torneo()

        response = self.client.get("/")

        self.assertContains(response, 'data-documento-seccion="otros"')
        self.assertContains(response, "Circular informativa")

    def test_gestion_documentos_muestra_solo_documentos_del_torneo_actual(self):
        self.client.force_login(self.usuario)
        self.seleccionar_torneo()
        response = self.client.get("/gestion/documentos/")

        self.assertContains(response, "Reglamento Veranero")
        self.assertNotContains(response, "Reglamento Copa Antigua")
        self.assertNotContains(response, "Reglamento Sin Torneo")

    def test_formulario_documento_permite_categoria_o_documento_general(self):
        categoria = Categoria.objects.create(
            torneo=self.torneo, nombre="Senior Master", edad_minima=40, edad_maxima=80,
        )
        Categoria.objects.create(
            torneo=self.otro_torneo, nombre="Categoría ajena", edad_minima=18, edad_maxima=80,
        )
        self.client.force_login(self.usuario)
        self.seleccionar_torneo()

        response = self.client.get("/gestion/documentos/nuevo/")

        self.assertContains(response, "Documento general del torneo (sin categoría)")
        self.assertContains(response, categoria.nombre)
        self.assertNotContains(response, "Categoría ajena")

    def test_gestion_documentos_filtra_por_categoria_y_generales(self):
        categoria = Categoria.objects.create(
            torneo=self.torneo, nombre="Senior Master", edad_minima=40, edad_maxima=80,
        )
        Documento.objects.create(
            torneo=self.torneo, categoria=categoria, tipo="COMUNICADO",
            titulo="Comunicado Senior", archivo="https://example.com/senior.pdf", activo=True,
        )
        self.client.force_login(self.usuario)
        self.seleccionar_torneo()

        por_categoria = self.client.get("/gestion/documentos/", {"categoria": categoria.id})
        generales = self.client.get("/gestion/documentos/", {"categoria": "general"})

        self.assertContains(por_categoria, "Comunicado Senior")
        self.assertNotContains(por_categoria, "Reglamento Veranero")
        self.assertContains(generales, "Reglamento Veranero")
        self.assertNotContains(generales, "Comunicado Senior")

    def test_no_permite_abrir_documento_de_otro_torneo_por_url_directa(self):
        self.seleccionar_torneo()

        response = self.client.get(f"/documentos/{self.documento_otro.id}/")

        self.assertEqual(response.status_code, 404)

    def test_no_permite_editar_documento_de_otro_torneo(self):
        self.client.force_login(self.usuario)
        self.seleccionar_torneo()

        response = self.client.get(f"/gestion/documentos/{self.documento_otro.id}/editar/")

        self.assertEqual(response.status_code, 404)

    def test_formulario_documento_no_ofrece_planilla_de_juego(self):
        self.client.force_login(self.usuario)
        self.seleccionar_torneo()

        response = self.client.get("/gestion/documentos/nuevo/")

        self.assertContains(response, "Nuevo documento")
        self.assertNotContains(response, "Planilla de juego")

    @patch("torneos.views.subir_documento_torneo", side_effect=DocumentoStorageError("Almacenamiento no disponible."))
    def test_error_de_almacenamiento_no_genera_pantalla_500(self, _subir):
        self.client.force_login(self.usuario)
        self.seleccionar_torneo()
        cantidad_inicial = Documento.objects.count()

        response = self.client.post("/gestion/documentos/nuevo/", {
            "torneo": self.torneo.id,
            "tipo": "REGLAMENTO",
            "titulo": "Documento sin cargar",
            "descripcion": "Prueba",
            "activo": "on",
            "archivo_subido": SimpleUploadedFile("reglamento.pdf", b"%PDF-1.4", content_type="application/pdf"),
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Almacenamiento no disponible")
        self.assertEqual(Documento.objects.count(), cantidad_inicial)


class AlmacenamientoDocumentosTests(TestCase):
    @patch.dict(os.environ, {
        "SUPABASE_STORAGE_BUCKET": "torneos-media",
        "SUPABASE_S3_ENDPOINT_URL": "https://example.supabase.co/storage/v1/s3",
        "SUPABASE_S3_ACCESS_KEY_ID": "access",
        "SUPABASE_S3_SECRET_ACCESS_KEY": "secret",
        "SUPABASE_S3_REGION_NAME": "us-east-1",
        "SUPABASE_PUBLIC_MEDIA_URL": "https://example.supabase.co/storage/v1/object/public/torneos-media",
    })
    @patch("boto3.client")
    def test_normaliza_nombre_con_enie_en_clave_supabase(self, cliente_boto):
        archivo = SimpleUploadedFile(
            "COMUNICADO 002 TORNEO FIN DE AÑO 2026.pdf",
            b"%PDF-1.4",
            content_type="application/pdf",
        )

        url = subir_documento_supabase(archivo, "COMUNICADO")

        llave = cliente_boto.return_value.upload_fileobj.call_args.args[2]
        self.assertTrue(llave.isascii())
        self.assertIn("ANO_2026.pdf", llave)
        self.assertNotIn("Ñ", llave)
        self.assertIn(llave, url)

    @patch("torneos.views.subir_documento_cloudinary", return_value="https://cloudinary.example/documento.pdf")
    @patch("torneos.views.subir_documento_supabase", side_effect=RuntimeError("PutObject rechazado"))
    def test_usa_cloudinary_si_supabase_rechaza_putobject(self, _supabase, cloudinary):
        archivo = SimpleUploadedFile("documento.pdf", b"%PDF-1.4", content_type="application/pdf")

        url = subir_documento_torneo(archivo, "REGLAMENTO")

        self.assertEqual(url, "https://cloudinary.example/documento.pdf")
        cloudinary.assert_called_once()

    @patch("torneos.views.subir_documento_cloudinary", side_effect=RuntimeError("Cloudinary rechazado"))
    @patch("torneos.views.subir_documento_supabase", side_effect=RuntimeError("PutObject rechazado"))
    def test_informa_error_seguro_si_ambos_almacenamientos_fallan(self, _supabase, _cloudinary):
        archivo = SimpleUploadedFile("documento.pdf", b"%PDF-1.4", content_type="application/pdf")

        with self.assertRaises(DocumentoStorageError):
            subir_documento_torneo(archivo, "REGLAMENTO")


class PlanillasJuegoUploadTests(TestCase):
    def setUp(self):
        self.torneo = Torneo.objects.create(nombre="Veranero", fecha_inicio=date(2026, 1, 1))
        self.categoria = Categoria.objects.create(nombre="Senior Master", edad_minima=40, edad_maxima=80, torneo=self.torneo)
        self.equipo_local = Equipo.objects.create(nombre="Niqueleros FC", categoria=self.categoria)
        self.equipo_visitante = Equipo.objects.create(nombre="Integracion 28", categoria=self.categoria)
        self.planillero = User.objects.create_user("planillero-docs", password="clave")
        self.partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo_local,
            equipo_visitante=self.equipo_visitante,
            fecha=date(2026, 6, 3),
            hora=time(16, 0),
            estado="FINALIZADO",
            numero_fecha="Fecha 1",
        )
        self.partido.planilleros.add(self.planillero)

    def archivo_prueba(self):
        return SimpleUploadedFile("planilla.jpg", b"imagen", content_type="image/jpeg")

    def crear_documento_planilla(self):
        return Documento.objects.create(
            tipo="PLANILLA_JUEGO",
            torneo=self.torneo,
            categoria=self.categoria,
            partido=self.partido,
            equipo_local=self.equipo_local,
            equipo_visitante=self.equipo_visitante,
            titulo="Planilla Fecha 1",
            archivo="https://example.com/fecha1.jpg",
            numero_fecha="Fecha 1",
            fecha_partido=self.partido.fecha,
            hora_partido=self.partido.hora,
            cargado_por=self.planillero,
        )

    def crear_partido_programado(self):
        partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo_local,
            equipo_visitante=self.equipo_visitante,
            fecha=date(2026, 6, 8),
            hora=time(18, 0),
            estado="PROGRAMADO",
            numero_fecha="Fecha 2",
        )
        partido.planilleros.add(self.planillero)
        return partido

    def test_login_planillero_muestra_acceso_exitoso_con_mis_partidos(self):
        response = self.client.post("/ingresar/", {
            "username": "planillero-docs",
            "password": "clave",
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Acceso exitoso")
        self.assertContains(response, "/planillero/partidos/")

    def test_planillero_ve_editor_de_partidos_asignados_editables(self):
        partido = self.crear_partido_programado()
        self.client.force_login(self.planillero)

        response = self.client.get("/planillero/partidos/")

        self.assertContains(response, "Editor juego")
        self.assertContains(response, f"/partido/{partido.id}/editor-movil/")
        self.assertContains(response, "Cargar planilla")

    def test_planillero_solo_ve_partidos_del_torneo_seleccionado(self):
        partido = self.crear_partido_programado()
        otro_torneo = Torneo.objects.create(nombre="Amistoso", fecha_inicio=date(2026, 7, 1))
        otra_categoria = Categoria.objects.create(nombre="Senior Master", edad_minima=40, edad_maxima=80, torneo=otro_torneo)
        otro_local = Equipo.objects.create(nombre="Paraiso", categoria=otra_categoria)
        otro_visitante = Equipo.objects.create(nombre="Riverenos", categoria=otra_categoria)
        otro_partido = Partido.objects.create(
            categoria=otra_categoria,
            equipo_local=otro_local,
            equipo_visitante=otro_visitante,
            fecha=date(2026, 7, 10),
            hora=time(18, 0),
            estado="PROGRAMADO",
            numero_fecha="Fecha 1",
        )
        otro_partido.planilleros.add(self.planillero)
        self.client.force_login(self.planillero)

        response = self.client.get(f"/planillero/partidos/?torneo={self.torneo.id}")

        self.assertContains(response, partido.equipo_local.nombre)
        self.assertNotContains(response, otro_partido.equipo_local.nombre)

    def test_planillero_no_ve_partido_finalizado_con_planilla_cargada(self):
        Documento.objects.create(
            tipo="PLANILLA_JUEGO",
            torneo=self.torneo,
            categoria=self.categoria,
            partido=self.partido,
            equipo_local=self.equipo_local,
            equipo_visitante=self.equipo_visitante,
            titulo="Planilla Fecha 1",
            archivo="https://example.com/fecha1.jpg",
            numero_fecha="Fecha 1",
            fecha_partido=date(2026, 6, 3),
            hora_partido=time(16, 0),
            cargado_por=self.planillero,
        )
        self.client.force_login(self.planillero)

        response = self.client.get(f"/planillero/partidos/?torneo={self.torneo.id}")

        self.assertNotContains(response, self.equipo_local.nombre)
        self.assertContains(response, "No tienes partidos asignados con ese filtro.")

    @patch("torneos.views.subir_documento_torneo", return_value="https://example.com/planilla.jpg")
    def test_planillero_puede_cargar_planilla_de_partido_asignado(self, upload_mock):
        self.client.force_login(self.planillero)

        response = self.client.post("/gestion/planillas-juego/nueva/", {
            "partido": self.partido.id,
            "categoria": self.categoria.id,
            "numero_fecha": "Fecha 1",
            "equipo_local": self.equipo_local.id,
            "equipo_visitante": self.equipo_visitante.id,
            "fecha_partido": "2026-06-03",
            "hora_partido": "16:00",
            "imagenes": self.archivo_prueba(),
        })

        self.assertEqual(response.status_code, 302)
        documento = Documento.objects.get(tipo="PLANILLA_JUEGO")
        self.assertEqual(documento.partido, self.partido)
        self.assertEqual(documento.categoria, self.categoria)
        self.assertEqual(documento.equipo_local, self.equipo_local)
        self.assertEqual(documento.equipo_visitante, self.equipo_visitante)
        self.assertEqual(documento.cargado_por, self.planillero)
        self.assertEqual(documento.archivo, "https://example.com/planilla.jpg")
        upload_mock.assert_called_once()

    @patch("torneos.views.subir_documento_torneo", return_value="https://example.com/planilla.jpg")
    def test_planillero_no_puede_cargar_planilla_de_partido_no_asignado(self, upload_mock):
        otro_planillero = User.objects.create_user("otro-planillero", password="clave")
        self.client.force_login(otro_planillero)

        response = self.client.post("/gestion/planillas-juego/nueva/", {
            "partido": self.partido.id,
            "categoria": self.categoria.id,
            "numero_fecha": "Fecha 1",
            "equipo_local": self.equipo_local.id,
            "equipo_visitante": self.equipo_visitante.id,
            "fecha_partido": "2026-06-03",
            "hora_partido": "16:00",
            "imagenes": self.archivo_prueba(),
        })

        self.assertEqual(response.status_code, 302)
        self.assertFalse(Documento.objects.filter(tipo="PLANILLA_JUEGO").exists())
        upload_mock.assert_not_called()

    @patch("torneos.views.subir_documento_torneo", return_value="https://example.com/planilla.jpg")
    def test_partido_del_fixture_autocompleta_datos_de_planilla(self, upload_mock):
        self.client.force_login(self.planillero)

        response = self.client.post("/gestion/planillas-juego/nueva/", {
            "partido": self.partido.id,
            "imagenes": self.archivo_prueba(),
        })

        self.assertEqual(response.status_code, 302)
        documento = Documento.objects.get(tipo="PLANILLA_JUEGO")
        self.assertEqual(documento.categoria, self.categoria)
        self.assertEqual(documento.equipo_local, self.equipo_local)
        self.assertEqual(documento.equipo_visitante, self.equipo_visitante)
        self.assertEqual(documento.numero_fecha, "Fecha 1")
        self.assertEqual(documento.fecha_partido, self.partido.fecha)
        self.assertEqual(documento.hora_partido, self.partido.hora)
        upload_mock.assert_called_once()

    def test_planillero_ve_partidos_asignados_aunque_no_tengan_planilla(self):
        self.client.force_login(self.planillero)

        response = self.client.get("/gestion/planillas-juego/")

        self.assertContains(response, "Niqueleros FC")
        self.assertContains(response, "Integracion 28")
        self.assertContains(response, "Sin planillas cargadas")
        self.assertContains(response, "Fecha 1")
        self.assertEqual(list(response.context["categorias"]), [self.categoria])
        self.assertEqual(list(response.context["partidos"]), [self.partido])

    def test_gestion_planillas_y_partidos_muestra_primero_el_proximo_programado(self):
        manana = date.today() + timedelta(days=1)
        pasado_manana = date.today() + timedelta(days=2)
        proximo = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo_local,
            equipo_visitante=self.equipo_visitante,
            fecha=manana,
            hora=time(10, 0),
            estado="PROGRAMADO",
            numero_fecha="Fecha 3",
        )
        posterior = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo_local,
            equipo_visitante=self.equipo_visitante,
            fecha=pasado_manana,
            hora=time(8, 0),
            estado="PROGRAMADO",
            numero_fecha="Fecha 4",
        )
        proximo.planilleros.add(self.planillero)
        posterior.planilleros.add(self.planillero)
        self.client.force_login(self.planillero)

        listado_planillas = self.client.get("/gestion/planillas-juego/")
        primer_partido_planillas = listado_planillas.context["grupos_planillas"][0].fechas[0].partidos[0].partido
        self.assertEqual(primer_partido_planillas, proximo)

        administrador = User.objects.create_superuser("admin-orden-partidos", "orden@example.com", "clave")
        self.client.force_login(administrador)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()
        listado_partidos = self.client.get("/gestion/partidos/")
        self.assertEqual(listado_partidos.context["partidos"][0], proximo)
        self.assertEqual(listado_partidos.context["partidos"][1], posterior)

    def test_planillero_no_puede_eliminar_planilla_cargada(self):
        documento = self.crear_documento_planilla()
        self.client.force_login(self.planillero)

        listado = self.client.get("/gestion/planillas-juego/")
        respuesta = self.client.post(f"/gestion/planillas-juego/{documento.id}/eliminar/")

        self.assertNotContains(listado, f"/gestion/planillas-juego/{documento.id}/eliminar/")
        self.assertEqual(respuesta.status_code, 302)
        self.assertTrue(Documento.objects.filter(id=documento.id).exists())

    @patch("torneos.views.eliminar_documento_almacenamiento")
    def test_administrador_puede_eliminar_planilla_cargada(self, eliminar_mock):
        documento = self.crear_documento_planilla()
        administrador = User.objects.create_superuser("admin-planillas", "admin@example.com", "clave")
        self.client.force_login(administrador)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

        listado = self.client.get("/gestion/planillas-juego/")
        respuesta = self.client.post(f"/gestion/planillas-juego/{documento.id}/eliminar/")

        self.assertContains(listado, f"/gestion/planillas-juego/{documento.id}/eliminar/")
        self.assertEqual(respuesta.status_code, 302)
        self.assertFalse(Documento.objects.filter(id=documento.id).exists())
        eliminar_mock.assert_called_once_with("https://example.com/fecha1.jpg")

    def test_lista_planillas_filtra_por_partido(self):
        rival = Equipo.objects.create(nombre="Riverenos", categoria=self.categoria)
        otro_partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo_local,
            equipo_visitante=rival,
            fecha=date(2026, 6, 10),
            hora=time(18, 0),
            estado="FINALIZADO",
            numero_fecha="Fecha 2",
        )
        otro_partido.planilleros.add(self.planillero)
        Documento.objects.create(
            tipo="PLANILLA_JUEGO",
            torneo=self.torneo,
            categoria=self.categoria,
            partido=self.partido,
            equipo_local=self.equipo_local,
            equipo_visitante=self.equipo_visitante,
            titulo="Planilla Fecha 1",
            archivo="https://example.com/fecha1.jpg",
            numero_fecha="Fecha 1",
            fecha_partido=date(2026, 6, 3),
            hora_partido=time(16, 0),
            cargado_por=self.planillero,
        )
        Documento.objects.create(
            tipo="PLANILLA_JUEGO",
            torneo=self.torneo,
            categoria=self.categoria,
            partido=otro_partido,
            equipo_local=self.equipo_local,
            equipo_visitante=rival,
            titulo="Planilla Fecha 2",
            archivo="https://example.com/fecha2.jpg",
            numero_fecha="Fecha 2",
            fecha_partido=date(2026, 6, 10),
            hora_partido=time(18, 0),
            cargado_por=self.planillero,
        )

        self.client.force_login(self.planillero)
        response = self.client.get(f"/gestion/planillas-juego/?partido={self.partido.id}")

        grupos = response.context["grupos_planillas"]
        self.assertEqual(len(grupos), 1)
        self.assertEqual(grupos[0].fechas[0].nombre, "Fecha 1")
        self.assertEqual(len(grupos[0].fechas[0].partidos), 1)
        self.assertEqual(grupos[0].fechas[0].partidos[0].partido, self.partido)
        self.assertEqual(grupos[0].fechas[0].partidos[0].documentos[0].titulo, "Planilla Fecha 1")


class SancionesTarjetasTests(TestCase):
    def setUp(self):
        self.torneo = Torneo.objects.create(
            nombre="Veranero",
            fecha_inicio=date(2026, 1, 1),
        )
        self.categoria = Categoria.objects.create(
            nombre="Senior",
            edad_minima=18,
            edad_maxima=60,
            torneo=self.torneo,
        )
        self.equipo = Equipo.objects.create(nombre="Paraiso", categoria=self.categoria)
        self.rival = Equipo.objects.create(nombre="Integracion", categoria=self.categoria)
        self.jugador = Jugador.objects.create(
            equipo=self.equipo,
            dorsal=10,
            nombres="Jugador Sancionado",
            cedula="100",
            fecha_nacimiento=date(1990, 1, 1),
        )

    def crear_partido(self, dia, fase="GRUPOS", estado="FINALIZADO"):
        return Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=self.rival,
            fecha=date(2026, 5, dia),
            hora=time(15, 0),
            estado=estado,
            fase=fase,
            numero_fecha=str(dia),
        )

    def assert_no_disponible_en(self, partido):
        _sincronizar_no_disponibles_por_tarjetas(partido)
        self.assertTrue(
            AlineacionPartido.objects.filter(
                partido=partido,
                jugador=self.jugador,
                equipo=self.equipo,
                rol="NO_DISPONIBLE",
            ).exists()
        )

    def test_roja_directa_deja_no_disponible_en_siguiente_partido(self):
        partido_origen = self.crear_partido(1)
        Tarjeta.objects.create(
            partido=partido_origen,
            jugador=self.jugador,
            equipo=self.equipo,
            tipo="ROJA",
        )
        siguiente = self.crear_partido(8, estado="PROGRAMADO")

        self.assert_no_disponible_en(siguiente)

    def test_roja_directa_bloquea_dos_partidos_y_se_habilita_en_el_tercero(self):
        partido_origen = self.crear_partido(1)
        Tarjeta.objects.create(
            partido=partido_origen, jugador=self.jugador, equipo=self.equipo,
            tipo="ROJA", origen_roja="DIRECTA",
        )
        primero = self.crear_partido(8, estado="PROGRAMADO")
        self.assert_no_disponible_en(primero)
        primero.estado = "FINALIZADO"
        primero.save(update_fields=["estado"])

        segundo = self.crear_partido(15, estado="PROGRAMADO")
        self.assert_no_disponible_en(segundo)
        segundo.estado = "FINALIZADO"
        segundo.save(update_fields=["estado"])

        tercero = self.crear_partido(22, estado="PROGRAMADO")
        self.assertEqual(_sincronizar_no_disponibles_por_tarjetas(tercero), {})

    def test_doble_amarilla_deja_no_disponible_en_siguiente_partido(self):
        partido_origen = self.crear_partido(1)
        Tarjeta.objects.create(
            partido=partido_origen, jugador=self.jugador, equipo=self.equipo,
            tipo="ROJA", origen_roja="DOBLE_AMARILLA",
        )
        siguiente = self.crear_partido(8, estado="PROGRAMADO")

        self.assert_no_disponible_en(siguiente)

        siguiente.estado = "FINALIZADO"
        siguiente.save(update_fields=["estado"])
        posterior = self.crear_partido(15, estado="PROGRAMADO")
        self.assertEqual(_sincronizar_no_disponibles_por_tarjetas(posterior), {})

    def test_tres_amarillas_en_partidos_distintos_de_fase_uno_sancionan(self):
        for dia in [1, 8, 15]:
            partido = self.crear_partido(dia)
            Tarjeta.objects.create(
                partido=partido,
                jugador=self.jugador,
                equipo=self.equipo,
                tipo="AMARILLA",
            )
        siguiente = self.crear_partido(22, fase="CUARTOS", estado="PROGRAMADO")

        self.assert_no_disponible_en(siguiente)


class ResumenPartidoOrdenTests(TestCase):
    def test_muestra_ultimo_movimiento_arriba(self):
        base = timezone.now()
        eventos = [
            SimpleNamespace(minuto=None, creado_en=base, orden=1),
            SimpleNamespace(minuto=15, creado_en=base, orden=3),
            SimpleNamespace(minuto=15, creado_en=base + timedelta(seconds=5), orden=2),
            SimpleNamespace(minuto=None, creado_en=base + timedelta(seconds=10), orden=4),
        ]

        ordenados = sorted(eventos, key=_clave_orden_evento_resumen, reverse=True)

        self.assertEqual([evento.orden for evento in ordenados], [2, 3, 4, 1])


class CronometroEventoTests(TestCase):
    def test_reanudar_suspendido_conserva_tiempo_periodo_y_marcador(self):
        torneo = Torneo.objects.create(nombre="Reprogramado", fecha_inicio=date(2026, 1, 1))
        categoria = Categoria.objects.create(nombre="Senior", edad_minima=18, edad_maxima=60, torneo=torneo)
        local = Equipo.objects.create(nombre="Local", categoria=categoria)
        visitante = Equipo.objects.create(nombre="Visitante", categoria=categoria)
        partido = Partido.objects.create(
            categoria=categoria,
            equipo_local=local,
            equipo_visitante=visitante,
            fecha=date(2026, 8, 30),
            hora=time(16, 0),
            cancha="Teresa Sierra",
            estado="SUSPENDIDO",
            periodo_en_vivo="ST",
            segundos_acumulados=(68 * 60) + 17,
            cronometro_pausado=True,
            goles_local=2,
            goles_visitante=1,
        )
        admin = User.objects.create_superuser("admin-reanudar", password="test")
        self.client.force_login(admin)

        respuesta = self.client.get(f"/partido/{partido.id}/cronometro/reanudar/")

        self.assertEqual(respuesta.status_code, 302)
        partido.refresh_from_db()
        self.assertEqual(partido.estado, "EN_JUEGO")
        self.assertEqual(partido.periodo_en_vivo, "ST")
        self.assertEqual(partido.segundos_acumulados, (68 * 60) + 17)
        self.assertEqual((partido.goles_local, partido.goles_visitante), (2, 1))
        self.assertFalse(partido.cronometro_pausado)
        self.assertIsNotNone(partido.inicio_en_vivo)

    def test_minuto_evento_coincide_con_minuto_visible_del_cronometro(self):
        partido = SimpleNamespace(
            estado="EN_JUEGO",
            segundos_acumulados=(16 * 60) + 42,
            inicio_en_vivo=None,
            cronometro_pausado=True,
        )

        self.assertEqual(_minuto_evento_en_vivo(partido), 16)

    def test_minuto_evento_no_baja_de_uno_al_inicio(self):
        partido = SimpleNamespace(
            estado="EN_JUEGO",
            segundos_acumulados=30,
            inicio_en_vivo=None,
            cronometro_pausado=True,
        )

        self.assertEqual(_minuto_evento_en_vivo(partido), 1)

    def test_sustitucion_sin_minuto_usa_cronometro(self):
        torneo = Torneo.objects.create(nombre="Veranero", fecha_inicio=date(2026, 1, 1))
        categoria = Categoria.objects.create(nombre="Senior", edad_minima=18, edad_maxima=60, torneo=torneo)
        equipo = Equipo.objects.create(nombre="Local", categoria=categoria)
        rival = Equipo.objects.create(nombre="Visitante", categoria=categoria)
        jugador_sale = Jugador.objects.create(equipo=equipo, nombres="Sale Uno", cedula="1", fecha_nacimiento=date(1990, 1, 1))
        jugador_entra = Jugador.objects.create(equipo=equipo, nombres="Entra Uno", cedula="2", fecha_nacimiento=date(1991, 1, 1))
        partido = Partido.objects.create(
            categoria=categoria,
            equipo_local=equipo,
            equipo_visitante=rival,
            fecha=date(2026, 6, 1),
            hora=time(16, 0),
            estado="EN_JUEGO",
            segundos_acumulados=(39 * 60) + 33,
            cronometro_pausado=True,
        )
        admin = User.objects.create_user("admin-crono", password="test", is_staff=True, is_superuser=True)
        AlineacionPartido.objects.create(
            partido=partido, equipo=equipo, jugador=jugador_sale, rol="TITULAR"
        )

        self.client.force_login(admin)
        respuesta = self.client.post(
            f"/partido/{partido.id}/agregar-sustitucion-movil/",
            {
                "equipo": str(equipo.id),
                "jugador_sale": str(jugador_sale.id),
                "jugador_entra": str(jugador_entra.id),
            },
        )

        self.assertEqual(respuesta.status_code, 302)
        sustitucion = SustitucionPartido.objects.get(partido=partido)
        self.assertEqual(sustitucion.minuto, 39)
        alineacion = AlineacionPartido.objects.get(partido=partido, jugador=jugador_entra)
        self.assertEqual(alineacion.equipo, equipo)
        self.assertEqual(alineacion.rol, "SUPLENTE")


class IncidenciasReglasEdadEnJuegoTests(TestCase):
    def setUp(self):
        self.torneo = Torneo.objects.create(nombre="Control en juego", fecha_inicio=date(2026, 1, 1))
        self.categoria = Categoria.objects.create(nombre="Senior", edad_minima=18, edad_maxima=80, torneo=self.torneo)
        ReglaEdadCategoria.objects.create(
            categoria=self.categoria,
            etiqueta="+50",
            edad_minima=50,
            minimo_titulares=1,
            orden=1,
        )
        self.equipo = Equipo.objects.create(nombre="Local", categoria=self.categoria)
        self.rival = Equipo.objects.create(nombre="Visitante", categoria=self.categoria)
        self.mayor = Jugador.objects.create(equipo=self.equipo, nombres="Mayor Titular", cedula="m1", fecha_nacimiento=date(1965, 1, 1))
        self.mayor_dos = Jugador.objects.create(equipo=self.equipo, nombres="Mayor Suplente", cedula="m2", fecha_nacimiento=date(1966, 1, 1))
        self.joven = Jugador.objects.create(equipo=self.equipo, nombres="Joven Titular", cedula="j1", fecha_nacimiento=date(1995, 1, 1))
        self.joven_dos = Jugador.objects.create(equipo=self.equipo, nombres="Joven Suplente", cedula="j2", fecha_nacimiento=date(1996, 1, 1))
        self.partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=self.rival,
            fecha=date(2026, 7, 1),
            hora=time(15, 0),
            estado="EN_JUEGO",
            periodo_en_vivo="PT",
            cronometro_pausado=True,
            segundos_acumulados=1200,
        )
        AlineacionPartido.objects.create(partido=self.partido, equipo=self.equipo, jugador=self.mayor, rol="TITULAR", posicion_cancha="DC")
        AlineacionPartido.objects.create(partido=self.partido, equipo=self.equipo, jugador=self.joven, rol="TITULAR", posicion_cancha="MC1")
        self.admin = User.objects.create_superuser("control-reglas", password="test")
        self.client.force_login(self.admin)

    def registrar_cambio(self, sale, entra):
        return self.client.post(
            f"/partido/{self.partido.id}/agregar-sustitucion-movil/",
            {"equipo": self.equipo.id, "jugador_sale": sale.id, "jugador_entra": entra.id},
        )

    def test_detecta_y_corrige_incidencia_sin_modificar_once_inicial(self):
        self.registrar_cambio(self.mayor, self.joven_dos)

        incidencia = IncidenciaReglaEdad.objects.get(partido=self.partido, equipo=self.equipo)
        self.assertEqual(incidencia.estado, "ABIERTA")
        self.assertFalse(incidencia.confirmada)
        self.assertTrue(any("+50" in error for error in incidencia.errores))
        alineacion_inicial = AlineacionPartido.objects.get(partido=self.partido, jugador=self.mayor)
        self.assertEqual(alineacion_inicial.rol, "TITULAR")
        self.assertEqual(alineacion_inicial.posicion_cancha, "DC")
        self.assertTrue(RegistroActividad.objects.filter(accion="ALERTA_REGLA_EDAD", usuario=self.admin).exists())

        self.partido.segundos_acumulados = 1230
        self.partido.save(update_fields=["segundos_acumulados"])
        self.registrar_cambio(self.joven, self.mayor_dos)

        incidencia.refresh_from_db()
        self.assertEqual(incidencia.estado, "CORREGIDA")
        self.assertEqual(incidencia.duracion_segundos, 30)
        self.assertFalse(incidencia.confirmada)
        self.assertTrue(RegistroActividad.objects.filter(accion="CORREGIR_REGLA_EDAD", usuario=self.admin).exists())

    def test_entretiempo_evalua_todos_los_cambios_al_iniciar_segundo_tiempo(self):
        self.partido.periodo_en_vivo = "ET"
        self.partido.save(update_fields=["periodo_en_vivo"])

        self.registrar_cambio(self.mayor, self.joven_dos)
        self.assertFalse(IncidenciaReglaEdad.objects.filter(partido=self.partido).exists())

        respuesta = self.client.get(f"/partido/{self.partido.id}/cronometro/segundo-tiempo/")

        self.assertEqual(respuesta.status_code, 302)
        self.partido.refresh_from_db()
        self.assertEqual(self.partido.segundos_acumulados, 45 * 60)
        self.assertEqual(_minuto_evento_en_vivo(self.partido), 45)
        incidencia = IncidenciaReglaEdad.objects.get(partido=self.partido, equipo=self.equipo)
        self.assertEqual(incidencia.periodo_inicio, "ST")

    def test_pulsar_segundo_tiempo_nuevamente_no_reinicia_el_reloj(self):
        inicio_original = timezone.now() - timedelta(minutes=3)
        self.partido.periodo_en_vivo = "ST"
        self.partido.cronometro_pausado = False
        self.partido.segundos_acumulados = 45 * 60
        self.partido.inicio_en_vivo = inicio_original
        self.partido.save(update_fields=[
            "periodo_en_vivo", "cronometro_pausado", "segundos_acumulados", "inicio_en_vivo",
        ])

        respuesta = self.client.get(f"/partido/{self.partido.id}/cronometro/segundo-tiempo/")

        self.assertEqual(respuesta.status_code, 302)
        self.partido.refresh_from_db()
        self.assertEqual(self.partido.inicio_en_vivo, inicio_original)
        self.assertGreaterEqual(_minuto_evento_en_vivo(self.partido), 48)

    def test_no_permite_sustitucion_entre_dos_jugadores_del_banco(self):
        respuesta = self.registrar_cambio(self.joven_dos, self.mayor_dos)

        self.assertEqual(respuesta.status_code, 302)
        self.assertFalse(SustitucionPartido.objects.filter(partido=self.partido).exists())

    def test_no_permite_que_entre_un_jugador_que_ya_esta_en_cancha(self):
        respuesta = self.registrar_cambio(self.joven, self.mayor)

        self.assertEqual(respuesta.status_code, 302)
        self.assertFalse(SustitucionPartido.objects.filter(partido=self.partido).exists())

    def test_senior_master_permite_que_un_jugador_salga_y_vuelva_a_entrar(self):
        self.categoria.nombre = "Senior Master"
        self.categoria.save(update_fields=["nombre"])

        self.registrar_cambio(self.mayor, self.joven_dos)
        self.registrar_cambio(self.joven_dos, self.mayor)

        self.assertEqual(SustitucionPartido.objects.filter(partido=self.partido).count(), 2)
        self.assertIn(self.mayor.id, jugadores_actuales_en_cancha(self.partido, self.equipo))

    def test_interbarrios_no_permite_reingreso_despues_de_salir(self):
        self.categoria.nombre = "Interbarrios"
        self.categoria.save(update_fields=["nombre"])

        self.registrar_cambio(self.mayor, self.joven_dos)
        respuesta = self.registrar_cambio(self.joven_dos, self.mayor)

        self.assertEqual(respuesta.status_code, 302)
        self.assertEqual(SustitucionPartido.objects.filter(partido=self.partido).count(), 1)
        self.assertNotIn(self.mayor.id, jugadores_actuales_en_cancha(self.partido, self.equipo))

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_confirma_infraccion_si_supera_sesenta_segundos_de_juego(self):
        self.registrar_cambio(self.mayor, self.joven_dos)
        self.partido.segundos_acumulados = 1261
        self.partido.save(update_fields=["segundos_acumulados"])

        respuesta = self.client.get(f"/partido/{self.partido.id}/editor-movil/")

        self.assertEqual(respuesta.status_code, 200)
        incidencia = IncidenciaReglaEdad.objects.get(partido=self.partido, equipo=self.equipo)
        self.assertTrue(incidencia.confirmada)
        auditoria = RegistroActividad.objects.get(accion="ALERTA_REGLA_EDAD", usuario=self.admin)
        self.assertIn("Mayor Titular", auditoria.descripcion)
        self.assertIn("Joven Suplente", auditoria.descripcion)
        self.assertIn("+50", auditoria.descripcion)
        confirmacion = RegistroActividad.objects.get(accion="CONFIRMAR_INFRACCION_REGLA_EDAD")
        self.assertIn("60 segundos", confirmacion.descripcion)
        self.assertIn("+50", confirmacion.descripcion)

    def test_sustitucion_convierte_en_suplente_al_jugador_no_disponible(self):
        torneo = Torneo.objects.create(nombre="Veranero", fecha_inicio=date(2026, 1, 1))
        categoria = Categoria.objects.create(nombre="Senior", edad_minima=18, edad_maxima=60, torneo=torneo)
        equipo = Equipo.objects.create(nombre="Local", categoria=categoria)
        rival = Equipo.objects.create(nombre="Visitante", categoria=categoria)
        jugador_sale = Jugador.objects.create(equipo=equipo, nombres="Sale Uno", cedula="11", fecha_nacimiento=date(1990, 1, 1))
        jugador_entra = Jugador.objects.create(equipo=equipo, nombres="Entra Uno", cedula="12", fecha_nacimiento=date(1991, 1, 1))
        partido = Partido.objects.create(
            categoria=categoria,
            equipo_local=equipo,
            equipo_visitante=rival,
            fecha=date(2026, 6, 1),
            hora=time(16, 0),
            estado="EN_JUEGO",
        )
        AlineacionPartido.objects.create(
            partido=partido,
            equipo=equipo,
            jugador=jugador_sale,
            rol="TITULAR",
        )
        AlineacionPartido.objects.create(
            partido=partido,
            equipo=equipo,
            jugador=jugador_entra,
            rol="NO_DISPONIBLE",
        )
        admin = User.objects.create_user("admin-suplente", password="test", is_staff=True, is_superuser=True)

        self.client.force_login(admin)
        respuesta = self.client.post(
            f"/partido/{partido.id}/agregar-sustitucion-movil/",
            {
                "equipo": str(equipo.id),
                "jugador_sale": str(jugador_sale.id),
                "jugador_entra": str(jugador_entra.id),
                "minuto": "10",
            },
        )

        self.assertEqual(respuesta.status_code, 302)
        alineacion = AlineacionPartido.objects.get(partido=partido, jugador=jugador_entra)
        self.assertEqual(alineacion.rol, "SUPLENTE")


class TablaPosicionesWoTests(TestCase):
    def test_partido_wo_suma_en_tabla(self):
        torneo = Torneo.objects.create(
            nombre="Veranero",
            fecha_inicio=date(2026, 1, 1),
        )
        categoria = Categoria.objects.create(
            nombre="Senior",
            edad_minima=18,
            edad_maxima=60,
            torneo=torneo,
        )
        ganador = Equipo.objects.create(nombre="Ganador WO", categoria=categoria)
        perdedor = Equipo.objects.create(nombre="No Presentado", categoria=categoria)
        Partido.objects.create(
            categoria=categoria,
            equipo_local=ganador,
            equipo_visitante=perdedor,
            fecha=date(2026, 5, 1),
            hora=time(15, 0),
            estado="WO",
            fase="GRUPOS",
            grupo="A",
            numero_fecha="1",
            goles_local=3,
            goles_visitante=-3,
        )

        tabla = construir_estructura(torneo)["Senior"]["grupos"]["A"]["tabla"]
        fila_ganador = next(fila for fila in tabla if fila["id"] == ganador.id)
        fila_perdedor = next(fila for fila in tabla if fila["id"] == perdedor.id)

        self.assertEqual(fila_ganador["pj"], 1)
        self.assertEqual(fila_ganador["pg"], 1)
        self.assertEqual(fila_ganador["pts"], 3)
        self.assertEqual(fila_ganador["gf"], 3)
        self.assertEqual(fila_ganador["gc"], -3)
        self.assertEqual(fila_ganador["dg"], 6)
        self.assertEqual(fila_perdedor["pp"], 1)
        self.assertEqual(fila_perdedor["pts"], 0)


class TablaPosicionesDesempateTarjetasTests(TestCase):
    def test_menos_tarjetas_desempata_equipos_con_igual_rendimiento(self):
        torneo = Torneo.objects.create(nombre="Fair Play", fecha_inicio=date(2026, 1, 1))
        categoria = Categoria.objects.create(
            nombre="Senior Disciplina",
            edad_minima=18,
            edad_maxima=60,
            torneo=torneo,
        )
        equipo_limpio = Equipo.objects.create(nombre="Equipo Limpio", categoria=categoria)
        equipo_sancionado = Equipo.objects.create(nombre="Equipo Sancionado", categoria=categoria)
        rival_limpio = Equipo.objects.create(nombre="Rival Limpio", categoria=categoria)
        rival_sancionado = Equipo.objects.create(nombre="Rival Sancionado", categoria=categoria)
        jugador = Jugador.objects.create(
            equipo=equipo_sancionado,
            nombres="Jugador Sancionado",
            cedula="DISC-1",
            fecha_nacimiento=date(1990, 1, 1),
        )
        datos_partido = {
            "categoria": categoria,
            "fecha": date(2026, 7, 26),
            "hora": time(15, 0),
            "estado": "FINALIZADO",
            "estadisticas_validadas": True,
            "fase": "GRUPOS",
            "grupo": "A",
            "goles_local": 1,
            "goles_visitante": 0,
        }
        Partido.objects.create(
            equipo_local=equipo_limpio,
            equipo_visitante=rival_limpio,
            **datos_partido,
        )
        partido_sancionado = Partido.objects.create(
            equipo_local=equipo_sancionado,
            equipo_visitante=rival_sancionado,
            **datos_partido,
        )
        Tarjeta.objects.create(
            partido=partido_sancionado,
            jugador=jugador,
            equipo=equipo_sancionado,
            tipo="AMARILLA",
        )

        tabla = construir_estructura(torneo)[categoria.nombre]["grupos"]["A"]["tabla"]
        posicion_limpio = next(i for i, fila in enumerate(tabla) if fila["id"] == equipo_limpio.id)
        posicion_sancionado = next(i for i, fila in enumerate(tabla) if fila["id"] == equipo_sancionado.id)

        self.assertLess(posicion_limpio, posicion_sancionado)
        self.assertEqual(tabla[posicion_limpio]["puntos_disciplina"], 0)
        self.assertEqual(tabla[posicion_sancionado]["puntos_disciplina"], 1)
        disciplina = construir_estructura(torneo)[categoria.nombre]["disciplina_equipos"]
        fila_disciplina = next(fila for fila in disciplina if fila["id"] == equipo_sancionado.id)
        self.assertEqual(fila_disciplina["ta"], 1)
        self.assertEqual(fila_disciplina["tr"], 0)
        self.assertEqual(fila_disciplina["puntos_disciplina"], 1)


class ReglasEdadCategoriaTests(TestCase):
    def setUp(self):
        self.torneo = Torneo.objects.create(
            nombre="Veranero",
            fecha_inicio=date(2026, 1, 1),
        )
        self.categoria = Categoria.objects.create(
            nombre="Senior Master",
            edad_minima=40,
            edad_maxima=80,
            torneo=self.torneo,
        )
        self.equipo = Equipo.objects.create(nombre="Paraiso", categoria=self.categoria)
        self.rival = Equipo.objects.create(nombre="Integracion", categoria=self.categoria)
        self.partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=self.rival,
            fecha=date(2026, 5, 1),
            hora=time(15, 0),
            estado="PROGRAMADO",
        )
        ReglaEdadCategoria.objects.create(
            categoria=self.categoria,
            etiqueta="+40",
            edad_minima=40,
            edad_maxima=44,
            maximo_titulares=4,
            orden=1,
        )
        ReglaEdadCategoria.objects.create(
            categoria=self.categoria,
            etiqueta="+45",
            edad_minima=45,
            edad_maxima=49,
            minimo_titulares=4,
            orden=2,
        )
        ReglaEdadCategoria.objects.create(
            categoria=self.categoria,
            etiqueta="+50",
            edad_minima=50,
            minimo_titulares=3,
            orden=3,
        )

    def crear_jugador(self, indice, nacimiento):
        return Jugador.objects.create(
            equipo=self.equipo,
            dorsal=indice,
            nombres=f"Jugador {indice}",
            cedula=f"EDAD{indice}",
            fecha_nacimiento=nacimiento,
        )

    def test_etiqueta_edad_se_calcula_con_fecha_del_partido(self):
        jugador = self.crear_jugador(1, date(1981, 5, 2))

        self.assertEqual(etiqueta_edad_jugador(jugador, self.categoria, self.partido.fecha), "+40")

    def test_edad_planilla_se_calcula_con_fecha_programada(self):
        fecha_programada = date(2026, 6, 7)
        nacimiento = date(1985, 6, 6)

        self.assertEqual(_edad(nacimiento, fecha_programada), "41")

    def test_valida_reglas_senior_master_con_reemplazos(self):
        jugadores = []
        for indice in range(1, 4):
            jugadores.append(self.crear_jugador(indice, date(1983, 1, 1)))
        for indice in range(4, 8):
            jugadores.append(self.crear_jugador(indice, date(1978, 1, 1)))
        for indice in range(8, 12):
            jugadores.append(self.crear_jugador(indice, date(1970, 1, 1)))

        errores = validar_reglas_edad_titulares(
            self.partido,
            self.equipo,
            [str(jugador.id) for jugador in jugadores],
        )

        self.assertEqual(errores, [])

    def test_reporta_maximo_de_cuarenta_en_cancha(self):
        jugadores = []
        for indice in range(1, 6):
            jugadores.append(self.crear_jugador(indice, date(1983, 1, 1)))
        for indice in range(6, 9):
            jugadores.append(self.crear_jugador(indice, date(1978, 1, 1)))
        for indice in range(9, 12):
            jugadores.append(self.crear_jugador(indice, date(1970, 1, 1)))

        errores = validar_reglas_edad_titulares(
            self.partido,
            self.equipo,
            [str(jugador.id) for jugador in jugadores],
        )

        self.assertTrue(any("+40" in error and "maximo 4" in error for error in errores))

    def test_cincuenta_reemplaza_cupo_de_cuarenta_y_cinco(self):
        jugadores = []
        for indice in range(1, 5):
            jugadores.append(self.crear_jugador(indice, date(1983, 1, 1)))
        for indice in range(5, 8):
            jugadores.append(self.crear_jugador(indice, date(1978, 1, 1)))
        for indice in range(8, 12):
            jugadores.append(self.crear_jugador(indice, date(1970, 1, 1)))

        errores = validar_reglas_edad_titulares(
            self.partido,
            self.equipo,
            [str(jugador.id) for jugador in jugadores],
        )

        self.assertEqual(errores, [])

    def test_reporta_regla_incompleta_con_once_titulares(self):
        jugadores = []
        for indice in range(1, 5):
            jugadores.append(self.crear_jugador(indice, date(1983, 1, 1)))
        for indice in range(5, 7):
            jugadores.append(self.crear_jugador(indice, date(1978, 1, 1)))
        for indice in range(7, 10):
            jugadores.append(self.crear_jugador(indice, date(1970, 1, 1)))
        for indice in range(10, 12):
            jugadores.append(self.crear_jugador(indice, date(1988, 1, 1)))

        errores = validar_reglas_edad_titulares(
            self.partido,
            self.equipo,
            [str(jugador.id) for jugador in jugadores],
        )

        self.assertTrue(any("+45" in error for error in errores))

    def test_reporta_regla_incompleta_con_menos_de_once_titulares(self):
        jugadores = []
        for indice in range(1, 5):
            jugadores.append(self.crear_jugador(indice, date(1983, 1, 1)))
        for indice in range(5, 8):
            jugadores.append(self.crear_jugador(indice, date(1978, 1, 1)))

        errores = validar_reglas_edad_titulares(
            self.partido,
            self.equipo,
            [str(jugador.id) for jugador in jugadores],
        )

        self.assertTrue(any("+50" in error and "minimo 3" in error for error in errores))

    def test_frontend_no_impone_maximo_por_defecto_de_cuarenta_y_cinco(self):
        reglas = reglas_edad_para_frontend(self.categoria)
        regla_45 = next(regla for regla in reglas if regla["etiqueta"] == "+45")

        self.assertIsNone(regla_45["maximo"])

    def test_cuarenta_y_cinco_puede_superar_ocho_si_no_tiene_maximo_configurado(self):
        jugadores = []
        for indice in range(1, 10):
            jugadores.append(self.crear_jugador(indice, date(1978, 1, 1)))
        for indice in range(10, 12):
            jugadores.append(self.crear_jugador(indice, date(1970, 1, 1)))

        errores = validar_reglas_edad_titulares(
            self.partido,
            self.equipo,
            [str(jugador.id) for jugador in jugadores],
        )

        self.assertFalse(any("+45" in error and "maximo" in error for error in errores))


class GestionCategoriaReglasTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_user("admin-categorias", password="test", is_staff=True, is_superuser=True)
        self.torneo = Torneo.objects.create(nombre="Veranero", fecha_inicio=date(2026, 1, 1))
        self.client.force_login(self.admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

    def datos_categoria(self):
        return {
            "nombre": "Senior Master",
            "descripcion": "Reglas senior",
            "edad_minima": "40",
            "edad_maxima": "80",
            "porcentaje_minimo_foraneos": "50",
            "reglas-TOTAL_FORMS": "3",
            "reglas-INITIAL_FORMS": "0",
            "reglas-MIN_NUM_FORMS": "0",
            "reglas-MAX_NUM_FORMS": "1000",
            "reglas-0-etiqueta": "+40",
            "reglas-0-edad_minima": "40",
            "reglas-0-edad_maxima": "44",
            "reglas-0-minimo_titulares": "0",
            "reglas-0-maximo_titulares": "4",
            "reglas-0-orden": "1",
            "reglas-0-activa": "on",
            "reglas-1-etiqueta": "+45",
            "reglas-1-edad_minima": "45",
            "reglas-1-edad_maxima": "49",
            "reglas-1-minimo_titulares": "5",
            "reglas-1-maximo_titulares": "",
            "reglas-1-orden": "2",
            "reglas-1-activa": "on",
            "reglas-2-etiqueta": "+50",
            "reglas-2-edad_minima": "50",
            "reglas-2-edad_maxima": "",
            "reglas-2-minimo_titulares": "2",
            "reglas-2-maximo_titulares": "",
            "reglas-2-orden": "3",
            "reglas-2-activa": "on",
        }

    def test_admin_crea_categoria_con_reglas_de_edad(self):
        respuesta = self.client.post("/gestion/categorias/nueva/", self.datos_categoria())

        self.assertEqual(respuesta.status_code, 302)
        categoria = Categoria.objects.get(nombre="Senior Master")
        reglas = {regla.etiqueta: regla for regla in categoria.reglas_edad.all()}
        self.assertEqual(reglas["+40"].maximo_titulares, 4)
        self.assertEqual(reglas["+45"].minimo_titulares, 5)
        self.assertIsNone(reglas["+45"].maximo_titulares)
        self.assertEqual(reglas["+50"].minimo_titulares, 2)


class PlanillasPDFTests(TestCase):
    def test_titulo_planilla_usa_nombre_y_descripcion_del_torneo(self):
        torneo = Torneo.objects.create(
            nombre="Torneo Amistoso Mata Mata.",
            descripcion="Senior Master",
            fecha_inicio=date(2026, 1, 1),
        )
        categoria = Categoria.objects.create(
            nombre="Senior",
            edad_minima=40,
            edad_maxima=80,
            torneo=torneo,
        )
        partido = Partido(
            categoria=categoria,
            equipo_local=Equipo(nombre="Local", categoria=categoria),
            equipo_visitante=Equipo(nombre="Visitante", categoria=categoria),
        )

        self.assertEqual(
            _titulo_planilla(partido),
            "PLANILLA DE JUEGO TORNEO AMISTOSO MATA MATA. SENIOR MASTER",
        )

    def test_planilla_usa_logos_del_torneo_en_el_encabezado(self):
        torneo = Torneo.objects.create(
            nombre="Veranero",
            fecha_inicio=date(2026, 1, 1),
            logo_izquierdo="torneos/veranero/logo_izquierdo.png",
            imagen_central="torneos/veranero/imagen_central.png",
            logo_derecho="torneos/veranero/logo_derecho.png",
        )
        categoria = Categoria.objects.create(
            nombre="Senior",
            edad_minima=40,
            edad_maxima=80,
            torneo=torneo,
        )
        partido = Partido(
            categoria=categoria,
            equipo_local=Equipo(nombre="Local", categoria=categoria),
            equipo_visitante=Equipo(nombre="Visitante", categoria=categoria),
        )

        sources = [source for *_, source in _header_image_sources(partido)]

        self.assertEqual(sources[0].name, "torneos/veranero/logo_izquierdo.png")
        self.assertEqual(sources[1].name, "torneos/veranero/imagen_central.png")
        self.assertEqual(sources[2].name, "torneos/veranero/logo_derecho.png")

    def test_planilla_oculta_dorsal_cero(self):
        self.assertEqual(_dorsal(0), "")
        self.assertEqual(_dorsal("0"), "")
        self.assertEqual(_dorsal(None), "")
        self.assertEqual(_dorsal(17), "17")

    def test_nombre_foraneo_no_duplica_marca_existente(self):
        jugador = SimpleNamespace(nombres="JUAN PEREZ (F) (F)", es_foraneo=True)

        self.assertEqual(_nombre_jugador_planilla(jugador), "Juan Perez (F)")

    def test_nombre_foraneo_agrega_una_sola_marca(self):
        jugador = SimpleNamespace(nombres="JUAN PEREZ", es_foraneo=True)

        self.assertEqual(_nombre_jugador_planilla(jugador), "Juan Perez (F)")

    def test_marca_de_agua_no_falla_sin_escudo(self):
        base = Image.new("RGB", (100, 100), "white")
        equipo = Equipo(nombre="Sin escudo")

        _draw_team_watermark(base, equipo, [0, 0, 100, 100])

        self.assertEqual(base.size, (100, 100))

    def test_marca_de_agua_amplia_escudo_pequeno(self):
        base = Image.new("RGB", (800, 800), "white")
        equipo = Equipo(nombre="Escudo pequeno", escudo="equipos/escudo.png")
        escudo = Image.new("RGBA", (40, 40), (255, 0, 0, 255))

        with patch("torneos.planillas_pdf._image_from_source", return_value=escudo):
            _draw_team_watermark(base, equipo, [0, 0, 800, 800], opacity=120)

        fondo = Image.new("RGB", base.size, "white")
        diferencia = Image.new("RGB", base.size)
        diferencia_pixels = diferencia.load()
        base_pixels = base.load()
        fondo_pixels = fondo.load()
        for y in range(base.height):
            for x in range(base.width):
                diferencia_pixels[x, y] = tuple(
                    abs(base_pixels[x, y][canal] - fondo_pixels[x, y][canal])
                    for canal in range(3)
                )
        caja = diferencia.getbbox()
        self.assertIsNotNone(caja)
        self.assertGreater(caja[2] - caja[0], 700)
        self.assertGreater(caja[3] - caja[1], 700)

    def test_planilla_usa_escudo_de_cada_equipo(self):
        torneo = Torneo.objects.create(nombre="Veranero", fecha_inicio=date(2026, 1, 1))
        categoria = Categoria.objects.create(
            nombre="Senior",
            edad_minima=40,
            edad_maxima=80,
            torneo=torneo,
        )
        equipo = Equipo(nombre="Paraiso", categoria=categoria, escudo="equipos/senior/paraiso.png")

        source = _team_shield_source(equipo)

        self.assertEqual(source.name, "equipos/senior/paraiso.png")


class ForaneosTests(TestCase):
    def setUp(self):
        parche_escudo = patch("torneos.views.escudo_url", return_value="")
        parche_escudo.start()
        self.addCleanup(parche_escudo.stop)
        self.torneo = Torneo.objects.create(
            nombre="IMCRED",
            fecha_inicio=date(2026, 1, 1),
        )
        self.categoria = Categoria.objects.create(
            nombre="Senior Master",
            edad_minima=40,
            edad_maxima=80,
            torneo=self.torneo,
            controlar_foraneos=True,
            porcentaje_minimo_foraneos=50,
        )
        self.equipo = Equipo.objects.create(nombre="Paraiso", categoria=self.categoria)
        self.rival = Equipo.objects.create(nombre="Integracion", categoria=self.categoria)
        self.foraneo = Jugador.objects.create(
            equipo=self.equipo,
            dorsal=9,
            nombres="Jugador Foraneo",
            cedula="F1",
            fecha_nacimiento=date(1970, 1, 1),
            es_foraneo=True,
        )
        self.titular = Jugador.objects.create(
            equipo=self.equipo,
            dorsal=10,
            nombres="Jugador Titular",
            cedula="T1",
            fecha_nacimiento=date(1970, 1, 1),
        )

    def crear_partido(self, dia):
        return Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=self.rival,
            fecha=date(2026, 5, dia),
            hora=time(15, 0),
            estado="FINALIZADO",
            fase="GRUPOS",
            numero_fecha=str(dia),
        )

    def test_minimo_foraneo_redondea_hacia_abajo(self):
        for dia in [1, 8, 15, 22, 29, 30, 31]:
            self.crear_partido(dia)

        fila = construir_estadisticas_foraneos(self.categoria)[0]

        self.assertEqual(fila["partidos_fase1"], 7)
        self.assertEqual(fila["minimo"], 3)

    def test_minimo_usa_los_partidos_del_fixture_de_primera_fase(self):
        partido_uno = self.crear_partido(1)
        partido_dos = self.crear_partido(8)
        Partido.objects.filter(id__in=[partido_uno.id, partido_dos.id]).update(
            estado="PROGRAMADO",
            estadisticas_validadas=False,
        )

        fila = construir_estadisticas_foraneos(self.categoria)[0]

        self.assertEqual(fila["partidos_fase1"], 2)
        self.assertEqual(fila["minimo"], 1)
        self.assertFalse(fila["cumple"])
        self.assertEqual(fila["estado"], "Pendiente")

    def test_cuenta_titular_y_suplente_que_entra(self):
        partido_uno = self.crear_partido(1)
        partido_dos = self.crear_partido(8)
        AlineacionPartido.objects.create(
            partido=partido_uno,
            equipo=self.equipo,
            jugador=self.foraneo,
            rol="TITULAR",
        )
        SustitucionPartido.objects.create(
            partido=partido_dos,
            equipo=self.equipo,
            jugador_sale=self.titular,
            jugador_entra=self.foraneo,
        )

        fila = construir_estadisticas_foraneos(self.categoria)[0]

        self.assertEqual(fila["jugados"], 2)
        self.assertTrue(fila["cumple"])

    def test_foraneo_sin_minimo_queda_bloqueado_en_fase_final(self):
        self.crear_partido(1)
        self.crear_partido(8)
        cuartos = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=self.rival,
            fecha=date(2026, 6, 1),
            hora=time(15, 0),
            estado="PROGRAMADO",
            fase="CUARTOS",
            numero_fecha="CUARTOS #1",
        )

        bloqueados = foraneos_no_habilitados_fase_final(cuartos, self.equipo)

        self.assertEqual(bloqueados[self.foraneo.id], {"jugados": 0, "minimo": 1})

    def test_foraneo_que_cumple_minimo_puede_jugar_fase_final(self):
        partido_uno = self.crear_partido(1)
        self.crear_partido(8)
        AlineacionPartido.objects.create(
            partido=partido_uno,
            equipo=self.equipo,
            jugador=self.foraneo,
            rol="TITULAR",
        )
        cuartos = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=self.rival,
            fecha=date(2026, 6, 1),
            hora=time(15, 0),
            estado="PROGRAMADO",
            fase="CUARTOS",
            numero_fecha="CUARTOS #1",
        )

        bloqueados = foraneos_no_habilitados_fase_final(cuartos, self.equipo)

        self.assertNotIn(self.foraneo.id, bloqueados)

    def test_planillero_no_puede_forzar_foraneo_bloqueado_como_suplente(self):
        self.crear_partido(1)
        self.crear_partido(8)
        cuartos = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=self.rival,
            fecha=date(2026, 6, 1),
            hora=time(15, 0),
            estado="PROGRAMADO",
            fase="CUARTOS",
            numero_fecha="CUARTOS #1",
        )
        planillero = User.objects.create_user("planillero-foraneos", password="test")
        cuartos.planilleros.add(planillero)
        self.client.force_login(planillero)

        respuesta = self.client.post(
            f"/partido/{cuartos.id}/guardar-alineacion-movil/",
            {
                "equipo": str(self.equipo.id),
                f"rol_{self.foraneo.id}": "SUPLENTE",
            },
        )

        self.assertEqual(respuesta.status_code, 302)
        alineacion = AlineacionPartido.objects.get(partido=cuartos, jugador=self.foraneo)
        self.assertEqual(alineacion.rol, "NO_DISPONIBLE")

    def test_planilla_conserva_foraneo_para_marcarlo_no_disponible(self):
        jugadores = _jugadores(self.equipo)

        self.assertIn(self.foraneo, jugadores)
        self.assertIn(self.titular, jugadores)


class JugadorFormTests(TestCase):
    def test_fecha_nacimiento_se_renderiza_en_formato_html_date(self):
        torneo = Torneo.objects.create(
            nombre="Veranero",
            fecha_inicio=date(2026, 1, 1),
        )
        categoria = Categoria.objects.create(
            nombre="Senior",
            edad_minima=18,
            edad_maxima=60,
            torneo=torneo,
        )
        equipo = Equipo.objects.create(nombre="Paraiso", categoria=categoria)
        jugador = Jugador.objects.create(
            equipo=equipo,
            dorsal=31,
            nombres="Ilario Rodriguez",
            cedula="15679719",
            fecha_nacimiento=date(1980, 5, 7),
        )

        html = str(JugadorForm(instance=jugador))

        self.assertIn('value="1980-05-07"', html)


class GestionJugadoresConservaFiltrosTests(TestCase):
    def setUp(self):
        self.torneo = Torneo.objects.create(nombre="Liga filtros", fecha_inicio=date(2026, 1, 1))
        self.categoria = Categoria.objects.create(
            nombre="Senior", torneo=self.torneo, edad_minima=18, edad_maxima=80,
        )
        self.equipo = Equipo.objects.create(nombre="Equipo Uno", categoria=self.categoria)
        self.jugador = Jugador.objects.create(
            equipo=self.equipo,
            nombres="Jugador Filtro",
            cedula="998877",
            fecha_nacimiento=date(1990, 1, 1),
        )
        self.admin = User.objects.create_superuser("admin-filtros-jugadores", password="test")
        self.client.force_login(self.admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()
        self.filtros = f"/gestion/jugadores/?categoria={self.categoria.id}&equipo={self.equipo.id}&q=Jugador"

    def test_editar_jugador_recibe_retorno_al_listado_filtrado(self):
        respuesta = self.client.get(
            f"/gestion/jugadores/{self.jugador.id}/editar/",
            {"volver": self.filtros},
        )

        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(respuesta.context["volver_href"], self.filtros)

    def test_eliminar_jugador_regresa_a_los_mismos_filtros(self):
        respuesta = self.client.post(
            f"/gestion/jugadores/{self.jugador.id}/eliminar/",
            {"volver": self.filtros},
        )

        self.assertRedirects(respuesta, self.filtros, fetch_redirect_response=False)


class PlanilleroPartidoTests(TestCase):
    def setUp(self):
        self.torneo = Torneo.objects.create(
            nombre="Veranero",
            fecha_inicio=date(2026, 1, 1),
        )
        self.categoria = Categoria.objects.create(
            nombre="Senior",
            edad_minima=18,
            edad_maxima=60,
            torneo=self.torneo,
        )
        self.local = Equipo.objects.create(nombre="Local", categoria=self.categoria)
        self.visitante = Equipo.objects.create(nombre="Visitante", categoria=self.categoria)
        self.jugador = Jugador.objects.create(
            equipo=self.local,
            dorsal=9,
            nombres="Planillero Gol",
            cedula="PG1",
            fecha_nacimiento=date(1990, 1, 1),
        )
        self.partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.local,
            equipo_visitante=self.visitante,
            fecha=date(2026, 5, 1),
            hora=time(15, 0),
            estado="PROGRAMADO",
        )
        self.planillero = User.objects.create_user("planillero", password="test")
        self.otro_usuario = User.objects.create_user("otro", password="test")
        self.admin = User.objects.create_user("admin", password="test", is_staff=True)
        AdminTorneo.objects.create(usuario=self.admin, torneo=self.torneo)
        self.partido.planilleros.add(self.planillero)

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_planillero_asignado_puede_abrir_editor(self):
        self.client.force_login(self.planillero)

        respuesta = self.client.get(f"/partido/{self.partido.id}/editor-movil/")

        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, "Resultado del partido")
        self.assertNotContains(respuesta, 'name="cancha"')

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_editor_planillero_vuelve_a_mis_partidos(self):
        self.client.force_login(self.planillero)

        respuesta = self.client.get(f"/partido/{self.partido.id}/editor-movil/")

        self.assertContains(respuesta, 'href="/planillero/partidos/"')
        self.assertContains(respuesta, "Mis partidos")

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_editor_movil_muestra_cronometro_desplegable(self):
        self.client.force_login(self.planillero)

        respuesta = self.client.get(f"/partido/{self.partido.id}/editor-movil/")

        self.assertContains(respuesta, '<details class="card cronometro-card">')
        self.assertContains(respuesta, "<summary>Cronometro en vivo</summary>")

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_tanda_penales_solo_se_habilita_en_eliminatoria_empatada(self):
        self.client.force_login(self.planillero)
        respuesta = self.client.post(f"/partido/{self.partido.id}/cronometro/penales/iniciar/")
        self.partido.refresh_from_db()
        self.assertEqual(respuesta.status_code, 302)
        self.assertNotEqual(self.partido.periodo_en_vivo, "PEN")

        self.partido.fase = "CUARTOS"
        self.partido.goles_local = self.partido.goles_visitante = 1
        self.partido.estado = "EN_JUEGO"
        self.partido.inicio_en_vivo = timezone.now()
        self.partido.cronometro_pausado = False
        self.partido.save()
        editor_antes = self.client.get(f"/partido/{self.partido.id}/editor-movil/")
        self.assertContains(editor_antes, 'id="seccion-penales" hidden')
        self.assertContains(editor_antes, "seccion.hidden = false")
        self.client.post(f"/partido/{self.partido.id}/cronometro/penales/preparar/")
        self.partido.refresh_from_db()
        self.assertEqual(self.partido.estado, "EN_JUEGO")
        self.assertEqual(self.partido.periodo_en_vivo, "PEN")
        self.assertTrue(self.partido.cronometro_pausado)
        self.assertIsNone(self.partido.inicio_en_vivo)
        self.assertIsNone(self.partido.equipo_inicia_penales_id)
        editor_preparado = self.client.get(f"/partido/{self.partido.id}/editor-movil/")
        self.assertNotContains(editor_preparado, 'id="seccion-penales" hidden')
        self.assertContains(editor_preparado, "Equipo que cobra primero")
        self.client.post(f"/partido/{self.partido.id}/cronometro/penales/iniciar/")
        self.partido.refresh_from_db()
        self.assertEqual(self.partido.periodo_en_vivo, "PEN")
        editor = self.client.get(f"/partido/{self.partido.id}/editor-movil/")
        self.assertContains(editor, "Siguiente cobro: Local")
        self.assertContains(editor, "no suman en la tabla de goleadores")
        self.assertNotContains(editor, 'id="seccion-penales" hidden')

    def test_sorteo_permite_que_visitante_cobre_primero(self):
        visitante = Jugador.objects.create(
            equipo=self.visitante, dorsal=10, nombres="Primer Cobrador Visitante", cedula="SORTEO10",
            fecha_nacimiento=date(1991, 1, 1),
        )
        AlineacionPartido.objects.create(
            partido=self.partido, equipo=self.visitante, jugador=visitante, rol="TITULAR",
        )
        self.partido.fase = "CUARTOS"
        self.partido.goles_local = self.partido.goles_visitante = 1
        self.partido.save()
        self.client.force_login(self.planillero)

        self.client.post(
            f"/partido/{self.partido.id}/cronometro/penales/iniciar/",
            {"equipo_inicia_penales": self.visitante.id},
        )
        self.client.post(
            f"/partido/{self.partido.id}/cronometro/penales/cobro/",
            {"jugador": visitante.id, "resultado": "GOL"},
        )

        self.partido.refresh_from_db()
        primer_cobro = CobroPenal.objects.get(partido=self.partido, orden=1)
        self.assertEqual(self.partido.equipo_inicia_penales_id, self.visitante.id)
        self.assertEqual(primer_cobro.equipo_id, self.visitante.id)

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_puede_cambiar_equipo_inicial_despues_de_activar_tanda(self):
        self.partido.fase = "CUARTOS"
        self.partido.estado = "EN_JUEGO"
        self.partido.periodo_en_vivo = "PEN"
        self.partido.goles_local = self.partido.goles_visitante = 1
        self.partido.equipo_inicia_penales = self.local
        self.partido.save()
        self.client.force_login(self.planillero)

        respuesta = self.client.post(
            f"/partido/{self.partido.id}/cronometro/penales/cambiar-equipo-inicial/",
            {"equipo_inicia_penales": self.visitante.id},
        )

        self.assertEqual(respuesta.status_code, 302)
        self.partido.refresh_from_db()
        self.assertEqual(self.partido.equipo_inicia_penales_id, self.visitante.id)
        self.assertEqual(_equipo_turno_tanda(self.partido, 0), self.visitante)
        editor = self.client.get(f"/partido/{self.partido.id}/editor-movil/")
        self.assertContains(editor, "Cambiar equipo que cobra primero")
        self.assertContains(
            editor,
            f'value="{self.visitante.id}" selected>{self.visitante.nombre}</option>',
        )

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_no_cambia_equipo_inicial_si_hay_cobros_registrados(self):
        visitante = Jugador.objects.create(
            equipo=self.visitante,
            dorsal=10,
            nombres="Cobrador Visitante",
            cedula="CAMBIO-INICIAL-10",
            fecha_nacimiento=date(1991, 1, 1),
        )
        self.partido.fase = "CUARTOS"
        self.partido.estado = "EN_JUEGO"
        self.partido.periodo_en_vivo = "PEN"
        self.partido.goles_local = self.partido.goles_visitante = 1
        self.partido.equipo_inicia_penales = self.local
        self.partido.save()
        CobroPenal.objects.create(
            partido=self.partido,
            equipo=self.local,
            jugador=self.jugador,
            orden=1,
            convertido=True,
        )
        self.client.force_login(self.planillero)

        respuesta = self.client.post(
            f"/partido/{self.partido.id}/cronometro/penales/cambiar-equipo-inicial/",
            {"equipo_inicia_penales": self.visitante.id},
            follow=True,
        )

        self.partido.refresh_from_db()
        self.assertEqual(self.partido.equipo_inicia_penales_id, self.local.id)
        self.assertContains(respuesta, "deshaz primero los cobros registrados")

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_puede_corregir_cobrador_despues_de_definir_tanda(self):
        locales = [self.jugador]
        for indice in range(2, 5):
            locales.append(Jugador.objects.create(
                equipo=self.local, dorsal=indice, nombres=f"Cobrador Local {indice}", cedula=f"CL{indice}",
                fecha_nacimiento=date(1990, indice, 1),
            ))
        visitante = Jugador.objects.create(
            equipo=self.visitante, dorsal=10, nombres="Cobrador Visitante", cedula="CVEDIT",
            fecha_nacimiento=date(1991, 1, 1),
        )
        for jugador in locales:
            AlineacionPartido.objects.create(partido=self.partido, equipo=self.local, jugador=jugador, rol="TITULAR")
        AlineacionPartido.objects.create(partido=self.partido, equipo=self.visitante, jugador=visitante, rol="TITULAR")
        self.partido.fase = "FINAL"
        self.partido.estado = "EN_JUEGO"
        self.partido.periodo_en_vivo = "PEN"
        self.partido.save()
        cobros = []
        for orden, jugador, equipo, convertido in (
            (1, locales[0], self.local, True), (2, visitante, self.visitante, False),
            (3, locales[1], self.local, True), (4, visitante, self.visitante, False),
            (5, locales[2], self.local, True), (6, visitante, self.visitante, False),
        ):
            cobros.append(CobroPenal.objects.create(
                partido=self.partido, equipo=equipo, jugador=jugador, orden=orden, convertido=convertido,
            ))
        self.client.force_login(self.planillero)

        self.client.post(
            f"/partido/cronometro/penales/cobro/{cobros[0].id}/modificar/",
            {"jugador": locales[3].id},
        )

        cobros[0].refresh_from_db()
        self.assertEqual(cobros[0].jugador_id, locales[3].id)
        self.assertTrue(cobros[0].convertido)
        editor = self.client.get(f"/partido/{self.partido.id}/editor-movil/")
        self.assertContains(editor, "Modificar cobrador", count=6)

    def test_cobros_actualizan_penales_pero_no_goleadores(self):
        visitante = Jugador.objects.create(
            equipo=self.visitante, dorsal=10, nombres="Cobrador Visitante", cedula="PV10",
            fecha_nacimiento=date(1991, 1, 1),
        )
        AlineacionPartido.objects.create(partido=self.partido, equipo=self.local, jugador=self.jugador, rol="TITULAR")
        AlineacionPartido.objects.create(partido=self.partido, equipo=self.visitante, jugador=visitante, rol="TITULAR")
        suplente = Jugador.objects.create(
            equipo=self.local, dorsal=12, nombres="Suplente No Elegible", cedula="SUP12",
            fecha_nacimiento=date(1992, 1, 1),
        )
        AlineacionPartido.objects.create(partido=self.partido, equipo=self.local, jugador=suplente, rol="SUPLENTE")
        self.partido.fase = "FINAL"
        self.partido.estado = "EN_JUEGO"
        self.partido.periodo_en_vivo = "PEN"
        self.partido.save()
        self.client.force_login(self.planillero)

        self.client.post(
            f"/partido/{self.partido.id}/cronometro/penales/cobro/",
            {"jugador": suplente.id, "resultado": "GOL"},
        )
        self.assertEqual(CobroPenal.objects.filter(partido=self.partido).count(), 0)

        for indice in range(6):
            jugador = self.jugador if indice % 2 == 0 else visitante
            resultado = "GOL" if indice % 2 == 0 else "FALLO"
            self.client.post(
                f"/partido/{self.partido.id}/cronometro/penales/cobro/",
                {"jugador": jugador.id, "resultado": resultado},
            )

        self.partido.refresh_from_db()
        self.assertEqual((self.partido.goles_local_penales, self.partido.goles_visitante_penales), (3, 0))
        self.assertEqual(CobroPenal.objects.filter(partido=self.partido).count(), 6)
        self.assertEqual(Gol.objects.filter(partido=self.partido).count(), 0)

        self.client.get(f"/partido/{self.partido.id}/cronometro/finalizar/")
        self.partido.refresh_from_db()
        self.assertEqual(self.partido.estado, "FINALIZADO")

    def test_no_repite_cobrador_hasta_agotar_jugadores_en_cancha(self):
        local_dos = Jugador.objects.create(
            equipo=self.local, dorsal=8, nombres="Segundo Local", cedula="L8",
            fecha_nacimiento=date(1990, 2, 1),
        )
        visitante_uno = Jugador.objects.create(
            equipo=self.visitante, dorsal=10, nombres="Primer Visitante", cedula="V10",
            fecha_nacimiento=date(1990, 3, 1),
        )
        visitante_dos = Jugador.objects.create(
            equipo=self.visitante, dorsal=11, nombres="Segundo Visitante", cedula="V11",
            fecha_nacimiento=date(1990, 4, 1),
        )
        for equipo, jugador in (
            (self.local, self.jugador), (self.local, local_dos),
            (self.visitante, visitante_uno), (self.visitante, visitante_dos),
        ):
            AlineacionPartido.objects.create(partido=self.partido, equipo=equipo, jugador=jugador, rol="TITULAR")
        self.partido.fase = "SEMIFINAL"
        self.partido.estado = "EN_JUEGO"
        self.partido.periodo_en_vivo = "PEN"
        self.partido.save()
        self.client.force_login(self.planillero)

        url = f"/partido/{self.partido.id}/cronometro/penales/cobro/"
        self.client.post(url, {"jugador": self.jugador.id, "resultado": "GOL"})
        self.client.post(url, {"jugador": visitante_uno.id, "resultado": "GOL"})
        self.client.post(url, {"jugador": self.jugador.id, "resultado": "GOL"})
        self.assertEqual(CobroPenal.objects.filter(partido=self.partido).count(), 2)

        self.client.post(url, {"jugador": local_dos.id, "resultado": "GOL"})
        self.assertEqual(CobroPenal.objects.filter(partido=self.partido).count(), 3)

    def test_muerte_subita_espera_el_cobro_del_rival(self):
        visitante = Jugador.objects.create(
            equipo=self.visitante, dorsal=10, nombres="Cobrador Visitante", cedula="MSV10",
            fecha_nacimiento=date(1991, 1, 1),
        )
        AlineacionPartido.objects.create(partido=self.partido, equipo=self.local, jugador=self.jugador, rol="TITULAR")
        AlineacionPartido.objects.create(partido=self.partido, equipo=self.visitante, jugador=visitante, rol="TITULAR")
        self.partido.fase = "FINAL"
        self.partido.estado = "EN_JUEGO"
        self.partido.periodo_en_vivo = "PEN"
        self.partido.save()
        self.client.force_login(self.planillero)
        url = f"/partido/{self.partido.id}/cronometro/penales/cobro/"

        for indice in range(11):
            jugador = self.jugador if indice % 2 == 0 else visitante
            self.client.post(url, {"jugador": jugador.id, "resultado": "GOL"})
        self.client.get(f"/partido/{self.partido.id}/cronometro/finalizar/")
        self.partido.refresh_from_db()
        self.assertEqual(self.partido.estado, "EN_JUEGO")

        self.client.post(url, {"jugador": visitante.id, "resultado": "FALLO"})
        self.client.get(f"/partido/{self.partido.id}/cronometro/finalizar/")
        self.partido.refresh_from_db()
        self.assertEqual(self.partido.estado, "FINALIZADO")

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_resumen_publico_muestra_goles_y_fallos_de_la_tanda(self):
        visitante = Jugador.objects.create(
            equipo=self.visitante, dorsal=10, nombres="Jugador Que Falla", cedula="RPF10",
            fecha_nacimiento=date(1991, 1, 1),
        )
        CobroPenal.objects.create(
            partido=self.partido, equipo=self.local, jugador=self.jugador, orden=1, convertido=True,
        )
        CobroPenal.objects.create(
            partido=self.partido, equipo=self.visitante, jugador=visitante, orden=2, convertido=False,
        )
        self.partido.goles_local_penales = 1
        self.partido.goles_visitante_penales = 0
        self.partido.save()

        respuesta = self.client.get(f"/partido/{self.partido.id}/live/")

        self.assertContains(respuesta, "Tanda de penales")
        self.assertContains(respuesta, self.jugador.nombres)
        self.assertContains(respuesta, "Anotó")
        self.assertContains(respuesta, visitante.nombres)
        self.assertContains(respuesta, "Falló")
        self.assertContains(respuesta, "timeline-event penal-cobro")
        self.assertContains(respuesta, "penalty-ball-miss")
        contenido = respuesta.content.decode("utf-8")
        self.assertLess(contenido.index(visitante.nombres), contenido.index(self.jugador.nombres))
        self.assertLess(contenido.index("Tanda de penales"), contenido.index("Fin de los 90 minutos"))

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_editor_movil_muestra_edad_en_alineacion(self):
        self.client.force_login(self.planillero)

        respuesta = self.client.get(f"/partido/{self.partido.id}/editor-movil/")

        self.assertContains(respuesta, "36 años")

    def test_gol_permite_minuto_manual(self):
        self.client.force_login(self.planillero)
        self.client.post(
            f"/partido/{self.partido.id}/agregar-gol-movil/",
            {"equipo": self.local.id, "jugador": self.jugador.id, "cantidad": 1, "minuto": 37},
        )
        self.assertEqual(Gol.objects.get(partido=self.partido).minuto, 37)

    def test_tarjeta_permite_minuto_manual(self):
        self.client.force_login(self.planillero)
        self.client.post(
            f"/partido/{self.partido.id}/agregar-tarjeta-movil/",
            {"equipo": self.local.id, "jugador": self.jugador.id, "tipo": "AMARILLA", "minuto": 42},
        )
        self.assertEqual(Tarjeta.objects.get(partido=self.partido).minuto, 42)

    def test_evento_rechaza_minuto_manual_negativo(self):
        self.client.force_login(self.planillero)
        self.client.post(
            f"/partido/{self.partido.id}/agregar-gol-movil/",
            {"equipo": self.local.id, "jugador": self.jugador.id, "cantidad": 1, "minuto": -1},
        )
        self.assertFalse(Gol.objects.filter(partido=self.partido).exists())

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_live_muestra_cuerpo_tecnico_y_avatares_en_sustituciones(self):
        self.local.director_tecnico = "Director Local"
        self.local.asistente_tecnico = "Asistente Local"
        self.local.delegado = "Delegado Local"
        self.local.administrador_app = "Administrador Local"
        self.local.save(update_fields=["director_tecnico", "asistente_tecnico", "delegado", "administrador_app"])
        entra = Jugador.objects.create(
            equipo=self.local,
            nombres="Jugador Entra",
            cedula="SUB-FOTO",
            fecha_nacimiento=date(1992, 1, 1),
        )
        SustitucionPartido.objects.create(
            partido=self.partido,
            equipo=self.local,
            jugador_sale=self.jugador,
            jugador_entra=entra,
            minuto=20,
        )

        respuesta = self.client.get(f"/partido/{self.partido.id}/live/")

        self.assertContains(respuesta, "Director Local")
        self.assertContains(respuesta, "Asistente Local")
        self.assertContains(respuesta, "Delegado Local")
        self.assertContains(respuesta, "Administrador Local")
        self.assertContains(respuesta, '<div class="staff-role">Delegado</div>')
        self.assertContains(respuesta, '<div class="staff-role">Admin App</div>')
        self.assertContains(respuesta, 'data-avatar-zoom')
        self.assertContains(respuesta, 'maximum-scale=6.0, user-scalable=yes')
        self.assertContains(respuesta, 'class="sub-player-avatar"', count=2)
        self.assertEqual(
            [item.jugador.id for item in respuesta.context["suplentes_local"]],
            [entra.id],
        )

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_live_marca_ganador_eliminatoria_y_muestra_hora_12(self):
        self.partido.fase = "CUARTOS"
        self.partido.numero_fecha = "CUARTOS #1"
        self.partido.estado = "FINALIZADO"
        self.partido.goles_local = 4
        self.partido.goles_visitante = 1
        self.partido.save(update_fields=[
            "fase", "numero_fecha", "estado", "goles_local", "goles_visitante",
        ])

        respuesta = self.client.get(f"/partido/{self.partido.id}/live/")

        self.assertTrue(respuesta.context["ganador_local"])
        self.assertFalse(respuesta.context["ganador_visitante"])
        self.assertContains(respuesta, "3:00 PM")
        self.assertContains(respuesta, 'class="live-winner-star"', count=1)

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_live_usa_selector_nativo_de_compartir_en_apk(self):
        respuesta = self.client.get(f"/partido/{self.partido.id}/live/")

        self.assertContains(respuesta, "AndroidDownloader.compartirEnlace")
        self.assertContains(respuesta, "navigator.share")

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_live_detecta_cambios_sin_recargar_constantemente(self):
        self.partido.estado = "EN_JUEGO"
        self.partido.save(update_fields=["estado"])

        pagina = self.client.get(f"/partido/{self.partido.id}/live/")
        revision_antes = self.client.get(
            f"/partido/{self.partido.id}/live/revision/"
        )
        Gol.objects.create(
            partido=self.partido,
            equipo=self.local,
            jugador=self.jugador,
            cantidad=1,
            minuto=12,
        )
        revision_despues = self.client.get(
            f"/partido/{self.partido.id}/live/revision/"
        )

        self.assertContains(pagina, "setInterval(revisarCambios, intervaloRevision)")
        self.assertContains(pagina, f"/partido/{self.partido.id}/live/revision/")
        self.assertEqual(revision_antes.status_code, 200)
        self.assertEqual(revision_despues.status_code, 200)
        self.assertNotEqual(
            revision_antes.json()["revision"],
            revision_despues.json()["revision"],
        )
        self.assertIn("no-cache", revision_despues["Cache-Control"])
        self.assertIn("ETag", revision_despues)

        sin_cambios = self.client.get(
            f"/partido/{self.partido.id}/live/revision/",
            {"revision": revision_despues.json()["revision"]},
        )
        self.assertEqual(sin_cambios.status_code, 204)
        self.assertEqual(sin_cambios.content, b"")

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_live_comprime_html_para_reducir_ancho_de_banda(self):
        sin_comprimir = self.client.get(f"/partido/{self.partido.id}/live/")
        comprimida = self.client.get(
            f"/partido/{self.partido.id}/live/",
            HTTP_ACCEPT_ENCODING="gzip",
        )

        self.assertEqual(comprimida.status_code, 200)
        self.assertEqual(comprimida.get("Content-Encoding"), "gzip")
        self.assertLess(len(comprimida.content), len(sin_comprimir.content))

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_live_muestra_aviso_penales_en_eliminatoria_empatada(self):
        self.partido.fase = "CUARTOS"
        self.partido.estado = "EN_JUEGO"
        self.partido.periodo_en_vivo = "PEN"
        self.partido.goles_local = 1
        self.partido.goles_visitante = 1
        self.partido.save(update_fields=[
            "fase", "estado", "periodo_en_vivo", "goles_local", "goles_visitante",
        ])

        respuesta = self.client.get(f"/partido/{self.partido.id}/live/")

        self.assertContains(respuesta, "PENALES")
        self.assertContains(respuesta, "&#128308; EN VIVO")
        self.assertNotContains(respuesta, 'class="live-clock"')

        partido_panel = next(
            item for item in construir_partidos_portada(self.torneo)
            if item["id"] == self.partido.id
        )
        tarjeta_panel = render_to_string(
            "partials/partido_portada_card.html",
            {"partido": partido_panel},
        )
        self.assertIn("EN VIVO", tarjeta_panel)
        self.assertNotIn('class="reloj-partido"', tarjeta_panel)

        self.client.force_login(self.planillero)
        editor = self.client.get(f"/partido/{self.partido.id}/editor-movil/")
        self.assertContains(editor, "abrirSeccionPenales()")
        self.assertContains(editor, "Equipo que cobra primero")
        self.assertNotContains(editor, "PENALES ACTIVO")

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_sustituciones_muestran_primero_el_cambio_mas_reciente(self):
        entra_temprano = Jugador.objects.create(
            equipo=self.local,
            nombres="Entra Temprano",
            cedula="SUB-ORDEN-20",
            fecha_nacimiento=date(1992, 1, 1),
        )
        sale_tarde = Jugador.objects.create(
            equipo=self.local,
            nombres="Sale Tarde",
            cedula="SUB-ORDEN-SALE-60",
            fecha_nacimiento=date(1991, 1, 1),
        )
        entra_tarde = Jugador.objects.create(
            equipo=self.local,
            nombres="Entra Tarde",
            cedula="SUB-ORDEN-60",
            fecha_nacimiento=date(1990, 1, 1),
        )
        cambio_temprano = SustitucionPartido.objects.create(
            partido=self.partido,
            equipo=self.local,
            jugador_sale=self.jugador,
            jugador_entra=entra_temprano,
            minuto=20,
        )
        cambio_tarde = SustitucionPartido.objects.create(
            partido=self.partido,
            equipo=self.local,
            jugador_sale=sale_tarde,
            jugador_entra=entra_tarde,
            minuto=60,
        )

        respuesta_live = self.client.get(f"/partido/{self.partido.id}/live/")
        self.assertEqual(
            [cambio.id for cambio in respuesta_live.context["sustituciones_local"]],
            [cambio_tarde.id, cambio_temprano.id],
        )

        self.client.force_login(self.planillero)
        respuesta_editor = self.client.get(f"/partido/{self.partido.id}/editor-movil/")
        self.assertEqual(
            [cambio.id for cambio in respuesta_editor.context["sustituciones"]],
            [cambio_tarde.id, cambio_temprano.id],
        )

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_live_no_repite_persona_con_varios_cargos(self):
        self.local.director_tecnico = "José Pérez"
        self.local.asistente_tecnico = "JOSE PEREZ"
        self.local.delegado = " jose  perez "
        self.local.administrador_app = "josé pérez"
        self.local.save(update_fields=["director_tecnico", "asistente_tecnico", "delegado", "administrador_app"])

        respuesta = self.client.get(f"/partido/{self.partido.id}/live/")

        personas = respuesta.context["cuerpo_tecnico_local"]
        self.assertEqual(len(personas), 1)
        self.assertEqual(personas[0].nombre, "José Pérez")
        self.assertEqual(personas[0].cargo, "Director técnico / Asistente técnico / Delegado / Admin App")

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_editor_movil_muestra_selector_de_equipos_debajo_de_cada_cancha(self):
        self.client.force_login(self.planillero)

        respuesta = self.client.get(f"/partido/{self.partido.id}/editor-movil/")

        self.assertContains(respuesta, "Cambiar equipo debajo de la cancha", count=2)

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_editor_y_live_usan_primer_nombre_y_primer_apellido(self):
        self.jugador.nombres = "Carlos Mario Galvis Padilla"
        self.jugador.save(update_fields=["nombres"])
        self.client.force_login(self.planillero)

        respuesta = self.client.get(f"/partido/{self.partido.id}/editor-movil/")

        self.assertContains(respuesta, "Carlos Galvis")
        self.assertContains(respuesta, 'data-nombre="Carlos Galvis"')
        self.assertNotContains(respuesta, 'data-nombre="Carlos Mario Galvis Padilla"')
        self.assertNotContains(respuesta, "Carlos Mario Galvis Padilla</option>")
        self.assertEqual(nombre_corto_jugador(self.jugador), "Carlos Galvis")
        self.assertEqual(nombre_resumen_jugador(self.jugador), "Carlos Galvis")

        self.jugador.nombres = "Carlos Galvis Padilla"
        self.assertEqual(nombre_corto_jugador(self.jugador), "Carlos Galvis")
        self.jugador.nombres = "Carlos Galvis"
        self.assertEqual(nombre_corto_jugador(self.jugador), "Carlos Galvis")

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_editor_movil_usa_edad_en_fecha_del_partido(self):
        hoy = date.today()
        self.jugador.fecha_nacimiento = hoy.replace(year=hoy.year - 41)
        self.jugador.save(update_fields=["fecha_nacimiento"])
        self.partido.fecha = hoy - timedelta(days=1)
        self.partido.save(update_fields=["fecha"])
        self.client.force_login(self.planillero)

        respuesta = self.client.get(f"/partido/{self.partido.id}/editor-movil/")

        self.assertContains(respuesta, "40 a")

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_editor_movil_muestra_rango_edad_en_selector_y_cancha(self):
        hoy = date.today()
        self.jugador.fecha_nacimiento = hoy.replace(year=hoy.year - 41)
        self.jugador.save(update_fields=["fecha_nacimiento"])
        ReglaEdadCategoria.objects.create(
            categoria=self.categoria,
            etiqueta="+40",
            edad_minima=40,
            edad_maxima=44,
        )
        self.client.force_login(self.planillero)

        respuesta = self.client.get(f"/partido/{self.partido.id}/editor-movil/")

        self.assertContains(respuesta, "(+40)")
        self.assertContains(respuesta, 'data-slot-edad')
        self.assertContains(respuesta, 'data-slot-dorsal')
        self.assertContains(respuesta, 'data-edad="+40"')
        self.assertContains(respuesta, 'data-etiqueta-edad="+40"')

    def test_guardar_alineacion_masiva_asigna_posicion_automatica_a_titular(self):
        self.client.force_login(self.planillero)

        respuesta = self.client.post(
            f"/partido/{self.partido.id}/guardar-alineacion-movil/",
            {
                "equipo": self.local.id,
                f"rol_{self.jugador.id}": "TITULAR",
            },
        )

        self.assertEqual(respuesta.status_code, 302)
        alineacion = AlineacionPartido.objects.get(partido=self.partido, equipo=self.local, jugador=self.jugador)
        self.assertEqual(alineacion.rol, "TITULAR")
        self.assertIn(alineacion.posicion_cancha, {codigo for codigo, _ in AlineacionPartido.POSICIONES_CANCHA})

    def test_planillero_puede_actualizar_dorsal_desde_alineacion(self):
        self.client.force_login(self.planillero)

        respuesta = self.client.post(
            f"/partido/{self.partido.id}/guardar-alineacion-movil/",
            {
                "equipo": self.local.id,
                f"rol_{self.jugador.id}": "TITULAR",
                f"dorsal_{self.jugador.id}": "27",
            },
        )

        self.assertEqual(respuesta.status_code, 302)
        self.jugador.refresh_from_db()
        self.assertEqual(self.jugador.dorsal, 27)
        self.assertTrue(RegistroActividad.objects.filter(
            usuario=self.planillero,
            accion="ACTUALIZAR_DORSALES_ALINEACION",
            objeto_id=self.local.id,
        ).exists())

    def test_dorsal_repetido_no_impide_guardar_alineacion(self):
        otro = Jugador.objects.create(
            equipo=self.local,
            dorsal=15,
            nombres="Otro Jugador",
            cedula="PG2",
            fecha_nacimiento=date(1991, 1, 1),
        )
        self.client.force_login(self.planillero)

        self.client.post(
            f"/partido/{self.partido.id}/guardar-alineacion-movil/",
            {
                "equipo": self.local.id,
                f"rol_{self.jugador.id}": "TITULAR",
                f"rol_{otro.id}": "SUPLENTE",
                f"dorsal_{self.jugador.id}": "15",
                f"dorsal_{otro.id}": "15",
            },
        )

        self.jugador.refresh_from_db()
        self.assertEqual(self.jugador.dorsal, 15)
        self.assertEqual(
            AlineacionPartido.objects.filter(partido=self.partido, equipo=self.local).count(),
            2,
        )

    def test_planillero_guarda_once_completo_de_titulares_en_vivo(self):
        jugadores = [self.jugador]
        for indice in range(2, 12):
            jugadores.append(Jugador.objects.create(
                equipo=self.local,
                dorsal=indice,
                nombres=f"Titular {indice}",
                cedula=f"ONCE-{indice}",
                fecha_nacimiento=date(1990, 1, 1),
            ))
        self.partido.estado = "EN_JUEGO"
        self.partido.save(update_fields=["estado"])
        self.client.force_login(self.planillero)
        datos = {"equipo": self.local.id}
        for jugador in jugadores:
            datos[f"rol_{jugador.id}"] = "TITULAR"
            datos[f"dorsal_{jugador.id}"] = str(jugador.dorsal or "")

        respuesta = self.client.post(
            f"/partido/{self.partido.id}/guardar-alineacion-movil/",
            datos,
        )

        self.assertEqual(respuesta.status_code, 302)
        self.assertEqual(
            AlineacionPartido.objects.filter(
                partido=self.partido,
                equipo=self.local,
                rol="TITULAR",
            ).count(),
            11,
        )

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_planillero_ve_mensaje_de_acceso_exitoso_al_ingresar(self):
        respuesta = self.client.post(
            f"/ingresar/?next=/partido/{self.partido.id}/editor-movil/",
            {
                "username": "planillero",
                "password": "test",
                "next": f"/partido/{self.partido.id}/editor-movil/",
            },
            follow=True,
        )

        self.assertContains(respuesta, "Acceso exitoso. Ya puedes diligenciar tus partidos asignados.")

    def test_login_no_precarga_usuario_y_desactiva_autocompletado(self):
        respuesta = self.client.get("/ingresar/")

        self.assertContains(respuesta, 'name="username"')
        self.assertContains(respuesta, 'autocomplete="off"')
        self.assertContains(respuesta, 'autocomplete="new-password"')
        self.assertNotContains(respuesta, 'value="pablo"')

    def test_usuario_no_asignado_no_puede_abrir_editor(self):
        self.client.force_login(self.otro_usuario)

        respuesta = self.client.get(f"/partido/{self.partido.id}/editor-movil/")

        self.assertEqual(respuesta.status_code, 403)

    def test_planillero_asignado_puede_registrar_gol(self):
        self.client.force_login(self.planillero)

        respuesta = self.client.post(
            f"/partido/{self.partido.id}/agregar-gol-movil/",
            {
                "equipo": self.local.id,
                "jugador": self.jugador.id,
                "cantidad": 1,
            },
        )

        self.assertEqual(respuesta.status_code, 302)
        self.assertTrue(Gol.objects.filter(partido=self.partido, jugador=self.jugador).exists())
        self.partido.refresh_from_db()
        self.assertFalse(self.partido.estadisticas_validadas)


    def test_planillero_registra_validacion_de_cedulas_de_titulares(self):
        segundo = Jugador.objects.create(
            equipo=self.local,
            dorsal=10,
            nombres="Sin Cedula",
            cedula="SC1",
            fecha_nacimiento=date(1991, 1, 1),
        )
        self.client.force_login(self.planillero)

        respuesta = self.client.post(
            f"/partido/{self.partido.id}/guardar-alineacion-movil/",
            {
                "equipo": self.local.id,
                f"rol_{self.jugador.id}": "TITULAR",
                f"posicion_{self.jugador.id}": "DC",
                f"rol_{segundo.id}": "TITULAR",
                f"posicion_{segundo.id}": "ED",
                "documento_validado": [str(self.jugador.id)],
            },
        )

        self.assertEqual(respuesta.status_code, 302)
        validado = AlineacionPartido.objects.get(partido=self.partido, jugador=self.jugador)
        pendiente = AlineacionPartido.objects.get(partido=self.partido, jugador=segundo)
        self.assertTrue(validado.documento_validado)
        self.assertEqual(validado.documento_validado_por, self.planillero)
        self.assertIsNotNone(validado.documento_validado_en)
        self.assertFalse(pendiente.documento_validado)
        solicitud = SolicitudValidacion.objects.get(tipo="ALINEACION", partido=self.partido, equipo=self.local)
        self.assertEqual(solicitud.estado, "PENDIENTE")
        self.assertIn(segundo.id, solicitud.datos["documentos_faltantes"])

    def test_planillero_no_puede_marcar_wo(self):
        self.client.force_login(self.planillero)

        self.client.post(
            f"/partido/{self.partido.id}/guardar-info-movil/",
            {
                "goles_local": 3,
                "goles_visitante": -3,
                "estado": "WO",
            },
        )

        self.partido.refresh_from_db()
        self.assertEqual(self.partido.estado, "PROGRAMADO")
        self.assertEqual(self.partido.goles_visitante, 0)

    def test_planillero_pierde_acceso_cuando_finaliza_partido(self):
        self.client.force_login(self.planillero)

        respuesta = self.client.post(
            f"/partido/{self.partido.id}/guardar-info-movil/",
            {
                "goles_local": 2,
                "goles_visitante": 1,
                "estado": "FINALIZADO",
            },
        )

        self.partido.refresh_from_db()
        self.assertEqual(self.partido.estado, "FINALIZADO")
        self.assertRedirects(respuesta, f"/partido/{self.partido.id}/live/", fetch_redirect_response=False)

        respuesta_editor = self.client.get(f"/partido/{self.partido.id}/editor-movil/")
        self.assertEqual(respuesta_editor.status_code, 403)

    def test_estadisticas_pendientes_no_entran_a_tabla_ni_goleadores(self):
        self.partido.estado = "FINALIZADO"
        self.partido.goles_local = 2
        self.partido.goles_visitante = 0
        self.partido.estadisticas_validadas = False
        self.partido.save()
        Gol.objects.create(partido=self.partido, equipo=self.local, jugador=self.jugador, cantidad=2)

        estructura = construir_estructura(self.torneo)
        datos = estructura["Senior"]["grupos"]["SIN GRUPO"]
        fila_local = next(fila for fila in datos["tabla"] if fila["id"] == self.local.id)

        self.assertEqual(fila_local["pj"], 0)
        self.assertEqual(estructura["Senior"]["goleadores_planilla"], [])

    def test_partido_en_vivo_actualiza_tabla_provisional_y_muestra_marcador(self):
        self.partido.estado = "EN_JUEGO"
        self.partido.goles_local = 2
        self.partido.goles_visitante = 1
        self.partido.estadisticas_validadas = False
        self.partido.save()

        estructura = construir_estructura(self.torneo)
        datos = estructura["Senior"]
        fila_local = next(
            fila for fila in datos["grupos"]["SIN GRUPO"]["tabla"]
            if fila["id"] == self.local.id
        )

        self.assertTrue(datos["hay_partidos_en_vivo"])
        self.assertEqual(fila_local["pj"], 1)
        self.assertEqual(fila_local["pg"], 1)
        self.assertEqual(fila_local["gf"], 2)
        self.assertEqual(fila_local["gc"], 1)
        self.assertEqual(fila_local["pts"], 3)
        self.assertEqual(fila_local["partidos_en_vivo"][0]["marcador"], "2-1")
        self.assertEqual(fila_local["partidos_en_vivo"][0]["estado_equipo"], "ganando")
        fila_visitante = next(
            fila for fila in datos["grupos"]["SIN GRUPO"]["tabla"]
            if fila["id"] == self.visitante.id
        )
        self.assertEqual(fila_visitante["partidos_en_vivo"][0]["marcador"], "2-1")
        self.assertEqual(fila_visitante["partidos_en_vivo"][0]["estado_equipo"], "perdiendo")

    def test_endpoint_en_vivo_devuelve_solo_datos_de_posiciones(self):
        self.partido.estado = "EN_JUEGO"
        self.partido.goles_local = 2
        self.partido.goles_visitante = 1
        self.partido.save()
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

        respuesta = self.client.get(
            "/actualizaciones/posiciones-en-vivo/",
            {"categoria": self.categoria.nombre},
        )

        self.assertEqual(respuesta.status_code, 200)
        datos = respuesta.json()
        self.assertTrue(datos["activo"])
        self.assertEqual(datos["tablas"][0]["clave"], "grupo:SIN GRUPO")
        self.assertEqual(datos["tablas"][0]["filas"][0]["id"], self.local.id)
        self.assertEqual(datos["tablas"][0]["filas"][0]["pts"], 3)
        self.assertEqual(
            datos["tablas"][0]["filas"][0]["partidos_en_vivo"][0]["marcador"],
            "2-1",
        )
        self.assertNotContains(respuesta, "<html")

    def test_admin_valida_estadisticas_y_entran_a_reportes(self):
        self.partido.estado = "FINALIZADO"
        self.partido.goles_local = 2
        self.partido.goles_visitante = 0
        self.partido.estadisticas_validadas = False
        self.partido.save()
        Gol.objects.create(partido=self.partido, equipo=self.local, jugador=self.jugador, cantidad=2)
        self.client.force_login(self.admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

        respuesta = self.client.post(f"/gestion/partidos/{self.partido.id}/validar-estadisticas/")

        self.assertEqual(respuesta.status_code, 302)
        self.partido.refresh_from_db()
        self.assertTrue(self.partido.estadisticas_validadas)
        estructura = construir_estructura(self.torneo)
        fila_local = next(fila for fila in estructura["Senior"]["grupos"]["SIN GRUPO"]["tabla"] if fila["id"] == self.local.id)
        self.assertEqual(fila_local["pj"], 1)
        self.assertEqual(estructura["Senior"]["goleadores_planilla"][0]["total"], 2)

    def test_autogol_aparece_como_ag_y_no_suma_al_total_del_goleador(self):
        self.partido.estado = "FINALIZADO"
        self.partido.estadisticas_validadas = True
        self.partido.save(update_fields=["estado", "estadisticas_validadas"])
        Gol.objects.create(
            partido=self.partido,
            equipo=self.local,
            jugador=self.jugador,
            cantidad=1,
            es_autogol=True,
        )

        goleadores = construir_estructura(self.torneo)["Senior"]["goleadores_planilla"]

        self.assertEqual(len(goleadores), 1)
        self.assertEqual(goleadores[0]["celdas"][0], "AG")
        self.assertEqual(goleadores[0]["total"], 0)

    def test_gol_normal_y_autogol_en_la_misma_fecha_solo_suman_el_gol_normal(self):
        self.partido.estado = "FINALIZADO"
        self.partido.estadisticas_validadas = True
        self.partido.save(update_fields=["estado", "estadisticas_validadas"])
        Gol.objects.create(
            partido=self.partido, equipo=self.local, jugador=self.jugador, cantidad=2,
        )
        Gol.objects.create(
            partido=self.partido, equipo=self.local, jugador=self.jugador,
            cantidad=1, es_autogol=True,
        )

        goleador = construir_estructura(self.torneo)["Senior"]["goleadores_planilla"][0]

        self.assertEqual(goleador["celdas"][0], "2 AG")
        self.assertEqual(goleador["total"], 2)


class OrdenFechasFaseUnoTests(TestCase):
    def test_fechas_fase_uno_se_ordenan_numericamente(self):
        torneo = Torneo.objects.create(nombre="Liga", fecha_inicio=date(2026, 1, 1))
        categoria = Categoria.objects.create(
            nombre="Unica", torneo=torneo, edad_minima=18, edad_maxima=80,
        )
        local = Equipo.objects.create(nombre="Local", categoria=categoria)
        visitante = Equipo.objects.create(nombre="Visitante", categoria=categoria)
        local.administrador_app = "Administrador Local"
        local.save(update_fields=["administrador_app"])
        for numero in ("Fecha 1", "Fecha 10", "Fecha 2", "Fecha 15", "Fecha 3"):
            Partido.objects.create(
                categoria=categoria,
                equipo_local=local,
                equipo_visitante=visitante,
                numero_fecha=numero,
                fase="GRUPOS",
                fecha=date(2026, 1, 2),
                hora=time(16, 0),
            )

        datos_categoria = construir_estructura(torneo)["Unica"]
        fechas = list(datos_categoria["partidos_por_fecha"])

        self.assertEqual(fechas, ["Fecha 1", "Fecha 2", "Fecha 3", "Fecha 10", "Fecha 15"])
        self.assertEqual(
            datos_categoria["columnas_planilla"][:5],
            ["Fecha 1", "Fecha 2", "Fecha 3", "Fecha 10", "Fecha 15"],
        )
        self.assertEqual(
            [columna["etiqueta"] for columna in datos_categoria["columnas_planilla_display"][:5]],
            ["F1", "F2", "F3", "F10", "F15"],
        )
        equipo_local = next(equipo for equipo in datos_categoria["equipos"] if equipo["id"] == local.id)
        self.assertEqual(equipo_local["administrador_app"], "Administrador Local")


class AdminTorneoPermisosTests(TestCase):
    def setUp(self):
        self.organizador = Organizador.objects.create(nombre="Liga Pahevi")
        self.otro_organizador = Organizador.objects.create(nombre="Otra Liga")
        self.torneo = Torneo.objects.create(nombre="Autorizado", organizador=self.organizador, fecha_inicio=date(2026, 1, 1))
        self.otro_torneo = Torneo.objects.create(nombre="Bloqueado", fecha_inicio=date(2026, 2, 1))
        self.categoria = Categoria.objects.create(
            nombre="Senior",
            edad_minima=18,
            edad_maxima=60,
            torneo=self.torneo,
        )
        self.otra_categoria = Categoria.objects.create(
            nombre="Libre",
            edad_minima=18,
            edad_maxima=60,
            torneo=self.otro_torneo,
        )
        self.equipo = Equipo.objects.create(nombre="Permitido", categoria=self.categoria)
        self.otro_equipo = Equipo.objects.create(nombre="Privado", categoria=self.otra_categoria)
        self.admin = User.objects.create_user("admin-torneo", password="test", is_staff=True)
        AdminTorneo.objects.create(usuario=self.admin, torneo=self.torneo, puede_editar=True, puede_validar=False, puede_programar=False)

    def test_admin_solo_ve_torneos_asignados(self):
        self.client.force_login(self.admin)

        respuesta = self.client.get("/gestion/torneos/")

        self.assertContains(respuesta, "Autorizado")
        self.assertNotContains(respuesta, "Bloqueado")

    def test_admin_no_puede_editar_equipo_de_otro_torneo_por_url(self):
        self.client.force_login(self.admin)

        respuesta = self.client.get(f"/gestion/equipos/{self.otro_equipo.id}/editar/")

        self.assertEqual(respuesta.status_code, 404)

    def test_admin_asignado_registra_actividad_al_editar_equipo(self):
        self.client.force_login(self.admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

        respuesta = self.client.post(
            f"/gestion/equipos/{self.equipo.id}/editar/",
            {
                "nombre": "Permitido FC",
                "categoria": self.categoria.id,
                "activo": "on",
            },
        )

        self.assertEqual(respuesta.status_code, 302)
        self.assertTrue(
            RegistroActividad.objects.filter(
                usuario=self.admin,
                torneo=self.torneo,
                accion="EDITAR",
                modelo="Equipo",
            ).exists()
        )

    def test_admin_sin_permiso_de_validar_no_valida_estadisticas(self):
        partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=Equipo.objects.create(nombre="Rival", categoria=self.categoria),
            fecha=date(2026, 5, 1),
            hora=time(15, 0),
            estado="FINALIZADO",
        )
        self.client.force_login(self.admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

        respuesta = self.client.post(f"/gestion/partidos/{partido.id}/validar-estadisticas/")

        self.assertEqual(respuesta.status_code, 403)

    def test_admin_principal_staff_puede_crear_organizadores(self):
        admin_principal = User.objects.create_user("admin-principal", password="test", is_staff=True)
        self.client.force_login(admin_principal)

        panel = self.client.get("/gestion/")
        respuesta = self.client.post(
            "/gestion/organizadores/nuevo/",
            {"nombre": "Nuevo Organizador", "activo": "on"},
        )

        self.assertContains(panel, "Organizadores")
        self.assertEqual(respuesta.status_code, 302)
        self.assertTrue(Organizador.objects.filter(nombre="Nuevo Organizador").exists())

    def test_admin_torneo_asignado_no_crea_organizadores(self):
        self.client.force_login(self.admin)

        respuesta = self.client.post(
            "/gestion/organizadores/nuevo/",
            {"nombre": "Organizador No Permitido", "activo": "on"},
        )

        self.assertNotEqual(respuesta.status_code, 200)
        self.assertFalse(Organizador.objects.filter(nombre="Organizador No Permitido").exists())

    def test_admin_principal_puede_crear_y_asignar_admin_de_organizador(self):
        admin_principal = User.objects.create_user("admin-principal-organizador", password="test", is_staff=True)
        self.client.force_login(admin_principal)

        respuesta = self.client.post(
            f"/gestion/organizadores/{self.organizador.id}/admins/",
            {
                "accion": "crear_admin",
                "username": "nuevo-admin-org",
                "first_name": "Nuevo",
                "last_name": "Admin",
                "email": "nuevo@example.com",
                "password": "clave-temporal-123",
                "puede_editar": "on",
                "puede_validar": "on",
                "puede_programar": "on",
                "activo": "on",
            },
        )

        usuario = User.objects.get(username="nuevo-admin-org")
        self.assertEqual(respuesta.status_code, 302)
        self.assertTrue(usuario.is_staff)
        self.assertTrue(AdminOrganizador.objects.filter(usuario=usuario, organizador=self.organizador).exists())

    def test_admin_principal_no_duplica_usuario_al_crear_admin_organizador(self):
        admin_principal = User.objects.create_user("admin-principal-duplicado", password="test", is_staff=True)
        User.objects.create_user("admin-existente", password="test", is_staff=True)
        self.client.force_login(admin_principal)

        respuesta = self.client.post(
            f"/gestion/organizadores/{self.organizador.id}/admins/",
            {
                "accion": "crear_admin",
                "username": "admin-existente",
                "password": "clave-temporal-123",
                "puede_editar": "on",
                "puede_validar": "on",
                "puede_programar": "on",
                "activo": "on",
            },
        )

        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, "Ya existe un usuario")
        self.assertFalse(AdminOrganizador.objects.filter(usuario__username="admin-existente", organizador=self.organizador).exists())

    def test_admin_de_organizador_ve_torneos_del_organizador(self):
        torneo_organizador = Torneo.objects.create(
            nombre="Segundo del organizador",
            organizador=self.organizador,
            fecha_inicio=date(2026, 3, 1),
        )
        admin = User.objects.create_user("admin-organizador", password="test", is_staff=True)
        AdminOrganizador.objects.create(usuario=admin, organizador=self.organizador)
        self.client.force_login(admin)

        respuesta = self.client.get("/gestion/torneos/")

        self.assertContains(respuesta, "Autorizado")
        self.assertContains(respuesta, torneo_organizador.nombre)
        self.assertNotContains(respuesta, "Bloqueado")

    def test_admin_de_organizador_accede_torneo_creado_despues(self):
        admin = User.objects.create_user("admin-organizador-futuro", password="test", is_staff=True)
        AdminOrganizador.objects.create(usuario=admin, organizador=self.organizador)
        torneo_nuevo = Torneo.objects.create(
            nombre="Torneo futuro",
            organizador=self.organizador,
            fecha_inicio=date(2026, 4, 1),
        )
        categoria_nueva = Categoria.objects.create(
            nombre="Futuro",
            edad_minima=18,
            edad_maxima=60,
            torneo=torneo_nuevo,
        )
        Equipo.objects.create(nombre="Equipo Futuro", categoria=categoria_nueva)
        self.client.force_login(admin)
        session = self.client.session
        session["torneo_id"] = torneo_nuevo.id
        session.save()

        respuesta = self.client.get("/gestion/equipos/")

        self.assertContains(respuesta, "Equipo Futuro")

    def test_admin_torneo_solo_planillas_entra_a_gestion_al_ingresar(self):
        admin = User.objects.create_user("admin-planillas", password="test")
        AdminTorneo.objects.create(
            usuario=admin,
            torneo=self.torneo,
            puede_editar=False,
            puede_validar=False,
            puede_programar=False,
            puede_descargar_planillas=True,
        )

        respuesta = self.client.post(
            "/ingresar/",
            {
                "username": "admin-planillas",
                "password": "test",
            },
        )

        self.assertRedirects(respuesta, "/gestion/", fetch_redirect_response=False)

    def test_admin_torneo_no_staff_solo_planillas_puede_ver_gestion(self):
        admin = User.objects.create_user("admin-planillas-no-staff", password="test")
        AdminTorneo.objects.create(
            usuario=admin,
            torneo=self.torneo,
            puede_editar=False,
            puede_validar=False,
            puede_programar=False,
            puede_descargar_planillas=True,
        )
        self.client.force_login(admin)

        respuesta = self.client.get("/gestion/")

        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, "Partidos")
        self.assertContains(respuesta, "Descargar planillas de impresion")

    def test_admin_programador_no_ve_editor_juego_en_partidos(self):
        admin = User.objects.create_user("admin-programa", password="test")
        AdminTorneo.objects.create(
            usuario=admin,
            torneo=self.torneo,
            puede_editar=False,
            puede_validar=False,
            puede_programar=True,
            puede_descargar_planillas=True,
        )
        Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=Equipo.objects.create(nombre="Rival", categoria=self.categoria),
            fecha=date(2026, 5, 1),
            hora=time(15, 0),
            estado="PROGRAMADO",
        )
        self.client.force_login(admin)

        respuesta = self.client.get("/gestion/partidos/")

        self.assertContains(respuesta, "Programar")
        self.assertNotContains(respuesta, "Editor juego")

    def test_editor_juego_desde_gestion_lleva_url_de_retorno(self):
        partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=Equipo.objects.create(nombre="Rival Editor", categoria=self.categoria),
            fecha=date(2026, 5, 1),
            hora=time(15, 0),
            estado="PROGRAMADO",
        )
        self.client.force_login(self.admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

        respuesta = self.client.get(f"/gestion/partidos/?categoria={self.categoria.id}&estado=PROGRAMADO")

        self.assertContains(
            respuesta,
            f"/partido/{partido.id}/editor-movil/?volver=/gestion/partidos/%3Fcategoria%3D{self.categoria.id}%26estado%3DPROGRAMADO",
        )

    def test_editor_juego_conserva_retorno_a_gestion_despues_de_guardar(self):
        rival = Equipo.objects.create(nombre="Rival Guarda", categoria=self.categoria)
        partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=rival,
            fecha=date(2026, 5, 1),
            hora=time(15, 0),
            estado="PROGRAMADO",
            numero_fecha="1",
            grupo="A",
            cancha="Teresa Sierra",
            fase="GRUPOS",
        )
        self.client.force_login(self.admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()
        volver = f"/gestion/partidos/?categoria={self.categoria.id}&estado=PROGRAMADO"

        respuesta = self.client.post(
            f"/partido/{partido.id}/guardar-info-movil/",
            {
                "volver": volver,
                "goles_local": "1",
                "goles_visitante": "0",
                "estado": "PROGRAMADO",
                "fecha": "2026-05-01",
                "hora": "15:00",
                "cancha": "Teresa Sierra",
                "numero_fecha": "1",
                "grupo": "A",
                "fase": "GRUPOS",
                "ajuste_puntos_local": "0",
                "ajuste_puntos_visitante": "0",
            },
        )

        self.assertRedirects(
            respuesta,
            f"/partido/{partido.id}/editor-movil/?volver=%2Fgestion%2Fpartidos%2F%3Fcategoria%3D{self.categoria.id}%26estado%3DPROGRAMADO",
            fetch_redirect_response=False,
        )

    def test_admin_programador_ve_opcion_eliminar_partido(self):
        admin = User.objects.create_user("admin-elimina-ui", password="test")
        AdminTorneo.objects.create(
            usuario=admin,
            torneo=self.torneo,
            puede_editar=False,
            puede_validar=False,
            puede_programar=True,
        )
        partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=Equipo.objects.create(nombre="Rival Eliminar", categoria=self.categoria),
            fecha=date(2026, 5, 1),
            hora=time(15, 0),
            estado="PROGRAMADO",
        )
        self.client.force_login(admin)

        respuesta = self.client.get("/gestion/partidos/")

        self.assertContains(respuesta, f'/gestion/partidos/{partido.id}/eliminar/')
        self.assertContains(respuesta, "Eliminar")

    def test_admin_programador_elimina_partido_y_registra_actividad(self):
        admin = User.objects.create_user("admin-elimina", password="test")
        AdminTorneo.objects.create(
            usuario=admin,
            torneo=self.torneo,
            puede_editar=False,
            puede_validar=False,
            puede_programar=True,
        )
        rival = Equipo.objects.create(nombre="Rival Borrado", categoria=self.categoria)
        jugador = Jugador.objects.create(
            equipo=self.equipo,
            dorsal=9,
            nombres="Jugador Borrado",
            cedula="9090",
            fecha_nacimiento=date(1990, 1, 1),
        )
        jugador_rival = Jugador.objects.create(
            equipo=rival,
            dorsal=10,
            nombres="Jugador Rival Borrado",
            cedula="9091",
            fecha_nacimiento=date(1991, 1, 1),
        )
        partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=rival,
            fecha=date(2026, 5, 1),
            hora=time(15, 0),
            estado="PROGRAMADO",
        )
        Gol.objects.create(partido=partido, equipo=self.equipo, jugador=jugador, cantidad=1)
        Tarjeta.objects.create(partido=partido, equipo=self.equipo, jugador=jugador, tipo="AMARILLA")
        AlineacionPartido.objects.create(partido=partido, equipo=self.equipo, jugador=jugador, rol="TITULAR")
        SustitucionPartido.objects.create(
            partido=partido,
            equipo=self.equipo,
            jugador_sale=jugador,
            jugador_entra=jugador_rival,
            minuto=20,
        )
        self.client.force_login(admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

        respuesta = self.client.post(f"/gestion/partidos/{partido.id}/eliminar/", {"next": "/gestion/partidos/"})

        self.assertEqual(respuesta.status_code, 302)
        self.assertFalse(Partido.objects.filter(id=partido.id).exists())
        self.assertFalse(Gol.objects.filter(partido_id=partido.id).exists())
        self.assertFalse(Tarjeta.objects.filter(partido_id=partido.id).exists())
        self.assertFalse(AlineacionPartido.objects.filter(partido_id=partido.id).exists())
        self.assertFalse(SustitucionPartido.objects.filter(partido_id=partido.id).exists())
        self.assertTrue(
            RegistroActividad.objects.filter(
                usuario=admin,
                torneo=self.torneo,
                accion="ELIMINAR",
                modelo="Partido",
            ).exists()
        )

    def test_admin_sin_permiso_programar_no_elimina_partido(self):
        partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=Equipo.objects.create(nombre="Rival Protegido", categoria=self.categoria),
            fecha=date(2026, 5, 1),
            hora=time(15, 0),
            estado="PROGRAMADO",
        )
        self.client.force_login(self.admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

        respuesta = self.client.post(f"/gestion/partidos/{partido.id}/eliminar/")

        self.assertEqual(respuesta.status_code, 403)
        self.assertTrue(Partido.objects.filter(id=partido.id).exists())

    def test_programar_partido_no_muestra_campos_de_resultado(self):
        admin = User.objects.create_user("admin-programa-form", password="test")
        AdminTorneo.objects.create(
            usuario=admin,
            torneo=self.torneo,
            puede_editar=False,
            puede_validar=False,
            puede_programar=True,
            puede_descargar_planillas=True,
        )
        partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=Equipo.objects.create(nombre="Rival Form", categoria=self.categoria),
            fecha=date(2026, 5, 1),
            hora=time(15, 0),
            estado="PROGRAMADO",
        )
        self.client.force_login(admin)

        respuesta = self.client.get(f"/gestion/partidos/{partido.id}/editar/")

        self.assertContains(respuesta, "Programar partido")
        self.assertNotContains(respuesta, 'name="goles_local"')
        self.assertNotContains(respuesta, 'name="goles_visitante"')
        self.assertNotContains(respuesta, 'name="ajuste_puntos_local"')
        self.assertNotContains(respuesta, 'name="ajuste_puntos_visitante"')
        self.assertNotContains(respuesta, 'name="goles_local_penales"')
        self.assertNotContains(respuesta, 'name="goles_visitante_penales"')
        self.assertNotContains(respuesta, 'name="estadisticas_validadas"')

    def test_editar_torneo_no_pide_fechas_ya_definidas(self):
        form = TorneoForm(instance=self.torneo)

        self.assertNotIn("fecha_inicio", form.fields)
        self.assertNotIn("fecha_fin", form.fields)

    def test_programar_partido_regresa_a_gestion_filtrada(self):
        admin = User.objects.create_user("admin-vuelve-filtro", password="test")
        AdminTorneo.objects.create(
            usuario=admin,
            torneo=self.torneo,
            puede_editar=False,
            puede_validar=False,
            puede_programar=True,
        )
        rival = Equipo.objects.create(nombre="Rival Retorno", categoria=self.categoria)
        partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=rival,
            fecha=date(2026, 5, 1),
            hora=time(15, 0),
            estado="PROGRAMADO",
            numero_fecha="1",
            grupo="A",
            cancha="Teresa Sierra",
            fase="GRUPOS",
        )
        self.client.force_login(admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

        volver = f"/gestion/partidos/?categoria={self.categoria.id}&estado=PROGRAMADO&q=Rival"
        volver_codificado = f"%2Fgestion%2Fpartidos%2F%3Fcategoria%3D{self.categoria.id}%26estado%3DPROGRAMADO%26q%3DRival"
        respuesta = self.client.post(
            f"/gestion/partidos/{partido.id}/editar/?volver={volver_codificado}",
            {
                "categoria": self.categoria.id,
                "equipo_local": self.equipo.id,
                "equipo_visitante": rival.id,
                "fecha": "2026-05-02",
                "hora": "16:00",
                "estado": "PROGRAMADO",
                "numero_fecha": "2",
                "grupo": "A",
                "cancha": "El Porvenir",
                "estado_programacion": "OFICIAL",
                "fase": "GRUPOS",
            },
        )

        self.assertRedirects(respuesta, volver, fetch_redirect_response=False)

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_admin_asignado_regresa_del_panel_a_gestion_sin_nuevo_login(self):
        admin = User.objects.create_user("admin-regresa", password="test")
        AdminTorneo.objects.create(
            usuario=admin,
            torneo=self.torneo,
            puede_editar=False,
            puede_validar=False,
            puede_programar=True,
            puede_descargar_planillas=True,
        )
        self.client.force_login(admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

        respuesta = self.client.get("/")

        self.assertContains(respuesta, 'href="/gestion/"')
        self.assertNotContains(respuesta, "Ingresar admin")
        self.assertIn("_auth_user_id", self.client.session)

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_admin_torneo_no_puede_descargar_programacion(self):
        self.client.force_login(self.admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

        panel = self.client.get("/")
        descarga_general = self.client.get("/descargar/programacion-general/")
        descarga_categoria = self.client.get("/descargar/programacion/Senior/")

        self.assertFalse(puede_descargar_programacion(self.admin))
        self.assertNotContains(panel, "Descargar programación")
        self.assertEqual(descarga_general.status_code, 302)
        self.assertEqual(descarga_categoria.status_code, 302)

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_admin_organizador_conserva_opcion_descargar_programacion(self):
        admin = User.objects.create_user("admin-organizador-descargas", password="test", is_staff=True)
        AdminOrganizador.objects.create(usuario=admin, organizador=self.organizador)
        self.client.force_login(admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

        respuesta = self.client.get("/")

        self.assertTrue(puede_descargar_programacion(admin))
        self.assertContains(respuesta, "Descargar programación")


class DescargaProgramacionFiltrosTests(TestCase):
    def setUp(self):
        self.torneo = Torneo.objects.create(nombre="Veranero", fecha_inicio=date(2026, 1, 1))
        self.categoria = Categoria.objects.create(nombre="Senior", edad_minima=18, edad_maxima=60, torneo=self.torneo)
        self.otra_categoria = Categoria.objects.create(nombre="Plus 50", edad_minima=50, edad_maxima=80, torneo=self.torneo)
        self.local = Equipo.objects.create(nombre="Local", categoria=self.categoria)
        self.visitante = Equipo.objects.create(nombre="Visitante", categoria=self.categoria)
        self.plus_local = Equipo.objects.create(nombre="Plus Local", categoria=self.otra_categoria)
        self.plus_visitante = Equipo.objects.create(nombre="Plus Visitante", categoria=self.otra_categoria)
        self.admin = User.objects.create_user("admin-programacion-descarga", password="test", is_staff=True, is_superuser=True)
        self.partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.local,
            equipo_visitante=self.visitante,
            fecha=date(2026, 7, 18),
            hora=time(16, 0),
            estado="PROGRAMADO",
            estado_programacion="OFICIAL",
            numero_fecha="Fecha 1",
            cancha="Teresa Sierra",
            grupo="A",
        )
        Partido.objects.create(
            categoria=self.otra_categoria,
            equipo_local=self.plus_local,
            equipo_visitante=self.plus_visitante,
            fecha=date(2026, 7, 19),
            hora=time(10, 0),
            estado="PROGRAMADO",
            estado_programacion="OFICIAL",
            numero_fecha="Fecha 2",
            cancha="El Porvenir",
            grupo="B",
        )
        self.client.force_login(self.admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

    def test_selector_muestra_filtros_de_programacion(self):
        respuesta = self.client.get("/descargar/programacion/?volver=/")

        self.assertContains(respuesta, "Todas las categorias")
        self.assertContains(respuesta, "Fecha 1")
        self.assertContains(respuesta, "18/07/2026")
        self.assertContains(respuesta, f'value="{self.categoria.id}"')
        self.assertContains(respuesta, "Descargar fixture para compartir (PNG)")
        self.assertContains(respuesta, "Descargar fixture completo (Excel)")

    def test_selector_incluye_cruces_aun_sin_hora_ni_cancha_definitivas(self):
        categoria_pendiente = Categoria.objects.create(
            nombre="Libre", edad_minima=18, edad_maxima=80, torneo=self.torneo,
        )
        local = Equipo.objects.create(nombre="Libre Local", categoria=categoria_pendiente)
        visitante = Equipo.objects.create(nombre="Libre Visitante", categoria=categoria_pendiente)
        Partido.objects.create(
            categoria=categoria_pendiente,
            equipo_local=local,
            equipo_visitante=visitante,
            numero_fecha="Fecha 10",
            fase="GRUPOS",
            fecha=date(2026, 8, 1),
            hora=time(0, 0),
            cancha="Por definir",
            estado_programacion="SUGERIDO",
        )

        respuesta = self.client.get("/descargar/programacion/?volver=/")

        self.assertContains(respuesta, f'value="{categoria_pendiente.id}"')
        self.assertContains(respuesta, "Libre")
        self.assertContains(respuesta, 'value="Fecha 10"')
        self.assertContains(respuesta, 'value="2026-08-01"')

    @patch("torneos.views.crear_imagen_desde_html")
    def test_fixture_compartible_omite_fecha_hora_y_cancha_de_futuros(self, crear_imagen):
        crear_imagen.return_value = HttpResponse(b"png", content_type="image/png")

        respuesta = self.client.get(
            "/descargar/fixture-compartible/",
            {"categoria": self.categoria.id},
        )

        self.assertEqual(respuesta.status_code, 200)
        html = crear_imagen.call_args.args[0]
        self.assertIn("FECHA 1", html.upper())
        self.assertIn("grid-template-columns:repeat(3", html)
        self.assertIn("LOCAL", html.upper())
        self.assertIn("VISITANTE", html.upper())
        self.assertIn("VS", html)
        self.assertIn('class="escudo-default"', html)
        self.assertNotIn("GRUPO A", html)
        self.assertNotIn("logo_imcred", html)
        self.assertNotIn("18/07/2026", html)
        self.assertNotIn("4:00 PM", html)
        self.assertNotIn("Teresa Sierra", html)

    @patch("torneos.views.crear_imagen_desde_html")
    def test_fixture_compartible_separa_los_partidos_por_grupo(self, crear_imagen):
        crear_imagen.return_value = HttpResponse(b"png", content_type="image/png")
        local_b = Equipo.objects.create(nombre="Local B", categoria=self.categoria)
        visitante_b = Equipo.objects.create(nombre="Visitante B", categoria=self.categoria)
        Partido.objects.create(
            categoria=self.categoria,
            equipo_local=local_b,
            equipo_visitante=visitante_b,
            fecha=date(2026, 7, 18),
            hora=time(0, 0),
            estado="PROGRAMADO",
            numero_fecha="Fecha 1",
            cancha="Por definir",
            grupo="B",
            fase="GRUPOS",
        )

        respuesta = self.client.get(
            "/descargar/fixture-compartible/",
            {"categoria": self.categoria.id},
        )

        self.assertEqual(respuesta.status_code, 200)
        html = crear_imagen.call_args.args[0]
        self.assertIn("GRUPO A", html)
        self.assertIn("GRUPO B", html)
        self.assertLess(html.index("GRUPO A"), html.index("GRUPO B"))

    @patch("torneos.views.crear_imagen_desde_html")
    def test_fixture_compartible_permite_descargar_un_solo_grupo(self, crear_imagen):
        crear_imagen.return_value = HttpResponse(b"png", content_type="image/png")
        local_b = Equipo.objects.create(nombre="Local B", categoria=self.categoria)
        visitante_b = Equipo.objects.create(nombre="Visitante B", categoria=self.categoria)
        Partido.objects.create(
            categoria=self.categoria,
            equipo_local=local_b,
            equipo_visitante=visitante_b,
            fecha=date(2026, 7, 18),
            hora=time(0, 0),
            estado="PROGRAMADO",
            numero_fecha="Fecha 1",
            cancha="Por definir",
            grupo="B",
            fase="GRUPOS",
        )

        respuesta = self.client.get(
            "/descargar/fixture-compartible/",
            {"categoria": self.categoria.id, "grupo": "B"},
        )

        self.assertEqual(respuesta.status_code, 200)
        html = crear_imagen.call_args.args[0]
        self.assertIn("GRUPO B", html)
        self.assertIn("LOCAL B", html.upper())
        self.assertNotIn(">LOCAL<", html.upper())
        self.assertIn("GRUPO_B", crear_imagen.call_args.args[1])

    @patch("torneos.views.crear_imagen_desde_html")
    def test_fixture_compartible_muestra_equipo_que_descansa_en_cada_fecha(self, crear_imagen):
        crear_imagen.return_value = HttpResponse(b"png", content_type="image/png")
        equipo_descanso = Equipo.objects.create(nombre="Equipo Descanso", categoria=self.categoria)
        Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.local,
            equipo_visitante=equipo_descanso,
            fecha=date(2026, 7, 25),
            hora=time(0, 0),
            estado="PROGRAMADO",
            numero_fecha="Fecha 2",
            cancha="Por definir",
            grupo="A",
            fase="GRUPOS",
        )

        respuesta = self.client.get(
            "/descargar/fixture-compartible/",
            {"categoria": self.categoria.id},
        )

        self.assertEqual(respuesta.status_code, 200)
        html = crear_imagen.call_args.args[0]
        html_mayuscula = html.upper()
        bloque_fecha_1 = html_mayuscula[html_mayuscula.index("FECHA 1"):html_mayuscula.index("FECHA 2")]
        self.assertIn("DESCANSA: EQUIPO DESCANSO", bloque_fecha_1)

    @patch("torneos.views.crear_imagen_desde_html")
    def test_programacion_incrusta_escudo_default_si_el_equipo_no_tiene_escudo(self, crear_imagen):
        crear_imagen.return_value = HttpResponse(b"png", content_type="image/png")

        respuesta = self.client.get(f"/descargar/programacion/{self.categoria.nombre}/")

        self.assertEqual(respuesta.status_code, 200)
        html = crear_imagen.call_args.args[0]
        self.assertGreaterEqual(html.count('class="escudo-default"'), 2)

    @patch("torneos.views.crear_imagen_desde_html")
    def test_fixture_extenso_usa_tres_fechas_por_fila_y_ancho_movil(self, crear_imagen):
        crear_imagen.return_value = HttpResponse(b"png", content_type="image/png")
        for numero in range(2, 11):
            Partido.objects.create(
                categoria=self.categoria,
                equipo_local=self.local,
                equipo_visitante=self.visitante,
                fecha=date(2026, 7, 18),
                hora=time(0, 0),
                numero_fecha=f"Fecha {numero}",
                fase="GRUPOS",
                cancha="Por definir",
            )

        respuesta = self.client.get(
            "/descargar/fixture-compartible/",
            {"categoria": self.categoria.id},
        )

        self.assertEqual(respuesta.status_code, 200)
        html = crear_imagen.call_args.args[0]
        self.assertIn("grid-template-columns:repeat(3", html)
        self.assertEqual(crear_imagen.call_args.args[2], 1080)
        self.assertIn("body{padding:18px 10px 28px}", html)
        self.assertIn("white-space:nowrap", html)

    def test_descarga_fixture_completo_incluye_todos_los_estados_y_columnas(self):
        self.partido.estado = "FINALIZADO"
        self.partido.goles_local = 2
        self.partido.goles_visitante = 1
        self.partido.save(update_fields=["estado", "goles_local", "goles_visitante"])

        respuesta = self.client.get("/descargar/fixture-completo/")

        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(
            respuesta["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        libro = load_workbook(BytesIO(respuesta.content))
        hoja = libro["Fixture completo"]
        self.assertEqual(hoja.max_row, 3)
        self.assertEqual(
            [celda.value for celda in hoja[1]],
            [
                "Categoría", "Fase", "Grupo", "Fecha fixture", "Fecha calendario",
                "Hora", "Cancha", "Equipo local", "Equipo visitante", "Estado", "Marcador",
            ],
        )
        filas = list(hoja.iter_rows(min_row=2, values_only=True))
        self.assertIn("Finalizado", [fila[9] for fila in filas])
        self.assertIn("2 - 1", [fila[10] for fila in filas])
        self.assertIn("4:00 PM", [fila[5] for fila in filas])

    def test_fixture_completo_se_puede_filtrar_por_categoria(self):
        respuesta = self.client.get(
            "/descargar/fixture-completo/",
            {"categoria": self.categoria.id},
        )

        libro = load_workbook(BytesIO(respuesta.content))
        filas = list(libro["Fixture completo"].iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(filas), 1)
        self.assertEqual(filas[0][0], "Senior")

    def test_etiquetas_de_valla_abrevian_fechas_y_fases(self):
        self.assertEqual(etiqueta_columna_planilla("1"), "F1")
        self.assertEqual(etiqueta_columna_planilla("Fecha 2"), "F2")
        self.assertEqual(etiqueta_columna_planilla("CUARTOS"), "CT")
        self.assertEqual(etiqueta_columna_planilla("SEMIFINAL"), "SM")
        self.assertEqual(etiqueta_columna_planilla("TERCER_PUESTO"), "TP")
        self.assertEqual(etiqueta_columna_planilla("FINAL"), "F")

    def test_constructor_filtra_programacion_por_categoria_fecha_y_dia(self):
        request = self.client.get("/descargar/programacion/").wsgi_request

        partidos = construir_partidos_programacion(
            request,
            categoria_obj=self.categoria,
            numero_fecha="Fecha 1",
            dia=date(2026, 7, 18),
        )

        self.assertEqual(len(partidos), 1)
        self.assertEqual(partidos[0]["local"], "Local")
        self.assertEqual(partidos[0]["hora_texto"], "4:00 PM")

    @patch("torneos.views.crear_imagen_desde_html")
    def test_programacion_incluye_suspendido_reprogramado_sin_cambiar_su_estado(self, crear_imagen):
        crear_imagen.return_value = HttpResponse(b"png", content_type="image/png")
        self.partido.estado = "SUSPENDIDO"
        self.partido.save(update_fields=["estado"])

        respuesta = self.client.get(f"/descargar/programacion/{self.categoria.nombre}/")

        self.assertEqual(respuesta.status_code, 200)
        html = crear_imagen.call_args.args[0]
        self.assertIn("REPROGRAMADO", html)
        self.assertNotIn(">SUSPENDIDO<", html)
        self.partido.refresh_from_db()
        self.assertEqual(self.partido.estado, "SUSPENDIDO")

    def test_programacion_normal_sigue_incluyendo_partido_programado(self):
        request = self.client.get("/descargar/programacion/").wsgi_request

        partidos = construir_partidos_programacion(request, self.categoria)

        self.assertEqual(len(partidos), 1)
        self.assertFalse(partidos[0]["reprogramado"])

    def test_descarga_por_dia_conserva_numero_de_fecha_en_el_titulo(self):
        partidos = [
            {"numero_fecha": "2"},
            {"numero_fecha": "2"},
        ]

        self.assertEqual(fechas_presentes_en_programacion(partidos), "2")

    def test_encabezado_cuartos_explica_cruces_de_semifinales(self):
        partidos = [
            {"numero_fecha": "CUARTOS #4"},
            {"numero_fecha": "CUARTOS #2"},
            {"numero_fecha": "CUARTOS #3"},
            {"numero_fecha": "CUARTOS #1"},
        ]

        self.assertEqual(
            fechas_presentes_en_programacion(partidos),
            "CUARTOS DE FINAL · SEMIFINAL 1: LLAVES 1 Y 4 · SEMIFINAL 2: LLAVES 2 Y 3",
        )

    def test_descarga_de_fechas_fase_incluye_finalizado_con_resultado(self):
        self.partido.estado = "FINALIZADO"
        self.partido.goles_local = 3
        self.partido.goles_visitante = 1
        self.partido.save(update_fields=["estado", "goles_local", "goles_visitante"])
        request = self.client.get("/descargar/programacion/").wsgi_request

        programacion_normal = construir_partidos_programacion(request, self.categoria)
        programacion_fechas_fase = construir_partidos_programacion(
            request,
            self.categoria,
            incluir_resultados=True,
        )

        self.assertEqual(programacion_normal, [])
        self.assertEqual(len(programacion_fechas_fase), 1)
        self.assertEqual(programacion_fechas_fase[0]["marcador_texto"], "3 - 1")
        self.assertTrue(programacion_fechas_fase[0]["finalizado"])

    def test_programacion_cuartos_especifica_cruces_de_semifinales(self):
        cuarto_uno = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.local,
            equipo_visitante=self.visitante,
            fecha=date(2026, 7, 26),
            hora=time(14, 0),
            estado="PROGRAMADO",
            estado_programacion="OFICIAL",
            numero_fecha="CUARTOS #1",
            cancha="Teresa Sierra",
            grupo="FINAL",
            fase="CUARTOS",
        )
        request = self.client.get("/descargar/programacion/").wsgi_request

        partidos = construir_partidos_programacion(
            request,
            categoria_obj=self.categoria,
            numero_fecha=cuarto_uno.numero_fecha,
            dia=cuarto_uno.fecha,
        )

        self.assertEqual(len(partidos), 1)
        self.assertEqual(
            partidos[0]["destino_eliminatoria"],
            "GANADOR LLAVE 1 vs GANADOR LLAVE 4 · SEMIFINAL 1",
        )


    def test_fases_finales_abiertas_excluyen_grupos_y_partidos_cerrados(self):
        semifinal = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.local,
            equipo_visitante=self.visitante,
            fecha=date(2026, 7, 28),
            hora=time(18, 0),
            estado="PROGRAMADO",
            estado_programacion="OFICIAL",
            numero_fecha="SEMIFINAL #1",
            cancha="Teresa Sierra",
            grupo="FINAL",
            fase="SEMIFINAL",
        )
        Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.local,
            equipo_visitante=self.visitante,
            fecha=date(2026, 7, 27),
            hora=time(16, 0),
            estado="FINALIZADO",
            estado_programacion="OFICIAL",
            numero_fecha="CUARTOS #1",
            cancha="Teresa Sierra",
            grupo="FINAL",
            fase="CUARTOS",
        )
        request = self.client.get("/descargar/programacion/").wsgi_request

        partidos = construir_partidos_programacion(
            request,
            categoria_obj=self.categoria,
            fases=["CUARTOS", "SEMIFINAL", "FINAL", "TERCER_PUESTO"],
        )

        self.assertEqual([partido["numero_fecha"] for partido in partidos], [semifinal.numero_fecha])

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_pestana_finales_usa_descarga_de_cuadro_completo(self):
        Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.local,
            equipo_visitante=self.visitante,
            fecha=date(2026, 7, 28),
            hora=time(18, 0),
            estado="PROGRAMADO",
            estado_programacion="OFICIAL",
            numero_fecha="SEMIFINAL #1",
            cancha="Teresa Sierra",
            grupo="FINAL",
            fase="SEMIFINAL",
        )
        respuesta = self.client.get("/", {"categoria": self.categoria.nombre})

        self.assertContains(respuesta, "Descargar cuadro completo de fases finales")
        self.assertContains(respuesta, "fase_final=1")
        self.assertContains(
            respuesta,
            "Semifinal 1: ganador llave 1 vs ganador llave 4",
        )

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_descarga_fase_final_incluye_partido_finalizado_y_marcador(self):
        cerrado_local = Equipo.objects.create(nombre="Cerrado Local", categoria=self.categoria)
        cerrado_visitante = Equipo.objects.create(nombre="Cerrado Visitante", categoria=self.categoria)
        Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.local,
            equipo_visitante=self.visitante,
            fecha=date(2026, 7, 28),
            hora=time(18, 0),
            estado="PROGRAMADO",
            estado_programacion="OFICIAL",
            numero_fecha="SEMIFINAL #1",
            cancha="Teresa Sierra",
            grupo="FINAL",
            fase="SEMIFINAL",
        )
        Partido.objects.create(
            categoria=self.categoria,
            equipo_local=cerrado_local,
            equipo_visitante=cerrado_visitante,
            fecha=date(2026, 7, 27),
            hora=time(16, 0),
            estado="FINALIZADO",
            estado_programacion="OFICIAL",
            numero_fecha="CUARTOS #1",
            cancha="Teresa Sierra",
            grupo="FINAL",
            fase="CUARTOS",
        )

        with patch(
            "torneos.views.crear_imagen_desde_html",
            return_value=HttpResponse("ok"),
        ) as generar:
            respuesta = self.client.get(
                f"/descargar/programacion/{self.categoria.nombre}/",
                {"fase_final": "1"},
            )

        self.assertEqual(respuesta.status_code, 200)
        html = generar.call_args.args[0]
        self.assertIn("Semifinal 1", html)
        self.assertIn("Local", html)
        self.assertIn("Visitante", html)
        self.assertIn("Cerrado Local", html)
        self.assertIn("Cerrado Visitante", html)
        self.assertIn("0 - 0", html)
        self.assertEqual(generar.call_args.args[2:4], (900, 1550))
        self.assertIn('class="enfrentamiento"', html)

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_descarga_dinamica_no_se_guarda_en_cache(self):
        respuesta = self.client.get(f"/descargar/programacion/{self.categoria.nombre}/")

        self.assertEqual(respuesta.status_code, 200)
        self.assertIn("no-store", respuesta["Cache-Control"])
        self.assertEqual(respuesta["Pragma"], "no-cache")

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_cuadro_final_incluye_cuartos_sugeridos_ya_finalizados(self):
        Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.local,
            equipo_visitante=self.visitante,
            fecha=date(2026, 8, 1),
            hora=time(16, 0),
            estado="FINALIZADO",
            estado_programacion="SUGERIDA",
            numero_fecha="CUARTOS #2",
            cancha="Teresa Sierra",
            grupo="FINAL",
            fase="CUARTOS",
            goles_local=1,
            goles_visitante=1,
            goles_local_penales=5,
            goles_visitante_penales=6,
        )

        with patch(
            "torneos.views.crear_imagen_desde_html",
            return_value=HttpResponse("ok"),
        ) as generar:
            respuesta = self.client.get(
                f"/descargar/programacion/{self.categoria.nombre}/",
                {"fase_final": "1"},
            )

        self.assertEqual(respuesta.status_code, 200)
        html = generar.call_args.args[0]
        self.assertIn("Llave 2", html)
        self.assertIn("1 - 1", html)
        self.assertIn("Penales: 5 - 6", html)
        self.assertIn('estrella-clasificado">&#9733;</b>Visitante', html)
        self.assertNotIn("Llave 2<br>por definir", html)


class FixtureProgramacionBalanceadaTests(TestCase):
    def setUp(self):
        self.torneo = Torneo.objects.create(nombre="Programacion", fecha_inicio=date(2026, 1, 1))
        self.categoria = Categoria.objects.create(
            nombre="Senior",
            edad_minima=18,
            edad_maxima=60,
            torneo=self.torneo,
        )
        self.equipos = [
            Equipo.objects.create(nombre=f"Equipo {indice}", categoria=self.categoria)
            for indice in range(1, 5)
        ]
        self.admin = User.objects.create_user("super", password="test", is_staff=True, is_superuser=True)

    def datos_fixture(self, programacion=False):
        datos = {
            "categoria": self.categoria.id,
            "grupos": 1,
            "reemplazar": "on",
        }
        for equipo in self.equipos:
            datos.setdefault("equipos_grupo_0", []).append(str(equipo.id))
        if programacion:
            datos.update({
                "generar_programacion": "on",
                "fecha_inicio_programacion": "2026-06-06",
                "canchas_programacion": "Principal\r\nPorvenir",
                "cancha_obligatoria": "Porvenir",
                "franjas_programacion": ["SAB_16", "SAB_18", "DOM_08", "DOM_10", "DOM_14", "DOM_16"],
            })
        return datos

    def test_tablas_distinguen_roja_doble_amarilla_y_roja_directa(self):
        partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipos[0],
            equipo_visitante=self.equipos[1],
            fecha=date(2026, 6, 6),
            hora=time(16),
            estado="FINALIZADO",
            estadisticas_validadas=True,
            numero_fecha="1",
            grupo="A",
        )
        jugador_doble = Jugador.objects.create(
            equipo=self.equipos[0], nombres="Doble Amarilla", cedula="DA-1", fecha_nacimiento=date(1990, 1, 1),
        )
        jugador_directa = Jugador.objects.create(
            equipo=self.equipos[1], nombres="Roja Directa", cedula="RD-1", fecha_nacimiento=date(1991, 1, 1),
        )
        Tarjeta.objects.create(
            partido=partido, jugador=jugador_doble, equipo=self.equipos[0],
            tipo="ROJA", origen_roja="DOBLE_AMARILLA",
        )
        Tarjeta.objects.create(
            partido=partido, jugador=jugador_directa, equipo=self.equipos[1],
            tipo="ROJA", origen_roja="DIRECTA",
        )

        datos = construir_estructura(self.torneo)[self.categoria.nombre]
        tarjetas = {fila["jugador"]: fila for fila in datos["tarjetas_planilla"]}
        alertas = {fila["jugador"]: fila for fila in datos["alertas_tarjetas"]}

        self.assertIn("R2A", tarjetas["Doble Amarilla"]["celdas"])
        self.assertIn("RD", tarjetas["Roja Directa"]["celdas"])
        self.assertEqual(tarjetas["Doble Amarilla"]["total_a"], 0)
        self.assertIn("SUSPENSIÓN 1 FECHA", alertas["Doble Amarilla"]["observacion"])
        self.assertIn("SUSPENSIÓN 2 FECHAS", alertas["Roja Directa"]["observacion"])

    def test_fixture_sin_programacion_mantiene_comportamiento_actual(self):
        self.client.force_login(self.admin)

        respuesta = self.client.post("/gestion/generar-fixture/", self.datos_fixture())

        self.assertEqual(respuesta.status_code, 200)
        partidos = Partido.objects.filter(categoria=self.categoria)
        self.assertEqual(partidos.count(), 6)
        self.assertTrue(all(partido.cancha == "" for partido in partidos))
        self.assertTrue(all(partido.hora == time(0, 0) for partido in partidos))
        self.assertTrue(all(partido.estado_programacion == "MANUAL" for partido in partidos))

    def test_fixture_grupo_ida_vuelta_duplica_fechas_e_invierte_localias(self):
        self.client.force_login(self.admin)
        datos = self.datos_fixture()
        datos["ida_vuelta"] = "1"

        respuesta = self.client.post("/gestion/generar-fixture/", datos)

        self.assertEqual(respuesta.status_code, 200)
        partidos = Partido.objects.filter(categoria=self.categoria).order_by("numero_fecha", "id")
        self.assertEqual(partidos.count(), 12)
        self.assertEqual(
            {int(partido.numero_fecha) for partido in partidos},
            set(range(1, 7)),
        )
        cruces = {
            (partido.equipo_local_id, partido.equipo_visitante_id)
            for partido in partidos
        }
        for local in self.equipos:
            for visitante in self.equipos:
                if local != visitante:
                    self.assertIn((local.id, visitante.id), cruces)

    def test_fixture_ida_vuelta_admite_un_grupo_de_diez_equipos(self):
        for indice in range(5, 11):
            self.equipos.append(
                Equipo.objects.create(nombre=f"Equipo {indice}", categoria=self.categoria)
            )
        self.client.force_login(self.admin)
        datos = self.datos_fixture()
        datos["ida_vuelta"] = "1"

        respuesta = self.client.post("/gestion/generar-fixture/", datos)

        self.assertEqual(respuesta.status_code, 200)
        partidos = Partido.objects.filter(categoria=self.categoria)
        self.assertEqual(partidos.count(), 90)
        self.assertEqual(
            {int(numero) for numero in partidos.values_list("numero_fecha", flat=True)},
            set(range(1, 19)),
        )
        for equipo in self.equipos:
            jugados = partidos.filter(Q(equipo_local=equipo) | Q(equipo_visitante=equipo)).count()
            self.assertEqual(jugados, 18)

    def test_portada_distingue_programados_reales_de_futuros_sin_programar(self):
        programado = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipos[0],
            equipo_visitante=self.equipos[1],
            fecha=date.today() + timedelta(days=30),
            hora=time(16, 0),
            cancha="Porvenir",
            estado="PROGRAMADO",
            numero_fecha="1",
            grupo="A",
        )
        futuro = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipos[2],
            equipo_visitante=self.equipos[3],
            fecha=date.today() + timedelta(days=30),
            hora=time(0, 0),
            cancha="",
            estado="PROGRAMADO",
            numero_fecha="1",
            grupo="A",
        )

        partidos_portada = {partido["id"]: partido for partido in construir_partidos_portada(self.torneo)}

        self.assertEqual(partidos_portada[programado.id]["bloque"], "PROGRAMADOS")
        self.assertEqual(partidos_portada[programado.id]["hora"], "4:00 PM")
        self.assertEqual(partidos_portada[programado.id]["fecha_corta"], programado.fecha.strftime("%d/%m/%Y"))
        self.assertIn(partidos_portada[programado.id]["dia_abreviado"], {"LUN", "MAR", "MIÉ", "JUE", "VIE", "SÁB", "DOM"})
        self.assertEqual(partidos_portada[futuro.id]["bloque"], "FUTUROS")
        self.assertEqual(partidos_portada[futuro.id]["hora"], "Por definir")

    def test_panel_oculta_grupo_a_si_categoria_tiene_un_solo_grupo(self):
        partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipos[0],
            equipo_visitante=self.equipos[1],
            fecha=date.today() + timedelta(days=2),
            hora=time(16, 0),
            cancha="Principal",
            estado="PROGRAMADO",
            numero_fecha="1",
            grupo="A",
            fase="GRUPOS",
        )

        portada = {item["id"]: item for item in construir_partidos_portada(self.torneo)}[partido.id]
        estructura = construir_estructura(self.torneo)[self.categoria.nombre]
        tarjeta = render_to_string("partials/partido_portada_card.html", {"partido": portada})

        self.assertFalse(portada["mostrar_grupo"])
        self.assertFalse(estructura["mostrar_grupos"])
        self.assertNotIn("Grupo A", tarjeta)

    def test_panel_mantiene_grupos_si_categoria_tiene_dos_o_mas(self):
        partido_a = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipos[0],
            equipo_visitante=self.equipos[1],
            fecha=date.today() + timedelta(days=2), hora=time(16), cancha="Principal",
            numero_fecha="1", grupo="A", fase="GRUPOS",
        )
        Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipos[2],
            equipo_visitante=self.equipos[3],
            fecha=date.today() + timedelta(days=2), hora=time(18), cancha="Principal",
            numero_fecha="1", grupo="B", fase="GRUPOS",
        )

        portada = {item["id"]: item for item in construir_partidos_portada(self.torneo)}[partido_a.id]
        estructura = construir_estructura(self.torneo)[self.categoria.nombre]
        tarjeta = render_to_string("partials/partido_portada_card.html", {"partido": portada})

        self.assertTrue(portada["mostrar_grupo"])
        self.assertTrue(estructura["mostrar_grupos"])
        self.assertIn("Grupo A", tarjeta)

    def test_portada_incluye_marcador_de_penales(self):
        partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipos[0],
            equipo_visitante=self.equipos[1],
            fecha=date.today(), hora=time(16, 0), cancha="Principal",
            estado="EN_JUEGO", fase="CUARTOS",
            goles_local=2, goles_visitante=2,
            goles_local_penales=6, goles_visitante_penales=5,
            periodo_en_vivo="PEN",
        )

        portada = {item["id"]: item for item in construir_partidos_portada(self.torneo)}[partido.id]

        self.assertTrue(portada["tiene_penales"])
        self.assertEqual(portada["goles_local_penales"], 6)
        self.assertEqual(portada["goles_visitante_penales"], 5)

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_panel_ordena_programados_por_fecha_y_hora_no_por_numero_fixture(self):
        posterior = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipos[0],
            equipo_visitante=self.equipos[1],
            fecha=date.today() + timedelta(days=10),
            hora=time(10, 0),
            cancha="Principal",
            estado="PROGRAMADO",
            numero_fecha="1",
            grupo="A",
        )
        proximo = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipos[2],
            equipo_visitante=self.equipos[3],
            fecha=date.today() + timedelta(days=2),
            hora=time(18, 0),
            cancha="Principal",
            estado="PROGRAMADO",
            numero_fecha="9",
            grupo="A",
        )
        self.client.force_login(self.admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

        respuesta = self.client.get("/")

        ids_ordenados = [partido["id"] for partido in respuesta.context["partidos_programados"]]
        self.assertLess(ids_ordenados.index(proximo.id), ids_ordenados.index(posterior.id))

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_panel_ordena_resultados_del_mas_antiguo_al_mas_reciente(self):
        antiguo = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipos[0],
            equipo_visitante=self.equipos[1],
            fecha=date.today(), hora=time(17, 0), cancha="Principal",
            estado="FINALIZADO", numero_fecha="1", grupo="A",
        )
        reciente = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipos[2],
            equipo_visitante=self.equipos[3],
            fecha=date.today(), hora=time(19, 0), cancha="Principal",
            estado="FINALIZADO", numero_fecha="1", grupo="A",
        )
        self.client.force_login(self.admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

        respuesta = self.client.get("/")

        ids_ordenados = [partido["id"] for partido in respuesta.context["partidos_resultados"]]
        self.assertLess(ids_ordenados.index(antiguo.id), ids_ordenados.index(reciente.id))

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_panel_del_torneo_finalizado_muestra_primero_el_ultimo_partido_de_la_final(self):
        primera_fase = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipos[0],
            equipo_visitante=self.equipos[1],
            fecha=date.today() - timedelta(days=20),
            hora=time(10, 0),
            cancha="Principal",
            estado="FINALIZADO",
            fase="GRUPOS",
            numero_fecha="1",
            grupo="A",
        )
        final = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipos[2],
            equipo_visitante=self.equipos[3],
            fecha=date.today(),
            hora=time(18, 0),
            cancha="Principal",
            estado="FINALIZADO",
            fase="FINAL",
            numero_fecha="FINAL",
            grupo="FINAL",
        )
        self.torneo.estado = "FINALIZADO"
        self.torneo.save(update_fields=["estado"])
        self.client.force_login(self.admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

        respuesta = self.client.get("/")

        resultados = respuesta.context["partidos_resultados"]
        self.assertEqual(resultados[0]["id"], final.id)
        self.assertGreater(
            [partido["id"] for partido in resultados].index(primera_fase.id),
            [partido["id"] for partido in resultados].index(final.id),
        )

    def test_descarga_goleadores_extensa_se_divide_en_paginas_legibles(self):
        self.client.force_login(self.admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()
        datos = {
            "goleadores_planilla": [
                {"jugador": f"Jugador {indice}", "equipo": "Equipo", "celdas": [], "total": 1}
                for indice in range(40)
            ],
            "columnas_planilla_display": [],
        }
        logos = {"logo_alcaldia": "", "logo_torneo": "", "logo_imcred": ""}

        with (
            patch("torneos.views.construir_estructura", return_value={self.categoria.nombre: datos}),
            patch("torneos.views.preparar_categoria_para_descarga", side_effect=lambda request, valor: valor),
            patch("torneos.views.logos_torneo", return_value=logos),
            patch("torneos.views.render_to_string", return_value="<html></html>") as renderizar,
            patch("torneos.views.crear_imagenes_desde_html", return_value=HttpResponse("ok")) as generar,
        ):
            respuesta = self.client.get(f"/descargar/goleadores/{self.categoria.nombre}/")

        self.assertEqual(respuesta.status_code, 200)
        paginas = generar.call_args.args[0]
        self.assertEqual(len(paginas), 3)
        contextos = [llamada.args[1] for llamada in renderizar.call_args_list]
        self.assertEqual(
            [len(contexto["datos_categoria"]["goleadores_planilla"]) for contexto in contextos],
            [18, 18, 4],
        )
        self.assertEqual([contexto["posicion_inicial"] for contexto in contextos], [0, 18, 36])

    def test_fixture_con_programacion_balancea_cancha_obligatoria(self):
        self.client.force_login(self.admin)

        respuesta = self.client.post("/gestion/generar-fixture/", self.datos_fixture(programacion=True))

        self.assertContains(respuesta, "Resumen de equidad")
        partidos = Partido.objects.filter(categoria=self.categoria)
        self.assertEqual(partidos.count(), 6)
        self.assertTrue(all(partido.cancha in ["Principal", "Porvenir"] for partido in partidos))
        self.assertTrue(all(partido.hora != time(0, 0) for partido in partidos))
        self.assertTrue(all(partido.estado_programacion == "SUGERIDA" for partido in partidos))

        apariciones_porvenir = {equipo.id: 0 for equipo in self.equipos}
        for partido in partidos.filter(cancha__iexact="Porvenir"):
            apariciones_porvenir[partido.equipo_local_id] += 1
            apariciones_porvenir[partido.equipo_visitante_id] += 1

        self.assertTrue(all(cantidad >= 1 for cantidad in apariciones_porvenir.values()))

    def test_descarga_programacion_excluye_partidos_sugeridos(self):
        self.client.force_login(self.admin)
        self.client.post("/gestion/generar-fixture/", self.datos_fixture(programacion=True))
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

        request = self.client.get("/gestion/partidos/").wsgi_request
        partidos = construir_partidos_programacion(request, self.categoria)

        self.assertEqual(partidos, [])

    def test_confirmar_programacion_vuelve_partido_oficial(self):
        self.client.force_login(self.admin)
        self.client.post("/gestion/generar-fixture/", self.datos_fixture(programacion=True))
        partido = Partido.objects.filter(categoria=self.categoria).first()

        respuesta = self.client.post(f"/gestion/partidos/{partido.id}/confirmar-programacion/")

        self.assertEqual(respuesta.status_code, 302)
        partido.refresh_from_db()
        self.assertEqual(partido.estado_programacion, "OFICIAL")


    def datos_fixture_mata_mata(self):
        while len(self.equipos) < 10:
            indice = len(self.equipos) + 1
            self.equipos.append(Equipo.objects.create(nombre=f"Equipo {indice}", categoria=self.categoria))

        return {
            "categoria": self.categoria.id,
            "tipo_fixture": "MATA_MATA_IDA_VUELTA",
            "grupos": 1,
            "reemplazar": "on",
        }

    def test_fixture_mata_mata_crea_parejas_ida_vuelta_por_sorteo(self):
        self.client.force_login(self.admin)

        respuesta = self.client.post("/gestion/generar-fixture/", self.datos_fixture_mata_mata())

        self.assertEqual(respuesta.status_code, 200)
        partidos = Partido.objects.filter(categoria=self.categoria, fase="GRUPOS")
        self.assertEqual(partidos.count(), 10)
        self.assertEqual(partidos.values("grupo").distinct().count(), 5)

        for grupo in partidos.values_list("grupo", flat=True).distinct():
            partidos_grupo = list(partidos.filter(grupo=grupo).order_by("numero_fecha"))
            self.assertEqual(len(partidos_grupo), 2)
            self.assertEqual(partidos_grupo[0].equipo_local, partidos_grupo[1].equipo_visitante)
            self.assertEqual(partidos_grupo[0].equipo_visitante, partidos_grupo[1].equipo_local)


    def test_fixture_mata_mata_permite_parejas_manuales(self):
        self.client.force_login(self.admin)
        datos = self.datos_fixture_mata_mata()
        for indice in range(5):
            datos[f"mata_local_{indice}"] = str(self.equipos[indice * 2].id)
            datos[f"mata_visitante_{indice}"] = str(self.equipos[indice * 2 + 1].id)

        respuesta = self.client.post("/gestion/generar-fixture/", datos)

        self.assertEqual(respuesta.status_code, 200)
        partidos_mata_1 = list(
            Partido.objects.filter(categoria=self.categoria, grupo="MATA 1", fase="GRUPOS").order_by("numero_fecha")
        )
        self.assertEqual(partidos_mata_1[0].equipo_local, self.equipos[0])
        self.assertEqual(partidos_mata_1[0].equipo_visitante, self.equipos[1])
        self.assertEqual(partidos_mata_1[1].equipo_local, self.equipos[1])
        self.assertEqual(partidos_mata_1[1].equipo_visitante, self.equipos[0])

    def test_fixture_mata_mata_muestra_tabla_general_de_todos_los_equipos(self):
        self.client.force_login(self.admin)
        datos = self.datos_fixture_mata_mata()
        for indice in range(5):
            datos[f"mata_local_{indice}"] = str(self.equipos[indice * 2].id)
            datos[f"mata_visitante_{indice}"] = str(self.equipos[indice * 2 + 1].id)
        self.client.post("/gestion/generar-fixture/", datos)

        for partido in Partido.objects.filter(categoria=self.categoria, fase="GRUPOS"):
            partido.goles_local = 2
            partido.goles_visitante = 0
            partido.estado = "FINALIZADO"
            partido.estadisticas_validadas = True
            partido.save(update_fields=["goles_local", "goles_visitante", "estado", "estadisticas_validadas"])

        tabla_general = construir_estructura(self.torneo)["Senior"]["tabla_general_mata_mata"]

        self.assertEqual(len(tabla_general), 10)
        self.assertEqual(sum(equipo["pj"] for equipo in tabla_general), 20)
        self.assertTrue(all(equipo["pts"] == 3 for equipo in tabla_general))

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_panel_mata_mata_muestra_descarga_tabla_general(self):
        self.client.force_login(self.admin)
        datos = self.datos_fixture_mata_mata()
        self.client.post("/gestion/generar-fixture/", datos)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

        for partido in Partido.objects.filter(categoria=self.categoria, fase="GRUPOS"):
            partido.goles_local = 2
            partido.goles_visitante = 0
            partido.estado = "FINALIZADO"
            partido.estadisticas_validadas = True
            partido.save(update_fields=["goles_local", "goles_visitante", "estado", "estadisticas_validadas"])

        respuesta = self.client.get(f"/?categoria={self.categoria.nombre}")

        self.assertContains(respuesta, "Descargar tabla general mata-mata")
        self.assertContains(respuesta, f"/descargar/tabla-general-mata-mata/{self.categoria.nombre}/")

    def test_mata_mata_genera_cuartos_con_sistema_oreja(self):
        self.client.force_login(self.admin)
        self.client.post("/gestion/generar-fixture/", self.datos_fixture_mata_mata())

        orden = {equipo.id: indice for indice, equipo in enumerate(self.equipos, start=1)}
        for partido in Partido.objects.filter(categoria=self.categoria, fase="GRUPOS"):
            if orden[partido.equipo_local_id] < orden[partido.equipo_visitante_id]:
                partido.goles_local = 3
                partido.goles_visitante = 0
            else:
                partido.goles_local = 0
                partido.goles_visitante = 3
            partido.estado = "FINALIZADO"
            partido.save(update_fields=["goles_local", "goles_visitante", "estado"])

        tabla_visible = construir_estructura(self.torneo)["Senior"]["tabla_general_mata_mata"]
        tabla_generador = tabla_general_mata_mata_ida_vuelta(self.categoria)
        self.assertEqual(
            [fila["id"] for fila in tabla_generador],
            [fila["id"] for fila in tabla_visible],
        )
        clasificados = tabla_visible[:8]

        respuesta = self.client.get(f"/generar-llaves/{self.categoria.nombre}/")

        self.assertEqual(respuesta.status_code, 302)
        cuartos = {
            partido.numero_fecha: partido
            for partido in Partido.objects.filter(categoria=self.categoria, fase="CUARTOS")
        }
        self.assertEqual(cuartos["CUARTOS #1"].equipo_local_id, clasificados[0]["id"])
        self.assertEqual(cuartos["CUARTOS #1"].equipo_visitante_id, clasificados[7]["id"])
        self.assertEqual(cuartos["CUARTOS #2"].equipo_local_id, clasificados[1]["id"])
        self.assertEqual(cuartos["CUARTOS #2"].equipo_visitante_id, clasificados[6]["id"])
        self.assertEqual(cuartos["CUARTOS #3"].equipo_local_id, clasificados[2]["id"])
        self.assertEqual(cuartos["CUARTOS #3"].equipo_visitante_id, clasificados[5]["id"])
        self.assertEqual(cuartos["CUARTOS #4"].equipo_local_id, clasificados[3]["id"])
        self.assertEqual(cuartos["CUARTOS #4"].equipo_visitante_id, clasificados[4]["id"])



class ImportacionPartidosPlanillerosTests(TestCase):
    def setUp(self):
        self.torneo = Torneo.objects.create(nombre="Veranero", fecha_inicio=date(2026, 1, 1))
        self.categoria = Categoria.objects.create(
            nombre="Senior Master",
            edad_minima=18,
            edad_maxima=60,
            torneo=self.torneo,
        )
        self.local = Equipo.objects.create(nombre="Local", categoria=self.categoria)
        self.visitante = Equipo.objects.create(nombre="Visitante", categoria=self.categoria)
        self.admin = User.objects.create_user("admin-importa", password="test", is_staff=True, is_superuser=True)

    def test_busca_planilleros_por_usuario_correo_y_nombre(self):
        usuario = User.objects.create_user(
            "planilla1",
            email="planilla1@example.com",
            password="test",
            first_name="Carlos",
            last_name="Planillero",
        )
        otro = User.objects.create_user("planilla2", email="planilla2@example.com", password="test")

        planilleros, no_encontrados = buscar_planilleros_excel(
            "planilla1; planilla2@example.com; Carlos Planillero; noexiste"
        )

        self.assertEqual(planilleros, [usuario, otro])
        self.assertEqual(no_encontrados, ["noexiste"])

    def test_busca_planillero_ignorando_espacios_en_usuario(self):
        usuario = User.objects.create_user("Planillero 1", password="test")

        planilleros, no_encontrados = buscar_planilleros_excel("planillero1")

        self.assertEqual(planilleros, [usuario])
        self.assertEqual(no_encontrados, [])

    def test_importacion_asigna_planillero_desde_columna_planilleros(self):
        planillero = User.objects.create_user("Planillero 1", password="test")
        workbook = Workbook()
        hoja = workbook.active
        hoja.append([
            "categoria",
            "equipo_local",
            "equipo_visitante",
            "fecha",
            "hora",
            "numero_fecha",
            "grupo",
            "cancha",
            "fase",
            "estado",
            "planilleros",
        ])
        hoja.append([
            self.categoria.nombre,
            self.local.nombre,
            self.visitante.nombre,
            date(2026, 6, 1),
            time(16, 0),
            "1",
            "A",
            "Porvenir",
            "GRUPOS",
            "PROGRAMADO",
            "planillero1",
        ])
        archivo = BytesIO()
        workbook.save(archivo)
        archivo.seek(0)

        self.client.force_login(self.admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()
        respuesta = self.client.post(
            "/gestion/partidos/importar/",
            {
                "archivo_excel": SimpleUploadedFile(
                    "partidos.xlsx",
                    archivo.read(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        self.assertEqual(respuesta.status_code, 302)
        partido = Partido.objects.get(categoria=self.categoria, equipo_local=self.local, equipo_visitante=self.visitante)
        self.assertEqual(list(partido.planilleros.all()), [planillero])


class ImportacionJugadoresPlanillaTests(TestCase):
    def setUp(self):
        self.torneo = Torneo.objects.create(nombre="Veranero", fecha_inicio=date(2026, 1, 1))
        self.categoria = Categoria.objects.create(
            nombre="Senior Master",
            edad_minima=18,
            edad_maxima=70,
            torneo=self.torneo,
        )
        self.equipo = Equipo.objects.create(nombre="NIQUELEROS FC", categoria=self.categoria)
        self.admin = User.objects.create_user("admin-jugadores", password="test", is_staff=True, is_superuser=True)

    def _archivo_planilla(self):
        workbook = Workbook()
        hoja = workbook.active
        hoja["D3"] = self.categoria.nombre
        hoja["I3"] = self.equipo.nombre
        hoja["C8"] = "JUGADOR NUEVO DE PRUEBA"
        hoja["D8"] = 10
        hoja["E8"] = 1
        hoja["F8"] = 1
        hoja["G8"] = 1980
        hoja["H8"] = "12345"
        archivo = BytesIO()
        workbook.save(archivo)
        archivo.seek(0)
        return archivo

    def test_importar_planilla_elimina_jugadores_que_no_vienen_en_excel(self):
        Jugador.objects.create(
            equipo=self.equipo,
            nombres="Jugador Viejo",
            cedula="99999",
            fecha_nacimiento=date(1981, 1, 1),
            estado="ACTIVO",
        )

        self.client.force_login(self.admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()
        archivo = self._archivo_planilla()
        respuesta = self.client.post(
            "/gestion/jugadores/importar-planilla/",
            {
                "archivo_excel": SimpleUploadedFile(
                    "jugadores.xlsx",
                    archivo.read(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        self.assertEqual(respuesta.status_code, 302)
        self.assertTrue(Jugador.objects.filter(equipo=self.equipo, cedula="12345").exists())
        self.assertEqual(
            Jugador.objects.get(equipo=self.equipo, cedula="12345").nombres,
            "Jugador Nuevo De Prueba",
        )
        self.assertFalse(Jugador.objects.filter(equipo=self.equipo, cedula="99999").exists())

    def test_importa_administrador_app_y_telefonos_del_cuerpo_tecnico(self):
        workbook = Workbook()
        hoja = workbook.active
        hoja["D3"] = self.categoria.nombre
        hoja["I3"] = self.equipo.nombre
        hoja["C39"] = "Director Nuevo"
        hoja["H39"] = "300 111 2233"
        hoja["C40"] = "Asistente Nuevo"
        hoja["H40"] = "301.222.3344"
        hoja["C41"] = "Administrador App Nuevo"
        hoja["H41"] = "302-333-4455"
        hoja["C8"] = "Jugador Nuevo"
        hoja["D8"] = 10
        hoja["E8"] = 1
        hoja["F8"] = 1
        hoja["G8"] = 1980
        hoja["H8"] = "12345"
        archivo = BytesIO()
        workbook.save(archivo)
        archivo.seek(0)

        self.client.force_login(self.admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()
        respuesta = self.client.post(
            "/gestion/jugadores/importar-planilla/",
            {
                "archivo_excel": SimpleUploadedFile(
                    "inscripcion.xlsx",
                    archivo.read(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        self.assertEqual(respuesta.status_code, 302)
        self.equipo.refresh_from_db()
        self.assertEqual(self.equipo.director_tecnico, "DIRECTOR NUEVO")
        self.assertEqual(self.equipo.telefono_dt, "3001112233")
        self.assertEqual(self.equipo.asistente_tecnico, "ASISTENTE NUEVO")
        self.assertEqual(self.equipo.telefono_at, "3012223344")
        self.assertEqual(self.equipo.administrador_app, "ADMINISTRADOR APP NUEVO")
        self.assertEqual(self.equipo.telefono_administrador_app, "302-333-4455")


class PartidoFormTests(TestCase):
    def setUp(self):
        self.torneo = Torneo.objects.create(nombre="Veranero", fecha_inicio=date(2026, 1, 1))
        self.categoria = Categoria.objects.create(
            nombre="Senior",
            edad_minima=18,
            edad_maxima=60,
            torneo=self.torneo,
        )
        self.local = Equipo.objects.create(nombre="Local", categoria=self.categoria)
        self.visitante = Equipo.objects.create(nombre="Visitante", categoria=self.categoria)
        self.planillero = User.objects.create_user("planillero1", password="test")
        self.planillero_no_asignado = User.objects.create_user("planillero2", password="test")
        self.admin = User.objects.create_user("admin-torneo", password="test", is_staff=True)
        self.admin_asignado = User.objects.create_user("admin-asignado", password="test", is_staff=True)
        self.partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.local,
            equipo_visitante=self.visitante,
            fecha=date(2026, 6, 1),
            hora=time(16, 0),
            estado="PROGRAMADO",
            numero_fecha="1",
            grupo="A",
            cancha="Porvenir",
        )
        self.partido.planilleros.add(self.planillero, self.admin_asignado)

    def test_fecha_se_renderiza_en_formato_html_date(self):
        form = PartidoForm(instance=self.partido, torneo=self.torneo)

        self.assertIn('value="2026-06-01"', str(form["fecha"]))

    def test_planilleros_muestra_solo_asignados_si_el_partido_ya_los_tiene(self):
        form = PartidoForm(instance=self.partido, torneo=self.torneo)
        usuarios = list(form.fields["planilleros"].queryset)

        self.assertIn(self.planillero, usuarios)
        self.assertIn(self.admin_asignado, usuarios)
        self.assertNotIn(self.planillero_no_asignado, usuarios)
        self.assertNotIn(self.admin, usuarios)


class GestionEquiposAccesoDelegadoMasivoTests(TestCase):
    def setUp(self):
        self.torneo = Torneo.objects.create(nombre="Veranero", fecha_inicio=date(2026, 1, 1))
        self.categoria = Categoria.objects.create(
            nombre="Senior",
            edad_minima=18,
            edad_maxima=60,
            torneo=self.torneo,
        )
        self.otra_categoria = Categoria.objects.create(
            nombre="Plus",
            edad_minima=18,
            edad_maxima=60,
            torneo=self.torneo,
        )
        self.delegado_uno = User.objects.create_user("delegado-uno", password="test")
        self.delegado_dos = User.objects.create_user("delegado-dos", password="test")
        self.equipo_uno = Equipo.objects.create(nombre="Equipo Uno", categoria=self.categoria, responsable=self.delegado_uno)
        self.equipo_dos = Equipo.objects.create(nombre="Equipo Dos", categoria=self.categoria, responsable=self.delegado_dos)
        self.equipo_otro = Equipo.objects.create(nombre="Equipo Otro", categoria=self.otra_categoria)
        self.admin = User.objects.create_user("admin-masivo", password="test", is_staff=True, is_superuser=True)

    def test_actualiza_vencimiento_sin_reemplazar_delegados(self):
        self.client.force_login(self.admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

        respuesta = self.client.post(
            "/gestion/equipos/acceso-delegado-masivo/",
            {
                "acceso_delegado_hasta": "2026-06-10T18:00",
                "categoria": str(self.categoria.id),
                "q": "",
            },
        )

        self.assertEqual(respuesta.status_code, 302)
        self.equipo_uno.refresh_from_db()
        self.equipo_dos.refresh_from_db()
        self.equipo_otro.refresh_from_db()
        self.assertEqual(self.equipo_uno.responsable, self.delegado_uno)
        self.assertEqual(self.equipo_dos.responsable, self.delegado_dos)
        self.assertIsNone(self.equipo_otro.responsable)
        self.assertEqual(timezone.localtime(self.equipo_uno.acceso_delegado_hasta).strftime("%Y-%m-%dT%H:%M"), "2026-06-10T18:00")
        self.assertEqual(timezone.localtime(self.equipo_dos.acceso_delegado_hasta).strftime("%Y-%m-%dT%H:%M"), "2026-06-10T18:00")
        self.assertIsNone(self.equipo_otro.acceso_delegado_hasta)


class GestionEquiposCrearDelegadosMasivoTests(TestCase):
    def setUp(self):
        self.torneo = Torneo.objects.create(nombre="Veranero", fecha_inicio=date(2026, 1, 1))
        self.categoria = Categoria.objects.create(
            nombre="Senior",
            edad_minima=18,
            edad_maxima=60,
            torneo=self.torneo,
        )
        self.equipo_uno = Equipo.objects.create(nombre="Equipo Uno", categoria=self.categoria)
        self.equipo_dos = Equipo.objects.create(nombre="Equipo Dos", categoria=self.categoria)
        self.delegado_existente = User.objects.create_user("delegado-existente", password="test")
        self.equipo_con_delegado = Equipo.objects.create(
            nombre="Equipo Con Delegado",
            categoria=self.categoria,
            responsable=self.delegado_existente,
        )
        self.admin = User.objects.create_user("admin-crea-delegados", password="test", is_staff=True, is_superuser=True)

    def test_crea_un_usuario_por_equipo_sin_responsable(self):
        self.client.force_login(self.admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

        respuesta = self.client.post(
            "/gestion/equipos/crear-delegados-masivo/",
            {
                "password_temporal": "Temporal123",
                "acceso_delegado_hasta": "2026-06-10T18:00",
                "categoria": str(self.categoria.id),
                "q": "",
            },
        )

        self.assertEqual(respuesta.status_code, 302)
        self.equipo_uno.refresh_from_db()
        self.equipo_dos.refresh_from_db()
        self.equipo_con_delegado.refresh_from_db()
        self.assertIsNotNone(self.equipo_uno.responsable)
        self.assertIsNotNone(self.equipo_dos.responsable)
        self.assertEqual(self.equipo_uno.responsable.username, "admin-equipouno")
        self.assertEqual(self.equipo_dos.responsable.username, "admin-equipodos")
        self.assertNotEqual(self.equipo_uno.responsable, self.equipo_dos.responsable)
        self.assertEqual(self.equipo_con_delegado.responsable, self.delegado_existente)
        self.assertTrue(self.equipo_uno.responsable.check_password("Temporal123"))
        self.assertEqual(timezone.localtime(self.equipo_uno.acceso_delegado_hasta).strftime("%Y-%m-%dT%H:%M"), "2026-06-10T18:00")


class GestionEquiposRenombrarDelegadosMasivoTests(TestCase):
    def setUp(self):
        self.torneo = Torneo.objects.create(nombre="Veranero", fecha_inicio=date(2026, 1, 1))
        self.categoria = Categoria.objects.create(
            nombre="Senior",
            edad_minima=18,
            edad_maxima=60,
            torneo=self.torneo,
        )
        self.usuario_largo = User.objects.create_user("delegado-senior-niqueleros-fc", password="Temporal123")
        self.equipo = Equipo.objects.create(
            nombre="Niqueleros FC",
            categoria=self.categoria,
            responsable=self.usuario_largo,
        )
        self.admin = User.objects.create_user("admin-renombra-delegados", password="test", is_staff=True, is_superuser=True)

    def test_renombra_usuario_largo_a_formato_corto_sin_cambiar_password(self):
        self.client.force_login(self.admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

        respuesta = self.client.post(
            "/gestion/equipos/renombrar-delegados-masivo/",
            {
                "categoria": str(self.categoria.id),
                "q": "",
            },
        )

        self.assertEqual(respuesta.status_code, 302)
        self.usuario_largo.refresh_from_db()
        self.assertEqual(self.usuario_largo.username, "admin-niquelerosfc")
        self.assertTrue(self.usuario_largo.check_password("Temporal123"))


class GestionEquiposPermisosDelegadosMasivoTests(TestCase):
    def setUp(self):
        self.torneo = Torneo.objects.create(nombre="Veranero", fecha_inicio=date(2026, 1, 1))
        self.categoria = Categoria.objects.create(
            nombre="Senior",
            edad_minima=18,
            edad_maxima=60,
            torneo=self.torneo,
        )
        self.otra_categoria = Categoria.objects.create(
            nombre="Máster",
            edad_minima=40,
            edad_maxima=80,
            torneo=self.torneo,
        )
        delegado = User.objects.create_user("delegado-permisos", password="test")
        otro_delegado = User.objects.create_user("otro-delegado-permisos", password="test")
        self.equipo = Equipo.objects.create(
            nombre="Equipo Senior",
            categoria=self.categoria,
            responsable=delegado,
            delegado_puede_editar_equipo=False,
            delegado_puede_cargar_fotos_jugadores=False,
        )
        self.equipo_otra_categoria = Equipo.objects.create(
            nombre="Equipo Máster",
            categoria=self.otra_categoria,
            responsable=otro_delegado,
            delegado_puede_editar_equipo=False,
            delegado_puede_cargar_fotos_jugadores=False,
        )
        self.admin = User.objects.create_superuser("admin-permisos", password="test")

    def test_habilita_permisos_solo_a_equipos_visibles_del_filtro(self):
        self.client.force_login(self.admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

        respuesta = self.client.post(
            "/gestion/equipos/permisos-delegados-masivo/",
            {
                "categoria": str(self.categoria.id),
                "q": "",
                "permiso_editar": "HABILITAR",
                "permiso_fotos": "HABILITAR",
            },
        )

        self.assertEqual(respuesta.status_code, 302)
        self.equipo.refresh_from_db()
        self.equipo_otra_categoria.refresh_from_db()
        self.assertTrue(self.equipo.delegado_puede_editar_equipo)
        self.assertTrue(self.equipo.delegado_puede_cargar_fotos_jugadores)
        self.assertFalse(self.equipo_otra_categoria.delegado_puede_editar_equipo)
        self.assertFalse(self.equipo_otra_categoria.delegado_puede_cargar_fotos_jugadores)
        self.assertTrue(RegistroActividad.objects.filter(
            accion="PERMISOS_DELEGADOS_MASIVO",
            usuario=self.admin,
        ).exists())

    def test_puede_mantener_edicion_y_habilitar_solo_fotos(self):
        self.client.force_login(self.admin)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

        self.client.post(
            "/gestion/equipos/permisos-delegados-masivo/",
            {
                "categoria": str(self.categoria.id),
                "q": "",
                "permiso_editar": "MANTENER",
                "permiso_fotos": "HABILITAR",
            },
        )

        self.equipo.refresh_from_db()
        self.assertFalse(self.equipo.delegado_puede_editar_equipo)
        self.assertTrue(self.equipo.delegado_puede_cargar_fotos_jugadores)


class DelegadoEquipoTests(TestCase):
    def setUp(self):
        self.torneo = Torneo.objects.create(
            nombre="Veranero",
            fecha_inicio=date(2026, 1, 1),
        )
        self.categoria = Categoria.objects.create(
            nombre="Senior",
            edad_minima=18,
            edad_maxima=60,
            torneo=self.torneo,
        )
        self.delegado = User.objects.create_user("delegado", password="test")
        self.otro_usuario = User.objects.create_user("otro-delegado", password="test")
        self.equipo = Equipo.objects.create(
            nombre="Niqueleros",
            categoria=self.categoria,
            responsable=self.delegado,
            acceso_delegado_hasta=timezone.now() + timedelta(days=2),
        )
        self.otro_equipo = Equipo.objects.create(
            nombre="Rival",
            categoria=self.categoria,
            responsable=self.otro_usuario,
            acceso_delegado_hasta=timezone.now() + timedelta(days=2),
        )
        self.jugador = Jugador.objects.create(
            equipo=self.equipo,
            dorsal=7,
            nombres="Jugador Uno",
            cedula="123",
            fecha_nacimiento=date(1990, 1, 1),
        )

    def test_delegado_con_acceso_vigente_puede_editar_su_equipo(self):
        self.client.force_login(self.delegado)

        respuesta = self.client.post(
            f"/delegado/equipos/{self.equipo.id}/editar/",
            {
                "delegado": "Pablo Mazo",
                "telefono": "300123",
                "director_tecnico": "DT Uno",
                "telefono_dt": "300456",
                "asistente_tecnico": "AT Uno",
                "telefono_at": "300789",
            },
        )

        self.assertEqual(respuesta.status_code, 302)
        self.equipo.refresh_from_db()
        self.assertEqual(self.equipo.delegado, "Pablo Mazo")
        self.assertEqual(self.equipo.director_tecnico, "DT Uno")
        self.assertTrue(RegistroActividad.objects.filter(
            usuario=self.delegado,
            torneo=self.torneo,
            accion="EDITAR_EQUIPO_DELEGADO",
            objeto_id=self.equipo.id,
        ).exists())

    def test_formulario_delegado_permite_fotos_del_cuerpo_tecnico(self):
        self.client.force_login(self.delegado)

        respuesta = self.client.get(f"/delegado/equipos/{self.equipo.id}/editar/")

        self.assertContains(respuesta, 'name="foto_director_tecnico"')
        self.assertContains(respuesta, 'name="foto_asistente_tecnico"')
        self.assertContains(respuesta, 'name="foto_delegado"')
        self.assertContains(respuesta, 'name="administrador_app"')
        self.assertContains(respuesta, 'name="foto_administrador_app"')

    def test_delegado_ve_mensaje_de_acceso_exitoso_al_ingresar(self):
        respuesta = self.client.post(
            "/ingresar/",
            {
                "username": "delegado",
                "password": "test",
            },
            follow=True,
        )

        self.assertContains(respuesta, "Acceso exitoso. Bienvenido al portal de delegados.")

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_delegado_regresa_del_panel_a_mis_equipos_sin_nuevo_login(self):
        self.client.force_login(self.delegado)
        session = self.client.session
        session["torneo_id"] = self.torneo.id
        session.save()

        respuesta = self.client.get("/")

        self.assertContains(respuesta, f'href="/delegado/equipos/"')
        self.assertIn("_auth_user_id", self.client.session)

    def test_delegado_con_next_de_admin_entra_a_mis_equipos(self):
        respuesta = self.client.post(
            "/ingresar/?next=/gestion/",
            {
                "username": "delegado",
                "password": "test",
                "next": "/gestion/",
            },
        )

        self.assertRedirects(respuesta, "/delegado/equipos/", fetch_redirect_response=False)

    def test_delegado_sin_acceso_vigente_tambien_entra_a_mis_equipos(self):
        self.equipo.acceso_delegado_hasta = None
        self.equipo.save(update_fields=["acceso_delegado_hasta"])

        respuesta = self.client.post(
            "/ingresar/?next=/gestion/",
            {
                "username": "delegado",
                "password": "test",
                "next": "/gestion/",
            },
            follow=True,
        )

        self.assertContains(respuesta, "Niqueleros")
        self.assertContains(respuesta, "Sin fecha de acceso asignada.")
        self.assertContains(respuesta, "Alineacion de partidos")
        self.assertContains(respuesta, "Edicion de equipo bloqueada")

    def test_mis_equipos_filtra_torneo_activo_y_oculta_historicos(self):
        torneo_viejo = Torneo.objects.create(
            nombre="Torneo anterior",
            fecha_inicio=date(2025, 1, 1),
            estado="FINALIZADO",
        )
        categoria_vieja = Categoria.objects.create(
            nombre="Plus 50",
            edad_minima=50,
            edad_maxima=80,
            torneo=torneo_viejo,
        )
        Equipo.objects.create(
            nombre="Congal historico",
            categoria=categoria_vieja,
            responsable=self.delegado,
            acceso_delegado_hasta=timezone.now() + timedelta(days=2),
        )
        self.client.force_login(self.delegado)
        session = self.client.session
        session["torneo_id"] = torneo_viejo.id
        session.save()

        respuesta = self.client.get("/delegado/equipos/")

        self.assertContains(respuesta, "Niqueleros")
        self.assertContains(respuesta, "Veranero - Senior")
        self.assertNotContains(respuesta, "Congal historico")
        self.assertNotContains(respuesta, "Plus 50")

    def test_mis_equipos_respeta_torneo_activo_seleccionado(self):
        torneo_interbarrios = Torneo.objects.create(
            nombre="Interbarrios 2026",
            fecha_inicio=date(2026, 2, 1),
            estado="ACTIVO",
        )
        categoria_interbarrios = Categoria.objects.create(
            nombre="Interbarrios",
            edad_minima=18,
            edad_maxima=60,
            torneo=torneo_interbarrios,
        )
        Equipo.objects.create(
            nombre="Congal Interbarrios",
            categoria=categoria_interbarrios,
            responsable=self.delegado,
            acceso_delegado_hasta=timezone.now() + timedelta(days=2),
        )
        self.client.force_login(self.delegado)
        session = self.client.session
        session["torneo_id"] = torneo_interbarrios.id
        session.save()

        respuesta = self.client.get("/delegado/equipos/")

        self.assertContains(respuesta, "Torneo activo:</strong> Interbarrios 2026", html=False)
        self.assertContains(respuesta, "Congal Interbarrios")
        self.assertContains(respuesta, "Interbarrios 2026 - Interbarrios")
        self.assertNotContains(respuesta, "Niqueleros")
        self.assertNotContains(respuesta, "Veranero - Senior")

    def test_delegado_no_accede_por_url_a_equipo_de_torneo_no_activo(self):
        torneo_viejo = Torneo.objects.create(
            nombre="Torneo viejo URL",
            fecha_inicio=date(2025, 1, 1),
            estado="FINALIZADO",
        )
        categoria_vieja = Categoria.objects.create(
            nombre="Senior viejo",
            edad_minima=18,
            edad_maxima=60,
            torneo=torneo_viejo,
        )
        equipo_viejo = Equipo.objects.create(
            nombre="Equipo viejo URL",
            categoria=categoria_vieja,
            responsable=self.delegado,
            acceso_delegado_hasta=timezone.now() + timedelta(days=2),
        )
        self.client.force_login(self.delegado)
        session = self.client.session
        session["torneo_id"] = torneo_viejo.id
        session.save()

        respuesta = self.client.get(f"/delegado/equipos/{equipo_viejo.id}/partidos/")

        self.assertEqual(respuesta.status_code, 404)

    def test_delegado_sin_acceso_a_edicion_de_equipo_puede_ver_partidos_de_alineacion(self):
        self.equipo.acceso_delegado_hasta = None
        self.equipo.save(update_fields=["acceso_delegado_hasta"])
        ahora_local = timezone.localtime()
        partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=self.otro_equipo,
            fecha=(ahora_local - timedelta(minutes=5)).date(),
            hora=(ahora_local - timedelta(minutes=5)).time(),
            estado="PROGRAMADO",
        )
        self.client.force_login(self.delegado)

        respuesta_lista = self.client.get(f"/delegado/equipos/{self.equipo.id}/partidos/")
        respuesta_guardar = self.client.post(
            f"/delegado/equipos/{self.equipo.id}/partidos/{partido.id}/alineacion/",
            {
                "cancha_DC": self.jugador.id,
            },
        )

        self.assertContains(respuesta_lista, "Cargar alineacion")
        self.assertEqual(respuesta_guardar.status_code, 302)
        self.assertTrue(
            AlineacionPartido.objects.filter(
                partido=partido,
                equipo=self.equipo,
                jugador=self.jugador,
                rol="TITULAR",
                posicion_cancha="DC",
            ).exists()
        )

    def test_delegado_no_ve_ni_abre_alineacion_de_partido_futuro(self):
        ahora_local = timezone.localtime()
        partido_futuro = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=self.otro_equipo,
            fecha=(ahora_local + timedelta(days=2)).date(),
            hora=(ahora_local + timedelta(days=2)).time(),
            estado="PROGRAMADO",
        )
        self.client.force_login(self.delegado)

        respuesta_lista = self.client.get(f"/delegado/equipos/{self.equipo.id}/partidos/")
        respuesta_directa = self.client.get(
            f"/delegado/equipos/{self.equipo.id}/partidos/{partido_futuro.id}/alineacion/"
        )

        self.assertNotContains(respuesta_lista, "Rival")
        self.assertContains(respuesta_lista, "No hay partidos con la ventana de alineaci")
        self.assertEqual(respuesta_directa.status_code, 403)

    def test_delegado_no_ve_ni_abre_partido_con_programacion_sugerida(self):
        ahora_local = timezone.localtime()
        partido_sugerido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=self.otro_equipo,
            fecha=ahora_local.date(),
            hora=ahora_local.time(),
            estado="PROGRAMADO",
            estado_programacion="SUGERIDA",
        )
        self.client.force_login(self.delegado)

        respuesta_lista = self.client.get(f"/delegado/equipos/{self.equipo.id}/partidos/")
        respuesta_directa = self.client.get(
            f"/delegado/equipos/{self.equipo.id}/partidos/{partido_sugerido.id}/alineacion/"
        )

        self.assertNotContains(respuesta_lista, "Rival")
        self.assertContains(respuesta_lista, "No hay partidos con la ventana de alineaci")
        self.assertEqual(respuesta_directa.status_code, 403)

    def test_delegado_no_ve_partido_programado_con_ventana_vencida(self):
        ahora_local = timezone.localtime()
        Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=self.otro_equipo,
            fecha=(ahora_local - timedelta(hours=1)).date(),
            hora=(ahora_local - timedelta(hours=1)).time(),
            estado="PROGRAMADO",
            estado_programacion="OFICIAL",
        )
        self.client.force_login(self.delegado)

        respuesta = self.client.get(f"/delegado/equipos/{self.equipo.id}/partidos/")

        self.assertNotContains(respuesta, "Cargar alineacion")
        self.assertContains(respuesta, "No hay partidos con la ventana de alineaci")

    def test_responsable_staff_tampoco_salta_ventana_de_partido_futuro(self):
        self.delegado.is_staff = True
        self.delegado.save(update_fields=["is_staff"])
        ahora_local = timezone.localtime()
        partido_futuro = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=self.otro_equipo,
            fecha=(ahora_local + timedelta(days=2)).date(),
            hora=(ahora_local + timedelta(days=2)).time(),
            estado="PROGRAMADO",
        )
        self.client.force_login(self.delegado)

        respuesta = self.client.get(
            f"/delegado/equipos/{self.equipo.id}/partidos/{partido_futuro.id}/alineacion/"
        )

        self.assertIn(respuesta.status_code, (403, 404))

    def test_editar_equipo_bloqueado_redirige_a_partidos_de_alineacion(self):
        self.equipo.acceso_delegado_hasta = None
        self.equipo.save(update_fields=["acceso_delegado_hasta"])
        self.client.force_login(self.delegado)

        respuesta = self.client.get(f"/delegado/equipos/{self.equipo.id}/editar/")

        self.assertRedirects(
            respuesta,
            f"/delegado/equipos/{self.equipo.id}/partidos/",
            fetch_redirect_response=False,
        )

    def test_delegado_puede_agregar_jugador_a_su_equipo(self):
        self.client.force_login(self.delegado)

        respuesta = self.client.post(
            f"/delegado/equipos/{self.equipo.id}/jugadores/nuevo/",
            {
                "dorsal": 10,
                "nombres": "Jugador Nuevo",
                "cedula": "456",
                "fecha_nacimiento": "1995-02-03",
                "telefono": "301000",
                "estado": "ACTIVO",
            },
        )

        self.assertEqual(respuesta.status_code, 302)
        jugador = Jugador.objects.get(equipo=self.equipo, cedula="456", nombres="Jugador Nuevo")
        self.assertTrue(RegistroActividad.objects.filter(
            usuario=self.delegado,
            torneo=self.torneo,
            accion="CREAR_JUGADOR_DELEGADO",
            objeto_id=jugador.id,
        ).exists())

    def test_delegado_con_permiso_solo_fotos_no_edita_datos(self):
        self.equipo.delegado_puede_editar_equipo = False
        self.equipo.delegado_puede_cargar_fotos_jugadores = True
        self.equipo.save(update_fields=["delegado_puede_editar_equipo", "delegado_puede_cargar_fotos_jugadores"])
        self.client.force_login(self.delegado)

        respuesta_lista = self.client.get("/delegado/equipos/")
        respuesta_editar = self.client.get(f"/delegado/equipos/{self.equipo.id}/editar/")
        respuesta_nuevo = self.client.get(f"/delegado/equipos/{self.equipo.id}/jugadores/nuevo/")
        respuesta_fotos = self.client.get(f"/delegado/equipos/{self.equipo.id}/fotos-jugadores/")

        self.assertContains(respuesta_lista, "Fotos jugadores")
        self.assertContains(respuesta_lista, "Edicion de equipo bloqueada")
        self.assertContains(respuesta_fotos, "Tomar foto")
        self.assertContains(respuesta_fotos, "Subir de galeria")
        self.assertContains(respuesta_fotos, 'accept="image/*"')
        self.assertContains(respuesta_fotos, 'capture="environment"')
        self.assertContains(respuesta_fotos, 'data-photo-resize="jugador"')
        self.assertContains(respuesta_fotos, "Cuerpo t&eacute;cnico y Admin App", html=True)
        self.assertContains(respuesta_fotos, 'name="cuerpo-foto_director_tecnico"')
        self.assertContains(respuesta_fotos, 'name="cuerpo-foto_asistente_tecnico"')
        self.assertContains(respuesta_fotos, 'name="cuerpo-foto_delegado"')
        self.assertContains(respuesta_fotos, 'name="cuerpo-foto_administrador_app"')
        self.assertRedirects(
            respuesta_editar,
            f"/delegado/equipos/{self.equipo.id}/fotos-jugadores/",
            fetch_redirect_response=False,
        )
        self.assertEqual(respuesta_nuevo.status_code, 404)

    def test_delegado_con_permiso_solo_fotos_puede_cargar_foto_de_jugador(self):
        self.equipo.delegado_puede_editar_equipo = False
        self.equipo.delegado_puede_cargar_fotos_jugadores = True
        self.equipo.save(update_fields=["delegado_puede_editar_equipo", "delegado_puede_cargar_fotos_jugadores"])
        self.client.force_login(self.delegado)
        imagen_bytes = BytesIO()
        Image.new("RGB", (8, 8), "#00ff66").save(imagen_bytes, format="PNG")
        imagen_bytes.seek(0)
        archivo = SimpleUploadedFile("jugador.png", imagen_bytes.read(), content_type="image/png")

        respuesta = self.client.post(
            f"/delegado/equipos/{self.equipo.id}/fotos-jugadores/",
            {f"jugador_{self.jugador.id}-foto": archivo},
        )

        self.assertRedirects(
            respuesta,
            f"/delegado/equipos/{self.equipo.id}/fotos-jugadores/",
            fetch_redirect_response=False,
        )
        self.jugador.refresh_from_db()
        self.assertTrue(self.jugador.foto.name)
        self.assertEqual(self.jugador.nombres, "Jugador Uno")
        registro = RegistroActividad.objects.get(
            usuario=self.delegado,
            torneo=self.torneo,
            accion="CARGAR_FOTOS_JUGADORES",
            objeto_id=self.equipo.id,
        )
        self.assertEqual(registro.datos["cantidad"], 1)

    def test_delegado_con_permiso_solo_fotos_puede_cargar_foto_cuerpo_tecnico(self):
        self.equipo.delegado_puede_editar_equipo = False
        self.equipo.delegado_puede_cargar_fotos_jugadores = True
        self.equipo.director_tecnico = "Director Uno"
        self.equipo.save(update_fields=["delegado_puede_editar_equipo", "delegado_puede_cargar_fotos_jugadores", "director_tecnico"])
        self.client.force_login(self.delegado)
        imagen_bytes = BytesIO()
        Image.new("RGB", (8, 8), "#00ff66").save(imagen_bytes, format="PNG")
        imagen_bytes.seek(0)
        archivo = SimpleUploadedFile("director.png", imagen_bytes.read(), content_type="image/png")

        respuesta = self.client.post(
            f"/delegado/equipos/{self.equipo.id}/fotos-jugadores/",
            {"cuerpo-foto_director_tecnico": archivo},
        )

        self.assertRedirects(
            respuesta,
            f"/delegado/equipos/{self.equipo.id}/fotos-jugadores/",
            fetch_redirect_response=False,
        )
        self.equipo.refresh_from_db()
        self.assertTrue(self.equipo.foto_director_tecnico.name)
        self.assertEqual(self.equipo.director_tecnico, "Director Uno")
        self.assertTrue(RegistroActividad.objects.filter(
            usuario=self.delegado,
            accion="CARGAR_FOTOS_JUGADORES",
            datos__cuerpo_tecnico=True,
        ).exists())

    def test_delegado_sin_permiso_de_fotos_no_abre_carga_de_fotos(self):
        self.equipo.delegado_puede_cargar_fotos_jugadores = False
        self.equipo.save(update_fields=["delegado_puede_cargar_fotos_jugadores"])
        self.client.force_login(self.delegado)

        respuesta = self.client.get(f"/delegado/equipos/{self.equipo.id}/fotos-jugadores/")

        self.assertEqual(respuesta.status_code, 403)

    def test_delegado_no_puede_editar_otro_equipo(self):
        self.client.force_login(self.delegado)

        respuesta = self.client.get(f"/delegado/equipos/{self.otro_equipo.id}/editar/")

        self.assertEqual(respuesta.status_code, 404)

    def test_delegado_pierde_acceso_al_vencer_plazo(self):
        self.equipo.acceso_delegado_hasta = timezone.now() - timedelta(minutes=1)
        self.equipo.save(update_fields=["acceso_delegado_hasta"])
        self.client.force_login(self.delegado)

        respuesta = self.client.get(f"/delegado/equipos/{self.equipo.id}/editar/")

        self.assertRedirects(
            respuesta,
            f"/delegado/equipos/{self.equipo.id}/partidos/",
            fetch_redirect_response=False,
        )

    def test_delegado_puede_guardar_alineacion_desde_hora_programada(self):
        ahora_local = timezone.localtime()
        partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=self.otro_equipo,
            fecha=(ahora_local - timedelta(hours=1)).date(),
            hora=(ahora_local - timedelta(hours=1)).time(),
            estado="PROGRAMADO",
        )
        self.client.force_login(self.delegado)

        respuesta = self.client.post(
            f"/delegado/equipos/{self.equipo.id}/partidos/{partido.id}/alineacion/",
            {
                f"rol_{self.jugador.id}": "TITULAR",
                f"posicion_{self.jugador.id}": "DC",
            },
        )

        self.assertEqual(respuesta.status_code, 302)
        self.assertTrue(
            AlineacionPartido.objects.filter(
                partido=partido,
                equipo=self.equipo,
                jugador=self.jugador,
                rol="TITULAR",
                posicion_cancha="DC",
            ).exists()
        )

    def test_delegado_asigna_posicion_automatica_si_marca_titular_desde_lista(self):
        ahora_local = timezone.localtime()
        partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=self.otro_equipo,
            fecha=(ahora_local - timedelta(hours=1)).date(),
            hora=(ahora_local - timedelta(hours=1)).time(),
            estado="PROGRAMADO",
        )
        self.client.force_login(self.delegado)

        respuesta = self.client.post(
            f"/delegado/equipos/{self.equipo.id}/partidos/{partido.id}/alineacion/",
            {
                f"rol_{self.jugador.id}": "TITULAR",
            },
        )

        self.assertEqual(respuesta.status_code, 302)
        alineacion = AlineacionPartido.objects.get(partido=partido, equipo=self.equipo, jugador=self.jugador)
        self.assertEqual(alineacion.rol, "TITULAR")
        self.assertIn(alineacion.posicion_cancha, {codigo for codigo, _ in AlineacionPartido.POSICIONES_CANCHA})

    def test_delegado_guarda_titular_desde_cancha_y_suplente_desde_banco(self):
        suplente = Jugador.objects.create(
            equipo=self.equipo,
            dorsal=8,
            nombres="Jugador Suplente",
            cedula="789",
            fecha_nacimiento=date(1991, 1, 1),
        )
        ahora_local = timezone.localtime()
        partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=self.otro_equipo,
            fecha=(ahora_local - timedelta(hours=1)).date(),
            hora=(ahora_local - timedelta(hours=1)).time(),
            estado="PROGRAMADO",
        )
        self.client.force_login(self.delegado)

        respuesta = self.client.post(
            f"/delegado/equipos/{self.equipo.id}/partidos/{partido.id}/alineacion/",
            {
                "cancha_DC": self.jugador.id,
                f"rol_{suplente.id}": "SUPLENTE",
            },
        )

        self.assertEqual(respuesta.status_code, 302)
        self.assertTrue(
            AlineacionPartido.objects.filter(
                partido=partido,
                equipo=self.equipo,
                jugador=self.jugador,
                rol="TITULAR",
                posicion_cancha="DC",
            ).exists()
        )
        self.assertTrue(
            AlineacionPartido.objects.filter(
                partido=partido,
                equipo=self.equipo,
                jugador=suplente,
                rol="SUPLENTE",
            ).exists()
        )

    def test_delegado_ve_edad_en_editor_de_alineacion(self):
        ahora_local = timezone.localtime()
        partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=self.otro_equipo,
            fecha=(ahora_local - timedelta(hours=1)).date(),
            hora=(ahora_local - timedelta(hours=1)).time(),
            estado="PROGRAMADO",
        )
        self.client.force_login(self.delegado)

        respuesta = self.client.get(f"/delegado/equipos/{self.equipo.id}/partidos/{partido.id}/alineacion/")

        self.assertContains(respuesta, texto_edad_jugador(self.jugador, self.categoria, partido.fecha))
        self.assertContains(respuesta, 'data-slot-player-age')
        self.assertContains(respuesta, f'data-edad="{texto_edad_jugador(self.jugador, self.categoria, partido.fecha)}"')
        self.assertContains(respuesta, f'data-etiqueta-edad="{etiqueta_edad_jugador(self.jugador, self.categoria, partido.fecha)}"')

    def test_delegado_ve_titular_marcado_en_lista_de_estados(self):
        ahora_local = timezone.localtime()
        partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=self.otro_equipo,
            fecha=(ahora_local - timedelta(hours=1)).date(),
            hora=(ahora_local - timedelta(hours=1)).time(),
            estado="PROGRAMADO",
        )
        AlineacionPartido.objects.create(
            partido=partido,
            equipo=self.equipo,
            jugador=self.jugador,
            rol="TITULAR",
            posicion_cancha="DC",
        )
        self.client.force_login(self.delegado)

        respuesta = self.client.get(f"/delegado/equipos/{self.equipo.id}/partidos/{partido.id}/alineacion/")

        self.assertContains(respuesta, 'value="TITULAR" data-titular-banco checked')
        self.assertNotContains(
            respuesta,
            f'name="rol_{self.jugador.id}" value="" checked',
        )

    def test_delegado_no_puede_guardar_alineacion_antes_de_hora_programada(self):
        ahora_local = timezone.localtime()
        partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=self.otro_equipo,
            fecha=(ahora_local + timedelta(hours=1, minutes=5)).date(),
            hora=(ahora_local + timedelta(hours=1, minutes=5)).time(),
            estado="PROGRAMADO",
        )
        self.client.force_login(self.delegado)

        respuesta = self.client.post(
            f"/delegado/equipos/{self.equipo.id}/partidos/{partido.id}/alineacion/",
            {
                f"rol_{self.jugador.id}": "TITULAR",
                f"posicion_{self.jugador.id}": "DC",
            },
        )

        self.assertEqual(respuesta.status_code, 403)
        self.assertFalse(AlineacionPartido.objects.filter(partido=partido).exists())

    def test_delegado_no_puede_guardar_alineacion_despues_de_quince_minutos_en_juego(self):
        ahora_local = timezone.localtime()
        partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=self.otro_equipo,
            fecha=ahora_local.date(),
            hora=(ahora_local - timedelta(hours=1)).time(),
            estado="EN_JUEGO",
            inicio_en_vivo=timezone.now() - timedelta(minutes=16),
        )
        self.client.force_login(self.delegado)

        respuesta = self.client.post(
            f"/delegado/equipos/{self.equipo.id}/partidos/{partido.id}/alineacion/",
            {
                f"rol_{self.jugador.id}": "TITULAR",
                f"posicion_{self.jugador.id}": "DC",
            },
        )

        self.assertEqual(respuesta.status_code, 403)
        self.assertFalse(AlineacionPartido.objects.filter(partido=partido).exists())

    def test_delegado_puede_abrir_alineacion_una_hora_antes(self):
        ahora_local = timezone.localtime()
        partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=self.otro_equipo,
            fecha=(ahora_local + timedelta(minutes=59)).date(),
            hora=(ahora_local + timedelta(minutes=59)).time(),
            estado="PROGRAMADO",
        )
        self.client.force_login(self.delegado)

        respuesta = self.client.get(
            f"/delegado/equipos/{self.equipo.id}/partidos/{partido.id}/alineacion/",
        )

        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, "Guardar borrador")
        self.assertContains(respuesta, "Enviar alineación definitiva")

    def test_borrador_persiste_y_no_cierra_la_alineacion(self):
        ahora_local = timezone.localtime()
        partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=self.otro_equipo,
            fecha=ahora_local.date(),
            hora=ahora_local.time(),
            estado="PROGRAMADO",
        )
        self.client.force_login(self.delegado)

        respuesta = self.client.post(
            f"/delegado/equipos/{self.equipo.id}/partidos/{partido.id}/alineacion/",
            {
                "accion": "guardar_borrador",
                f"rol_{self.jugador.id}": "TITULAR",
                f"posicion_{self.jugador.id}": "DC",
            },
        )

        self.assertRedirects(
            respuesta,
            f"/delegado/equipos/{self.equipo.id}/partidos/{partido.id}/alineacion/",
            fetch_redirect_response=False,
        )
        self.assertTrue(AlineacionPartido.objects.filter(partido=partido, jugador=self.jugador).exists())
        self.assertFalse(EntregaAlineacionPartido.objects.filter(partido=partido, equipo=self.equipo).exists())
        registro = RegistroActividad.objects.get(
            usuario=self.delegado,
            accion="GUARDAR_BORRADOR_ALINEACION",
            objeto_id=partido.id,
        )
        self.assertEqual(registro.torneo, self.torneo)
        self.assertEqual(registro.datos["equipo_id"], self.equipo.id)
        self.assertEqual(registro.datos["titulares"], 1)
        self.assertEqual(
            self.client.get(f"/delegado/equipos/{self.equipo.id}/partidos/{partido.id}/alineacion/").status_code,
            200,
        )

    def test_delegado_puede_actualizar_dorsal_desde_borrador_alineacion(self):
        ahora_local = timezone.localtime()
        partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=self.otro_equipo,
            fecha=ahora_local.date(),
            hora=ahora_local.time(),
            estado="PROGRAMADO",
        )
        self.client.force_login(self.delegado)

        respuesta = self.client.post(
            f"/delegado/equipos/{self.equipo.id}/partidos/{partido.id}/alineacion/",
            {
                "accion": "guardar_borrador",
                f"rol_{self.jugador.id}": "SUPLENTE",
                f"dorsal_{self.jugador.id}": "18",
            },
        )

        self.assertEqual(respuesta.status_code, 302)
        self.jugador.refresh_from_db()
        self.assertEqual(self.jugador.dorsal, 18)
        self.assertTrue(RegistroActividad.objects.filter(
            usuario=self.delegado,
            accion="ACTUALIZAR_DORSALES_ALINEACION",
            objeto_id=self.equipo.id,
        ).exists())

    def test_envio_definitivo_cierra_inmediatamente_acceso_del_delegado(self):
        ahora_local = timezone.localtime()
        partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=self.otro_equipo,
            fecha=ahora_local.date(),
            hora=ahora_local.time(),
            estado="PROGRAMADO",
        )
        self.client.force_login(self.delegado)

        respuesta = self.client.post(
            f"/delegado/equipos/{self.equipo.id}/partidos/{partido.id}/alineacion/",
            {
                "accion": "enviar_definitiva",
                f"rol_{self.jugador.id}": "TITULAR",
                f"posicion_{self.jugador.id}": "DC",
            },
        )

        self.assertRedirects(
            respuesta,
            f"/delegado/equipos/{self.equipo.id}/partidos/",
            fetch_redirect_response=False,
        )
        entrega = EntregaAlineacionPartido.objects.get(partido=partido, equipo=self.equipo)
        self.assertEqual(entrega.enviada_por, self.delegado)
        bloqueado = self.client.get(
            f"/delegado/equipos/{self.equipo.id}/partidos/{partido.id}/alineacion/",
        )
        self.assertEqual(bloqueado.status_code, 403)
        self.assertContains(bloqueado, "definitiva ya fue enviada", status_code=403)
        listado = self.client.get(f"/delegado/equipos/{self.equipo.id}/partidos/")
        self.assertNotContains(
            listado,
            f"/delegado/equipos/{self.equipo.id}/partidos/{partido.id}/alineacion/",
        )

    def test_delegado_asignado_como_planillero_no_abre_editor_completo(self):
        ahora_local = timezone.localtime()
        partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=self.otro_equipo,
            fecha=(ahora_local - timedelta(hours=1)).date(),
            hora=(ahora_local - timedelta(hours=1)).time(),
            estado="PROGRAMADO",
        )
        partido.planilleros.add(self.delegado)
        self.client.force_login(self.delegado)

        respuesta = self.client.get(f"/partido/{partido.id}/editor-movil/")

        self.assertRedirects(
            respuesta,
            f"/delegado/equipos/{self.equipo.id}/partidos/{partido.id}/alineacion/",
            fetch_redirect_response=False,
        )

    @override_settings(STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    })
    def test_partido_en_vivo_muestra_solo_boton_de_alineacion_a_delegado(self):
        ahora_local = timezone.localtime()
        partido = Partido.objects.create(
            categoria=self.categoria,
            equipo_local=self.equipo,
            equipo_visitante=self.otro_equipo,
            fecha=(ahora_local - timedelta(hours=1)).date(),
            hora=(ahora_local - timedelta(hours=1)).time(),
            estado="PROGRAMADO",
        )
        partido.planilleros.add(self.delegado)
        self.client.force_login(self.delegado)

        respuesta = self.client.get(f"/partido/{partido.id}/live/")

        self.assertContains(respuesta, "Cargar alineaci")
        self.assertNotContains(respuesta, "Goles <span>")
        self.assertNotContains(respuesta, "Tarjetas <span>")


class JugadorCedulaPorEquipoTests(TransactionTestCase):
    def test_permite_misma_cedula_en_categorias_distintas(self):
        torneo = Torneo.objects.create(
            nombre="Veranero",
            fecha_inicio=date(2026, 1, 1),
        )
        senior = Categoria.objects.create(
            nombre="Senior Master",
            edad_minima=18,
            edad_maxima=60,
            torneo=torneo,
        )
        plus = Categoria.objects.create(
            nombre="Plus 50",
            edad_minima=50,
            edad_maxima=80,
            torneo=torneo,
        )
        equipo_senior = Equipo.objects.create(nombre="Niqueleros", categoria=senior)
        equipo_plus = Equipo.objects.create(nombre="Niqueleros", categoria=plus)

        Jugador.objects.create(
            equipo=equipo_senior,
            dorsal=7,
            nombres="Jugador Compartido",
            cedula="12345",
            fecha_nacimiento=date(1970, 1, 1),
        )
        Jugador.objects.create(
            equipo=equipo_plus,
            dorsal=7,
            nombres="Jugador Compartido",
            cedula="12345",
            fecha_nacimiento=date(1970, 1, 1),
        )

        self.assertEqual(Jugador.objects.filter(cedula="12345").count(), 2)

    def test_no_permite_misma_cedula_dos_veces_en_el_mismo_equipo(self):
        torneo = Torneo.objects.create(
            nombre="Veranero",
            fecha_inicio=date(2026, 1, 1),
        )
        categoria = Categoria.objects.create(
            nombre="Senior Master",
            edad_minima=18,
            edad_maxima=60,
            torneo=torneo,
        )
        equipo = Equipo.objects.create(nombre="Niqueleros", categoria=categoria)

        Jugador.objects.create(
            equipo=equipo,
            dorsal=7,
            nombres="Jugador Uno",
            cedula="12345",
            fecha_nacimiento=date(1970, 1, 1),
        )

        with self.assertRaises(ValidationError):
            Jugador.objects.create(
                equipo=equipo,
                dorsal=8,
                nombres="Jugador Dos",
                cedula="12345",
                fecha_nacimiento=date(1975, 1, 1),
            )

    def test_no_permite_misma_cedula_en_otro_equipo_de_la_misma_categoria(self):
        torneo = Torneo.objects.create(
            nombre="Veranero",
            fecha_inicio=date(2026, 1, 1),
        )
        categoria = Categoria.objects.create(
            nombre="Senior Master",
            edad_minima=18,
            edad_maxima=60,
            torneo=torneo,
        )
        niqueleros = Equipo.objects.create(nombre="Niqueleros", categoria=categoria)
        otro_equipo = Equipo.objects.create(nombre="Otro Equipo", categoria=categoria)

        Jugador.objects.create(
            equipo=niqueleros,
            dorsal=7,
            nombres="Jugador Uno",
            cedula="12345",
            fecha_nacimiento=date(1970, 1, 1),
        )

        with self.assertRaises(ValidationError):
            Jugador.objects.create(
                equipo=otro_equipo,
                dorsal=8,
                nombres="Jugador Dos",
                cedula="12345",
                fecha_nacimiento=date(1975, 1, 1),
            )


class AsignacionMultiplePlanilleroTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser(
            username="admin-multiple", password="clave-segura", email="admin@example.com"
        )
        self.planillero = User.objects.create_user(
            username="planillero-multiple", password="clave-segura",
            first_name="Ana", last_name="Planillera",
        )
        self.planillero_previo = User.objects.create_user(
            username="planillero-previo", password="clave-segura"
        )
        self.torneo = Torneo.objects.create(nombre="Torneo activo", fecha_inicio=date(2026, 1, 1))
        self.categoria = Categoria.objects.create(
            nombre="Única", torneo=self.torneo, edad_minima=18, edad_maxima=80
        )
        local = Equipo.objects.create(nombre="Local", categoria=self.categoria)
        visitante = Equipo.objects.create(nombre="Visitante", categoria=self.categoria)
        self.partido_uno = Partido.objects.create(
            categoria=self.categoria, equipo_local=local, equipo_visitante=visitante,
            fecha=date(2026, 1, 10), hora=time(16),
        )
        self.partido_dos = Partido.objects.create(
            categoria=self.categoria, equipo_local=visitante, equipo_visitante=local,
            fecha=date(2026, 1, 17), hora=time(16),
        )
        self.partido_uno.planilleros.add(self.planillero_previo)

        otro_torneo = Torneo.objects.create(nombre="Otro torneo", fecha_inicio=date(2026, 2, 1))
        otra_categoria = Categoria.objects.create(
            nombre="Otra", torneo=otro_torneo, edad_minima=18, edad_maxima=80
        )
        otro_local = Equipo.objects.create(nombre="Otro local", categoria=otra_categoria)
        otro_visitante = Equipo.objects.create(nombre="Otro visitante", categoria=otra_categoria)
        self.partido_otro_torneo = Partido.objects.create(
            categoria=otra_categoria, equipo_local=otro_local, equipo_visitante=otro_visitante,
            fecha=date(2026, 2, 10), hora=time(16),
        )

        self.client.force_login(self.admin)
        sesion = self.client.session
        sesion["torneo_id"] = self.torneo.id
        sesion.save()

    def test_muestra_control_de_asignacion_multiple(self):
        respuesta = self.client.get("/gestion/partidos/")

        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, "Asignar planillero a varios partidos")
        self.assertContains(respuesta, self.planillero.username)
        self.assertContains(
            respuesta,
            'type="checkbox" name="partidos"',
            count=2,
        )

    def test_asigna_varios_partidos_sin_borrar_asignaciones_previas(self):
        respuesta = self.client.post(
            "/gestion/partidos/asignar-planillero/",
            {
                "planillero": self.planillero.id,
                "partidos": [
                    self.partido_uno.id,
                    self.partido_dos.id,
                    self.partido_otro_torneo.id,
                ],
            },
        )

        self.assertEqual(respuesta.status_code, 302)
        self.assertTrue(self.partido_uno.planilleros.filter(id=self.planillero.id).exists())
        self.assertTrue(self.partido_dos.planilleros.filter(id=self.planillero.id).exists())
        self.assertTrue(self.partido_uno.planilleros.filter(id=self.planillero_previo.id).exists())
        self.assertFalse(
            self.partido_otro_torneo.planilleros.filter(id=self.planillero.id).exists()
        )


class ControlReemplazosJugadoresTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser("admin-reemplazos", "admin@example.com", "clave-segura")
        self.torneo = Torneo.objects.create(nombre="Torneo reemplazos", fecha_inicio=date(2026, 1, 1))
        self.categoria = Categoria.objects.create(
            nombre="Única", torneo=self.torneo, edad_minima=18, edad_maxima=80,
            controlar_reemplazos_jugadores=True,
        )
        self.equipo_a = Equipo.objects.create(nombre="Equipo A", categoria=self.categoria)
        self.equipo_b = Equipo.objects.create(nombre="Equipo B", categoria=self.categoria)
        self.equipo_c = Equipo.objects.create(nombre="Equipo C", categoria=self.categoria)
        self.jugador = Jugador.objects.create(
            equipo=self.equipo_a, nombres="Jugador Saliente", cedula="1001",
            fecha_nacimiento=date(1990, 1, 1),
        )
        self.client.force_login(self.admin)
        sesion = self.client.session
        sesion["torneo_id"] = self.torneo.id
        sesion.save()

    def datos_nuevo(self, cedula="2001"):
        return {
            "nombres": "Jugador Nuevo", "cedula": cedula,
            "fecha_nacimiento": "1992-02-02", "dorsal": "9",
        }

    def test_bloqueo_de_fecha_tres_es_individual_por_equipo(self):
        Partido.objects.create(
            categoria=self.categoria, equipo_local=self.equipo_a, equipo_visitante=self.equipo_b,
            numero_fecha="Fecha 3", fase="GRUPOS", fecha=date(2026, 1, 20), hora=time(16), estado="FINALIZADO",
        )

        self.assertTrue(tercera_fecha_iniciada(self.equipo_a))
        self.assertTrue(tercera_fecha_iniciada(self.equipo_b))
        self.assertFalse(tercera_fecha_iniciada(self.equipo_c))

    def test_antes_de_fecha_tres_reemplaza_jugador_que_no_piso_cancha(self):
        respuesta = self.client.post(
            f"/gestion/jugadores/{self.jugador.id}/reemplazar/", self.datos_nuevo()
        )

        self.assertEqual(respuesta.status_code, 302)
        self.jugador.refresh_from_db()
        self.assertEqual(self.jugador.estado, "RETIRADO")
        self.assertTrue(Jugador.objects.filter(equipo=self.equipo_a, cedula="2001", estado="ACTIVO").exists())

    def test_jugador_que_piso_cancha_exige_fuerza_mayor_y_soporte(self):
        partido = Partido.objects.create(
            categoria=self.categoria, equipo_local=self.equipo_a, equipo_visitante=self.equipo_b,
            numero_fecha="Fecha 1", fase="GRUPOS", fecha=date(2026, 1, 5), hora=time(16), estado="FINALIZADO",
        )
        AlineacionPartido.objects.create(partido=partido, equipo=self.equipo_a, jugador=self.jugador, rol="TITULAR")
        self.assertTrue(politica_reemplazo_jugador(self.jugador)["requiere_fuerza_mayor"])

        respuesta = self.client.post(
            f"/gestion/jugadores/{self.jugador.id}/reemplazar/", self.datos_nuevo()
        )
        self.assertEqual(respuesta.status_code, 200)
        self.jugador.refresh_from_db()
        self.assertEqual(self.jugador.estado, "ACTIVO")

        datos = self.datos_nuevo()
        datos.update({
            "motivo": "LESION", "justificacion": "Lesión certificada por el médico.",
            "soporte": SimpleUploadedFile("soporte.pdf", b"%PDF-1.4 soporte", content_type="application/pdf"),
        })
        respuesta = self.client.post(f"/gestion/jugadores/{self.jugador.id}/reemplazar/", datos)
        self.assertEqual(respuesta.status_code, 302)
        self.jugador.refresh_from_db()
        self.assertEqual(self.jugador.estado, "RETIRADO")
