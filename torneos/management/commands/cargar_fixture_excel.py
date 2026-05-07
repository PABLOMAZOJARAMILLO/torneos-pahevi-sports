from django.core.management.base import BaseCommand
from torneos.models import Categoria, Equipo, Partido
from openpyxl import load_workbook
from datetime import datetime, time


class Command(BaseCommand):
    help = 'Carga partidos del fixture desde Excel'

    def add_arguments(self, parser):
        parser.add_argument('archivo_excel', type=str)

    def handle(self, *args, **options):
        archivo_excel = options['archivo_excel']

        wb = load_workbook(archivo_excel, data_only=True)
        ws = wb.active

        creados = 0
        omitidos = 0

        for fila in range(2, ws.max_row + 1):
            categoria_nombre = ws[f'A{fila}'].value
            local_nombre = ws[f'B{fila}'].value
            visitante_nombre = ws[f'C{fila}'].value
            fecha_valor = ws[f'D{fila}'].value
            hora_valor = ws[f'E{fila}'].value
            cancha = ws[f'F{fila}'].value

            if not categoria_nombre or not local_nombre or not visitante_nombre:
                continue

            try:
                categoria = Categoria.objects.get(nombre__iexact=str(categoria_nombre).strip())
                local = Equipo.objects.get(nombre__iexact=str(local_nombre).strip(), categoria=categoria)
                visitante = Equipo.objects.get(nombre__iexact=str(visitante_nombre).strip(), categoria=categoria)
            except Exception:
                self.stdout.write(self.style.WARNING(
                    f'Fila {fila}: categoría o equipo no encontrado. Omitido.'
                ))
                omitidos += 1
                continue

            try:
                if isinstance(fecha_valor, datetime):
                    fecha = fecha_valor.date()
                else:
                    fecha = datetime.strptime(str(fecha_valor), '%Y-%m-%d').date()

                if isinstance(hora_valor, time):
                    hora = hora_valor
                elif isinstance(hora_valor, datetime):
                    hora = hora_valor.time()
                else:
                    hora = datetime.strptime(str(hora_valor), '%H:%M').time()
            except Exception:
                self.stdout.write(self.style.WARNING(
                    f'Fila {fila}: fecha u hora inválida. Omitido.'
                ))
                omitidos += 1
                continue

            existe = Partido.objects.filter(
                categoria=categoria,
                equipo_local=local,
                equipo_visitante=visitante,
                fecha=fecha,
                hora=hora
            ).exists()

            if existe:
                self.stdout.write(self.style.WARNING(
                    f'Fila {fila}: partido ya existe. Omitido.'
                ))
                omitidos += 1
                continue

            Partido.objects.create(
                categoria=categoria,
                equipo_local=local,
                equipo_visitante=visitante,
                fecha=fecha,
                hora=hora,
                estado='PROGRAMADO',
                observaciones=f'Cancha: {cancha}' if cancha else ''
            )

            creados += 1

        self.stdout.write(self.style.SUCCESS(
            f'Carga finalizada. Partidos creados: {creados}. Omitidos: {omitidos}.'
        ))