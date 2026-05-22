from django.core.management.base import BaseCommand
from torneos.models import Equipo, Jugador
from openpyxl import load_workbook
from datetime import date


class Command(BaseCommand):
    help = 'Carga jugadores desde la planilla oficial de inscripción'

    def add_arguments(self, parser):
        parser.add_argument('equipo_id', type=int)
        parser.add_argument('archivo_excel', type=str)

    def handle(self, *args, **options):
        equipo_id = options['equipo_id']
        archivo_excel = options['archivo_excel']

        equipo = Equipo.objects.get(id=equipo_id)

        wb = load_workbook(archivo_excel, data_only=True)
        ws = wb.active

        jugadores_creados = 0
        jugadores_omitidos = 0

        for fila in range(8, ws.max_row + 1):
            nombre = ws[f'C{fila}'].value
            dorsal = ws[f'D{fila}'].value
            dia = ws[f'E{fila}'].value
            mes = ws[f'F{fila}'].value
            anio = ws[f'G{fila}'].value
            cedula = ws[f'H{fila}'].value

            if not nombre or not cedula:
                continue

            try:
                fecha_nacimiento = date(int(anio), int(mes), int(dia))
            except Exception:
                self.stdout.write(self.style.WARNING(
                    f'Fila {fila}: fecha inválida, jugador omitido'
                ))
                jugadores_omitidos += 1
                continue

            cedula = str(cedula).strip()
            nombre = str(nombre).strip()

            if Jugador.objects.filter(cedula=cedula).exists():
                self.stdout.write(self.style.WARNING(
                    f'Omitido: {nombre} ya existe con cédula {cedula}'
                ))
                jugadores_omitidos += 1
                continue

            Jugador.objects.create(
                equipo=equipo,
                dorsal=dorsal,
                nombres=nombre,
                cedula=cedula,
                fecha_nacimiento=fecha_nacimiento,
                estado='ACTIVO'
            )

            jugadores_creados += 1

        self.stdout.write(self.style.SUCCESS(
            f'Carga finalizada. Jugadores creados: {jugadores_creados}. Omitidos: {jugadores_omitidos}.'
        ))
