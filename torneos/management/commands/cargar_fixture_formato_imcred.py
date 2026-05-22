from django.core.management.base import BaseCommand
from torneos.models import Categoria, Equipo, Partido
from openpyxl import load_workbook
from datetime import date, time


class Command(BaseCommand):
    help = 'Carga fixture desde el formato de torneo'

    def add_arguments(self, parser):
        parser.add_argument('archivo_excel', type=str)
        parser.add_argument('grupo', type=str)

    def handle(self, *args, **options):
        archivo_excel = options['archivo_excel']
        grupo = options['grupo'].upper()

        wb = load_workbook(archivo_excel, data_only=True)
        ws = wb.active

        categoria_actual = None
        numero_fecha_actual = None

        creados = 0
        omitidos = 0

        for fila in range(1, ws.max_row + 1):
            valor_a = ws[f'A{fila}'].value
            valor_b = ws[f'B{fila}'].value
            valor_c = ws[f'C{fila}'].value
            valor_d = ws[f'D{fila}'].value

            valores_fila = [valor_a, valor_b, valor_c, valor_d]

            if valor_a and str(valor_a).strip().upper() not in ['CATEGORIA']:
                categoria_actual = str(valor_a).strip()

            encontro_fecha = False
            for valor in valores_fila:
                if valor and 'FECHA' in str(valor).upper():
                    numero_fecha_actual = str(valor).strip()
                    encontro_fecha = True
                    break

            if encontro_fecha:
                continue

            if not valor_c or str(valor_c).strip().upper() != 'VS':
                continue

            if not valor_b or not valor_d or not categoria_actual:
                continue

            try:
                categoria = Categoria.objects.get(nombre__iexact=categoria_actual)
                local = Equipo.objects.get(nombre__iexact=str(valor_b).strip(), categoria=categoria)
                visitante = Equipo.objects.get(nombre__iexact=str(valor_d).strip(), categoria=categoria)
            except Exception:
                self.stdout.write(self.style.WARNING(
                    f'Fila {fila}: no encontré categoría/equipo: {categoria_actual} | {valor_b} vs {valor_d}'
                ))
                omitidos += 1
                continue

            existe = Partido.objects.filter(
                categoria=categoria,
                equipo_local=local,
                equipo_visitante=visitante,
                numero_fecha=numero_fecha_actual,
                grupo=grupo
            ).exists()

            if existe:
                omitidos += 1
                continue

            Partido.objects.create(
                categoria=categoria,
                equipo_local=local,
                equipo_visitante=visitante,
                numero_fecha=numero_fecha_actual,
                grupo=grupo,
                fecha=date.today(),
                hora=time(0, 0),
                estado='PROGRAMADO'
            )

            creados += 1

        self.stdout.write(self.style.SUCCESS(
            f'Fixture cargado. Grupo {grupo}. Partidos creados: {creados}. Omitidos: {omitidos}.'
        ))
