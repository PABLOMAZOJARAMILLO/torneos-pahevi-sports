from collections import defaultdict
import base64
import csv
from io import BytesIO
from types import SimpleNamespace
from datetime import date, datetime, time, timedelta
from concurrent.futures import ThreadPoolExecutor, TimeoutError
import os
import random
import re
import uuid
import zipfile
from urllib.parse import quote, urlencode

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import login as auth_login, logout, update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.views import LoginView
from django.http import FileResponse, HttpResponse, HttpResponseForbidden
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import connection, transaction
from django.db.models import Count, Prefetch, Q, Sum
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.staticfiles import finders
from django.template.loader import render_to_string
from django.templatetags.static import static
from django.utils.html import escape
from django.utils.http import url_has_allowed_host_and_scheme
from django.utils.dateparse import parse_date, parse_datetime
from django.utils.text import slugify
from html2image import Html2Image
import requests
from django.views.decorators.http import require_POST
from openpyxl import load_workbook

from .forms import TorneoForm, OrganizadorForm, CategoriaForm, ReglaEdadCategoriaFormSet, DocumentoForm, PlanillaJuegoUploadForm, EquipoForm, EquipoDelegadoForm, EquipoReinscripcionForm, JugadorForm, JugadorDelegadoForm, JugadorFotoDelegadoForm, PartidoForm, PartidoProgramacionForm, AdminTorneoForm, AdminOrganizadorForm, CrearAdminOrganizadorForm
from .models import Torneo, Organizador, Categoria, Documento, Equipo, Partido, Gol, Tarjeta, Jugador, AlineacionPartido, EntregaAlineacionPartido, SustitucionPartido, IncidenciaReglaEdad, ReglaEdadCategoria, AdminTorneo, AdminOrganizador, RegistroActividad, VisitaPublicaDiaria, SolicitudValidacion, limpiar_ruta_cloudinary
from .planillas_pdf import generar_planilla_juego_pdf, nombre_archivo_planilla
from django.utils import timezone

def puede_gestionar_organizadores(user):
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    if not user.is_staff:
        return False
    if user.has_perm("torneos.add_organizador") or user.has_perm("torneos.change_organizador"):
        return True

    tiene_torneos = tabla_disponible("torneos_admintorneo") and AdminTorneo.objects.filter(usuario=user, activo=True).exists()
    tiene_organizadores = (
        tabla_disponible("torneos_adminorganizador")
        and AdminOrganizador.objects.filter(usuario=user, activo=True).exists()
    )
    return not tiene_torneos and not tiene_organizadores


def es_editor_torneo(user):
    if not user.is_authenticated:
        return False
    if user.is_superuser or puede_gestionar_organizadores(user):
        return True
    if not tabla_disponible("torneos_admintorneo"):
        return user.is_staff
    tiene_torneos = AdminTorneo.objects.filter(usuario=user, activo=True).exists()
    tiene_organizadores = (
        tabla_disponible("torneos_adminorganizador")
        and AdminOrganizador.objects.filter(usuario=user, activo=True).exists()
    )
    return tiene_torneos or tiene_organizadores


def es_superadmin(user):
    return user.is_authenticated and user.is_superuser


def puede_descargar_programacion(user):
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    return (
        tabla_disponible("torneos_adminorganizador")
        and AdminOrganizador.objects.filter(usuario=user, activo=True).exists()
    )


def puede_diligenciar_partido(user, partido):
    if user.is_authenticated and user.is_superuser:
        return True
    if user.is_authenticated and user.is_staff and partido:
        if not tabla_disponible("torneos_admintorneo"):
            return True
        return usuario_puede_editar_torneo(user, partido.categoria.torneo if partido.categoria_id else None)
    if not user.is_authenticated or not partido:
        return False
    if partido.estado == "FINALIZADO":
        return False
    return partido.planilleros.filter(id=user.id).exists()


def puede_cargar_planillas_juego(user):
    if not user.is_authenticated:
        return False
    if es_editor_torneo(user):
        return True
    return user.partidos_planillero.exists()


def es_planillero_asignado(user):
    return bool(user.is_authenticated and not es_editor_torneo(user) and user.partidos_planillero.exists())


def torneos_asignados_planillero(user):
    if not user.is_authenticated:
        return Torneo.objects.none()
    torneo_ids = user.partidos_planillero.values_list("categoria__torneo_id", flat=True)
    return Torneo.objects.filter(
        id__in=torneo_ids,
    ).distinct().order_by("-fecha_inicio", "nombre")


def torneo_actual_planillero(request):
    torneos = torneos_asignados_planillero(request.user)
    torneo_id = request.GET.get("torneo") or request.session.get("torneo_planillero_id")
    torneo = torneos.filter(id=torneo_id).first() if torneo_id else None

    if not torneo:
        torneo = torneos.filter(estado="ACTIVO").first() or torneos.first()

    if torneo:
        request.session["torneo_planillero_id"] = torneo.id

    return torneo, torneos


def denegar_partido_no_autorizado():
    return HttpResponseForbidden("No tienes permiso para editar este partido.")


def equipos_delegado_vigentes(user):
    if not user.is_authenticated:
        return Equipo.objects.none()
    return Equipo.objects.select_related("categoria").filter(
        responsable=user,
        acceso_delegado_hasta__gte=timezone.now(),
    )


def equipos_delegado_asignados(user, torneo=None):
    if not user.is_authenticated:
        return Equipo.objects.none()
    equipos = Equipo.objects.select_related("categoria", "categoria__torneo").filter(responsable=user)
    if torneo:
        equipos = equipos.filter(categoria__torneo=torneo)
    return equipos


def torneos_delegado_asignados(user):
    if not user.is_authenticated:
        return Torneo.objects.none()
    return Torneo.objects.filter(categorias__equipos__responsable=user).distinct().order_by("-fecha_inicio", "nombre")


def torneo_actual_delegado(request):
    torneos = torneos_delegado_asignados(request.user)
    torneo_id = request.GET.get("torneo") or request.session.get("torneo_id")
    torneo = None

    if torneo_id:
        torneo = torneos.filter(id=torneo_id, estado="ACTIVO").first()

    if not torneo:
        torneo = torneos.filter(estado="ACTIVO").first()

    if torneo:
        request.session["torneo_id"] = torneo.id

    return torneo


def equipos_alineacion_delegado_actual(request):
    torneo = torneo_actual_delegado(request)
    if not torneo:
        return Equipo.objects.none()
    return equipos_alineacion_para_usuario(request.user).filter(categoria__torneo=torneo)


def equipos_editables_delegado_actual(request):
    torneo = torneo_actual_delegado(request)
    if not torneo:
        return Equipo.objects.none()
    return equipos_editables_para_usuario(request.user).filter(categoria__torneo=torneo)


def puede_editar_equipo_delegado(user, equipo):
    if user.is_authenticated and user.is_superuser:
        return True
    if user.is_authenticated and user.is_staff:
        if not tabla_disponible("torneos_admintorneo"):
            return True
        return usuario_puede_editar_torneo(user, equipo.categoria.torneo if equipo.categoria_id else None)
    return bool(
        user.is_authenticated
        and equipo.responsable_id == user.id
        and equipo.acceso_delegado_vigente()
        and equipo.delegado_puede_editar_equipo
    )


def puede_cargar_fotos_jugadores_delegado(user, equipo):
    if user.is_authenticated and user.is_superuser:
        return True
    if user.is_authenticated and user.is_staff:
        if not tabla_disponible("torneos_admintorneo"):
            return True
        return usuario_puede_editar_torneo(user, equipo.categoria.torneo if equipo.categoria_id else None)
    return bool(
        user.is_authenticated
        and equipo.responsable_id == user.id
        and equipo.acceso_delegado_vigente()
        and equipo.delegado_puede_cargar_fotos_jugadores
    )


def equipos_editables_para_usuario(user):
    equipos = Equipo.objects.select_related("categoria")
    if user.is_authenticated and user.is_superuser:
        return equipos
    if user.is_authenticated and user.is_staff:
        if not tabla_disponible("torneos_admintorneo"):
            return equipos
        filtro = Q(categoria__torneo__admins_asignados__usuario=user, categoria__torneo__admins_asignados__activo=True)
        if tabla_disponible("torneos_adminorganizador"):
            filtro |= Q(categoria__torneo__organizador__admins_asignados__usuario=user, categoria__torneo__organizador__admins_asignados__activo=True)
        return equipos.filter(filtro).distinct()
    return equipos_delegado_vigentes(user).filter(delegado_puede_editar_equipo=True)


def equipos_con_fotos_delegado_actual(request):
    torneo = torneo_actual_delegado(request)
    if not torneo:
        return Equipo.objects.none()
    return equipos_delegado_vigentes(request.user).filter(
        categoria__torneo=torneo,
        delegado_puede_cargar_fotos_jugadores=True,
    )


def equipos_alineacion_para_usuario(user):
    equipos = Equipo.objects.select_related("categoria")
    if user.is_authenticated and user.is_superuser:
        return equipos
    if user.is_authenticated and user.is_staff:
        if not tabla_disponible("torneos_admintorneo"):
            return equipos
        filtro = Q(categoria__torneo__admins_asignados__usuario=user, categoria__torneo__admins_asignados__activo=True)
        if tabla_disponible("torneos_adminorganizador"):
            filtro |= Q(categoria__torneo__organizador__admins_asignados__usuario=user, categoria__torneo__organizador__admins_asignados__activo=True)
        return equipos.filter(filtro).distinct()
    return equipos_delegado_asignados(user)


def inicio_programado_partido(partido):
    if not partido.fecha or not partido.hora:
        return None
    inicio = datetime.combine(partido.fecha, partido.hora)
    if timezone.is_naive(inicio):
        return timezone.make_aware(inicio, timezone.get_current_timezone())
    return inicio


def ventana_alineacion_delegado(partido, equipo=None, ahora=None):
    ahora = ahora or timezone.now()
    if equipo and EntregaAlineacionPartido.objects.filter(partido=partido, equipo=equipo).exists():
        return False, "La alineación definitiva ya fue enviada."

    inicio = inicio_programado_partido(partido)
    if not inicio:
        return False, "Sin fecha u hora programada."
    apertura = inicio - timedelta(hours=1)
    if ahora < apertura:
        return False, f"Disponible desde {apertura.strftime('%d/%m/%Y %H:%M')}."

    if partido.estado == "PROGRAMADO":
        return True, "Disponible desde una hora antes y hasta 15 minutos después del inicio real."
    if partido.estado == "EN_JUEGO":
        inicio_real = partido.inicio_en_vivo or inicio
        cierre = inicio_real + timedelta(minutes=15)
        if ahora <= cierre:
            return True, f"Disponible hasta {cierre.strftime('%H:%M')}."
        return False, "La ventana de 15 minutos después del inicio ya finalizó."
    return False, "Disponible solo antes del partido o durante sus primeros 15 minutos."


def partido_pertenece_equipo(partido, equipo):
    return equipo.id in [partido.equipo_local_id, partido.equipo_visitante_id]


def puede_editar_alineacion_delegado(user, partido, equipo):
    if user.is_authenticated and user.is_superuser:
        return partido_pertenece_equipo(partido, equipo)
    if user.is_authenticated and user.is_staff:
        if not tabla_disponible("torneos_admintorneo"):
            return partido_pertenece_equipo(partido, equipo)
        return partido_pertenece_equipo(partido, equipo) and usuario_puede_editar_torneo(user, partido.categoria.torneo if partido.categoria_id else None)
    if not user.is_authenticated or equipo.responsable_id != user.id:
        return False
    if not partido_pertenece_equipo(partido, equipo):
        return False
    habilitado, _ = ventana_alineacion_delegado(partido, equipo)
    return habilitado


def partidos_alineacion_para_equipo(equipo):
    partidos = Partido.objects.select_related(
        "categoria",
        "equipo_local",
        "equipo_visitante",
    ).filter(
        categoria=equipo.categoria,
    ).filter(
        Q(equipo_local=equipo) | Q(equipo_visitante=equipo)
    ).filter(
        estado__in=["PROGRAMADO", "EN_JUEGO"]
    ).order_by("fecha", "hora", "id")

    entregados = set(
        EntregaAlineacionPartido.objects.filter(equipo=equipo).values_list("partido_id", flat=True)
    )
    items = []
    for partido in partidos:
        if partido.id in entregados:
            continue
        habilitado, motivo = ventana_alineacion_delegado(partido, equipo)
        items.append(SimpleNamespace(
            partido=partido,
            rival=partido.equipo_visitante if partido.equipo_local_id == equipo.id else partido.equipo_local,
            habilitado=habilitado,
            motivo=motivo,
        ))
    return items


def equipo_delegado_para_partido(user, partido):
    if not user.is_authenticated or not partido or es_editor_torneo(user):
        return None
    return equipos_delegado_asignados(user).filter(
        id__in=[partido.equipo_local_id, partido.equipo_visitante_id],
    ).first()


def url_alineacion_delegado_si_aplica(user, partido):
    equipo = equipo_delegado_para_partido(user, partido)
    if not equipo:
        return ""
    habilitado, _ = ventana_alineacion_delegado(partido, equipo)
    if not habilitado:
        return ""
    return reverse("delegado_alineacion_partido", args=[equipo.id, partido.id])


class IngresoTorneosView(LoginView):
    template_name = "registration/login.html"

    def get_success_url(self):
        if (
            self.request.user.is_authenticated
            and not es_editor_torneo(self.request.user)
            and equipos_delegado_asignados(self.request.user).exists()
        ):
            return reverse("delegado_mis_equipos")
        if (
            self.request.user.is_authenticated
            and not es_editor_torneo(self.request.user)
            and self.request.user.partidos_planillero.exists()
        ):
            return reverse("planillero_mis_partidos")
        return super().get_success_url()

    def form_valid(self, form):
        auth_login(self.request, form.get_user())
        user = self.request.user
        registrar_actividad(
            self.request,
            "INICIAR_SESION",
            descripcion=f"{user.username} inicio sesion.",
            datos={"ruta": self.request.path},
        )
        if es_editor_torneo(user):
            mensaje = "Acceso exitoso. Bienvenido al panel de gestion."
            acciones = [
                ("Gestion", reverse("gestion_panel")),
                ("Partidos", reverse("gestion_partidos")),
                ("Planillas de juego", reverse("gestion_planillas_juego")),
                ("Panel principal", reverse("panel")),
            ]
        elif equipos_delegado_asignados(user).exists():
            mensaje = "Acceso exitoso. Bienvenido al portal de delegados."
            acciones = [
                ("Mis equipos", reverse("delegado_mis_equipos")),
                ("Panel principal", reverse("panel")),
            ]
        elif user.partidos_planillero.exists():
            mensaje = "Acceso exitoso. Ya puedes diligenciar tus partidos asignados."
            acciones = [
                ("Mis partidos", reverse("planillero_mis_partidos")),
                ("Planillas de juego", reverse("gestion_planillas_juego")),
                ("Panel principal", reverse("panel")),
            ]
        else:
            mensaje = "Acceso exitoso."
            acciones = [("Panel principal", reverse("panel"))]

        acciones.append(("Cambiar mi contraseña", reverse("cambiar_contrasena")))

        botones = "".join(
            f'<a class="btn" href="{escape(url)}">{escape(texto)}</a>'
            for texto, url in acciones
        )
        return HttpResponse(f"""
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Acceso exitoso</title>
    <style>
        * {{ box-sizing: border-box; }}
        body {{
            margin: 0;
            min-height: 100vh;
            display: grid;
            place-items: center;
            padding: 18px;
            background: radial-gradient(circle at top, #12345a, #07111f);
            color: #ffffff;
            font-family: Arial, sans-serif;
            font-weight: 800;
        }}
        .card {{
            width: min(100%, 460px);
            background: #101d30;
            border: 1px solid rgba(0, 255, 102, 0.35);
            border-radius: 18px;
            padding: 22px;
            box-shadow: 0 20px 44px rgba(0, 0, 0, 0.35);
        }}
        h1 {{
            margin: 0 0 10px;
            color: #00ff66;
            font-size: clamp(28px, 7vw, 40px);
            text-transform: uppercase;
            text-align: center;
        }}
        p {{
            margin: 0 0 18px;
            color: #dbeafe;
            text-align: center;
        }}
        .actions {{
            display: grid;
            gap: 10px;
        }}
        .btn {{
            display: flex;
            align-items: center;
            justify-content: center;
            min-height: 48px;
            border-radius: 999px;
            background: #00e565;
            color: #03110a;
            text-decoration: none;
            font-weight: 900;
        }}
    </style>
</head>
<body>
    <main class="card">
        <h1>Acceso exitoso</h1>
        <p>{escape(mensaje)}</p>
        <div class="actions">{botones}</div>
    </main>
</body>
</html>
""")

    def get_default_redirect_url(self):
        if self.request.user.is_authenticated and es_editor_torneo(self.request.user):
            return reverse("gestion_panel")
        if (
            self.request.user.is_authenticated
            and not es_editor_torneo(self.request.user)
            and equipos_delegado_asignados(self.request.user).exists()
        ):
            return reverse("delegado_mis_equipos")
        if (
            self.request.user.is_authenticated
            and not es_editor_torneo(self.request.user)
            and self.request.user.partidos_planillero.exists()
        ):
            return reverse("planillero_mis_partidos")
        return super().get_default_redirect_url()


ESTADOS_PLANILLERO_PARTIDO = {"PROGRAMADO", "EN_JUEGO", "FINALIZADO", "SUSPENDIDO"}


def entero_post(request, campo, predeterminado=0, minimo=None):
    try:
        valor = int(request.POST.get(campo, predeterminado) or predeterminado)
    except (TypeError, ValueError):
        valor = predeterminado
    if minimo is not None:
        valor = max(valor, minimo)
    return valor


def tabla_disponible(nombre_tabla):
    try:
        return nombre_tabla in connection.introspection.table_names()
    except Exception:
        return False


def torneos_para_usuario(request):
    # No usamos select_related ni filtros por Organizador aqui porque Render puede
    # servir el codigo nuevo unos segundos antes de aplicar la migracion.
    torneos = Torneo.objects.order_by("-fecha_inicio", "nombre")
    usuario = getattr(request, "user", None)

    if not usuario or not usuario.is_authenticated:
        return torneos

    if usuario.is_superuser:
        return torneos

    if tabla_disponible("torneos_admintorneo"):
        filtro = Q(admins_asignados__usuario=usuario, admins_asignados__activo=True)
        if tabla_disponible("torneos_adminorganizador"):
            filtro |= Q(organizador__admins_asignados__usuario=usuario, organizador__admins_asignados__activo=True)
        torneos_asignados = torneos.filter(filtro).distinct()
        if torneos_asignados.exists() or usuario.is_staff:
            return torneos_asignados

    return torneos


def permisos_torneo_usuario(user, torneo):
    if not user.is_authenticated:
        return None
    if user.is_superuser:
        return SimpleNamespace(
            puede_editar=True,
            puede_validar=True,
            puede_programar=True,
            puede_descargar_planillas=True,
            activo=True,
        )
    if not torneo or not tabla_disponible("torneos_admintorneo"):
        return None
    permiso_torneo = AdminTorneo.objects.filter(usuario=user, torneo=torneo, activo=True).first()
    permiso_organizador = None
    if (
        tabla_disponible("torneos_adminorganizador")
        and getattr(torneo, "organizador_id", None)
    ):
        permiso_organizador = AdminOrganizador.objects.filter(
            usuario=user,
            organizador_id=torneo.organizador_id,
            activo=True,
        ).first()

    if permiso_torneo and permiso_organizador:
        return SimpleNamespace(
            puede_editar=permiso_torneo.puede_editar or permiso_organizador.puede_editar,
            puede_validar=permiso_torneo.puede_validar or permiso_organizador.puede_validar,
            puede_programar=permiso_torneo.puede_programar or permiso_organizador.puede_programar,
            puede_descargar_planillas=(
                getattr(permiso_torneo, "puede_descargar_planillas", False)
                or getattr(permiso_organizador, "puede_descargar_planillas", False)
            ),
            activo=True,
        )

    return permiso_torneo or permiso_organizador


def usuario_puede_editar_torneo(user, torneo):
    permisos = permisos_torneo_usuario(user, torneo)
    return bool(permisos and permisos.puede_editar)


def usuario_puede_validar_torneo(user, torneo):
    permisos = permisos_torneo_usuario(user, torneo)
    return bool(permisos and permisos.puede_validar)


def usuario_puede_programar_torneo(user, torneo):
    permisos = permisos_torneo_usuario(user, torneo)
    return bool(permisos and permisos.puede_programar)


def usuario_puede_descargar_planillas_torneo(user, torneo):
    permisos = permisos_torneo_usuario(user, torneo)
    return bool(permisos and getattr(permisos, "puede_descargar_planillas", False))


def usuario_solo_descarga_planillas(user, torneo):
    permisos = permisos_torneo_usuario(user, torneo)
    return bool(
        permisos
        and getattr(permisos, "puede_descargar_planillas", False)
        and not permisos.puede_editar
        and not permisos.puede_validar
        and not permisos.puede_programar
    )


def denegar_permiso_torneo():
    return HttpResponseForbidden("No tienes permiso para manipular este torneo.")


def puede_gestionar_torneo(request, torneo, permiso="editar"):
    if request.user.is_superuser:
        return True
    if not tabla_disponible("torneos_admintorneo"):
        return True
    if permiso == "validar":
        return usuario_puede_validar_torneo(request.user, torneo)
    if permiso == "programar":
        return usuario_puede_programar_torneo(request.user, torneo)
    if permiso == "descargar_planillas":
        return usuario_puede_descargar_planillas_torneo(request.user, torneo)
    return usuario_puede_editar_torneo(request.user, torneo)


def solicitudes_validacion_para_usuario(user, estado="PENDIENTE"):
    if not tabla_disponible("torneos_solicitudvalidacion"):
        return SolicitudValidacion.objects.none()
    solicitudes = SolicitudValidacion.objects.select_related(
        "torneo",
        "partido",
        "partido__equipo_local",
        "partido__equipo_visitante",
        "equipo",
        "jugador",
        "creado_por",
    ).order_by("-creado_en")
    if estado:
        solicitudes = solicitudes.filter(estado=estado)
    if user.is_superuser:
        return solicitudes
    if not user.is_staff or not tabla_disponible("torneos_admintorneo"):
        return solicitudes.none()

    filtro = Q(torneo__admins_asignados__usuario=user, torneo__admins_asignados__activo=True, torneo__admins_asignados__puede_validar=True)
    if tabla_disponible("torneos_adminorganizador"):
        filtro |= Q(torneo__organizador__admins_asignados__usuario=user, torneo__organizador__admins_asignados__activo=True, torneo__organizador__admins_asignados__puede_validar=True)
    return solicitudes.filter(filtro).distinct()


def crear_solicitud_validacion(tipo, titulo, descripcion="", user=None, torneo=None, partido=None, equipo=None, jugador=None, datos=None):
    if not tabla_disponible("torneos_solicitudvalidacion"):
        return None
    if not torneo:
        if partido and partido.categoria_id:
            torneo = partido.categoria.torneo
        elif equipo and equipo.categoria_id:
            torneo = equipo.categoria.torneo

    filtros = {"tipo": tipo, "estado": "PENDIENTE"}
    if tipo == "ALINEACION" and partido and equipo:
        filtros["partido"] = partido
        filtros["equipo"] = equipo
    elif partido:
        filtros["partido"] = partido
    elif jugador:
        filtros["jugador"] = jugador
    elif equipo:
        filtros["equipo"] = equipo

    solicitud = SolicitudValidacion.objects.filter(**filtros).first()
    if solicitud:
        solicitud.titulo = titulo
        solicitud.descripcion = descripcion
        solicitud.creado_por = user if getattr(user, "is_authenticated", False) else solicitud.creado_por
        solicitud.torneo = torneo
        solicitud.partido = partido
        solicitud.equipo = equipo
        solicitud.jugador = jugador
        solicitud.datos = datos or solicitud.datos
        solicitud.creado_en = timezone.now()
        solicitud.save(update_fields=[
            "titulo",
            "descripcion",
            "creado_por",
            "torneo",
            "partido",
            "equipo",
            "jugador",
            "datos",
            "creado_en",
        ])
        return solicitud

    return SolicitudValidacion.objects.create(
        tipo=tipo,
        titulo=titulo,
        descripcion=descripcion,
        creado_por=user if getattr(user, "is_authenticated", False) else None,
        torneo=torneo,
        partido=partido,
        equipo=equipo,
        jugador=jugador,
        datos=datos or {},
    )


def ip_cliente(request):
    encabezado = request.META.get("HTTP_X_FORWARDED_FOR")
    if encabezado:
        return encabezado.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR")


def registrar_actividad(request, accion, objeto=None, torneo=None, descripcion="", datos=None):
    if not tabla_disponible("torneos_registroactividad"):
        return

    if not torneo and objeto is not None:
        torneo = torneo_de_objeto(objeto)
    if not torneo:
        torneo = torneo_de_actividad_request(request)

    datos_registro = dict(datos or {})
    datos_registro.setdefault("tipo_usuario", tipo_usuario_actividad(request))

    RegistroActividad.objects.create(
        usuario=request.user if request.user.is_authenticated else None,
        torneo=torneo,
        accion=accion,
        modelo=objeto.__class__.__name__ if objeto is not None else "",
        objeto_id=getattr(objeto, "id", None),
        objeto_repr=str(objeto)[:255] if objeto is not None else "",
        descripcion=descripcion,
        datos=datos_registro,
        ip=ip_cliente(request),
        user_agent=(request.META.get("HTTP_USER_AGENT") or "")[:300],
    )
    request._actividad_registrada = True


def torneo_de_actividad_request(request):
    torneo_id = request.session.get("torneo_id") if hasattr(request, "session") else None
    if torneo_id:
        torneo = Torneo.objects.filter(id=torneo_id).first()
        if torneo:
            return torneo

    coincidencia = getattr(request, "resolver_match", None)
    parametros = (getattr(coincidencia, "kwargs", {}) or {}) if coincidencia else {}
    partido_id = parametros.get("partido_id")
    if partido_id:
        torneo = Torneo.objects.filter(categorias__partido__id=partido_id).first()
        if torneo:
            return torneo

    equipo_id = parametros.get("equipo_id") or request.POST.get("equipo")
    if equipo_id:
        torneo = Torneo.objects.filter(categorias__equipos__id=equipo_id).first()
        if torneo:
            return torneo

    jugador_id = parametros.get("jugador_id")
    if jugador_id:
        torneo = Torneo.objects.filter(categorias__equipos__jugadores__id=jugador_id).first()
        if torneo:
            return torneo

    user = getattr(request, "user", None)
    if getattr(user, "is_authenticated", False):
        torneos_delegado = list(
            Torneo.objects.filter(categorias__equipos__responsable=user)
            .distinct()
            .order_by("id")[:2]
        )
        if len(torneos_delegado) == 1:
            return torneos_delegado[0]
    return None


def tipo_usuario_actividad(request):
    cache = getattr(request, "_tipo_usuario_actividad", None)
    if cache:
        return cache

    user = request.user
    if not getattr(user, "is_authenticated", False):
        tipo = "Público"
    elif user.is_superuser:
        tipo = "Superadministrador"
    elif equipos_delegado_asignados(user).exists():
        tipo = "Delegado"
    elif user.partidos_planillero.exists():
        tipo = "Planillero"
    elif es_editor_torneo(user):
        tipo = "Administrador"
    else:
        tipo = "Usuario"

    request._tipo_usuario_actividad = tipo
    return tipo


def torneo_de_objeto(objeto):
    if isinstance(objeto, Torneo):
        return objeto
    if isinstance(objeto, Categoria):
        return objeto.torneo
    if isinstance(objeto, Documento):
        return objeto.torneo
    if isinstance(objeto, Equipo):
        return objeto.categoria.torneo if objeto.categoria_id else None
    if isinstance(objeto, Jugador):
        return objeto.equipo.categoria.torneo if objeto.equipo_id and objeto.equipo.categoria_id else None
    if isinstance(objeto, Partido):
        return objeto.categoria.torneo if objeto.categoria_id else None
    return None


def organizador_seguro(torneo):
    if torneo.organizador_id:
        try:
            return torneo.organizador
        except Exception:
            return None
    return None


def nombre_organizador_torneo(torneo):
    organizador = organizador_seguro(torneo)
    if organizador:
        return organizador.nombre.strip() or torneo.nombre
    return torneo.nombre


def organizadores_para_portal(torneos):
    organizadores = {}
    independientes = []

    for torneo in torneos:
        organizador = organizador_seguro(torneo)
        if organizador:
            grupo = organizadores.get(torneo.organizador_id)
            if not grupo:
                grupo = SimpleNamespace(
                    id=torneo.organizador_id,
                    nombre=nombre_organizador_torneo(torneo),
                    logo=organizador.logo or torneo.logo_portada,
                    torneos=[],
                    es_organizador=True,
                )
                organizadores[torneo.organizador_id] = grupo
            if not grupo.logo and torneo.logo_portada:
                grupo.logo = torneo.logo_portada
            grupo.torneos.append(torneo)
        else:
            independientes.append(SimpleNamespace(
                id=None,
                nombre=torneo.nombre,
                logo=torneo.logo_portada,
                torneos=[torneo],
                es_organizador=False,
                torneo_id=torneo.id,
            ))

    return list(organizadores.values()) + independientes


def cerrar_sesion(request):
    if request.user.is_authenticated:
        registrar_actividad(
            request,
            "CERRAR_SESION",
            descripcion=f"{request.user.username} cerro sesion.",
            datos={"ruta": request.path},
        )
    logout(request)
    return redirect("panel")


@login_required
def cambiar_contrasena(request):
    form = PasswordChangeForm(request.user, request.POST or None)
    if request.method == "POST" and form.is_valid():
        usuario = form.save()
        update_session_auth_hash(request, usuario)
        registrar_actividad(
            request,
            "CAMBIAR_CONTRASENA",
            descripcion=f"{usuario.username} cambió su propia contraseña.",
            datos={"ruta": request.path},
        )
        messages.success(request, "Contraseña actualizada correctamente. Tu sesión continúa abierta.")
        return redirect("cambiar_contrasena")

    if es_editor_torneo(request.user):
        volver_url = reverse("gestion_panel")
    elif equipos_delegado_asignados(request.user).exists():
        volver_url = reverse("delegado_mis_equipos")
    elif request.user.partidos_planillero.exists():
        volver_url = reverse("planillero_mis_partidos")
    else:
        volver_url = reverse("panel")

    return render(request, "registration/cambiar_contrasena.html", {
        "form": form,
        "volver_url": volver_url,
        "volver_panel_url": volver_url,
        "volver_panel_text": "Volver",
    })


def service_worker(request):
    sw_path = finders.find("sw.js")
    if not sw_path:
        return HttpResponse("", content_type="application/javascript")

    with open(sw_path, "r", encoding="utf-8") as archivo:
        response = HttpResponse(archivo.read(), content_type="application/javascript")
    response["Service-Worker-Allowed"] = "/"
    return response


def torneo_actual(request, auto_seleccionar=True):
    torneos = torneos_para_usuario(request)
    torneo_id = request.GET.get("torneo") or request.session.get("torneo_id")
    torneo = None

    if torneo_id:
        torneo = torneos.filter(id=torneo_id).first()

    if not torneo and auto_seleccionar:
        torneo = torneos.filter(estado="ACTIVO").first() or torneos.first()

    if torneo:
        request.session["torneo_id"] = torneo.id

    return torneo


def listar_imagenes_cloudinary(max_results=80):
    imagenes = []

    if getattr(settings, "USE_CLOUDINARY_STORAGE", False):
        try:
            import cloudinary.api

            configurar = getattr(default_storage, "_configure", None)
            if configurar:
                configurar()

            respuesta = cloudinary.api.resources(
                resource_type="image",
                type="upload",
                max_results=max_results,
            )
        except Exception as exc:
            print(f"No se pudieron listar imagenes de Cloudinary: {exc}")
        else:
            for recurso in respuesta.get("resources", []):
                public_id = recurso.get("public_id", "")
                url = recurso.get("secure_url") or recurso.get("url")
                if not public_id or not url:
                    continue
                if public_id.startswith("documentos/"):
                    continue

                carpeta = public_id.split("/", 1)[0] if "/" in public_id else "General"
                imagenes.append({
                    "public_id": public_id,
                    "url": url,
                    "carpeta": carpeta,
                    "nombre": public_id.rsplit("/", 1)[-1],
                })

    if imagenes:
        return imagenes

    return listar_imagenes_usadas()


def listar_imagenes_usadas():
    imagenes = []
    vistos = set()

    def agregar(nombre, url):
        nombre = str(nombre or "").strip()
        url = str(url or "").strip()
        if not nombre or nombre in vistos:
            return
        vistos.add(nombre)
        imagenes.append({
            "public_id": nombre,
            "url": url,
            "carpeta": nombre.split("/", 1)[0] if "/" in nombre else "General",
            "nombre": nombre.rsplit("/", 1)[-1],
        })

    for equipo in Equipo.objects.exclude(escudo="").exclude(escudo__isnull=True).order_by("nombre"):
        try:
            agregar(equipo.escudo.name, equipo.escudo.url)
        except Exception:
            continue

    for jugador in Jugador.objects.exclude(foto="").exclude(foto__isnull=True).order_by("nombres"):
        try:
            agregar(jugador.foto.name, jugador.foto.url)
        except Exception:
            continue

    return imagenes


def url_imagen_cloudinary(public_id):
    if not public_id:
        return ""

    try:
        configurar = getattr(default_storage, "_configure", None)
        if configurar:
            configurar()

        import cloudinary.utils

        return cloudinary.utils.cloudinary_url(
            str(public_id),
            resource_type="image",
            secure=True,
        )[0]
    except Exception:
        return ""


@login_required
@user_passes_test(es_editor_torneo)
def gestion_biblioteca_cloudinary(request):
    torneo = torneo_actual(request)
    if not puede_gestionar_torneo(request, torneo, "editar"):
        return denegar_permiso_torneo()
    imagenes = listar_imagenes_cloudinary(500)
    q = request.GET.get("q", "").strip()
    equipos = Equipo.objects.select_related("categoria").order_by("categoria__nombre", "nombre")
    jugadores = Jugador.objects.select_related("equipo", "equipo__categoria").order_by("equipo__nombre", "nombres")

    if torneo:
        equipos = equipos.filter(categoria__torneo=torneo)
        jugadores = jugadores.filter(equipo__categoria__torneo=torneo)

    if q:
        imagenes = [
            imagen for imagen in imagenes
            if q.lower() in imagen["public_id"].lower()
        ]

    return render(request, "gestion/biblioteca_cloudinary.html", {
        "imagenes": imagenes,
        "q": q,
        "equipos": equipos,
        "jugadores": jugadores,
    })


@login_required
@user_passes_test(es_editor_torneo)
def gestion_asignar_imagen_cloudinary(request):
    if request.method != "POST":
        return redirect("gestion_biblioteca_cloudinary")

    torneo = torneo_actual(request)
    if not puede_gestionar_torneo(request, torneo, "editar"):
        return denegar_permiso_torneo()
    public_id = (request.POST.get("public_id") or "").strip()
    tipo = request.POST.get("tipo")
    objeto_id = request.POST.get("objeto_id")

    if not public_id or not tipo or not objeto_id:
        messages.error(request, "Selecciona una imagen y un destino.")
        return redirect("gestion_biblioteca_cloudinary")

    if tipo == "equipo":
        equipos = Equipo.objects.select_related("categoria")
        if torneo:
            equipos = equipos.filter(categoria__torneo=torneo)
        equipo = get_object_or_404(equipos, id=objeto_id)
        equipo.escudo = public_id
        equipo.save(update_fields=["escudo"])
        registrar_actividad(request, "ASIGNAR_IMAGEN", equipo, descripcion=f"Asigno imagen al equipo {equipo.nombre}.")
        messages.success(request, f"Imagen asignada al equipo {equipo.nombre}.")
        return redirect("gestion_equipo_editar", equipo_id=equipo.id)

    if tipo == "jugador":
        jugadores = Jugador.objects.select_related("equipo", "equipo__categoria")
        if torneo:
            jugadores = jugadores.filter(equipo__categoria__torneo=torneo)
        jugador = get_object_or_404(jugadores, id=objeto_id)
        jugador.foto = public_id
        jugador.save(update_fields=["foto"])
        registrar_actividad(request, "ASIGNAR_IMAGEN", jugador, descripcion=f"Asigno imagen al jugador {jugador.nombres}.")
        messages.success(request, f"Imagen asignada al jugador {jugador.nombres}.")
        return redirect("gestion_jugador_editar", jugador_id=jugador.id)

    messages.error(request, "Destino no valido.")
    return redirect("gestion_biblioteca_cloudinary")


def aplicar_imagen_cloudinary(instancia, campo, public_id, archivo_subido):
    public_id = (public_id or "").strip()
    if public_id and not archivo_subido:
        setattr(instancia, campo, public_id)


def subir_imagen_torneo_cloudinary(archivo, torneo, campo):
    if not archivo:
        return ""
    if not getattr(settings, "USE_CLOUDINARY_STORAGE", False):
        return ""

    import cloudinary
    import cloudinary.uploader

    cloudinary_url = getattr(settings, "CLOUDINARY_URL", "").strip()
    if cloudinary_url:
        os.environ["CLOUDINARY_URL"] = cloudinary_url

    cloudinary.config(secure=True)

    torneo_nombre = limpiar_ruta_cloudinary(getattr(torneo, "nombre", "SIN_TORNEO"))
    archivo.seek(0)
    resultado = cloudinary.uploader.upload(
        archivo,
        resource_type="image",
        folder=f"torneos/{torneo_nombre}",
        public_id=campo,
        overwrite=True,
        invalidate=True,
    )
    return resultado.get("public_id") or ""


def aplicar_imagenes_torneo_cloudinary(torneo, archivos):
    for campo in ("logo_portada", "logo_izquierdo", "imagen_central", "logo_derecho"):
        public_id = subir_imagen_torneo_cloudinary(archivos.get(campo), torneo, campo)
        if public_id:
            setattr(torneo, campo, public_id)


def documentos_publicos_por_tipo(torneo=None):
    documentos = Documento.objects.select_related(
        "categoria", "equipo_local", "equipo_visitante",
    ).filter(activo=True).order_by("tipo", "-creado_en", "titulo")
    if torneo:
        documentos = documentos.filter(torneo=torneo)
    else:
        documentos = documentos.none()

    return {
        "reglamentos": documentos.filter(tipo="REGLAMENTO"),
        "resoluciones": documentos.filter(tipo="RESOLUCION"),
        "demandas": documentos.filter(tipo="DEMANDA"),
        "comunicados": documentos.filter(tipo="COMUNICADO"),
        "planillas": documentos.filter(tipo="PLANILLA_JUEGO"),
        "otros": documentos.filter(tipo="OTRO"),
    }


def listar_documentos_cloudinary_por_tipo(max_results=500):
    documentos = {
        "reglamentos": [],
        "resoluciones": [],
        "demandas": [],
        "comunicados": [],
        "otros": [],
    }

    if not getattr(settings, "USE_CLOUDINARY_STORAGE", False):
        return documentos

    try:
        configurar = getattr(default_storage, "_configure", None)
        if configurar:
            configurar()

        import cloudinary.api

        recursos = []
        for resource_type in ("raw", "image"):
            respuesta = cloudinary.api.resources(
                resource_type=resource_type,
                type="upload",
                prefix="documentos/",
                max_results=max_results,
            )
            recursos.extend(respuesta.get("resources", []))
    except Exception as exc:
        print(f"No se pudieron recuperar documentos de Cloudinary: {exc}")
        return documentos

    tipos = {
        "REGLAMENTO": "reglamentos",
        "RESOLUCION": "resoluciones",
        "DEMANDA": "demandas",
        "COMUNICADO": "comunicados",
        "OTRO": "otros",
    }

    vistos = set()
    for recurso in recursos:
        public_id = recurso.get("public_id", "")
        url = recurso.get("secure_url") or recurso.get("url")
        if not public_id or not url or public_id in vistos:
            continue

        partes = public_id.split("/")
        if len(partes) < 3:
            continue

        tipo = partes[1].upper()
        llave = tipos.get(tipo)
        if not llave:
            continue

        vistos.add(public_id)
        titulo = partes[-1].rsplit(".", 1)[0].replace("_", " ").replace("-", " ").strip().upper()
        documentos[llave].append(SimpleNamespace(
            id=None,
            titulo=titulo or tipo,
            descripcion="",
            archivo=url,
            tipo=tipo,
        ))

    for llave in documentos:
        documentos[llave].sort(key=lambda item: item.titulo)

    return documentos


def documento_visible_en_torneo_actual(request, documento_id):
    torneo = torneo_actual(request, auto_seleccionar=False)
    documentos = Documento.objects.filter(id=documento_id, activo=True)
    if torneo:
        documentos = documentos.filter(torneo=torneo)
    else:
        documentos = documentos.none()
    return get_object_or_404(documentos)


def documento_publico(request, documento_id):
    documento = documento_visible_en_torneo_actual(request, documento_id)
    if documento.tipo == "PLANILLA_JUEGO":
        return redirect(documento.archivo)
    archivo_url = request.build_absolute_uri(reverse("documento_archivo_publico", args=[documento.id]))
    visor_url = f"https://docs.google.com/gview?embedded=1&url={quote(archivo_url, safe='')}"
    return redirect(visor_url)


def documento_archivo_publico(request, documento_id):
    documento = documento_visible_en_torneo_actual(request, documento_id)
    respuesta = requests.get(documento.archivo, timeout=20)
    respuesta.raise_for_status()
    content_type = respuesta.headers.get("Content-Type") or "application/pdf"
    response = HttpResponse(respuesta.content, content_type=content_type)
    nombre = limpiar_ruta_cloudinary(documento.titulo) or f"documento-{documento.id}"
    response["Content-Disposition"] = f'inline; filename="{nombre}.pdf"'
    return response


def subir_documento_supabase(archivo, tipo):
    import boto3
    from urllib.parse import quote

    bucket = os.getenv("SUPABASE_STORAGE_BUCKET", "torneos-media").strip()
    endpoint_url = os.getenv("SUPABASE_S3_ENDPOINT_URL", "").strip()
    access_key = os.getenv("SUPABASE_S3_ACCESS_KEY_ID", "").strip()
    secret_key = os.getenv("SUPABASE_S3_SECRET_ACCESS_KEY", "").strip()
    region_name = os.getenv("SUPABASE_S3_REGION_NAME", "us-east-1").strip()
    public_base = os.getenv("SUPABASE_PUBLIC_MEDIA_URL", "").strip().rstrip("/")

    if not all([bucket, endpoint_url, access_key, secret_key, public_base]):
        return ""

    nombre_archivo = limpiar_ruta_cloudinary(os.path.splitext(archivo.name)[0])
    extension = os.path.splitext(archivo.name)[1].lower() or ".pdf"
    llave = f"documentos/{limpiar_ruta_cloudinary(tipo)}/{uuid.uuid4().hex}_{nombre_archivo}{extension}"

    cliente = boto3.client(
        "s3",
        endpoint_url=endpoint_url,
        aws_access_key_id=access_key,
        aws_secret_access_key=secret_key,
        region_name=region_name,
    )

    archivo.seek(0)
    cliente.upload_fileobj(
        archivo,
        bucket,
        llave,
        ExtraArgs={
            "ContentType": getattr(archivo, "content_type", "application/octet-stream"),
        },
    )
    return f"{public_base}/{quote(llave, safe='/')}"


def subir_documento_cloudinary(archivo, tipo):
    import cloudinary
    import cloudinary.uploader

    cloudinary_url = getattr(settings, "CLOUDINARY_URL", "").strip()
    if cloudinary_url:
        os.environ["CLOUDINARY_URL"] = cloudinary_url

    cloudinary.config(secure=True)

    archivo.seek(0)
    resultado = cloudinary.uploader.upload(
        archivo,
        resource_type="raw",
        folder=f"documentos/{limpiar_ruta_cloudinary(tipo)}",
    )
    return resultado.get("secure_url") or resultado["url"]


def subir_documento_torneo(archivo, tipo):
    url_supabase = subir_documento_supabase(archivo, tipo)
    if url_supabase:
        return url_supabase

    return subir_documento_cloudinary(archivo, tipo)


def limpiar_nombre(nombre):
    nombre = str(nombre).strip()
    nombre = re.sub(r'[\\/*?:"<>|]', '', nombre)
    return nombre.replace(' ', '_').upper()


def limpiar_texto_excel(valor):
    return "" if valor is None else str(valor).strip()


def limpiar_cedula_excel(valor):
    if valor is None:
        return ""

    valor = str(valor).strip()

    if valor.endswith(".0"):
        valor = valor[:-2]

    return valor.replace(".", "").replace(",", "").replace(" ", "")


def limpiar_entero_excel(valor):
    if valor in [None, ""]:
        return None

    try:
        return int(float(valor))
    except Exception:
        return None


def normalizar_anio_excel(anio):
    anio = limpiar_entero_excel(anio)

    if anio is None:
        return None

    if anio < 100:
        return 2000 + anio if anio <= 30 else 1900 + anio

    return anio


def construir_fecha_excel(dia, mes, anio):
    dia = limpiar_entero_excel(dia)
    mes = limpiar_entero_excel(mes)
    anio = normalizar_anio_excel(anio)

    if not dia or not mes or not anio:
        return None

    try:
        return date(anio, mes, dia)
    except Exception:
        return None


def obtener_hoja_planilla_excel(workbook):
    nombres = [
        "Planilla inscripcion",
        "Planilla inscripción",
        "PLANILLA INSCRIPCION",
        "PLANILLA INSCRIPCIÓN",
        "Inscripcion",
        "Inscripción",
    ]

    for nombre in nombres:
        if nombre in workbook.sheetnames:
            return workbook[nombre]

    return workbook.active


def normalizar_encabezado_excel(valor):
    valor = limpiar_nombre(limpiar_texto_excel(valor)).lower()
    return valor.replace("_", "")


def valor_por_encabezado(row, indices, *nombres):
    for nombre in nombres:
        indice = indices.get(normalizar_encabezado_excel(nombre))

        if indice is not None:
            return row[indice].value

    return None


def encabezado_existe(indices, *nombres):
    return any(normalizar_encabezado_excel(nombre) in indices for nombre in nombres)


def separar_planilleros_excel(valor):
    texto = limpiar_texto_excel(valor)
    if not texto:
        return []

    return [
        item.strip()
        for item in re.split(r"[;,\n\r]+", texto)
        if item.strip()
    ]


def normalizar_usuario_excel(valor):
    return re.sub(r"[^a-z0-9]+", "", limpiar_texto_excel(valor).lower())


def buscar_planilleros_excel(valor):
    usuarios = []
    no_encontrados = []

    for identificador in separar_planilleros_excel(valor):
        usuario = User.objects.filter(
            Q(username__iexact=identificador) | Q(email__iexact=identificador),
            is_active=True,
        ).first()

        if not usuario:
            identificador_normalizado = normalizar_usuario_excel(identificador)
            for candidato in User.objects.filter(is_active=True):
                valores_candidato = [
                    candidato.username,
                    candidato.email,
                    candidato.get_full_name(),
                    f"{candidato.first_name}{candidato.last_name}",
                ]
                if any(normalizar_usuario_excel(valor) == identificador_normalizado for valor in valores_candidato):
                    usuario = candidato
                    break

        if usuario:
            if usuario not in usuarios:
                usuarios.append(usuario)
        else:
            no_encontrados.append(identificador)

    return usuarios, no_encontrados


def construir_fecha_partido_excel(valor):
    if not valor:
        return None

    if isinstance(valor, datetime):
        return valor.date()

    if isinstance(valor, date):
        return valor

    texto = limpiar_texto_excel(valor)

    for formato in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            pass

    return None


def construir_hora_partido_excel(valor):
    if not valor:
        return None

    if isinstance(valor, datetime):
        return valor.time().replace(second=0, microsecond=0)

    if isinstance(valor, time):
        return valor.replace(second=0, microsecond=0)

    texto = limpiar_texto_excel(valor).upper().replace(".", "")

    for formato in ("%H:%M", "%I:%M %p", "%I %p"):
        try:
            return datetime.strptime(texto, formato).time().replace(second=0, microsecond=0)
        except ValueError:
            pass

    return None


def escudo_estatico_url(nombre_archivo):
    if not nombre_archivo:
        return ""

    ruta = f"torneos/escudos/{nombre_archivo}"

    if finders.find(ruta):
        return static(ruta)

    return ""


def escudo_default_url():
    return static("torneos/img/logo_imcred.png")


def escudo_url(equipo):
    if not equipo:
        return escudo_default_url()

    if equipo.escudo:
        try:
            if equipo.escudo.url:
                return equipo.escudo.url
        except Exception:
            pass

        try:
            if equipo.escudo.storage.exists(equipo.escudo.name):
                return equipo.escudo.url
        except Exception:
            pass

        nombre_archivo = os.path.basename(equipo.escudo.name).replace(" ", "_")
        escudo = escudo_estatico_url(nombre_archivo)

        if escudo:
            return escudo

    nombre_equipo = limpiar_nombre(equipo.nombre)

    for extension in ("png", "jpg", "jpeg", "webp"):
        escudo = escudo_estatico_url(f"{nombre_equipo}.{extension}")

        if escudo:
            return escudo

    return escudo_default_url()

def url_absoluta(request, url):
    if not url:
        return ""

    url = str(url)

    if url.startswith("http://") or url.startswith("https://"):
        return url

    return request.build_absolute_uri(url)


def rutas_logos(request):
    return {
        "logo_alcaldia": request.build_absolute_uri(static("torneos/img/logo_alcaldia.png")),
        "logo_app": request.build_absolute_uri(static("torneos/img/logo_app.png")),
        "logo_torneo": request.build_absolute_uri(static("torneos/img/logo_torneo.png")),
        "logo_imcred": request.build_absolute_uri(static("torneos/img/logo_imcred.png")),
    }


def url_campo_imagen(campo):
    if not campo:
        return ""
    try:
        return campo.url
    except Exception:
        return ""


def logos_torneo(request, torneo=None):
    logos = rutas_logos(request)
    if not torneo:
        return logos
    return {
        "logo_alcaldia": url_campo_imagen(torneo.logo_izquierdo),
        "logo_torneo": url_campo_imagen(torneo.imagen_central),
        "logo_imcred": url_campo_imagen(torneo.logo_derecho),
    }


def estructura_base_categoria():
    return {
        "grupos": {},
        "tabla_general_mata_mata": [],
        "partidos_por_fecha": {},
        "columnas_planilla": [],
        "goleadores_planilla": [],
        "tarjetas_planilla": [],
        "valla_planilla": [],
        "alertas_tarjetas": [],
        "foraneos": [],
        "controlar_foraneos": False,
        "porcentaje_foraneos": 50,
        "equipos": [],
        "llaves": {
            "cuartos": [],
            "semifinal": [],
            "final": [],
            "tercer_puesto": [],
        },
    }


def nombre_columna_partido(partido):
    fase = partido.fase or "GRUPOS"

    if fase == "GRUPOS":
        return partido.numero_fecha or "SIN FECHA"

    return fase


def etiqueta_columna_planilla(columna):
    etiquetas = {
        "CUARTOS": "CTOS",
        "SEMIFINAL": "SF",
        "TERCER_PUESTO": "TP",
        "FINAL": "F",
    }
    return etiquetas.get(columna, columna)


ESTADOS_PARTIDO_CERRADO = ["FINALIZADO", "DECIDIDO_COMITE", "WO"]


def _marcar_estadisticas_pendientes(partido, user=None):
    if user is not None and es_editor_torneo(user):
        return
    partido.estadisticas_validadas = False
    partido.estadisticas_validadas_en = None
    partido.estadisticas_validadas_por = None
    partido.save(update_fields=[
        "estadisticas_validadas",
        "estadisticas_validadas_en",
        "estadisticas_validadas_por",
    ])
    crear_solicitud_validacion(
        "ESTADISTICAS",
        f"Validar estadisticas: {partido.equipo_local} vs {partido.equipo_visitante}",
        descripcion="Las acciones del partido fueron modificadas y deben validarse antes de quedar oficiales en estadisticas.",
        user=user,
        partido=partido,
        equipo=partido.equipo_local,
        datos={"partido_id": partido.id},
    )


def _validar_estadisticas_partido(partido, user):
    partido.estadisticas_validadas = True
    partido.estadisticas_validadas_en = timezone.now()
    partido.estadisticas_validadas_por = user
    partido.save(update_fields=[
        "estadisticas_validadas",
        "estadisticas_validadas_en",
        "estadisticas_validadas_por",
    ])
    if tabla_disponible("torneos_solicitudvalidacion"):
        SolicitudValidacion.objects.filter(
            tipo="ESTADISTICAS",
            partido=partido,
            estado="PENDIENTE",
        ).update(
            estado="VALIDADO",
            resuelto_por=user,
            resuelto_en=timezone.now(),
        )


def construir_estadisticas_foraneos(categoria):
    if not categoria or not categoria.controlar_foraneos:
        return []

    partidos_fase1 = Partido.objects.filter(
        categoria=categoria,
        fase="GRUPOS",
    )
    total_partidos_por_equipo = defaultdict(int)
    partidos_por_equipo = defaultdict(set)
    for partido in partidos_fase1.only("id", "equipo_local_id", "equipo_visitante_id"):
        partidos_por_equipo[partido.equipo_local_id].add(partido.id)
        partidos_por_equipo[partido.equipo_visitante_id].add(partido.id)
    for equipo_id, partidos_ids in partidos_por_equipo.items():
        total_partidos_por_equipo[equipo_id] = len(partidos_ids)

    partidos_jugados_por_jugador = defaultdict(set)
    alineaciones = AlineacionPartido.objects.filter(
        partido__categoria=categoria,
        partido__fase="GRUPOS",
        partido__estado__in=ESTADOS_PARTIDO_CERRADO,
        rol="TITULAR",
        jugador__es_foraneo=True,
    ).values_list("jugador_id", "partido_id")
    for jugador_id, partido_id in alineaciones:
        partidos_jugados_por_jugador[jugador_id].add(partido_id)

    sustituciones = SustitucionPartido.objects.filter(
        partido__categoria=categoria,
        partido__fase="GRUPOS",
        partido__estado__in=ESTADOS_PARTIDO_CERRADO,
        jugador_entra__es_foraneo=True,
    ).values_list("jugador_entra_id", "partido_id")
    for jugador_id, partido_id in sustituciones:
        partidos_jugados_por_jugador[jugador_id].add(partido_id)

    porcentaje = categoria.porcentaje_minimo_foraneos or 0
    filas = []
    jugadores = Jugador.objects.filter(
        equipo__categoria=categoria,
        es_foraneo=True,
    ).select_related("equipo").only(
        "id",
        "nombres",
        "equipo_id",
        "equipo__id",
        "equipo__nombre",
        "equipo__escudo",
    ).order_by("equipo__nombre", "nombres")
    for jugador in jugadores:
        total_fase1 = total_partidos_por_equipo.get(jugador.equipo_id, 0)
        minimo = int((total_fase1 * porcentaje) / 100)
        jugados = len(partidos_jugados_por_jugador.get(jugador.id, set()))
        filas.append({
            "jugador": jugador.nombres,
            "equipo": jugador.equipo.nombre,
            "escudo": escudo_url(jugador.equipo),
            "partidos_fase1": total_fase1,
            "jugados": jugados,
            "minimo": minimo,
            "porcentaje": porcentaje,
            "cumple": jugados >= minimo,
            "estado": "Habilitado" if jugados >= minimo else "Pendiente",
        })

    return sorted(filas, key=lambda fila: (fila["cumple"], fila["equipo"], fila["jugador"]))


def construir_estructura(torneo=None):
    estructura = {}

    categorias = Categoria.objects.all().order_by("nombre")
    if torneo:
        categorias = categorias.filter(torneo=torneo)

    for categoria in categorias:
        estructura[categoria.nombre] = estructura_base_categoria()
        estructura[categoria.nombre]["controlar_foraneos"] = categoria.controlar_foraneos
        estructura[categoria.nombre]["porcentaje_foraneos"] = categoria.porcentaje_minimo_foraneos
        estructura[categoria.nombre]["foraneos"] = construir_estadisticas_foraneos(categoria)

    partidos = Partido.objects.select_related(
        "categoria",
        "equipo_local",
        "equipo_visitante"
    ).order_by(
        "categoria__nombre",
        "grupo",
        "numero_fecha",
        "fase",
        "fecha",
        "hora"
    )
    if torneo:
        partidos = partidos.filter(categoria__torneo=torneo)

    columnas_por_categoria = defaultdict(list)

    for partido in partidos:
        categoria = partido.categoria.nombre
        grupo = partido.grupo or "SIN GRUPO"
        fecha = partido.numero_fecha or "SIN FECHA"
        columna = nombre_columna_partido(partido)

        estructura.setdefault(categoria, estructura_base_categoria())

        if columna not in columnas_por_categoria[categoria]:
            columnas_por_categoria[categoria].append(columna)

        estructura[categoria]["grupos"].setdefault(grupo, {
            "fechas": {},
            "tabla": {}
        })

        datos_grupo = estructura[categoria]["grupos"][grupo]

        datos_grupo["fechas"].setdefault(fecha, [])
        datos_grupo["fechas"][fecha].append(partido)

        if partido.fase == "GRUPOS":
            estructura[categoria]["partidos_por_fecha"].setdefault(fecha, [])
            estructura[categoria]["partidos_por_fecha"][fecha].append({
                "grupo": grupo,
                "partido": partido,
                "escudo_local": escudo_url(partido.equipo_local),
                "escudo_visitante": escudo_url(partido.equipo_visitante),
            })

        for equipo in [partido.equipo_local, partido.equipo_visitante]:
            if equipo:
                datos_grupo["tabla"].setdefault(equipo.id, {
                    "id": equipo.id,
                    "equipo": equipo.nombre,
                    "escudo": escudo_url(equipo),
                    "pj": 0,
                    "pg": 0,
                    "pe": 0,
                    "pp": 0,
                    "gf": 0,
                    "gc": 0,
                    "dg": 0,
                    "pts": 0,
                })

        if partido.estado in ESTADOS_PARTIDO_CERRADO and partido.estadisticas_validadas:
            gl = partido.goles_local or 0
            gv = partido.goles_visitante or 0

            local = datos_grupo["tabla"][partido.equipo_local_id]
            visitante = datos_grupo["tabla"][partido.equipo_visitante_id]

            local["pj"] += 1
            visitante["pj"] += 1

            local["gf"] += gl
            local["gc"] += gv

            visitante["gf"] += gv
            visitante["gc"] += gl

            if gl > gv:
                local["pg"] += 1
                local["pts"] += 3
                visitante["pp"] += 1

            elif gl < gv:
                visitante["pg"] += 1
                visitante["pts"] += 3
                local["pp"] += 1

            else:
                local["pe"] += 1
                visitante["pe"] += 1
                local["pts"] += 1
                visitante["pts"] += 1

            local["pts"] += partido.ajuste_puntos_local or 0
            visitante["pts"] += partido.ajuste_puntos_visitante or 0

    for categoria, datos_categoria in estructura.items():
        tabla_general_mata = {}
        for grupo, datos_grupo in datos_categoria["grupos"].items():
            for equipo in datos_grupo["tabla"].values():
                equipo["dg"] = equipo["gf"] - equipo["gc"]

                if str(grupo).startswith("MATA "):
                    acumulado = tabla_general_mata.setdefault(equipo["id"], {
                        "id": equipo["id"],
                        "equipo": equipo["equipo"],
                        "escudo": equipo["escudo"],
                        "pj": 0,
                        "pg": 0,
                        "pe": 0,
                        "pp": 0,
                        "gf": 0,
                        "gc": 0,
                        "dg": 0,
                        "pts": 0,
                    })
                    for campo in ["pj", "pg", "pe", "pp", "gf", "gc", "pts"]:
                        acumulado[campo] += equipo[campo]
                    acumulado["dg"] = acumulado["gf"] - acumulado["gc"]

            datos_grupo["tabla"] = sorted(
                datos_grupo["tabla"].values(),
                key=lambda x: (x["pts"], x["dg"], x["gf"]),
                reverse=True
            )

        datos_categoria["tabla_general_mata_mata"] = sorted(
            tabla_general_mata.values(),
            key=lambda x: (x["pts"], x["dg"], x["gf"], x["equipo"]),
            reverse=True,
        )

    goleadores_temp = defaultdict(lambda: defaultdict(lambda: {
        "jugador": "",
        "equipo": "",
        "escudo": "",
        "valores": defaultdict(int),
        "total": 0,
    }))

    goles_qs = Gol.objects.select_related(
        "partido__categoria",
        "jugador",
        "equipo",
        "partido"
    )
    if torneo:
        goles_qs = goles_qs.filter(partido__categoria__torneo=torneo)
    goles_qs = goles_qs.filter(partido__estadisticas_validadas=True)

    for gol in goles_qs:
        if gol.partido.estado not in ESTADOS_PARTIDO_CERRADO:
            continue

        categoria = gol.partido.categoria.nombre
        columna = nombre_columna_partido(gol.partido)
        jugador = gol.jugador.nombres
        cantidad = gol.cantidad or 1

        data = goleadores_temp[categoria][jugador]
        data["jugador"] = jugador
        data["equipo"] = gol.equipo.nombre
        data["escudo"] = escudo_url(gol.equipo)
        data["valores"][columna] += cantidad
        data["total"] += cantidad

        if columna not in columnas_por_categoria[categoria]:
            columnas_por_categoria[categoria].append(columna)

    tarjetas_temp = defaultdict(lambda: defaultdict(lambda: {
        "jugador": "",
        "equipo": "",
        "escudo": "",
        "valores": defaultdict(str),
        "total_a": 0,
        "total_r": 0,
        "total": 0,
    }))

    tarjetas_qs = Tarjeta.objects.select_related(
        "partido__categoria",
        "jugador",
        "equipo",
        "partido"
    )
    if torneo:
        tarjetas_qs = tarjetas_qs.filter(partido__categoria__torneo=torneo)
    tarjetas_qs = tarjetas_qs.filter(partido__estadisticas_validadas=True)

    for tarjeta in tarjetas_qs:
        if tarjeta.partido.estado not in ESTADOS_PARTIDO_CERRADO:
            continue

        categoria = tarjeta.partido.categoria.nombre
        columna = nombre_columna_partido(tarjeta.partido)
        jugador = tarjeta.jugador.nombres

        data = tarjetas_temp[categoria][jugador]
        data["jugador"] = jugador
        data["equipo"] = tarjeta.equipo.nombre
        data["escudo"] = escudo_url(tarjeta.equipo)

        actual = data["valores"][columna]

        if tarjeta.tipo == "AMARILLA":
            data["total_a"] += 1
            data["valores"][columna] = (actual + " A").strip()

        elif tarjeta.tipo == "ROJA":
            data["total_r"] += 1
            data["valores"][columna] = (actual + " R").strip()

        data["total"] += 1

        if columna not in columnas_por_categoria[categoria]:
            columnas_por_categoria[categoria].append(columna)

    valla_temp = defaultdict(lambda: defaultdict(lambda: {
        "equipo": "",
        "escudo": "",
        "valores": defaultdict(str),
        "pj": 0,
        "gc": 0,
        "promedio": 0,
    }))

    for partido in partidos:
        if partido.estado not in ESTADOS_PARTIDO_CERRADO or not partido.estadisticas_validadas:
            continue

        categoria = partido.categoria.nombre
        columna = nombre_columna_partido(partido)

        gl = partido.goles_local or 0
        gv = partido.goles_visitante or 0

        local = partido.equipo_local.nombre
        visitante = partido.equipo_visitante.nombre

        data_local = valla_temp[categoria][local]
        data_local["equipo"] = local
        data_local["escudo"] = escudo_url(partido.equipo_local)
        data_local["valores"][columna] = str(gv)
        data_local["pj"] += 1
        data_local["gc"] += gv

        data_visitante = valla_temp[categoria][visitante]
        data_visitante["equipo"] = visitante
        data_visitante["escudo"] = escudo_url(partido.equipo_visitante)
        data_visitante["valores"][columna] = str(gl)
        data_visitante["pj"] += 1
        data_visitante["gc"] += gl

    alertas_temp = defaultdict(lambda: defaultdict(lambda: {
        "jugador": "",
        "equipo": "",
        "escudo": "",
        "amarillas_grupos": 0,
        "amarillas_finales": 0,
        "rojas_total": 0,
    }))

    alertas_tarjetas_qs = Tarjeta.objects.select_related(
        "partido__categoria",
        "jugador",
        "equipo",
        "partido"
    )
    if torneo:
        alertas_tarjetas_qs = alertas_tarjetas_qs.filter(partido__categoria__torneo=torneo)
    alertas_tarjetas_qs = alertas_tarjetas_qs.filter(partido__estadisticas_validadas=True)

    for tarjeta in alertas_tarjetas_qs:
        if tarjeta.partido.estado not in ESTADOS_PARTIDO_CERRADO:
            continue

        categoria = tarjeta.partido.categoria.nombre
        jugador_id = tarjeta.jugador.id
        fase = tarjeta.partido.fase or "GRUPOS"

        data = alertas_temp[categoria][jugador_id]
        data["jugador"] = tarjeta.jugador.nombres
        data["equipo"] = tarjeta.equipo.nombre
        data["escudo"] = escudo_url(tarjeta.equipo)

        if tarjeta.tipo == "AMARILLA":
            if fase == "GRUPOS":
                data["amarillas_grupos"] += 1
            else:
                data["amarillas_finales"] += 1

        elif tarjeta.tipo == "ROJA":
            data["rojas_total"] += 1

    for categoria, datos_categoria in estructura.items():
        columnas_ordenadas = []

        for col in columnas_por_categoria[categoria]:
            if col not in columnas_ordenadas:
                columnas_ordenadas.append(col)

        columnas_ordenadas = []

        for col in columnas_por_categoria[categoria]:
           if col not in columnas_ordenadas:
                columnas_ordenadas.append(col)

        fases_finales = ["CUARTOS", "SEMIFINAL", "TERCER_PUESTO", "FINAL"]

        columnas_ordenadas = [
            col for col in columnas_ordenadas
            if col not in fases_finales
        ]

        for fase_fija in fases_finales:
            columnas_ordenadas.append(fase_fija)

        datos_categoria["columnas_planilla"] = columnas_ordenadas
        datos_categoria["columnas_planilla_display"] = [
            {
                "valor": col,
                "etiqueta": etiqueta_columna_planilla(col),
            }
            for col in columnas_ordenadas
        ]

        goleadores = []

        for jugador in goleadores_temp[categoria].values():
            fila = {
                "jugador": jugador["jugador"],
                "equipo": jugador["equipo"],
                "escudo": jugador["escudo"],
                "celdas": [],
                "total": jugador["total"],
            }

            for col in columnas_ordenadas:
                fila["celdas"].append(jugador["valores"].get(col, ""))

            goleadores.append(fila)

        datos_categoria["goleadores_planilla"] = sorted(
            goleadores,
            key=lambda x: x["total"],
            reverse=True
        )

        tarjetas = []

        for jugador in tarjetas_temp[categoria].values():
            fila = {
                "jugador": jugador["jugador"],
                "equipo": jugador["equipo"],
                "escudo": jugador["escudo"],
                "celdas": [],
                "total_a": jugador["total_a"],
                "total_r": jugador["total_r"],
                "total": jugador["total"],
            }

            for col in columnas_ordenadas:
                fila["celdas"].append(jugador["valores"].get(col, ""))

            tarjetas.append(fila)

        datos_categoria["tarjetas_planilla"] = sorted(
            tarjetas,
            key=lambda x: x["total"],
            reverse=True
        )

        vallas = []

        for equipo in valla_temp[categoria].values():
            promedio = round(equipo["gc"] / equipo["pj"], 2) if equipo["pj"] > 0 else 0

            fila = {
                "equipo": equipo["equipo"],
                "escudo": equipo["escudo"],
                "celdas": [],
                "pj": equipo["pj"],
                "gc": equipo["gc"],
                "promedio": promedio,
            }

            for col in columnas_ordenadas:
                fila["celdas"].append(equipo["valores"].get(col, ""))

            vallas.append(fila)

        datos_categoria["valla_planilla"] = sorted(
            vallas,
            key=lambda x: (x["promedio"], x["gc"])
        )

        alertas = []

        for data in alertas_temp[categoria].values():
            observaciones = []

            if data["amarillas_grupos"] >= 3:
                observaciones.append("SUSPENSIÓN 1 FECHA POR 3 AMARILLAS EN GRUPOS")
            elif data["amarillas_grupos"] == 2:
                observaciones.append("ALERTA: A 1 AMARILLA DE SUSPENSIÓN EN GRUPOS")

            if data["amarillas_finales"] >= 3:
                observaciones.append("SUSPENSIÓN 1 FECHA POR 3 AMARILLAS EN FASE FINAL")
            elif data["amarillas_finales"] == 2:
                observaciones.append("ALERTA: A 1 AMARILLA DE SUSPENSIÓN EN FASE FINAL")

            if data["rojas_total"] >= 3:
                observaciones.append("SANCIÓN: RESTO DEL TORNEO POR 3 ROJAS")
            elif data["rojas_total"] == 2:
                observaciones.append("ALERTA: A 1 ROJA DE SANCIÓN POR RESTO DEL TORNEO")

            if observaciones:
                alertas.append({
                    "jugador": data["jugador"],
                    "equipo": data["equipo"],
                    "escudo": data["escudo"],
                    "amarillas_grupos": data["amarillas_grupos"],
                    "amarillas_finales": data["amarillas_finales"],
                    "rojas_total": data["rojas_total"],
                    "observacion": " / ".join(observaciones),
                })

        datos_categoria["alertas_tarjetas"] = sorted(
            alertas,
            key=lambda x: (
                x["rojas_total"],
                x["amarillas_grupos"],
                x["amarillas_finales"]
            ),
            reverse=True
        )

        # ==================================================
    # LLAVES DE FASE FINAL
    # ==================================================

    for categoria_nombre, datos_categoria in estructura.items():
        partidos_finales = Partido.objects.filter(
            categoria__nombre=categoria_nombre,
            fase__in=["CUARTOS", "SEMIFINAL", "FINAL", "TERCER_PUESTO"]
        ).select_related(
            "equipo_local",
            "equipo_visitante"
        ).order_by(
            "fase",
            "numero_fecha",
            "id"
        )
        if torneo:
            partidos_finales = partidos_finales.filter(categoria__torneo=torneo)

        llaves = {
            "cuartos": [],
            "semifinal": [],
            "final": [],
            "tercer_puesto": [],
        }

        for p in partidos_finales:
            gl = p.goles_local if p.goles_local is not None else 0
            gv = p.goles_visitante if p.goles_visitante is not None else 0
            pl = p.goles_local_penales if p.goles_local_penales is not None else 0
            pv = p.goles_visitante_penales if p.goles_visitante_penales is not None else 0

            ganador_local = False
            ganador_visitante = False

            if p.estado in ESTADOS_PARTIDO_CERRADO:
                if gl > gv:
                    ganador_local = True
                elif gv > gl:
                    ganador_visitante = True
                else:
                    if pl > pv:
                        ganador_local = True
                    elif pv > pl:
                        ganador_visitante = True

            item = {
                "id": p.id,
                "local": p.equipo_local.nombre if p.equipo_local else "Por definir",
                "visitante": p.equipo_visitante.nombre if p.equipo_visitante else "Por definir",
                "escudo_local": escudo_url(p.equipo_local),
                "escudo_visitante": escudo_url(p.equipo_visitante),
                "gl": gl,
                "gv": gv,
                "pl": pl,
                "pv": pv,
                "tiene_penales": gl == gv and (pl > 0 or pv > 0),
                "ganador_local": ganador_local,
                "ganador_visitante": ganador_visitante,
                "estado": p.estado,
                "estado_display": p.get_estado_display(),
                "numero_fecha": p.numero_fecha,
                "numero_llave": int(re.search(r"\d+", p.numero_fecha or "0").group()) if re.search(r"\d+", p.numero_fecha or "") else 0,
                "fecha": p.fecha,
                "hora": p.hora,
                "cancha": p.cancha,
                "inicio_en_vivo": p.inicio_en_vivo,
                "cronometro_pausado": p.cronometro_pausado,
                "periodo_en_vivo": p.periodo_en_vivo,
                "segundos_vivos": segundos_vivos_partido(p),
            }

            if p.fase == "CUARTOS":
                llaves["cuartos"].append(item)
            elif p.fase == "SEMIFINAL":
                llaves["semifinal"].append(item)
            elif p.fase == "FINAL":
                llaves["final"].append(item)
            elif p.fase == "TERCER_PUESTO":
                llaves["tercer_puesto"].append(item)

        orden_cuartos = {1: 1, 4: 2, 2: 3, 3: 4}
        llaves["cuartos"] = sorted(
            llaves["cuartos"],
            key=lambda item: (orden_cuartos.get(item["numero_llave"], item["numero_llave"] or 99), item["id"] or 0)
        )
        llaves["semifinal"] = sorted(
            llaves["semifinal"],
            key=lambda item: (item["numero_llave"] or 99, item["id"] or 0)
        )

        datos_categoria["llaves"] = llaves

    # ==================================================
    # EQUIPOS Y JUGADORES POR CATEGORÍA
    # ==================================================

    for categoria_nombre, datos_categoria in estructura.items():
        equipos_categoria = Equipo.objects.filter(
            categoria__nombre=categoria_nombre,
            activo=True
        ).prefetch_related("jugadores").order_by("nombre")
        if torneo:
            equipos_categoria = equipos_categoria.filter(categoria__torneo=torneo)

        lista_equipos = []

        for equipo_obj in equipos_categoria:
            jugadores = []

            for jugador in equipo_obj.jugadores.all().order_by("dorsal", "nombres"):
                edad = ""

                if jugador.fecha_nacimiento:
                    hoy = date.today()
                    edad = hoy.year - jugador.fecha_nacimiento.year - (
                        (hoy.month, hoy.day) < (
                            jugador.fecha_nacimiento.month,
                            jugador.fecha_nacimiento.day
                        )
                    )

                jugadores.append({
                    "dorsal": jugador.dorsal,
                    "nombres": jugador.nombres,
                    "cedula": jugador.cedula,
                    "estado": jugador.estado,
                    "foto": jugador.foto.url if jugador.foto else "",
                    "edad": edad,
                })

            lista_equipos.append({
                "id": equipo_obj.id,
                "nombre": equipo_obj.nombre,
                "escudo": escudo_url(equipo_obj),
                "jugadores": jugadores,
                "director_tecnico": equipo_obj.director_tecnico,
                "asistente_tecnico": equipo_obj.asistente_tecnico,
                "delegado": equipo_obj.delegado,
            })

        datos_categoria["equipos"] = lista_equipos
    return estructura


def preparar_categoria_para_descarga(request, datos_categoria):
    if not datos_categoria:
        return datos_categoria

    for grupo, datos_grupo in datos_categoria["grupos"].items():
        for equipo in datos_grupo["tabla"]:
            equipo["escudo"] = url_absoluta(request, equipo.get("escudo"))

    for g in datos_categoria["goleadores_planilla"]:
        g["escudo"] = url_absoluta(request, g.get("escudo"))

    for t in datos_categoria["tarjetas_planilla"]:
        t["escudo"] = url_absoluta(request, t.get("escudo"))

    for v in datos_categoria["valla_planilla"]:
        v["escudo"] = url_absoluta(request, v.get("escudo"))

    return datos_categoria


def segundos_vivos_partido(partido):
    segundos = partido.segundos_acumulados or 0
    if (
        partido.estado == "EN_JUEGO"
        and partido.inicio_en_vivo
        and not partido.cronometro_pausado
    ):
        diferencia = timezone.now() - partido.inicio_en_vivo
        segundos += max(0, int(diferencia.total_seconds()))
    return segundos


def foto_jugador_url(jugador):
    if jugador and jugador.foto:
        try:
            return jugador.foto.url
        except Exception:
            return ""
    return ""


def iniciales_jugador(jugador):
    nombre = (getattr(jugador, "nombres", "") or "").strip()
    partes = [parte[0] for parte in nombre.split()[:2] if parte]
    return "".join(partes).upper() or "J"


def _nombre_primer_apellido(jugador):
    nombre = (getattr(jugador, "nombres", "") or "").strip()
    partes = [parte for parte in nombre.split() if parte]
    conectores_apellido = {"de", "del", "la", "las", "los", "da", "das", "do", "dos", "van", "von", "y"}
    if len(partes) >= 4:
        apellido = [partes[2]]
        idx = 3
        while apellido[-1].lower() in conectores_apellido and idx < len(partes):
            apellido.append(partes[idx])
            idx += 1
        return f"{partes[0]} {' '.join(apellido)}"
    if len(partes) >= 3:
        apellido = [partes[1]]
        idx = 2
        while apellido[-1].lower() in conectores_apellido and idx < len(partes):
            apellido.append(partes[idx])
            idx += 1
        return f"{partes[0]} {' '.join(apellido)}"
    if len(partes) >= 2:
        return f"{partes[0]} {partes[1]}"
    return nombre or "Jugador"


def nombre_corto_jugador(jugador):
    nombre = (getattr(jugador, "nombres", "") or "").strip()
    return " ".join(nombre.split()[:3]) or "Jugador"


def nombre_resumen_jugador(jugador):
    return nombre_corto_jugador(jugador)


def edad_jugador_en_fecha(jugador, fecha_referencia=None):
    if not jugador or not jugador.fecha_nacimiento:
        return None
    fecha = fecha_referencia or date.today()
    return fecha.year - jugador.fecha_nacimiento.year - (
        (fecha.month, fecha.day) < (jugador.fecha_nacimiento.month, jugador.fecha_nacimiento.day)
    )


def reglas_edad_categoria(categoria):
    if not categoria:
        return []
    reglas_cache = getattr(categoria, "_reglas_edad_cache", None)
    if reglas_cache is None:
        reglas_cache = list(
            categoria.reglas_edad.filter(activa=True).order_by("orden", "edad_minima", "id")
        )
        categoria._reglas_edad_cache = reglas_cache
    return reglas_cache


def regla_edad_jugador(jugador, categoria=None, fecha_referencia=None):
    categoria = categoria or getattr(getattr(jugador, "equipo", None), "categoria", None)
    edad = edad_jugador_en_fecha(jugador, fecha_referencia)
    for regla in reglas_edad_categoria(categoria):
        if regla.coincide_con_edad(edad):
            return regla
    return None


def etiqueta_edad_jugador(jugador, categoria=None, fecha_referencia=None):
    regla = regla_edad_jugador(jugador, categoria, fecha_referencia)
    return regla.etiqueta if regla else ""


def texto_edad_jugador(jugador, categoria=None, fecha_referencia=None):
    etiqueta = etiqueta_edad_jugador(jugador, categoria, fecha_referencia)
    if etiqueta:
        return etiqueta
    edad = edad_jugador_en_fecha(jugador, fecha_referencia)
    if edad is None:
        return ""
    return f"{edad} años"


def reglas_edad_para_frontend(categoria):
    return [
        {
            "etiqueta": regla.etiqueta,
            "edad_minima": regla.edad_minima,
            "minimo": regla.minimo_titulares or 0,
            "maximo": regla.maximo_titulares,
        }
        for regla in reglas_edad_categoria(categoria)
        if regla.minimo_titulares or regla.maximo_titulares is not None
    ]


def validar_reglas_edad_titulares(partido, equipo, titulares_ids):
    reglas = [
        regla for regla in reglas_edad_categoria(partido.categoria)
        if regla.minimo_titulares or regla.maximo_titulares is not None
    ]
    if not reglas:
        return []

    jugadores = Jugador.objects.filter(id__in=titulares_ids, equipo=equipo)
    conteos = {regla.id: 0 for regla in reglas}
    for jugador in jugadores:
        regla = regla_edad_jugador(jugador, partido.categoria, partido.fecha)
        if regla and regla.id in conteos:
            conteos[regla.id] += 1

    errores = []
    reglas_ordenadas = sorted(reglas, key=lambda regla: (regla.edad_minima, regla.id))
    for regla in reglas:
        cantidad = conteos.get(regla.id, 0)
        maximo_titulares = regla.maximo_titulares
        if maximo_titulares is not None and cantidad > maximo_titulares:
            errores.append(
                f"{regla.etiqueta}: maximo {maximo_titulares} en cancha, tienes {cantidad}."
            )

    sobrantes_mayores = 0
    for regla in reversed(reglas_ordenadas):
        cantidad = conteos.get(regla.id, 0) + sobrantes_mayores
        if regla.minimo_titulares and cantidad < regla.minimo_titulares:
            errores.append(
                f"{regla.etiqueta}: minimo {regla.minimo_titulares} en cancha, tienes {cantidad}."
            )
        sobrantes_mayores = max(0, cantidad - regla.minimo_titulares)
    return errores


TOLERANCIA_REGLA_EDAD_SEGUNDOS = 60


def categoria_permite_reingresos(categoria):
    """Senior Master y Plus 50 usan sustituciones libres con reingreso."""
    return slugify(categoria.nombre or "") in {"senior-master", "plus-50"}


def jugadores_actuales_en_cancha(partido, equipo):
    """Calcula quienes juegan sin modificar la alineación inicial registrada."""
    jugadores_ids = set(
        AlineacionPartido.objects.filter(
            partido=partido,
            equipo=equipo,
            rol="TITULAR",
        ).values_list("jugador_id", flat=True)
    )
    sustituciones = SustitucionPartido.objects.filter(
        partido=partido,
        equipo=equipo,
    ).order_by("creado_en", "id")
    for cambio in sustituciones:
        jugadores_ids.discard(cambio.jugador_sale_id)
        jugadores_ids.add(cambio.jugador_entra_id)
    return jugadores_ids


def actualizar_incidencia_regla_edad(partido, equipo, request=None, sustitucion=None, permitir_crear=False):
    incidencia = IncidenciaReglaEdad.objects.filter(
        partido=partido,
        equipo=equipo,
        estado="ABIERTA",
    ).order_by("-id").first()

    # En el descanso se pueden registrar varios cambios; se evalúan juntos al iniciar el ST.
    if partido.periodo_en_vivo == "ET":
        return incidencia

    segundos = segundos_vivos_partido(partido)
    jugadores_ids = jugadores_actuales_en_cancha(partido, equipo)
    errores = validar_reglas_edad_titulares(partido, equipo, jugadores_ids)

    if errores:
        if not incidencia and permitir_crear:
            incidencia = IncidenciaReglaEdad.objects.create(
                partido=partido,
                equipo=equipo,
                sustitucion_inicio=sustitucion,
                errores=errores,
                segundo_inicio=segundos,
                minuto_inicio=max(segundos // 60, 1),
                periodo_inicio=partido.periodo_en_vivo or "",
                creada_por=request.user if request and request.user.is_authenticated else None,
            )
            if request:
                detalle_cambio = ""
                if sustitucion:
                    detalle_cambio = f" Cambio: entró {sustitucion.jugador_entra.nombres} por {sustitucion.jugador_sale.nombres}."
                registrar_actividad(
                    request,
                    "ALERTA_REGLA_EDAD",
                    partido,
                    descripcion=(
                        f"Minuto {max(segundos // 60, 1)}: posible infracción de reglas de edad de {equipo.nombre}."
                        f"{detalle_cambio} {' '.join(errores)}"
                    ),
                    datos={
                        "equipo_id": equipo.id,
                        "sustitucion_id": getattr(sustitucion, "id", None),
                        "segundo_inicio": segundos,
                        "errores": errores,
                    },
                )
        elif incidencia:
            cambios = []
            se_confirmo = False
            if incidencia.errores != errores:
                incidencia.errores = errores
                cambios.append("errores")
            if not incidencia.confirmada and segundos - incidencia.segundo_inicio >= TOLERANCIA_REGLA_EDAD_SEGUNDOS:
                incidencia.confirmada = True
                cambios.append("confirmada")
                se_confirmo = True
            if cambios:
                incidencia.save(update_fields=cambios)
            if se_confirmo and request:
                registrar_actividad(
                    request,
                    "CONFIRMAR_INFRACCION_REGLA_EDAD",
                    partido,
                    descripcion=(
                        f"Minuto {max(segundos // 60, 1)}: se confirmó la infracción de reglas de edad de "
                        f"{equipo.nombre} al superar {TOLERANCIA_REGLA_EDAD_SEGUNDOS} segundos. "
                        f"{' '.join(incidencia.errores)}"
                    ),
                    datos={
                        "equipo_id": equipo.id,
                        "incidencia_id": incidencia.id,
                        "duracion_segundos": segundos - incidencia.segundo_inicio,
                        "errores": incidencia.errores,
                    },
                )
        return incidencia

    if incidencia:
        duracion = max(segundos - incidencia.segundo_inicio, 0)
        incidencia.estado = "CORREGIDA"
        incidencia.segundo_fin = segundos
        incidencia.minuto_fin = max(segundos // 60, 1)
        incidencia.finalizada_en = timezone.now()
        incidencia.duracion_segundos = duracion
        incidencia.confirmada = incidencia.confirmada or duracion >= TOLERANCIA_REGLA_EDAD_SEGUNDOS
        incidencia.corregida_por = request.user if request and request.user.is_authenticated else None
        incidencia.save(update_fields=[
            "estado", "segundo_fin", "minuto_fin", "finalizada_en",
            "duracion_segundos", "confirmada", "corregida_por",
        ])
        if request:
            registrar_actividad(
                request,
                "CORREGIR_REGLA_EDAD",
                partido,
                descripcion=f"Se corrigió la alineación en juego de {equipo.nombre} después de {duracion} segundos.",
                datos={
                    "equipo_id": equipo.id,
                    "incidencia_id": incidencia.id,
                    "duracion_segundos": duracion,
                    "confirmada": incidencia.confirmada,
                },
            )
    return incidencia


def construir_partidos_portada(torneo=None):
    partidos = Partido.objects.filter(
        fecha__isnull=False,
    ).select_related(
        "categoria",
        "equipo_local",
        "equipo_visitante",
    )
    if torneo:
        partidos = partidos.filter(categoria__torneo=torneo)

    estados_visibles = [
        "PROGRAMADO",
        "EN_JUEGO",
        "FINALIZADO",
        "DECIDIDO_COMITE",
        "WO",
        "APLAZADO",
        "SUSPENDIDO",
    ]
    tarjetas_por_partido = defaultdict(int)
    goles_por_partido = defaultdict(int)

    tarjetas_eventos = Tarjeta.objects.all()
    goles_eventos = Gol.objects.all()
    if torneo:
        tarjetas_eventos = tarjetas_eventos.filter(partido__categoria__torneo=torneo)
        goles_eventos = goles_eventos.filter(partido__categoria__torneo=torneo)

    for item in tarjetas_eventos.values("partido_id"):
        tarjetas_por_partido[item["partido_id"]] += 1

    for item in goles_eventos.values("partido_id"):
        goles_por_partido[item["partido_id"]] += 1

    partidos_portada = []

    for partido in partidos:
        if partido.estado not in estados_visibles:
            continue

        cancha_normalizada = (partido.cancha or "").strip().lower()
        programacion_completa = (
            bool(cancha_normalizada)
            and cancha_normalizada != "por definir"
            and partido.hora
            and partido.hora != time(0, 0)
        )

        if partido.estado in ESTADOS_PARTIDO_CERRADO:
            bloque = "RESULTADOS RECIENTES"
            orden_bloque = 0
            orden_fecha = partido.fecha.toordinal()
        elif partido.estado == "EN_JUEGO" or programacion_completa:
            bloque = "PROGRAMADOS"
            orden_bloque = 1
            orden_fecha = partido.fecha.toordinal()
        else:
            bloque = "FUTUROS"
            orden_bloque = 2
            orden_fecha = partido.fecha.toordinal()

        fase = partido.fase or "GRUPOS"
        gl = partido.goles_local or 0
        gv = partido.goles_visitante or 0
        pl = partido.goles_local_penales or 0
        pv = partido.goles_visitante_penales or 0
        ganador_local = False
        ganador_visitante = False

        if fase != "GRUPOS" and partido.estado in ESTADOS_PARTIDO_CERRADO:
            if gl > gv:
                ganador_local = True
            elif gv > gl:
                ganador_visitante = True
            elif pl > pv:
                ganador_local = True
            elif pv > pl:
                ganador_visitante = True

        categoria_nombre = partido.categoria.nombre
        categoria_mayuscula = categoria_nombre.upper()
        if "PLUS" in categoria_mayuscula:
            categoria_clase = "cat-plus"
        elif "SENIOR" in categoria_mayuscula:
            categoria_clase = "cat-senior"
        elif "INTER" in categoria_mayuscula:
            categoria_clase = "cat-interbarrios"
        else:
            categoria_clase = "cat-general"

        partidos_portada.append({
            "id": partido.id,
            "bloque": bloque,
            "orden_bloque": orden_bloque,
            "orden_fecha": orden_fecha,
            "categoria": categoria_nombre,
            "categoria_clase": categoria_clase,
            "grupo": partido.grupo or "",
            "fase": fase,
            "numero_fecha": partido.numero_fecha or "",
            "estado": partido.estado,
            "estado_display": partido.get_estado_display(),
            "fecha": partido.fecha,
            "hora": partido.hora.strftime("%H:%M") if partido.hora else "Por definir",
            "hora_orden": partido.hora or time(0, 0),
            "cancha": partido.cancha or "Por definir",
            "local": partido.equipo_local.nombre,
            "visitante": partido.equipo_visitante.nombre,
            "escudo_local": escudo_url(partido.equipo_local),
            "escudo_visitante": escudo_url(partido.equipo_visitante),
            "goles_local": gl,
            "goles_visitante": gv,
            "goles_local_penales": pl,
            "goles_visitante_penales": pv,
            "tiene_penales": pl > 0 or pv > 0,
            "ganador_local": ganador_local,
            "ganador_visitante": ganador_visitante,
            "eventos": goles_por_partido[partido.id] + tarjetas_por_partido[partido.id],
            "inicio_en_vivo": partido.inicio_en_vivo,
            "cronometro_pausado": partido.cronometro_pausado,
            "segundos_acumulados": partido.segundos_acumulados,
            "segundos_vivos": segundos_vivos_partido(partido),
            "periodo_en_vivo": partido.periodo_en_vivo,
        })

    return sorted(
        partidos_portada,
        key=lambda partido: (
            partido["orden_bloque"],
            partido["orden_fecha"],
            partido["hora_orden"],
            partido["categoria"],
        )
    )


def panel_principal(request):
    if request.GET.get("portal") == "1":
        request.session.pop("torneo_id", None)

    torneo = torneo_actual(request, auto_seleccionar=False)
    torneos_menu = torneos_para_usuario(request)
    if not torneo:
        organizador_id = request.GET.get("organizador")
        organizador_actual = None
        torneos_portal = torneos_menu
        if organizador_id and str(organizador_id).isdigit():
            torneos_portal = torneos_menu.filter(organizador_id=organizador_id)
            try:
                organizador = Organizador.objects.filter(id=organizador_id, activo=True).first()
            except Exception:
                organizador = None
            if organizador:
                organizador_actual = SimpleNamespace(
                    id=organizador.id,
                    nombre=organizador.nombre,
                )

        logos = rutas_logos(request)
        return render(request, "portal_torneos.html", {
            "torneos_menu": torneos_portal,
            "organizadores_portal": organizadores_para_portal(torneos_menu),
            "organizador_actual": organizador_actual,
            "tiene_gestion_torneo": es_editor_torneo(request.user),
            "tiene_equipos_delegado": equipos_delegado_asignados(request.user).exists(),
            "logo_alcaldia": logos["logo_alcaldia"],
            "logo_app": logos["logo_app"],
            "logo_torneo": logos["logo_torneo"],
            "logo_imcred": logos["logo_imcred"],
        })

    if torneo.organizador_id:
        torneos_menu = torneos_menu.filter(organizador_id=torneo.organizador_id)
    else:
        torneos_menu = torneos_menu.filter(id=torneo.id)

    categoria_seleccionada = request.GET.get("categoria", "").strip()
    categorias = Categoria.objects.order_by("nombre")
    if torneo:
        categorias = categorias.filter(torneo=torneo)

    if categoria_seleccionada:
        estructura_total = construir_estructura(torneo)
        estructura = {
            nombre: datos
            for nombre, datos in estructura_total.items()
            if nombre == categoria_seleccionada
        }
    else:
        estructura = {
            categoria.nombre: estructura_base_categoria()
            for categoria in categorias
        }

    logos = logos_torneo(request, torneo)
    partidos_portada = construir_partidos_portada(torneo)
    documentos = documentos_publicos_por_tipo(torneo)
    planillas_publicas = list(documentos["planillas"])
    categorias_planillas = {}
    for documento in planillas_publicas:
        categoria_nombre = documento.categoria.nombre if documento.categoria else "Sin categoría"
        fecha_nombre = documento.numero_fecha or "Sin fecha del fixture"
        categoria = categorias_planillas.setdefault(categoria_nombre, {})
        categoria.setdefault(fecha_nombre, []).append(documento)
    planillas_agrupadas = [
        SimpleNamespace(
            nombre=categoria_nombre,
            fechas=[
                SimpleNamespace(
                    nombre=fecha_nombre,
                    etiqueta=(
                        fecha_nombre.upper()
                        if str(fecha_nombre).strip().lower().startswith("fecha")
                        else f"FECHA {fecha_nombre}"
                    ),
                    documentos=documentos_fecha,
                )
                for fecha_nombre, documentos_fecha in fechas.items()
            ],
        )
        for categoria_nombre, fechas in categorias_planillas.items()
    ]
    fechas_grupos = sorted(
        {
            p["numero_fecha"]
            for p in partidos_portada
            if p["numero_fecha"]
        },
        key=lambda valor: (
            int(valor) if str(valor).isdigit() else 9999,
            str(valor),
        )
    )
    partidos_fechas = [
        {
            "key": f"fecha-{re.sub(r'[^a-zA-Z0-9_-]+', '-', str(fecha)).strip('-').lower()}",
            "label": f"Fecha {fecha}",
            "partidos": [p for p in partidos_portada if p["numero_fecha"] == fecha],
        }
        for fecha in fechas_grupos
    ]
    partidos_cuartos = [p for p in partidos_portada if p["fase"] == "CUARTOS"]
    partidos_semifinal = [p for p in partidos_portada if p["fase"] == "SEMIFINAL"]
    partidos_final = [p for p in partidos_portada if p["fase"] in ["FINAL", "TERCER_PUESTO"]]

    def orden_fecha_fixture(partido):
        numero_fecha = partido["numero_fecha"]
        return (
            int(numero_fecha) if str(numero_fecha).isdigit() else 9999,
            str(numero_fecha),
        )

    partidos_resultados = sorted(
        [p for p in partidos_portada if p["bloque"] == "RESULTADOS RECIENTES"],
        key=lambda p: (
            p["orden_fecha"],
            p["hora_orden"],
            p["categoria"],
            p["grupo"],
        ),
    )
    partidos_programados = sorted(
        [p for p in partidos_portada if p["bloque"] == "PROGRAMADOS"],
        key=lambda p: (
            0 if p["estado"] == "EN_JUEGO" else 1,
            orden_fecha_fixture(p),
            p["orden_fecha"],
            p["hora_orden"],
            p["categoria"],
            p["grupo"],
        ),
    )
    partidos_futuros = sorted(
        [p for p in partidos_portada if p["bloque"] == "FUTUROS"],
        key=lambda p: (
            orden_fecha_fixture(p),
            p["orden_fecha"],
            p["hora_orden"],
            p["categoria"],
            p["grupo"],
        ),
    )

    return render(request, "panel_principal.html", {
        "estructura": estructura,
        "torneos_menu": torneos_menu,
        "torneo_seleccionado": torneo,
        "organizador_portal_id": torneo.organizador_id,
        "categorias_menu": categorias,
        "categoria_seleccionada": categoria_seleccionada,
        "renderizar_categorias_detalle": bool(categoria_seleccionada),
        "partidos_portada": partidos_portada,
        "partidos_resultados": partidos_resultados,
        "partidos_programados": partidos_programados,
        "partidos_futuros": partidos_futuros,
        "partidos_fechas": partidos_fechas,
        "partidos_cuartos": partidos_cuartos,
        "partidos_semifinal": partidos_semifinal,
        "partidos_final": partidos_final,
        "reglamentos": documentos["reglamentos"],
        "resoluciones": documentos["resoluciones"],
        "demandas": documentos["demandas"],
        "comunicados": documentos["comunicados"],
        "planillas": planillas_publicas,
        "planillas_agrupadas": planillas_agrupadas,
        "otros": documentos["otros"],
        "logo_alcaldia": logos["logo_alcaldia"],
        "logo_torneo": logos["logo_torneo"],
        "logo_imcred": logos["logo_imcred"],
        "tiene_gestion_torneo": es_editor_torneo(request.user),
        "puede_descargar_programacion": puede_descargar_programacion(request.user),
        "tiene_equipos_delegado": equipos_delegado_asignados(request.user).exists(),
    })


def detalle_partido_publico(request, partido_id):
    return redirect("partido_live", partido_id=partido_id)


def url_retorno_descarga(request):
    return (
        request.GET.get("volver")
        or request.META.get("HTTP_REFERER")
        or reverse("panel")
    )


def url_retorno_gestion(request, fallback_name):
    fallback = reverse(fallback_name)
    volver_url = (request.POST.get("volver") or request.GET.get("volver") or "").strip()
    if volver_url and url_has_allowed_host_and_scheme(
        volver_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return volver_url
    return fallback


def respuesta_descarga_sin_partidos(request, mensaje):
    volver_url = escape(url_retorno_descarga(request))
    mensaje = escape(mensaje)
    return HttpResponse(f"""
<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Sin programación</title>
<style>
body {{
    margin: 0;
    min-height: 100vh;
    display: grid;
    place-items: center;
    background: #eef3f9;
    color: #061426;
    font-family: Arial, sans-serif;
    padding: 24px;
    box-sizing: border-box;
}}
.caja {{
    width: min(92vw, 520px);
    background: #ffffff;
    border: 1px solid #cbd5e1;
    border-radius: 18px;
    padding: 24px;
    box-shadow: 0 18px 40px rgba(15, 23, 42, 0.12);
    text-align: center;
}}
p {{
    font-size: 20px;
    font-weight: 800;
    margin: 0 0 18px;
}}
a {{
    display: inline-flex;
    align-items: center;
    justify-content: center;
    min-height: 44px;
    padding: 0 22px;
    border-radius: 999px;
    background: #00e676;
    color: #061426;
    font-weight: 900;
    text-decoration: none;
}}
</style>
</head>
<body>
<div class="caja">
    <p>{mensaje}</p>
    <a href="{volver_url}">Volver al panel</a>
</div>
<script>
setTimeout(function() {{
    window.location.href = "{volver_url}";
}}, 3500);
</script>
</body>
</html>
""")


def crear_imagen_desde_html(html, nombre_archivo, ancho=1600, alto=1800, volver_url="/"):
    return render(None, "descargas/auto_descarga.html", {
        "contenido_html": html,
        "nombre_archivo": nombre_archivo,
        "ancho": ancho,
        "alto": alto,
        "volver_url": volver_url,
    })
    carpeta_media = os.path.join(os.getcwd(), "media", "descargas")
    os.makedirs(carpeta_media, exist_ok=True)

    ruta_archivo = os.path.join(carpeta_media, nombre_archivo)

    hti = Html2Image(output_path=carpeta_media)
    hti.browser.flags = [
        "--no-sandbox",
        "--disable-dev-shm-usage",
        "--allow-file-access-from-files",
    ]

    hti.screenshot(
        html_str=html,
        save_as=nombre_archivo,
        size=(ancho, alto)
    )

    return FileResponse(
        open(ruta_archivo, "rb"),
        as_attachment=True,
        filename=nombre_archivo,
        content_type="image/png"
    )


def crear_imagenes_desde_html(paginas, volver_url="/"):
    return render(None, "descargas/auto_descarga.html", {
        "paginas": paginas,
        "volver_url": volver_url,
    })


@login_required
@user_passes_test(es_editor_torneo)
def descargar_tabla_grupo(request, categoria, grupo):
    torneo = torneo_actual(request)
    estructura = construir_estructura(torneo)
    datos_categoria = estructura.get(categoria)

    if not datos_categoria:
        return HttpResponse("Categoría no encontrada")

    datos_categoria = preparar_categoria_para_descarga(request, datos_categoria)
    datos_grupo = datos_categoria["grupos"].get(grupo)

    if not datos_grupo:
        return HttpResponse("Grupo no encontrado")

    logos = logos_torneo(request, torneo)

    html = render_to_string("descargas/tabla_grupo.html", {
        "categoria": categoria,
        "grupo": grupo,
        "datos_grupo": datos_grupo,
        "logo_alcaldia": logos["logo_alcaldia"],
        "logo_torneo": logos["logo_torneo"],
        "logo_imcred": logos["logo_imcred"],
    })

    nombre = limpiar_nombre(f"TABLA_{categoria}_{grupo}.png")
    return crear_imagen_desde_html(html, nombre, 1600, 1200, url_retorno_descarga(request))


@login_required
@user_passes_test(es_editor_torneo)
def descargar_tabla_general_mata_mata(request, categoria):
    torneo = torneo_actual(request)
    estructura = construir_estructura(torneo)
    datos_categoria = estructura.get(categoria)

    if not datos_categoria:
        return HttpResponse("Categoría no encontrada")

    datos_categoria = preparar_categoria_para_descarga(request, datos_categoria)
    tabla_general = datos_categoria.get("tabla_general_mata_mata") or []

    if not tabla_general:
        return HttpResponse("Tabla general mata-mata no encontrada")

    logos = logos_torneo(request, torneo)

    html = render_to_string("descargas/tabla_grupo.html", {
        "categoria": categoria,
        "grupo": "General mata-mata",
        "datos_grupo": {"tabla": tabla_general},
        "logo_alcaldia": logos["logo_alcaldia"],
        "logo_torneo": logos["logo_torneo"],
        "logo_imcred": logos["logo_imcred"],
    })

    nombre = limpiar_nombre(f"TABLA_GENERAL_MATA_MATA_{categoria}.png")
    return crear_imagen_desde_html(html, nombre, 1600, 1200, url_retorno_descarga(request))


@login_required
@user_passes_test(es_editor_torneo)
def descargar_goleadores_categoria(request, categoria):
    torneo = torneo_actual(request)
    estructura = construir_estructura(torneo)
    datos_categoria = estructura.get(categoria)

    if not datos_categoria:
        return HttpResponse("Categoría no encontrada")

    datos_categoria = preparar_categoria_para_descarga(request, datos_categoria)
    logos = logos_torneo(request, torneo)

    html = render_to_string("descargas/goleadores_categoria.html", {
        "categoria": categoria,
        "datos_categoria": datos_categoria,
        "logo_alcaldia": logos["logo_alcaldia"],
        "logo_torneo": logos["logo_torneo"],
        "logo_imcred": logos["logo_imcred"],
    })

    nombre = limpiar_nombre(f"GOLEADORES_{categoria}.png")
    return crear_imagen_desde_html(html, nombre, 1800, 2000, url_retorno_descarga(request))


@login_required
@user_passes_test(es_editor_torneo)
def descargar_tarjetas_categoria(request, categoria):
    torneo = torneo_actual(request)
    estructura = construir_estructura(torneo)
    datos_categoria = estructura.get(categoria)

    if not datos_categoria:
        return HttpResponse("Categoría no encontrada")

    datos_categoria = preparar_categoria_para_descarga(request, datos_categoria)
    logos = logos_torneo(request, torneo)

    html = render_to_string("descargas/tarjetas_categoria.html", {
        "categoria": categoria,
        "datos_categoria": datos_categoria,
        "logo_alcaldia": logos["logo_alcaldia"],
        "logo_torneo": logos["logo_torneo"],
        "logo_imcred": logos["logo_imcred"],
    })

    nombre = limpiar_nombre(f"TARJETAS_{categoria}.png")
    return crear_imagen_desde_html(html, nombre, 1800, 2000, url_retorno_descarga(request))


@login_required
@user_passes_test(es_editor_torneo)
def descargar_valla_categoria(request, categoria):
    torneo = torneo_actual(request)
    estructura = construir_estructura(torneo)
    datos_categoria = estructura.get(categoria)

    if not datos_categoria:
        return HttpResponse("Categoría no encontrada")

    datos_categoria = preparar_categoria_para_descarga(request, datos_categoria)
    logos = logos_torneo(request, torneo)

    html = render_to_string("descargas/valla_categoria.html", {
        "categoria": categoria,
        "datos_categoria": datos_categoria,
        "logo_alcaldia": logos["logo_alcaldia"],
        "logo_torneo": logos["logo_torneo"],
        "logo_imcred": logos["logo_imcred"],
    })

    nombre = limpiar_nombre(f"VALLA_MENOS_VENCIDA_{categoria}.png")
    return crear_imagen_desde_html(html, nombre, 1800, 1800, url_retorno_descarga(request))


@login_required
@user_passes_test(es_editor_torneo)
def descargar_foraneos_categoria(request, categoria):
    torneo = torneo_actual(request)
    estructura = construir_estructura(torneo)
    datos_categoria = estructura.get(categoria)

    if not datos_categoria:
        return HttpResponse("Categoria no encontrada")

    datos_categoria = preparar_categoria_para_descarga(request, datos_categoria)
    logos = logos_torneo(request, torneo)

    foraneos = list(datos_categoria.get("foraneos") or [])
    tamano_pagina = 22
    total_paginas = max(1, (len(foraneos) + tamano_pagina - 1) // tamano_pagina)
    paginas = []
    for indice_pagina in range(total_paginas):
        inicio = indice_pagina * tamano_pagina
        filas = [
            SimpleNamespace(numero=inicio + indice + 1, datos=fila)
            for indice, fila in enumerate(foraneos[inicio:inicio + tamano_pagina])
        ]
        datos_pagina = dict(datos_categoria)
        datos_pagina["foraneos_paginados"] = filas
        html = render_to_string("descargas/foraneos_categoria.html", {
            "categoria": categoria,
            "datos_categoria": datos_pagina,
            "pagina_actual": indice_pagina + 1,
            "total_paginas": total_paginas,
            "logo_alcaldia": logos["logo_alcaldia"],
            "logo_torneo": logos["logo_torneo"],
            "logo_imcred": logos["logo_imcred"],
            "tiene_equipos_delegado": equipos_delegado_asignados(request.user).exists(),
        })
        paginas.append({
            "contenido_html": html,
            "nombre_archivo": limpiar_nombre(
                f"FORANEOS_{categoria}_PAGINA_{indice_pagina + 1}_DE_{total_paginas}.png"
            ),
        })

    return crear_imagenes_desde_html(paginas, url_retorno_descarga(request))


@login_required
@user_passes_test(es_editor_torneo)
def descargar_imagen(request, categoria):
    torneo = torneo_actual(request)
    estructura_total = construir_estructura(torneo)

    if categoria not in estructura_total:
        return HttpResponse("Categoría no encontrada")

    estructura = {
        categoria: estructura_total[categoria]
    }

    logos = logos_torneo(request, torneo)

    html = render_to_string("panel_principal.html", {
        "estructura": estructura,
        "logo_alcaldia": logos["logo_alcaldia"],
        "logo_torneo": logos["logo_torneo"],
        "logo_imcred": logos["logo_imcred"],
    })

    nombre = limpiar_nombre(f"PANEL_{categoria}.png")
    return crear_imagen_desde_html(html, nombre, 1600, 2800, url_retorno_descarga(request))


def es_fixture_mata_mata_ida_vuelta(categoria):
    return Partido.objects.filter(
        categoria=categoria,
        fase="GRUPOS",
        grupo__startswith="MATA ",
    ).exists()


def tabla_general_mata_mata_ida_vuelta(categoria):
    partidos = Partido.objects.filter(
        categoria=categoria,
        fase="GRUPOS",
        grupo__startswith="MATA ",
    ).select_related("equipo_local", "equipo_visitante")

    if not partidos.exists():
        return []

    if partidos.exclude(estado__in=ESTADOS_PARTIDO_CERRADO).exists():
        return []

    tabla = {}

    def fila(equipo):
        return tabla.setdefault(equipo.id, {
            "id": equipo.id,
            "equipo": equipo.nombre,
            "pj": 0,
            "pg": 0,
            "pe": 0,
            "pp": 0,
            "gf": 0,
            "gc": 0,
            "dg": 0,
            "pts": 0,
        })

    for partido in partidos:
        local = fila(partido.equipo_local)
        visitante = fila(partido.equipo_visitante)
        gl = partido.goles_local or 0
        gv = partido.goles_visitante or 0
        local["pj"] += 1
        visitante["pj"] += 1
        local["gf"] += gl
        local["gc"] += gv
        visitante["gf"] += gv
        visitante["gc"] += gl

        if gl > gv:
            local["pg"] += 1
            visitante["pp"] += 1
            local["pts"] += 3
        elif gv > gl:
            visitante["pg"] += 1
            local["pp"] += 1
            visitante["pts"] += 3
        else:
            local["pe"] += 1
            visitante["pe"] += 1
            local["pts"] += 1
            visitante["pts"] += 1

    for item in tabla.values():
        item["dg"] = item["gf"] - item["gc"]

    return sorted(
        tabla.values(),
        key=lambda item: (item["pts"], item["dg"], item["gf"], item["equipo"]),
        reverse=True,
    )


def crear_partidos_mata_mata_desde_parejas(parejas):
    partidos = []

    for indice, (local, visitante) in enumerate(parejas, start=1):
        grupo = f"MATA {indice}"
        partidos.append((grupo, "1", local, visitante))
        partidos.append((grupo, "2", visitante, local))

    return partidos


def crear_partidos_mata_mata_ida_vuelta(categoria, equipos):
    equipos_sorteados = list(equipos)
    random.shuffle(equipos_sorteados)
    parejas = [
        (equipos_sorteados[indice], equipos_sorteados[indice + 1])
        for indice in range(0, len(equipos_sorteados), 2)
    ]
    return crear_partidos_mata_mata_desde_parejas(parejas)


def parejas_mata_mata_desde_formulario(equipos, request_post):
    equipos_por_id = {str(equipo.id): equipo for equipo in equipos}
    cantidad_parejas = len(equipos) // 2
    parejas = []
    seleccionados = []
    hay_manual = False

    for indice in range(cantidad_parejas):
        local_id = request_post.get(f"mata_local_{indice}") or ""
        visitante_id = request_post.get(f"mata_visitante_{indice}") or ""

        if local_id or visitante_id:
            hay_manual = True

        if not local_id and not visitante_id:
            continue

        if not local_id or not visitante_id:
            return True, [], "Completa ambos equipos en cada pareja manual o dejala totalmente vacia."

        if local_id == visitante_id:
            return True, [], "Un equipo no puede jugar contra si mismo en una pareja mata-mata."

        if local_id not in equipos_por_id or visitante_id not in equipos_por_id:
            return True, [], "Hay un equipo seleccionado que no pertenece a esta categoria."

        parejas.append((equipos_por_id[local_id], equipos_por_id[visitante_id]))
        seleccionados.extend([local_id, visitante_id])

    if not hay_manual:
        return False, [], ""

    repetidos = {equipo_id for equipo_id in seleccionados if seleccionados.count(equipo_id) > 1}
    if repetidos:
        nombres = ", ".join(equipos_por_id[equipo_id].nombre for equipo_id in repetidos)
        return True, [], f"No repitas equipos en el sorteo manual: {nombres}."

    faltantes = [equipo.nombre for equipo_id, equipo in equipos_por_id.items() if equipo_id not in seleccionados]
    if faltantes:
        return True, [], "Faltan equipos por emparejar en el sorteo manual: " + ", ".join(faltantes) + "."

    return True, parejas, ""


def grupo_completo(categoria, grupo):
    partidos = Partido.objects.filter(
        categoria=categoria,
        grupo=grupo,
        fase="GRUPOS"
    )

    if not partidos.exists():
        return False

    for partido in partidos:
        if partido.estado not in ESTADOS_PARTIDO_CERRADO:
            return False

    return True


def obtener_tabla_categoria_grupo(categoria_nombre, grupo, torneo=None):
    estructura = construir_estructura(torneo)
    datos_categoria = estructura.get(categoria_nombre)

    if not datos_categoria:
        return []

    datos_grupo = datos_categoria["grupos"].get(grupo)

    if not datos_grupo:
        return []

    return datos_grupo["tabla"]


def crear_o_actualizar_cuarto(categoria, numero, local, visitante):
    partido, creado = Partido.objects.get_or_create(
        categoria=categoria,
        fase="CUARTOS",
        numero_fecha=f"CUARTOS #{numero}",
        defaults={
            "grupo": "FINAL",
            "equipo_local": local,
            "equipo_visitante": visitante,
            "goles_local": 0,
            "goles_visitante": 0,
            "estado": "PROGRAMADO",
            "fecha": date.today(),
            "hora": time(0, 0),
            "cancha": "Por definir",
        }
    )

    if partido.estado in ESTADOS_PARTIDO_CERRADO:
        return partido

    partido.grupo = "FINAL"
    partido.equipo_local = local
    partido.equipo_visitante = visitante
    partido.estado = "PROGRAMADO"

    if not partido.fecha:
        partido.fecha = date.today()

    if not partido.hora:
        partido.hora = time(0, 0)

    if not partido.cancha:
        partido.cancha = "Por definir"

    partido.save()

    return partido


@login_required
@user_passes_test(es_editor_torneo)
def generar_llaves_cuartos(request, categoria):
    torneo = torneo_actual(request)
    categorias = Categoria.objects.filter(nombre=categoria)
    if torneo:
        categorias = categorias.filter(torneo=torneo)
    categoria_obj = categorias.first()

    if not categoria_obj:
        messages.error(request, "Categoría no encontrada.")
        return redirect("panel")

    if es_fixture_mata_mata_ida_vuelta(categoria_obj):
        tabla_general = tabla_general_mata_mata_ida_vuelta(categoria_obj)

        if not tabla_general:
            messages.error(request, "La primera ronda mata-mata todavia tiene partidos pendientes.")
            return redirect("panel")

        if len(tabla_general) < 8:
            messages.error(request, "Se necesitan al menos 8 equipos clasificados para generar los cuartos.")
            return redirect("panel")

        clasificados = tabla_general[:8]
        equipos_clasificados = {
            equipo.id: equipo
            for equipo in Equipo.objects.filter(
                categoria=categoria_obj,
                id__in=[fila["id"] for fila in clasificados],
            )
        }

        crear_o_actualizar_cuarto(categoria_obj, 1, equipos_clasificados[clasificados[0]["id"]], equipos_clasificados[clasificados[7]["id"]])
        crear_o_actualizar_cuarto(categoria_obj, 2, equipos_clasificados[clasificados[1]["id"]], equipos_clasificados[clasificados[6]["id"]])
        crear_o_actualizar_cuarto(categoria_obj, 3, equipos_clasificados[clasificados[2]["id"]], equipos_clasificados[clasificados[5]["id"]])
        crear_o_actualizar_cuarto(categoria_obj, 4, equipos_clasificados[clasificados[3]["id"]], equipos_clasificados[clasificados[4]["id"]])

        messages.success(request, f"Llaves mata-mata generadas para {categoria}: 1 vs 8, 2 vs 7, 3 vs 6 y 4 vs 5.")
        return redirect("panel")

    # PLUS 50: un solo grupo
    if categoria.upper() == "PLUS 50":
        tabla_general = obtener_tabla_categoria_grupo(categoria, "A", torneo)

        if len(tabla_general) < 4:
            messages.error(request, "No hay suficientes equipos para generar las semifinales de PLUS 50.")
            return redirect("panel")

        primero = Equipo.objects.get(nombre=tabla_general[0]["equipo"], categoria=categoria_obj)
        segundo = Equipo.objects.get(nombre=tabla_general[1]["equipo"], categoria=categoria_obj)
        tercero = Equipo.objects.get(nombre=tabla_general[2]["equipo"], categoria=categoria_obj)
        cuarto = Equipo.objects.get(nombre=tabla_general[3]["equipo"], categoria=categoria_obj)

        crear_o_actualizar_cuarto(categoria_obj, 1, primero, cuarto)
        crear_o_actualizar_cuarto(categoria_obj, 2, segundo, tercero)

        messages.success(request, f"Llaves generadas correctamente para {categoria}.")
        return redirect("panel")

    # SENIOR MASTER u otras categorías con grupos A y B
    if not grupo_completo(categoria_obj, "A"):
        messages.error(request, f"El Grupo A de {categoria} todavía tiene partidos pendientes.")
        return redirect("panel")

    if not grupo_completo(categoria_obj, "B"):
        messages.error(request, f"El Grupo B de {categoria} todavía tiene partidos pendientes.")
        return redirect("panel")

    tabla_a = obtener_tabla_categoria_grupo(categoria, "A", torneo)
    tabla_b = obtener_tabla_categoria_grupo(categoria, "B", torneo)

    if len(tabla_a) < 4 or len(tabla_b) < 4:
        messages.error(request, "No hay suficientes equipos para generar los cuartos.")
        return redirect("panel")

    primero_a = Equipo.objects.get(nombre=tabla_a[0]["equipo"], categoria=categoria_obj)
    segundo_a = Equipo.objects.get(nombre=tabla_a[1]["equipo"], categoria=categoria_obj)
    tercero_a = Equipo.objects.get(nombre=tabla_a[2]["equipo"], categoria=categoria_obj)
    cuarto_a = Equipo.objects.get(nombre=tabla_a[3]["equipo"], categoria=categoria_obj)

    primero_b = Equipo.objects.get(nombre=tabla_b[0]["equipo"], categoria=categoria_obj)
    segundo_b = Equipo.objects.get(nombre=tabla_b[1]["equipo"], categoria=categoria_obj)
    tercero_b = Equipo.objects.get(nombre=tabla_b[2]["equipo"], categoria=categoria_obj)
    cuarto_b = Equipo.objects.get(nombre=tabla_b[3]["equipo"], categoria=categoria_obj)

    crear_o_actualizar_cuarto(categoria_obj, 1, primero_a, cuarto_b)
    crear_o_actualizar_cuarto(categoria_obj, 2, segundo_a, tercero_b)
    crear_o_actualizar_cuarto(categoria_obj, 3, primero_b, cuarto_a)
    crear_o_actualizar_cuarto(categoria_obj, 4, segundo_b, tercero_a)

    messages.success(request, f"Llaves de cuartos generadas correctamente para {categoria}.")
    return redirect("panel")

def ganador_partido(partido):
    if partido.estado not in ESTADOS_PARTIDO_CERRADO:
        return None

    gl = partido.goles_local or 0
    gv = partido.goles_visitante or 0

    if gl > gv:
        return partido.equipo_local

    if gv > gl:
        return partido.equipo_visitante

    # Si el partido quedó empatado, decide por penales
    pl = partido.goles_local_penales or 0
    pv = partido.goles_visitante_penales or 0

    if pl > pv:
        return partido.equipo_local

    if pv > pl:
        return partido.equipo_visitante

    return None


def crear_o_actualizar_partido_final(categoria, fase, numero_fecha, local, visitante):
    partido, creado = Partido.objects.get_or_create(
        categoria=categoria,
        fase=fase,
        numero_fecha=numero_fecha,
        defaults={
            "grupo": "FINAL",
            "equipo_local": local,
            "equipo_visitante": visitante,
            "goles_local": 0,
            "goles_visitante": 0,
            "estado": "PROGRAMADO",
            "fecha": date.today(),
            "hora": time(0, 0),
            "cancha": "Por definir",
        }
    )

    if partido.estado in ESTADOS_PARTIDO_CERRADO:
        return partido

    partido.grupo = "FINAL"
    partido.equipo_local = local
    partido.equipo_visitante = visitante
    partido.estado = "PROGRAMADO"

    if not partido.fecha:
        partido.fecha = date.today()

    if not partido.hora:
        partido.hora = time(0, 0)

    if not partido.cancha:
        partido.cancha = "Por definir"

    partido.save()
    return partido


@login_required
@user_passes_test(es_editor_torneo)
def generar_semifinales(request, categoria):
    torneo = torneo_actual(request)
    categorias = Categoria.objects.filter(nombre=categoria)
    if torneo:
        categorias = categorias.filter(torneo=torneo)
    categoria_obj = categorias.first()

    if not categoria_obj:
        messages.error(request, "Categoría no encontrada.")
        return redirect("panel")

    q1 = Partido.objects.filter(categoria=categoria_obj, fase="CUARTOS", numero_fecha="CUARTOS #1").first()
    q2 = Partido.objects.filter(categoria=categoria_obj, fase="CUARTOS", numero_fecha="CUARTOS #2").first()
    q3 = Partido.objects.filter(categoria=categoria_obj, fase="CUARTOS", numero_fecha="CUARTOS #3").first()
    q4 = Partido.objects.filter(categoria=categoria_obj, fase="CUARTOS", numero_fecha="CUARTOS #4").first()

    if not all([q1, q2, q3, q4]):
        messages.error(request, "Primero debes generar los cuartos de final.")
        return redirect("panel")

    g1 = ganador_partido(q1)
    g2 = ganador_partido(q2)
    g3 = ganador_partido(q3)
    g4 = ganador_partido(q4)

    if not all([g1, g2, g3, g4]):
        messages.error(request, "Todos los cuartos deben estar finalizados para generar semifinales.")
        return redirect("panel")

    crear_o_actualizar_partido_final(categoria_obj, "SEMIFINAL", "SEMIFINAL #1", g1, g4)
    crear_o_actualizar_partido_final(categoria_obj, "SEMIFINAL", "SEMIFINAL #2", g2, g3)

    messages.success(request, f"Semifinales generadas correctamente para {categoria}.")
    return redirect("panel")


@login_required
@user_passes_test(es_editor_torneo)
def generar_final(request, categoria):
    torneo = torneo_actual(request)
    categorias = Categoria.objects.filter(nombre=categoria)
    if torneo:
        categorias = categorias.filter(torneo=torneo)
    categoria_obj = categorias.first()

    if not categoria_obj:
        messages.error(request, "Categoría no encontrada.")
        return redirect("panel")

    sf1 = Partido.objects.filter(categoria=categoria_obj, fase="SEMIFINAL", numero_fecha="SEMIFINAL #1").first()
    sf2 = Partido.objects.filter(categoria=categoria_obj, fase="SEMIFINAL", numero_fecha="SEMIFINAL #2").first()

    if not all([sf1, sf2]):
        messages.error(request, "Primero debes generar las semifinales.")
        return redirect("panel")

    g1 = ganador_partido(sf1)
    g2 = ganador_partido(sf2)

    if not all([g1, g2]):
        messages.error(request, "Las semifinales deben estar finalizadas para generar la final.")
        return redirect("panel")

    crear_o_actualizar_partido_final(categoria_obj, "FINAL", "FINAL", g1, g2)

    messages.success(request, f"Final generada correctamente para {categoria}.")
    return redirect("panel")

def perdedor_partido(partido):
    if partido.estado not in ESTADOS_PARTIDO_CERRADO:
        return None

    gl = partido.goles_local or 0
    gv = partido.goles_visitante or 0

    if gl > gv:
        return partido.equipo_visitante

    if gv > gl:
        return partido.equipo_local

    pl = partido.goles_local_penales or 0
    pv = partido.goles_visitante_penales or 0

    if pl > pv:
        return partido.equipo_visitante

    if pv > pl:
        return partido.equipo_local

    return None


@login_required
@user_passes_test(es_editor_torneo)
def generar_tercer_puesto(request, categoria):
    torneo = torneo_actual(request)
    categorias = Categoria.objects.filter(nombre=categoria)
    if torneo:
        categorias = categorias.filter(torneo=torneo)
    categoria_obj = categorias.first()

    if not categoria_obj:
        messages.error(request, "Categoría no encontrada.")
        return redirect("panel")

    sf1 = Partido.objects.filter(
        categoria=categoria_obj,
        fase="SEMIFINAL",
        numero_fecha="SEMIFINAL #1"
    ).first()

    sf2 = Partido.objects.filter(
        categoria=categoria_obj,
        fase="SEMIFINAL",
        numero_fecha="SEMIFINAL #2"
    ).first()

    if not all([sf1, sf2]):
        messages.error(request, "Primero debes generar las semifinales.")
        return redirect("panel")

    perdedor_1 = perdedor_partido(sf1)
    perdedor_2 = perdedor_partido(sf2)

    if not all([perdedor_1, perdedor_2]):
        messages.error(request, "Las semifinales deben estar finalizadas para generar el tercer puesto.")
        return redirect("panel")

    crear_o_actualizar_partido_final(
        categoria_obj,
        "TERCER_PUESTO",
        "TERCER PUESTO",
        perdedor_1,
        perdedor_2
    )

    messages.success(request, f"Partido por tercer puesto generado correctamente para {categoria}.")
    return redirect("panel")
def construir_partidos_programacion(request, categoria_obj=None, numero_fecha="", dia=None, incluir_resultados=False):
    dias_semana = {
        0: "LUNES",
        1: "MARTES",
        2: "MIÉRCOLES",
        3: "JUEVES",
        4: "VIERNES",
        5: "SÁBADO",
        6: "DOMINGO",
    }
    meses = {
        1: "ENERO",
        2: "FEBRERO",
        3: "MARZO",
        4: "ABRIL",
        5: "MAYO",
        6: "JUNIO",
        7: "JULIO",
        8: "AGOSTO",
        9: "SEPTIEMBRE",
        10: "OCTUBRE",
        11: "NOVIEMBRE",
        12: "DICIEMBRE",
    }

    torneo = torneo_actual(request)
    partidos = Partido.objects.filter(
        estado_programacion__in=["MANUAL", "OFICIAL"],
        fecha__isnull=False,
        hora__isnull=False,
        cancha__isnull=False,
    ).exclude(
        cancha=""
    ).exclude(
        cancha__iexact="Por definir"
    ).exclude(
        hora=time(0, 0)
    ).select_related(
        "categoria",
        "equipo_local",
        "equipo_visitante"
    ).order_by(
        "fecha",
        "cancha",
        "hora",
        "categoria__nombre",
        "grupo",
        "fase"
    )

    if torneo:
        partidos = partidos.filter(categoria__torneo=torneo)

    if not incluir_resultados:
        partidos = partidos.filter(estado="PROGRAMADO")

    if categoria_obj:
        partidos = partidos.filter(categoria=categoria_obj)

    if numero_fecha:
        partidos = partidos.filter(numero_fecha__iexact=numero_fecha)

    if dia:
        partidos = partidos.filter(fecha=dia)

    partidos_programacion = []

    for p in partidos:
        dia_semana = ""

        if p.fecha:
            dia_semana = dias_semana[p.fecha.weekday()]

        fase = p.fase or "GRUPOS"
        fecha_corta = ""

        if p.fecha:
            fecha_corta = f"{dia_semana} {p.fecha.day} {meses[p.fecha.month]}"

        hora_12 = ""

        if p.hora:
            hora = p.hora.hour
            periodo = "AM" if hora < 12 else "PM"
            hora_12 = hora % 12 or 12
            minuto = f":{p.hora.minute:02d}"
            hora_12 = f"{hora_12}{minuto} {periodo}"

        categoria_nombre = p.categoria.nombre if p.categoria else ""
        categoria_normalizada = limpiar_nombre(categoria_nombre)
        color_categoria = "#f1db19"
        color_texto_categoria = "#111827"

        if "PLUS_50" in categoria_normalizada:
            color_categoria = "#075985"
            color_texto_categoria = "#ffffff"
        elif "INTERBARRIOS" in categoria_normalizada:
            color_categoria = "#166534"
            color_texto_categoria = "#ffffff"
        elif "SENIOR" in categoria_normalizada:
            color_categoria = "#f1db19"
            color_texto_categoria = "#111827"

        partidos_programacion.append({
            "categoria": categoria_nombre,
            "color_categoria": color_categoria,
            "color_texto_categoria": color_texto_categoria,
            "bloque": f"{fecha_corta} / CANCHA {p.cancha}",
            "hora_texto": hora_12,
            "numero_fecha": p.numero_fecha or "",
            "grupo": p.grupo,
            "fase": fase,
            "estado": p.estado,
            "finalizado": p.estado == "FINALIZADO",
            "marcador_texto": f"{p.goles_local} - {p.goles_visitante}" if p.estado == "FINALIZADO" else "VS",
            "dia_semana": dia_semana,
            "fecha": p.fecha,
            "hora": p.hora,
            "cancha": p.cancha,
            "local": p.equipo_local.nombre if p.equipo_local else "POR DEFINIR",
            "visitante": p.equipo_visitante.nombre if p.equipo_visitante else "POR DEFINIR",
            "escudo_local": url_absoluta(request, escudo_url(p.equipo_local)),
            "escudo_visitante": url_absoluta(request, escudo_url(p.equipo_visitante)),
        })

    return partidos_programacion


def medidas_programacion(cantidad):
    if cantidad > 8:
        filas = (cantidad + 1) // 2
        alto_disponible = 1600
        espacio_entre_tarjetas = 12 * max(filas - 1, 0)
        alto_tarjeta = max(95, min(270, (alto_disponible - espacio_entre_tarjetas) // filas))

        if alto_tarjeta >= 240:
            fuente_detalle = 18
            fuente_grupo = 19
            fuente_equipo = 21
            escudo = 56
        elif alto_tarjeta >= 180:
            fuente_detalle = 15
            fuente_grupo = 16
            fuente_equipo = 18
            escudo = 44
        else:
            fuente_detalle = 13
            fuente_grupo = 14
            fuente_equipo = 15
            escudo = 34

        return {
            "ancho": 1080,
            "alto": 1920,
            "compacta": True,
            "alto_tarjeta": alto_tarjeta,
            "fuente_detalle": fuente_detalle,
            "fuente_grupo": fuente_grupo,
            "fuente_equipo": fuente_equipo,
            "escudo": escudo,
        }

    if cantidad <= 4:
        alto = 1920
    elif cantidad <= 8:
        alto = 2850
    else:
        alto = 650 + (cantidad * 270)

    return {
        "ancho": 1080,
        "alto": alto,
        "compacta": False,
        "alto_tarjeta": 0,
        "fuente_detalle": 0,
        "fuente_grupo": 0,
        "fuente_equipo": 0,
        "escudo": 0,
    }


def filtros_descarga_programacion(request):
    torneo = torneo_actual(request)
    categoria_id = (request.GET.get("categoria") or "").strip()
    numero_fecha = (request.GET.get("fecha_fixture") or "").strip()
    dia = fecha_desde_texto(request.GET.get("dia"))
    categoria_obj = None

    if categoria_id:
        categorias = Categoria.objects.filter(id=categoria_id)
        if torneo:
            categorias = categorias.filter(torneo=torneo)
        categoria_obj = categorias.first()

    return torneo, categoria_obj, numero_fecha, dia


def titulo_descarga_programacion(categoria_obj=None, numero_fecha="", dia=None):
    partes = []
    if categoria_obj:
        partes.append(categoria_obj.nombre)
    else:
        partes.append("TODAS LAS CATEGORIAS")
    if numero_fecha:
        partes.append(str(numero_fecha).upper())
    if dia:
        partes.append(f"DIA {dia.strftime('%d/%m/%Y')}")
    return " - ".join(partes)


@login_required
@user_passes_test(puede_descargar_programacion)
def seleccionar_descarga_programacion(request):
    torneo = torneo_actual(request)
    partidos = Partido.objects.filter(
        estado="PROGRAMADO",
        estado_programacion__in=["MANUAL", "OFICIAL"],
        fecha__isnull=False,
        hora__isnull=False,
        cancha__isnull=False,
    ).exclude(cancha="").exclude(cancha__iexact="Por definir").exclude(hora=time(0, 0))
    if torneo:
        partidos = partidos.filter(categoria__torneo=torneo)

    categorias = Categoria.objects.filter(partido__in=partidos).distinct().order_by("nombre")
    fechas_fixture = partidos.exclude(numero_fecha="").values_list("numero_fecha", flat=True).distinct().order_by("numero_fecha")
    dias = partidos.values_list("fecha", flat=True).distinct().order_by("fecha")

    return render(request, "gestion/descargar_programacion.html", {
        "torneo_seleccionado": torneo,
        "categorias": categorias,
        "fechas_fixture": fechas_fixture,
        "dias": dias,
        "volver_url": url_retorno_descarga(request),
    })


@login_required
@user_passes_test(puede_descargar_programacion)
def descargar_programacion_categoria(request, categoria):
    torneo = torneo_actual(request)
    categorias = Categoria.objects.filter(nombre=categoria)
    if torneo:
        categorias = categorias.filter(torneo=torneo)
    categoria_obj = categorias.first()
    numero_fecha = (request.GET.get("fecha_fixture") or "").strip()
    dia = fecha_desde_texto(request.GET.get("dia"))

    if not categoria_obj:
        return HttpResponse("Categoría no encontrada")

    partidos_programacion = construir_partidos_programacion(
        request,
        categoria_obj,
        numero_fecha,
        dia,
        incluir_resultados=True,
    )

    if not partidos_programacion:
        return respuesta_descarga_sin_partidos(request, "No hay partidos programados con fecha, hora y cancha para esta categoria.")

    logos = logos_torneo(request, torneo)
    cantidad = len(partidos_programacion)
    medidas = medidas_programacion(cantidad)

    html = render_to_string("descargas/programacion_categoria.html", {
        "categoria": titulo_descarga_programacion(categoria_obj, numero_fecha, dia),
        "partidos": partidos_programacion,
        "ancho": medidas["ancho"],
        "compacta": medidas["compacta"],
        "alto_tarjeta": medidas["alto_tarjeta"],
        "fuente_detalle": medidas["fuente_detalle"],
        "fuente_grupo": medidas["fuente_grupo"],
        "fuente_equipo": medidas["fuente_equipo"],
        "escudo": medidas["escudo"],
        "logo_alcaldia": logos["logo_alcaldia"],
        "logo_torneo": logos["logo_torneo"],
        "logo_imcred": logos["logo_imcred"],
    })

    nombre = limpiar_nombre(f"PROGRAMACION_PARTIDOS_PROGRAMADOS_{titulo_descarga_programacion(categoria_obj, numero_fecha, dia)}.png")
    return crear_imagen_desde_html(html, nombre, medidas["ancho"], medidas["alto"], url_retorno_descarga(request))


@login_required
@user_passes_test(puede_descargar_programacion)
def descargar_programacion_general(request):
    torneo, categoria_obj, numero_fecha, dia = filtros_descarga_programacion(request)
    partidos_programacion = construir_partidos_programacion(request, categoria_obj, numero_fecha, dia)

    if not partidos_programacion:
        return respuesta_descarga_sin_partidos(request, "No hay partidos programados con fecha, hora y cancha asignada.")

    logos = logos_torneo(request, torneo)
    cantidad = len(partidos_programacion)
    medidas = medidas_programacion(cantidad)
    titulo_programacion = titulo_descarga_programacion(categoria_obj, numero_fecha, dia)

    html = render_to_string("descargas/programacion_categoria.html", {
        "categoria": titulo_programacion,
        "mostrar_categoria": not categoria_obj,
        "partidos": partidos_programacion,
        "ancho": medidas["ancho"],
        "compacta": medidas["compacta"],
        "alto_tarjeta": medidas["alto_tarjeta"],
        "fuente_detalle": medidas["fuente_detalle"],
        "fuente_grupo": medidas["fuente_grupo"],
        "fuente_equipo": medidas["fuente_equipo"],
        "escudo": medidas["escudo"],
        "logo_alcaldia": logos["logo_alcaldia"],
        "logo_torneo": logos["logo_torneo"],
        "logo_imcred": logos["logo_imcred"],
    })

    nombre = limpiar_nombre(f"PROGRAMACION_{titulo_programacion}.png")
    return crear_imagen_desde_html(html, nombre, medidas["ancho"], medidas["alto"], url_retorno_descarga(request))


# ======================================================
# EDITOR MÓVIL PROFESIONAL DE PARTIDOS
# ======================================================

def _jugadores_del_partido(partido):
    jugadores_local = Jugador.objects.filter(
        equipo=partido.equipo_local,
        estado='ACTIVO'
    ).order_by('dorsal', 'nombres')

    jugadores_visitante = Jugador.objects.filter(
        equipo=partido.equipo_visitante,
        estado='ACTIVO'
    ).order_by('dorsal', 'nombres')

    return jugadores_local, jugadores_visitante


def _validar_jugador_equipo(jugador, equipo, partido):
    equipos_validos = [partido.equipo_local_id, partido.equipo_visitante_id]
    return jugador.equipo_id == equipo.id and equipo.id in equipos_validos


def _volver_editor_partido_url(request, partido):
    fallback = (
        reverse("planillero_mis_partidos")
        if es_planillero_asignado(request.user)
        else f"{reverse('panel')}?torneo={partido.categoria.torneo_id}"
    )
    volver_url = (request.POST.get("volver") or request.GET.get("volver") or "").strip()
    if volver_url and url_has_allowed_host_and_scheme(
        volver_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return volver_url
    return fallback


def _url_editor_tab(partido_id, tab, volver_url=""):
    url = reverse('editor_partido_movil', args=[partido_id])
    if volver_url:
        url = f"{url}?volver={quote(volver_url, safe='')}"
    return f"{url}#{tab}"


def _url_editor_partido(request, partido, tab=""):
    volver_url = _volver_editor_partido_url(request, partido)
    if tab:
        return _url_editor_tab(partido.id, tab, volver_url)
    return f"{reverse('editor_partido_movil', args=[partido.id])}?volver={quote(volver_url, safe='')}"


POSICIONES_ALINEACION_DEFAULT = [codigo for codigo, _ in AlineacionPartido.POSICIONES_CANCHA]


def _orden_partido(partido):
    return (
        partido.fecha or date.min,
        partido.hora or time.min,
        partido.id or 0,
    )


def _partidos_equipo_antes(partido, equipo):
    fecha_partido, hora_partido, _ = _orden_partido(partido)
    return Partido.objects.filter(
        categoria=partido.categoria,
        estado__in=ESTADOS_PARTIDO_CERRADO,
        estadisticas_validadas=True,
    ).filter(
        Q(equipo_local=equipo) | Q(equipo_visitante=equipo)
    ).filter(
        Q(fecha__lt=fecha_partido) |
        Q(fecha=fecha_partido, hora__lt=hora_partido) |
        Q(fecha=fecha_partido, hora=hora_partido, id__lt=partido.id)
    ).order_by("fecha", "hora", "id")


def _jugadores_sancionados_por_tarjetas(partido):
    sancionados = {}

    for equipo in [partido.equipo_local, partido.equipo_visitante]:
        partidos_previos = list(_partidos_equipo_antes(partido, equipo))
        if not partidos_previos:
            continue

        ultimo_partido = partidos_previos[-1]
        tarjetas_previas = Tarjeta.objects.filter(
            partido__in=partidos_previos,
            equipo=equipo,
        ).select_related("partido", "jugador", "equipo")

        tarjetas_por_jugador_partido = defaultdict(lambda: {"A": 0, "R": 0})
        amarillas_grupos_por_jugador = defaultdict(set)

        for tarjeta in tarjetas_previas:
            resumen = tarjetas_por_jugador_partido[(tarjeta.jugador_id, tarjeta.partido_id)]
            if tarjeta.tipo == "AMARILLA":
                resumen["A"] += 1
                if (tarjeta.partido.fase or "GRUPOS") == "GRUPOS":
                    amarillas_grupos_por_jugador[tarjeta.jugador_id].add(tarjeta.partido_id)
            elif tarjeta.tipo == "ROJA":
                resumen["R"] += 1

        jugadores_con_tarjeta = {
            jugador_id
            for jugador_id, _ in tarjetas_por_jugador_partido.keys()
        }

        for jugador_id in jugadores_con_tarjeta:
            tarjetas_ultimo = tarjetas_por_jugador_partido.get((jugador_id, ultimo_partido.id), {"A": 0, "R": 0})
            motivo = ""

            if tarjetas_ultimo["R"] > 0:
                motivo = "roja directa"
            elif tarjetas_ultimo["A"] >= 2:
                motivo = "doble amarilla"
            elif (ultimo_partido.fase or "GRUPOS") == "GRUPOS" and tarjetas_ultimo["A"] > 0:
                amarillas_hasta_ultimo = len(amarillas_grupos_por_jugador[jugador_id])
                amarillas_antes_ultimo = amarillas_hasta_ultimo - 1
                if amarillas_antes_ultimo < 3 <= amarillas_hasta_ultimo:
                    motivo = "3 amarillas en fase 1"

            if motivo:
                sancionados[jugador_id] = {
                    "equipo_id": equipo.id,
                    "motivo": motivo,
                    "partido_origen": ultimo_partido,
                }

    return sancionados


def _sincronizar_no_disponibles_por_tarjetas(partido):
    if partido.estado in ESTADOS_PARTIDO_CERRADO:
        return {}

    sancionados = _jugadores_sancionados_por_tarjetas(partido)
    if not sancionados:
        return {}

    alineaciones = AlineacionPartido.objects.filter(
        partido=partido,
        jugador_id__in=sancionados.keys(),
    )
    alineaciones_por_jugador = {alineacion.jugador_id: alineacion for alineacion in alineaciones}
    nuevas = []

    for jugador_id, data in sancionados.items():
        alineacion = alineaciones_por_jugador.get(jugador_id)
        if alineacion:
            if alineacion.rol != "NO_DISPONIBLE" or alineacion.posicion_cancha:
                alineacion.rol = "NO_DISPONIBLE"
                alineacion.posicion_cancha = ""
                alineacion.equipo_id = data["equipo_id"]
                alineacion.save(update_fields=["rol", "posicion_cancha", "equipo"])
        else:
            nuevas.append(AlineacionPartido(
                partido=partido,
                equipo_id=data["equipo_id"],
                jugador_id=jugador_id,
                rol="NO_DISPONIBLE",
                posicion_cancha="",
            ))

    if nuevas:
        AlineacionPartido.objects.bulk_create(nuevas, ignore_conflicts=True)

    return sancionados


def _marcar_roles_alineacion(jugadores, alineaciones_por_jugador, partido=None):
    for jugador in jugadores:
        alineacion = alineaciones_por_jugador.get(jugador.id)
        jugador.rol_alineacion = alineacion.rol if alineacion else ""
        jugador.posicion_alineacion = alineacion.posicion_cancha if alineacion else ""
        jugador.documento_validado_alineacion = bool(alineacion and alineacion.documento_validado)
        jugador.foto_alineacion = foto_jugador_url(jugador)
        jugador.iniciales_alineacion = iniciales_jugador(jugador)
        fecha_referencia = partido.fecha if partido else date.today()
        jugador.etiqueta_edad = etiqueta_edad_jugador(
            jugador,
            partido.categoria if partido else None,
            fecha_referencia,
        )
        jugador.texto_edad = texto_edad_jugador(
            jugador,
            partido.categoria if partido else None,
            fecha_referencia,
        )
    return jugadores



def asignar_posiciones_titulares_automaticas(seleccionados, indice_rol=1, indice_posicion=2):
    posiciones_usadas = {
        item[indice_posicion]
        for item in seleccionados
        if item[indice_rol] == "TITULAR" and item[indice_posicion]
    }
    posiciones_libres = [
        codigo
        for codigo, _ in AlineacionPartido.POSICIONES_CANCHA
        if codigo not in posiciones_usadas
    ]
    random.shuffle(posiciones_libres)

    nuevos = []
    for item in seleccionados:
        valores = list(item)
        if valores[indice_rol] == "TITULAR" and not valores[indice_posicion] and posiciones_libres:
            valores[indice_posicion] = posiciones_libres.pop(0)
        nuevos.append(tuple(valores))
    return nuevos


def preparar_dorsales_alineacion(request, equipo, jugadores):
    """Valida los dorsales publicados y devuelve los jugadores que deben actualizarse."""
    jugadores_editables = {jugador.id for jugador in jugadores}
    jugadores = list(Jugador.objects.filter(equipo=equipo).only("id", "nombres", "dorsal"))
    dorsales_propuestos = {jugador.id: jugador.dorsal for jugador in jugadores}
    errores = []

    for jugador in jugadores:
        llave = f"dorsal_{jugador.id}"
        if jugador.id not in jugadores_editables or llave not in request.POST:
            continue
        valor = (request.POST.get(llave) or "").strip()
        if not valor:
            dorsales_propuestos[jugador.id] = None
        elif not valor.isdigit() or not 1 <= int(valor) <= 999:
            # Un dorsal antiguo o mal digitado nunca debe impedir guardar la alineación.
            # Se conserva el valor que ya tenía el jugador.
            continue
        else:
            dorsales_propuestos[jugador.id] = int(valor)

    if errores:
        return [], errores

    actualizados = []
    for jugador in jugadores:
        nuevo_dorsal = dorsales_propuestos[jugador.id]
        if jugador.dorsal != nuevo_dorsal:
            jugador.dorsal = nuevo_dorsal
            actualizados.append(jugador)
    return actualizados, []


def guardar_dorsales_alineacion(request, equipo, jugadores_actualizados):
    if not jugadores_actualizados:
        return
    Jugador.objects.bulk_update(jugadores_actualizados, ["dorsal"])
    registrar_actividad(
        request,
        "ACTUALIZAR_DORSALES_ALINEACION",
        equipo,
        descripcion=(
            f"Actualizó dorsales desde la alineación de {equipo.nombre}: "
            + ", ".join(
                f"{jugador.nombres} #{jugador.dorsal}" if jugador.dorsal else f"{jugador.nombres} sin dorsal"
                for jugador in jugadores_actualizados
            )
        ),
        datos={
            "equipo_id": equipo.id,
            "jugadores": [
                {"jugador_id": jugador.id, "dorsal": jugador.dorsal}
                for jugador in jugadores_actualizados
            ],
        },
    )

def _registrar_alertas_validacion_alineacion(partido, equipo, user, seleccionados, errores_edad):
    titulares = [int(jugador_id) for jugador_id, rol, _, _ in seleccionados if rol == "TITULAR"]
    documentos_faltantes = [int(jugador_id) for jugador_id, rol, _, documento_ok in seleccionados if rol == "TITULAR" and not documento_ok]

    if not documentos_faltantes and not errores_edad:
        return

    jugadores = {
        jugador.id: jugador
        for jugador in Jugador.objects.filter(id__in=titulares).only("id", "nombres", "cedula", "dorsal")
    }
    faltantes = [jugadores[jugador_id] for jugador_id in documentos_faltantes if jugador_id in jugadores]
    nombres_faltantes = [jugador.nombres for jugador in faltantes]
    partes = []

    if nombres_faltantes:
        partes.append("Titulares sin documento validado: " + ", ".join(nombres_faltantes) + ".")
    if errores_edad:
        partes.append("Reglas de edad: " + " ".join(errores_edad))

    crear_solicitud_validacion(
        "ALINEACION",
        f"Validar alineacion: {equipo.nombre}",
        descripcion=" ".join(partes),
        user=user,
        partido=partido,
        equipo=equipo,
        datos={
            "equipo_id": equipo.id,
            "partido_id": partido.id,
            "titulares": [
                {
                    "jugador_id": jugador_id,
                    "nombre": jugadores[jugador_id].nombres if jugador_id in jugadores else "",
                    "cedula": jugadores[jugador_id].cedula if jugador_id in jugadores else "",
                    "documento_validado": jugador_id not in documentos_faltantes,
                }
                for jugador_id in titulares
            ],
            "documentos_faltantes": documentos_faltantes,
            "errores_edad": errores_edad,
        },
    )


def _ordenar_titulares_cancha(items):
    usadas = {item.posicion for item in items if item.posicion}
    disponibles = [posicion for posicion in POSICIONES_ALINEACION_DEFAULT if posicion not in usadas]
    for item in items:
        if not item.posicion:
            item.posicion = disponibles.pop(0) if disponibles else ""
    return sorted(items, key=lambda item: AlineacionPartido.ORDEN_POSICIONES_CANCHA.get(item.posicion, 99))


def _recalcular_marcador_por_goles(partido):
    goles_local = 0
    goles_visitante = 0

    for gol in Gol.objects.filter(partido=partido).only("equipo_id", "cantidad", "es_autogol"):
        cantidad = max(gol.cantidad or 1, 1)
        if gol.es_autogol:
            if gol.equipo_id == partido.equipo_local_id:
                goles_visitante += cantidad
            elif gol.equipo_id == partido.equipo_visitante_id:
                goles_local += cantidad
        elif gol.equipo_id == partido.equipo_local_id:
            goles_local += cantidad
        elif gol.equipo_id == partido.equipo_visitante_id:
            goles_visitante += cantidad

    partido.goles_local = goles_local
    partido.goles_visitante = goles_visitante
    partido.save(update_fields=["goles_local", "goles_visitante"])


def _minuto_evento_en_vivo(partido):
    if partido.estado != "EN_JUEGO":
        return None

    segundos = partido.segundos_acumulados or 0
    if partido.inicio_en_vivo and not partido.cronometro_pausado:
        segundos += max(int((timezone.now() - partido.inicio_en_vivo).total_seconds()), 0)

    return max(segundos // 60, 1)


def _clave_orden_evento_resumen(evento):
    creado_en = getattr(evento, "creado_en", None)
    return (
        evento.minuto is not None,
        evento.minuto if evento.minuto is not None else -1,
        creado_en.timestamp() if creado_en else 0,
        evento.orden or 0,
    )


def _grupo_tiempo_evento_resumen(evento):
    minuto = evento.minuto
    if minuto is None:
        return "primero", ""
    if minuto >= 46:
        return "segundo", ""
    return "primero", ""


def _agrupar_eventos_resumen_live(eventos):
    grupos = {
        "segundo": [],
        "primero": [],
    }
    titulos = {
        "segundo": "",
        "primero": "",
    }
    for evento in eventos:
        codigo, _titulo = _grupo_tiempo_evento_resumen(evento)
        grupos[codigo].append(evento)

    resultado = []
    for codigo in ("segundo", "primero"):
        if not grupos[codigo]:
            continue
        resultado.append(SimpleNamespace(
            codigo=codigo,
            titulo=titulos[codigo],
            eventos=grupos[codigo],
            mostrar_descanso=codigo == "primero" and bool(grupos["segundo"]),
        ))
    return resultado


@login_required
def editor_partido_movil(request, partido_id):
    partido = get_object_or_404(
        Partido.objects.select_related('categoria', 'equipo_local', 'equipo_visitante'),
        id=partido_id
    )
    equipo_delegado = equipo_delegado_para_partido(request.user, partido)
    if equipo_delegado:
        if puede_editar_alineacion_delegado(request.user, partido, equipo_delegado):
            return redirect("delegado_alineacion_partido", equipo_id=equipo_delegado.id, partido_id=partido.id)
        _, motivo = ventana_alineacion_delegado(partido, equipo_delegado)
        return HttpResponseForbidden(f"Los delegados solo pueden editar la alineacion de su equipo. {motivo}")
    if not puede_diligenciar_partido(request.user, partido):
        return denegar_partido_no_autorizado()
    sancionados_tarjetas = _sincronizar_no_disponibles_por_tarjetas(partido)

    jugadores_local, jugadores_visitante = _jugadores_del_partido(partido)

    goles = Gol.objects.filter(partido=partido).select_related('jugador', 'equipo').order_by('equipo__nombre', 'jugador__nombres')
    tarjetas = Tarjeta.objects.filter(partido=partido).select_related('jugador', 'equipo').order_by('equipo__nombre', 'tipo', 'jugador__nombres')
    alineaciones = AlineacionPartido.objects.filter(partido=partido).select_related('jugador', 'equipo').order_by('equipo__nombre', 'rol', 'jugador__nombres')
    sustituciones = SustitucionPartido.objects.filter(partido=partido).select_related('equipo', 'jugador_sale', 'jugador_entra').order_by('equipo__nombre', 'minuto', 'id')
    for equipo_partido in (partido.equipo_local, partido.equipo_visitante):
        actualizar_incidencia_regla_edad(partido, equipo_partido, request=request)
    incidencias_reglas_edad = IncidenciaReglaEdad.objects.filter(partido=partido).select_related(
        "equipo", "sustitucion_inicio", "sustitucion_inicio__jugador_sale", "sustitucion_inicio__jugador_entra",
    ).order_by("-iniciada_en", "-id")
    alineaciones_por_jugador = {alineacion.jugador_id: alineacion for alineacion in alineaciones}
    jugadores_local = _marcar_roles_alineacion(jugadores_local, alineaciones_por_jugador, partido)
    jugadores_visitante = _marcar_roles_alineacion(jugadores_visitante, alineaciones_por_jugador, partido)
    jugadores_en_cancha_local = jugadores_actuales_en_cancha(partido, partido.equipo_local)
    jugadores_en_cancha_visitante = jugadores_actuales_en_cancha(partido, partido.equipo_visitante)
    jugadores_que_salieron = set(
        SustitucionPartido.objects.filter(partido=partido).values_list("jugador_sale_id", flat=True)
    )
    permite_reingresos = categoria_permite_reingresos(partido.categoria)
    for jugador in jugadores_local:
        jugador.en_cancha_actual = jugador.id in jugadores_en_cancha_local
        jugador.puede_entrar_actual = (
            not jugador.en_cancha_actual
            and (permite_reingresos or jugador.id not in jugadores_que_salieron)
        )
    for jugador in jugadores_visitante:
        jugador.en_cancha_actual = jugador.id in jugadores_en_cancha_visitante
        jugador.puede_entrar_actual = (
            not jugador.en_cancha_actual
            and (permite_reingresos or jugador.id not in jugadores_que_salieron)
        )
    volver_url = _volver_editor_partido_url(request, partido)

    return render(request, 'editor_partido_movil.html', {
        'partido': partido,
        'jugadores_local': jugadores_local,
        'jugadores_visitante': jugadores_visitante,
        'goles': goles,
        'tarjetas': tarjetas,
        'alineaciones': alineaciones,
        'sustituciones': sustituciones,
        'incidencias_reglas_edad': incidencias_reglas_edad,
        'tolerancia_regla_edad_segundos': TOLERANCIA_REGLA_EDAD_SEGUNDOS,
        'segundos_vivos': segundos_vivos_partido(partido),
        'incidencia_reloj_activo': partido.estado == "EN_JUEGO" and not partido.cronometro_pausado and partido.periodo_en_vivo != "ET",
        'estados_partido': (
            Partido.ESTADOS
            if es_editor_torneo(request.user)
            else [(valor, etiqueta) for valor, etiqueta in Partido.ESTADOS if valor in ESTADOS_PLANILLERO_PARTIDO]
        ),
        'fases_partido': Partido.FASES,
        'posiciones_cancha': AlineacionPartido.POSICIONES_CANCHA,
        'reglas_edad_alineacion': reglas_edad_para_frontend(partido.categoria),
        'sancionados_tarjetas': sancionados_tarjetas,
        'puede_editar_programacion': es_editor_torneo(request.user),
        'ajuste_puntos_local_abs': abs(partido.ajuste_puntos_local or 0),
        'ajuste_puntos_visitante_abs': abs(partido.ajuste_puntos_visitante or 0),
        'ajuste_puntos_local_signo': '-' if (partido.ajuste_puntos_local or 0) < 0 else '+',
        'ajuste_puntos_visitante_signo': '-' if (partido.ajuste_puntos_visitante or 0) < 0 else '+',
        'editor_volver_url': volver_url,
        'editor_volver_text': "Mis partidos" if es_planillero_asignado(request.user) else "Panel",
        'editor_live_url': f"{reverse('partido_live', args=[partido.id])}?volver={quote(volver_url, safe='')}",
    })


@login_required
@require_POST
def guardar_info_partido_movil(request, partido_id):
    partido = get_object_or_404(Partido, id=partido_id)
    if not puede_diligenciar_partido(request.user, partido):
        return denegar_partido_no_autorizado()

    minimo_goles = None if es_editor_torneo(request.user) else 0
    partido.goles_local = entero_post(request, 'goles_local', 0, minimo_goles)
    partido.goles_visitante = entero_post(request, 'goles_visitante', 0, minimo_goles)
    estado_solicitado = request.POST.get('estado') or partido.estado
    if es_editor_torneo(request.user) or estado_solicitado in ESTADOS_PLANILLERO_PARTIDO:
        partido.estado = estado_solicitado

    if partido.estado == "EN_JUEGO" and not partido.inicio_en_vivo:
        partido.inicio_en_vivo = timezone.now()

    partido.goles_local_penales = entero_post(request, 'goles_local_penales', 0, 0)
    partido.goles_visitante_penales = entero_post(request, 'goles_visitante_penales', 0, 0)
    partido.observaciones = request.POST.get('observaciones') or ''

    if es_editor_torneo(request.user):
        partido.fecha = request.POST.get('fecha') or partido.fecha
        partido.hora = request.POST.get('hora') or partido.hora
        partido.cancha = request.POST.get('cancha') or ''
        partido.numero_fecha = request.POST.get('numero_fecha') or ''
        partido.grupo = request.POST.get('grupo') or ''
        partido.fase = request.POST.get('fase') or partido.fase
        partido.ajuste_puntos_local = entero_post(request, 'ajuste_puntos_local', 0)
        partido.ajuste_puntos_visitante = entero_post(request, 'ajuste_puntos_visitante', 0)
        partido.observacion_comite = request.POST.get('observacion_comite') or ''
    partido.save()
    _marcar_estadisticas_pendientes(partido, request.user)

    messages.success(request, 'Partido actualizado correctamente.')
    if not es_editor_torneo(request.user) and partido.estado == "FINALIZADO":
        return redirect('partido_live', partido_id=partido.id)
    return redirect(_url_editor_partido(request, partido))


@login_required
@require_POST
def agregar_gol_movil(request, partido_id):
    partido = get_object_or_404(Partido, id=partido_id)
    if not puede_diligenciar_partido(request.user, partido):
        return denegar_partido_no_autorizado()
    jugador_id = request.POST.get('jugador')
    equipo_id = request.POST.get('equipo')
    cantidad = request.POST.get('cantidad') or 1
    es_autogol = request.POST.get('es_autogol') == '1'
    es_penal = request.POST.get('es_penal') == '1'
    try:
        cantidad = max(int(cantidad), 1)
    except (TypeError, ValueError):
        cantidad = 1

    if jugador_id and equipo_id:
        jugador = get_object_or_404(Jugador, id=jugador_id)
        equipo = get_object_or_404(Equipo, id=equipo_id)

        if _validar_jugador_equipo(jugador, equipo, partido):
            Gol.objects.create(
                partido=partido,
                jugador=jugador,
                equipo=equipo,
                cantidad=cantidad,
                es_autogol=es_autogol,
                es_penal=es_penal,
                minuto=_minuto_evento_en_vivo(partido),
            )
            _recalcular_marcador_por_goles(partido)
            _marcar_estadisticas_pendientes(partido, request.user)
            if es_autogol:
                messages.success(request, 'Autogol agregado correctamente.')
            elif es_penal:
                messages.success(request, 'Gol de penal agregado correctamente.')
            else:
                messages.success(request, 'Gol agregado correctamente.')
        else:
            messages.error(request, 'El jugador no pertenece al equipo seleccionado.')

    return redirect(_url_editor_partido(request, partido))


@login_required
@require_POST
def agregar_tarjeta_movil(request, partido_id):
    partido = get_object_or_404(Partido, id=partido_id)
    if not puede_diligenciar_partido(request.user, partido):
        return denegar_partido_no_autorizado()
    jugador_id = request.POST.get('jugador')
    equipo_id = request.POST.get('equipo')
    tipo = request.POST.get('tipo')

    if jugador_id and equipo_id and tipo:
        jugador = get_object_or_404(Jugador, id=jugador_id)
        equipo = get_object_or_404(Equipo, id=equipo_id)

        if _validar_jugador_equipo(jugador, equipo, partido):
            Tarjeta.objects.create(
                partido=partido,
                jugador=jugador,
                equipo=equipo,
                tipo=tipo,
                minuto=_minuto_evento_en_vivo(partido),
            )
            _marcar_estadisticas_pendientes(partido, request.user)
            messages.success(request, 'Tarjeta agregada correctamente.')
        else:
            messages.error(request, 'El jugador no pertenece al equipo seleccionado.')

    return redirect(_url_editor_partido(request, partido))


@login_required
@require_POST
def agregar_alineacion_movil(request, partido_id):
    partido = get_object_or_404(Partido, id=partido_id)
    if not puede_diligenciar_partido(request.user, partido):
        return denegar_partido_no_autorizado()
    sancionados_tarjetas = _sincronizar_no_disponibles_por_tarjetas(partido)
    jugador_id = request.POST.get('jugador')
    equipo_id = request.POST.get('equipo')
    rol = request.POST.get('rol') or 'TITULAR'
    posicion_cancha = request.POST.get('posicion_cancha') or ''
    posiciones_validas = {codigo for codigo, _ in AlineacionPartido.POSICIONES_CANCHA}
    if posicion_cancha not in posiciones_validas:
        posicion_cancha = ''

    if jugador_id and equipo_id:
        jugador = get_object_or_404(Jugador, id=jugador_id)
        equipo = get_object_or_404(Equipo, id=equipo_id)

        if _validar_jugador_equipo(jugador, equipo, partido):
            if jugador.id in sancionados_tarjetas and rol != "NO_DISPONIBLE":
                messages.error(request, 'Este jugador esta sancionado por tarjetas y queda como no disponible.')
                return redirect(_url_editor_partido(request, partido, "alineacion"))
            AlineacionPartido.objects.update_or_create(
                partido=partido,
                jugador=jugador,
                defaults={
                    'equipo': equipo,
                    'rol': rol,
                    'posicion_cancha': posicion_cancha if rol == 'TITULAR' else '',
                    'documento_validado': False,
                    'documento_validado_por': None,
                    'documento_validado_en': None,
                }
            )
            _marcar_estadisticas_pendientes(partido, request.user)
            messages.success(request, 'Jugador agregado a la alineación.')
        else:
            messages.error(request, 'El jugador no pertenece al equipo seleccionado.')

    return redirect(_url_editor_partido(request, partido))


@login_required
@require_POST
def guardar_alineacion_masiva_movil(request, partido_id):
    partido = get_object_or_404(Partido, id=partido_id)
    if not puede_diligenciar_partido(request.user, partido):
        return denegar_partido_no_autorizado()
    sancionados_tarjetas = _sincronizar_no_disponibles_por_tarjetas(partido)
    equipo_id = request.POST.get("equipo")
    equipo = get_object_or_404(Equipo, id=equipo_id)

    if equipo.id not in [partido.equipo_local_id, partido.equipo_visitante_id]:
        messages.error(request, "Ese equipo no pertenece al partido.")
        return redirect(_url_editor_partido(request, partido, "alineacion"))

    jugadores_equipo = list(Jugador.objects.filter(equipo=equipo).only("id", "nombres", "dorsal"))
    jugadores_validos = {str(jugador.id) for jugador in jugadores_equipo}
    dorsales_actualizados, errores_dorsales = preparar_dorsales_alineacion(request, equipo, jugadores_equipo)
    if errores_dorsales:
        messages.error(request, " ".join(errores_dorsales))
        return redirect(_url_editor_partido(request, partido, "alineacion"))
    roles_validos = {"TITULAR", "SUPLENTE", "NO_DISPONIBLE"}
    posiciones_validas = {codigo for codigo, _ in AlineacionPartido.POSICIONES_CANCHA}
    sancionados_equipo = {
        str(jugador_id)
        for jugador_id, data in sancionados_tarjetas.items()
        if data["equipo_id"] == equipo.id
    }
    posiciones_usadas = set()
    documentos_validados = set(request.POST.getlist("documento_validado"))
    seleccionados = []

    for llave, rol in request.POST.items():
        if not llave.startswith("rol_") or rol not in roles_validos:
            continue

        jugador_id = llave.replace("rol_", "", 1)
        if jugador_id in jugadores_validos:
            posicion = request.POST.get(f"posicion_{jugador_id}") or ""
            if jugador_id in sancionados_equipo:
                rol = "NO_DISPONIBLE"
            if rol == "TITULAR":
                if posicion not in posiciones_validas:
                    posicion = ""
                if posicion and posicion in posiciones_usadas:
                    messages.error(request, "No repitas la misma posición en la cancha.")
                    return redirect(_url_editor_partido(request, partido, "alineacion"))
                if posicion:
                    posiciones_usadas.add(posicion)
            else:
                posicion = ""
            documento_validado = rol == "TITULAR" and jugador_id in documentos_validados
            seleccionados.append((jugador_id, rol, posicion, documento_validado))

    seleccionados_ids = {jugador_id for jugador_id, _, _, _ in seleccionados}
    for jugador_id in sancionados_equipo - seleccionados_ids:
        if jugador_id in jugadores_validos:
            seleccionados.append((jugador_id, "NO_DISPONIBLE", "", False))

    titulares = [jugador_id for jugador_id, rol, _, _ in seleccionados if rol == "TITULAR"]
    if len(titulares) > 11:
        messages.error(request, "Solo puedes seleccionar 11 titulares por equipo.")
        return redirect(_url_editor_partido(request, partido, "alineacion"))
    errores_edad = validar_reglas_edad_titulares(partido, equipo, titulares)
    seleccionados = asignar_posiciones_titulares_automaticas(seleccionados, indice_rol=1, indice_posicion=2)

    with transaction.atomic():
        guardar_dorsales_alineacion(request, equipo, dorsales_actualizados)
        AlineacionPartido.objects.filter(partido=partido, equipo=equipo).delete()
        ahora_validacion = timezone.now()
        nuevas_alineaciones = [
            AlineacionPartido(
                partido=partido,
                equipo=equipo,
                jugador_id=jugador_id,
                rol=rol,
                posicion_cancha=posicion,
                documento_validado=documento_validado,
                documento_validado_por=request.user if documento_validado else None,
                documento_validado_en=ahora_validacion if documento_validado else None,
            )
            for jugador_id, rol, posicion, documento_validado in seleccionados
        ]
        AlineacionPartido.objects.bulk_create(nuevas_alineaciones)
    _marcar_estadisticas_pendientes(partido, request.user)
    _registrar_alertas_validacion_alineacion(partido, equipo, request.user, seleccionados, errores_edad)

    documentos_faltantes = [jugador_id for jugador_id, rol, _, documento_ok in seleccionados if rol == "TITULAR" and not documento_ok]
    if documentos_faltantes:
        messages.warning(
            request,
            f"Cedulas pendientes por validar en titulares: {len(documentos_faltantes)}. Quedo registro para validacion."
        )

    if sancionados_equipo:
        messages.warning(
            request,
            "Los jugadores sancionados por tarjetas quedaron como no disponibles."
        )
    if errores_edad:
        messages.warning(
            request,
            "Advertencia de reglas de edad: " + " ".join(errores_edad)
        )
    messages.success(
        request,
        f"Alineacion de {equipo.nombre} guardada: {len(titulares)} titulares, "
        f"{sum(1 for _, rol, _, _ in seleccionados if rol == 'SUPLENTE')} suplentes."
    )
    return redirect(_url_editor_partido(request, partido, "alineacion"))


@login_required
@require_POST
def agregar_sustitucion_movil(request, partido_id):
    partido = get_object_or_404(Partido, id=partido_id)
    if not puede_diligenciar_partido(request.user, partido):
        return denegar_partido_no_autorizado()
    equipo_id = request.POST.get('equipo')
    jugador_sale_id = request.POST.get('jugador_sale')
    jugador_entra_id = request.POST.get('jugador_entra')
    minuto = request.POST.get('minuto') or _minuto_evento_en_vivo(partido)
    observacion = request.POST.get('observacion') or ''

    if equipo_id and jugador_sale_id and jugador_entra_id:
        equipo = get_object_or_404(Equipo, id=equipo_id)
        jugador_sale = get_object_or_404(Jugador, id=jugador_sale_id)
        jugador_entra = get_object_or_404(Jugador, id=jugador_entra_id)

        jugadores_en_cancha = jugadores_actuales_en_cancha(partido, equipo)
        jugadores_validos = (
            _validar_jugador_equipo(jugador_sale, equipo, partido)
            and _validar_jugador_equipo(jugador_entra, equipo, partido)
        )

        if not jugadores_validos:
            messages.error(request, 'Los jugadores deben pertenecer al equipo seleccionado.')
        elif jugador_sale.id == jugador_entra.id:
            messages.error(request, 'El jugador que entra debe ser distinto del jugador que sale.')
        elif jugador_sale.id not in jugadores_en_cancha:
            messages.error(request, f'{jugador_sale.nombres} no puede salir porque no está actualmente en cancha.')
        elif jugador_entra.id in jugadores_en_cancha:
            messages.error(request, f'{jugador_entra.nombres} no puede entrar porque ya está actualmente en cancha.')
        elif (
            not categoria_permite_reingresos(partido.categoria)
            and SustitucionPartido.objects.filter(
                partido=partido,
                equipo=equipo,
                jugador_sale=jugador_entra,
            ).exists()
        ):
            messages.error(
                request,
                f'{jugador_entra.nombres} no puede volver a ingresar en esta categoría después de haber salido.',
            )
        else:
            with transaction.atomic():
                sustitucion = SustitucionPartido.objects.create(
                    partido=partido,
                    equipo=equipo,
                    jugador_sale=jugador_sale,
                    jugador_entra=jugador_entra,
                    minuto=minuto,
                    observacion=observacion
                )
                alineacion_entra, creada = AlineacionPartido.objects.get_or_create(
                    partido=partido,
                    jugador=jugador_entra,
                    defaults={
                        "equipo": equipo,
                        "rol": "SUPLENTE",
                        "posicion_cancha": "",
                    },
                )
                if not creada and alineacion_entra.rol == "NO_DISPONIBLE":
                    alineacion_entra.equipo = equipo
                    alineacion_entra.rol = "SUPLENTE"
                    alineacion_entra.posicion_cancha = ""
                    alineacion_entra.save(update_fields=["equipo", "rol", "posicion_cancha"])
                actualizar_incidencia_regla_edad(
                    partido,
                    equipo,
                    request=request,
                    sustitucion=sustitucion,
                    permitir_crear=True,
                )
            _marcar_estadisticas_pendientes(partido, request.user)
            messages.success(request, 'Sustitución agregada correctamente.')

    return redirect(_url_editor_partido(request, partido))


@login_required
@require_POST
def eliminar_gol_movil(request, gol_id):
    gol = get_object_or_404(Gol, id=gol_id)
    partido_id = gol.partido_id
    partido = gol.partido
    if not puede_diligenciar_partido(request.user, partido):
        return denegar_partido_no_autorizado()
    gol.delete()
    _recalcular_marcador_por_goles(partido)
    _marcar_estadisticas_pendientes(partido, request.user)
    messages.success(request, 'Gol eliminado.')
    return redirect(_url_editor_partido(request, partido))


@login_required
@require_POST
def eliminar_tarjeta_movil(request, tarjeta_id):
    tarjeta = get_object_or_404(Tarjeta, id=tarjeta_id)
    partido_id = tarjeta.partido_id
    if not puede_diligenciar_partido(request.user, tarjeta.partido):
        return denegar_partido_no_autorizado()
    tarjeta.delete()
    _marcar_estadisticas_pendientes(tarjeta.partido, request.user)
    messages.success(request, 'Tarjeta eliminada.')
    return redirect(_url_editor_partido(request, tarjeta.partido))


@login_required
@require_POST
def eliminar_alineacion_movil(request, alineacion_id):
    alineacion = get_object_or_404(AlineacionPartido, id=alineacion_id)
    partido_id = alineacion.partido_id
    if not puede_diligenciar_partido(request.user, alineacion.partido):
        return denegar_partido_no_autorizado()
    alineacion.delete()
    _marcar_estadisticas_pendientes(alineacion.partido, request.user)
    messages.success(request, 'Jugador eliminado de la alineación.')
    return redirect(_url_editor_partido(request, alineacion.partido))


@login_required
@require_POST
def eliminar_sustitucion_movil(request, sustitucion_id):
    sustitucion = get_object_or_404(SustitucionPartido, id=sustitucion_id)
    partido_id = sustitucion.partido_id
    if not puede_diligenciar_partido(request.user, sustitucion.partido):
        return denegar_partido_no_autorizado()
    partido = sustitucion.partido
    equipo = sustitucion.equipo
    sustitucion.delete()
    actualizar_incidencia_regla_edad(partido, equipo, request=request)
    _marcar_estadisticas_pendientes(sustitucion.partido, request.user)
    messages.success(request, 'Sustitución eliminada.')
    return redirect(_url_editor_partido(request, sustitucion.partido))

def lista_equipos(request):
    equipos = Equipo.objects.select_related(
        'categoria'
    ).order_by(
        'categoria__nombre',
        'nombre'
    )

    return render(request, 'equipos/lista_equipos.html', {
        'equipos': equipos
    })


def detalle_equipo(request, equipo_id):
    equipo = get_object_or_404(
        Equipo.objects.select_related('categoria'),
        id=equipo_id
    )

    jugadores = Jugador.objects.filter(
        equipo=equipo
    ).order_by(
        'dorsal',
        'nombres'
    )

    return render(request, 'equipos/detalle_equipo.html', {
        'equipo': equipo,
        'jugadores': jugadores
    })

@login_required
def mis_equipos(request):
    torneo = torneo_actual_delegado(request)
    equipos_qs = equipos_delegado_asignados(request.user, torneo=torneo) if torneo else Equipo.objects.none()
    equipos = list(equipos_qs.order_by('categoria__nombre', 'nombre'))
    ahora = timezone.now()
    for equipo in equipos:
        equipo.acceso_vigente_delegado = equipo.acceso_delegado_vigente()
        equipo.puede_editar_datos_delegado = puede_editar_equipo_delegado(request.user, equipo)
        equipo.puede_cargar_fotos_delegado = puede_cargar_fotos_jugadores_delegado(request.user, equipo)
        if not equipo.acceso_delegado_hasta:
            equipo.estado_acceso_delegado = "Sin fecha de acceso asignada."
        elif equipo.acceso_delegado_hasta < ahora:
            equipo.estado_acceso_delegado = f"Acceso vencido el {timezone.localtime(equipo.acceso_delegado_hasta).strftime('%d/%m/%Y %H:%M')}."
        else:
            equipo.estado_acceso_delegado = f"Disponible hasta {timezone.localtime(equipo.acceso_delegado_hasta).strftime('%d/%m/%Y %H:%M')}."

    return render(request, 'equipos/mis_equipos.html', {
        'equipos': equipos,
        'torneo_actual': torneo,
    })


@login_required
def delegado_equipo_editar(request, equipo_id):
    equipo = get_object_or_404(equipos_alineacion_delegado_actual(request), id=equipo_id)
    if not puede_editar_equipo_delegado(request.user, equipo):
        if puede_cargar_fotos_jugadores_delegado(request.user, equipo):
            messages.warning(request, "La edicion de datos esta bloqueada. Solo puedes cargar fotos de jugadores.")
            return redirect("delegado_fotos_jugadores", equipo_id=equipo.id)
        messages.warning(request, "La edicion del equipo esta bloqueada. Puedes cargar la alineacion de partidos desde aqui.")
        return redirect("delegado_partidos_equipo", equipo_id=equipo.id)

    form = EquipoDelegadoForm(request.POST or None, request.FILES or None, instance=equipo)
    jugadores = equipo.jugadores.order_by("dorsal", "nombres")

    if request.method == "POST" and form.is_valid():
        equipo = form.save(commit=False)
        aplicar_imagen_cloudinary(
            equipo,
            "escudo",
            request.POST.get("imagen_cloudinary"),
            request.FILES.get("escudo"),
        )
        equipo.save()
        crear_solicitud_validacion(
            "EQUIPO",
            f"Validar cambios del equipo {equipo.nombre}",
            descripcion=f"El delegado actualizo datos generales del equipo {equipo.nombre}.",
            user=request.user,
            equipo=equipo,
            datos={"equipo_id": equipo.id},
        )
        registrar_actividad(
            request,
            "EDITAR_EQUIPO_DELEGADO",
            equipo,
            descripcion=f"El delegado actualizó los datos del equipo {equipo.nombre}.",
            datos={"equipo_id": equipo.id, "escudo_actualizado": bool(request.FILES.get("escudo") or request.POST.get("imagen_cloudinary"))},
        )
        messages.success(request, "Equipo actualizado correctamente.")
        return redirect("delegado_equipo_editar", equipo_id=equipo.id)

    return render(request, "equipos/delegado_equipo_formulario.html", {
        "titulo": f"Editar equipo: {equipo.nombre}",
        "equipo": equipo,
        "form": form,
        "jugadores": jugadores,
        "escudo_actual": escudo_url(equipo),
        "cloudinary_images": listar_imagenes_cloudinary(),
        "cloudinary_label": "Seleccionar escudo existente de Cloudinary",
    })


@login_required
def delegado_fotos_jugadores(request, equipo_id):
    equipo = get_object_or_404(equipos_alineacion_delegado_actual(request), id=equipo_id)
    if not puede_cargar_fotos_jugadores_delegado(request.user, equipo):
        return HttpResponseForbidden("No tienes permiso para cargar fotos de jugadores de este equipo.")

    jugadores = list(equipo.jugadores.order_by("dorsal", "nombres"))

    if request.method == "POST":
        actualizados = 0
        for jugador in jugadores:
            form = JugadorFotoDelegadoForm(
                request.POST,
                request.FILES,
                instance=jugador,
                prefix=f"jugador_{jugador.id}",
            )
            if form.is_valid() and form.cleaned_data.get("foto"):
                form.save()
                actualizados += 1
        if actualizados:
            crear_solicitud_validacion(
                "JUGADOR",
                f"Validar fotos de jugadores: {equipo.nombre}",
                descripcion=f"El delegado cargo o actualizo {actualizados} foto(s) de jugadores en {equipo.nombre}.",
                user=request.user,
                equipo=equipo,
                datos={"equipo_id": equipo.id, "accion": "FOTOS_JUGADORES", "cantidad": actualizados},
            )
            registrar_actividad(
                request,
                "CARGAR_FOTOS_JUGADORES",
                equipo,
                descripcion=f"El delegado cargó o actualizó {actualizados} foto(s) de jugadores de {equipo.nombre}.",
                datos={"equipo_id": equipo.id, "cantidad": actualizados},
            )
            messages.success(request, f"Fotos actualizadas: {actualizados}.")
        else:
            request._actividad_registrada = True
            messages.info(request, "No seleccionaste fotos nuevas para cargar.")
        return redirect("delegado_fotos_jugadores", equipo_id=equipo.id)

    filas = [
        {
            "jugador": jugador,
            "form": JugadorFotoDelegadoForm(instance=jugador, prefix=f"jugador_{jugador.id}"),
            "foto_url": foto_jugador_url(jugador),
        }
        for jugador in jugadores
    ]

    return render(request, "equipos/delegado_fotos_jugadores.html", {
        "titulo": f"Fotos de jugadores: {equipo.nombre}",
        "equipo": equipo,
        "filas": filas,
    })


@login_required
def delegado_partidos_equipo(request, equipo_id):
    equipo = get_object_or_404(equipos_alineacion_delegado_actual(request), id=equipo_id)

    return render(request, "equipos/delegado_partidos_equipo.html", {
        "equipo": equipo,
        "torneo_actual": equipo.categoria.torneo if equipo.categoria_id else None,
        "partidos_alineacion": partidos_alineacion_para_equipo(equipo),
    })


@login_required
def delegado_alineacion_partido(request, equipo_id, partido_id):
    equipo = get_object_or_404(equipos_alineacion_delegado_actual(request), id=equipo_id)
    partido = get_object_or_404(
        Partido.objects.select_related("categoria", "equipo_local", "equipo_visitante"),
        id=partido_id,
        categoria=equipo.categoria,
    )
    if not partido_pertenece_equipo(partido, equipo):
        return HttpResponseForbidden("Este equipo no pertenece al partido.")
    if not puede_editar_alineacion_delegado(request.user, partido, equipo):
        _, motivo = ventana_alineacion_delegado(partido, equipo)
        return HttpResponseForbidden(f"No puedes editar esta alineacion en este momento. {motivo}")

    sancionados_tarjetas = _sincronizar_no_disponibles_por_tarjetas(partido)
    jugadores = list(Jugador.objects.filter(equipo=equipo, estado="ACTIVO").order_by("dorsal", "nombres"))
    alineaciones = AlineacionPartido.objects.filter(partido=partido, equipo=equipo).select_related("jugador")
    alineaciones_por_jugador = {alineacion.jugador_id: alineacion for alineacion in alineaciones}
    jugadores = _marcar_roles_alineacion(jugadores, alineaciones_por_jugador, partido)
    sancionados_equipo = {
        jugador_id: data
        for jugador_id, data in sancionados_tarjetas.items()
        if data["equipo_id"] == equipo.id
    }

    if request.method == "POST":
        accion_envio = request.POST.get("accion") or "guardar_borrador"
        es_definitiva = accion_envio == "enviar_definitiva"
        jugadores_validos = {str(jugador.id) for jugador in jugadores}
        roles_validos = {"TITULAR", "SUPLENTE", "NO_DISPONIBLE"}
        posiciones_validas = {codigo for codigo, _ in AlineacionPartido.POSICIONES_CANCHA}
        posiciones_usadas = set()
        jugadores_en_cancha = {}
        seleccionados = []
        dorsales_actualizados, errores_dorsales = preparar_dorsales_alineacion(request, equipo, jugadores)
        if errores_dorsales:
            messages.error(request, " ".join(errores_dorsales))
            return redirect("delegado_alineacion_partido", equipo_id=equipo.id, partido_id=partido.id)

        for posicion in posiciones_validas:
            jugador_id = request.POST.get(f"cancha_{posicion}") or ""
            if not jugador_id:
                continue
            if jugador_id not in jugadores_validos:
                continue
            if jugador_id in jugadores_en_cancha:
                messages.error(request, "No repitas el mismo jugador en la cancha.")
                return redirect("delegado_alineacion_partido", equipo_id=equipo.id, partido_id=partido.id)
            if int(jugador_id) in sancionados_equipo:
                messages.error(request, "Un jugador sancionado no puede quedar como titular.")
                return redirect("delegado_alineacion_partido", equipo_id=equipo.id, partido_id=partido.id)
            jugadores_en_cancha[jugador_id] = posicion
            posiciones_usadas.add(posicion)

        for jugador in jugadores:
            jugador_id = str(jugador.id)
            rol = request.POST.get(f"rol_{jugador_id}") or ""
            if jugador_id in jugadores_en_cancha:
                rol = "TITULAR"
            if rol not in roles_validos:
                continue
            if jugador_id not in jugadores_validos:
                continue
            if jugador.id in sancionados_equipo:
                rol = "NO_DISPONIBLE"
            posicion = jugadores_en_cancha.get(jugador_id) or request.POST.get(f"posicion_{jugador_id}") or ""
            if rol == "TITULAR":
                if posicion not in posiciones_validas:
                    posicion = ""
                if posicion and posicion not in jugadores_en_cancha.values() and posicion in posiciones_usadas:
                    messages.error(request, "No repitas la misma posicion en la cancha.")
                    return redirect("delegado_alineacion_partido", equipo_id=equipo.id, partido_id=partido.id)
                if posicion:
                    posiciones_usadas.add(posicion)
            else:
                posicion = ""
            seleccionados.append((jugador.id, rol, posicion))

        seleccionados_ids = {jugador_id for jugador_id, _, _ in seleccionados}
        for jugador_id in sancionados_equipo.keys() - seleccionados_ids:
            seleccionados.append((jugador_id, "NO_DISPONIBLE", ""))

        titulares = [jugador_id for jugador_id, rol, _ in seleccionados if rol == "TITULAR"]
        if len(titulares) > 11:
            messages.error(request, "Solo puedes seleccionar 11 titulares.")
            return redirect("delegado_alineacion_partido", equipo_id=equipo.id, partido_id=partido.id)

        errores_edad = validar_reglas_edad_titulares(partido, equipo, titulares)
        seleccionados = asignar_posiciones_titulares_automaticas(seleccionados, indice_rol=1, indice_posicion=2)
        with transaction.atomic():
            guardar_dorsales_alineacion(request, equipo, dorsales_actualizados)
            AlineacionPartido.objects.filter(partido=partido, equipo=equipo).delete()
            AlineacionPartido.objects.bulk_create([
                AlineacionPartido(partido=partido, equipo=equipo, jugador_id=jugador_id, rol=rol, posicion_cancha=posicion)
                for jugador_id, rol, posicion in seleccionados
            ])
        _marcar_estadisticas_pendientes(partido, request.user)

        if es_definitiva:
            EntregaAlineacionPartido.objects.get_or_create(
                partido=partido,
                equipo=equipo,
                defaults={"enviada_por": request.user},
            )
            registrar_actividad(
                request,
                "ENVIAR_ALINEACION_DEFINITIVA",
                partido,
                descripcion=f"El delegado envió la alineación definitiva de {equipo.nombre}.",
                datos={"equipo_id": equipo.id, "partido_id": partido.id},
            )
        else:
            registrar_actividad(
                request,
                "GUARDAR_BORRADOR_ALINEACION",
                partido,
                descripcion=f"El delegado guardó un borrador de alineación de {equipo.nombre}.",
                datos={
                    "equipo_id": equipo.id,
                    "partido_id": partido.id,
                    "titulares": len(titulares),
                    "suplentes": sum(1 for _, rol, _ in seleccionados if rol == "SUPLENTE"),
                    "no_disponibles": sum(1 for _, rol, _ in seleccionados if rol == "NO_DISPONIBLE"),
                },
            )

        if sancionados_equipo:
            messages.warning(request, "Los jugadores sancionados quedaron como no disponibles.")
        if errores_edad:
            messages.warning(request, "Advertencia de reglas de edad: " + " ".join(errores_edad))
        if es_definitiva:
            messages.success(request, f"Alineación definitiva enviada para {equipo.nombre}.")
            return redirect("delegado_partidos_equipo", equipo_id=equipo.id)
        messages.success(request, f"Borrador de alineación guardado para {equipo.nombre}.")
        return redirect("delegado_alineacion_partido", equipo_id=equipo.id, partido_id=partido.id)

    return render(request, "equipos/delegado_alineacion_partido.html", {
        "equipo": equipo,
        "partido": partido,
        "jugadores": jugadores,
        "posiciones_cancha": AlineacionPartido.POSICIONES_CANCHA,
        "reglas_edad_alineacion": reglas_edad_para_frontend(partido.categoria),
        "sancionados_tarjetas": sancionados_equipo,
    })


@login_required
def delegado_jugador_nuevo(request, equipo_id):
    equipo = get_object_or_404(equipos_editables_delegado_actual(request), id=equipo_id)
    if not puede_editar_equipo_delegado(request.user, equipo):
        return HttpResponseForbidden("El acceso a este equipo ya no esta vigente.")

    form = JugadorDelegadoForm(
        request.POST or None,
        request.FILES or None,
        permitir_foto=puede_cargar_fotos_jugadores_delegado(request.user, equipo),
    )

    if request.method == "POST" and form.is_valid():
        jugador = form.save(commit=False)
        jugador.equipo = equipo
        jugador.nombres = jugador.nombres.upper()
        jugador.save()
        crear_solicitud_validacion(
            "JUGADOR",
            f"Validar jugador nuevo: {jugador.nombres}",
            descripcion=f"El delegado agrego a {jugador.nombres} en {equipo.nombre}.",
            user=request.user,
            equipo=equipo,
            jugador=jugador,
            datos={"equipo_id": equipo.id, "jugador_id": jugador.id, "accion": "CREAR"},
        )
        registrar_actividad(
            request,
            "CREAR_JUGADOR_DELEGADO",
            jugador,
            descripcion=f"El delegado agregó a {jugador.nombres} en {equipo.nombre}.",
            datos={"equipo_id": equipo.id, "jugador_id": jugador.id, "foto_cargada": bool(request.FILES.get("foto"))},
        )
        messages.success(request, "Jugador agregado correctamente.")
        return redirect("delegado_equipo_editar", equipo_id=equipo.id)

    return render(request, "equipos/delegado_jugador_formulario.html", {
        "titulo": f"Agregar jugador: {equipo.nombre}",
        "equipo": equipo,
        "form": form,
    })


@login_required
def delegado_jugador_editar(request, jugador_id):
    jugador = get_object_or_404(
        Jugador.objects.select_related("equipo", "equipo__categoria").filter(equipo__in=equipos_editables_delegado_actual(request)),
        id=jugador_id,
    )
    if not puede_editar_equipo_delegado(request.user, jugador.equipo):
        return HttpResponseForbidden("No tienes permiso para editar este jugador.")

    form = JugadorDelegadoForm(
        request.POST or None,
        request.FILES or None,
        instance=jugador,
        permitir_foto=puede_cargar_fotos_jugadores_delegado(request.user, jugador.equipo),
    )

    if request.method == "POST" and form.is_valid():
        jugador = form.save(commit=False)
        jugador.nombres = jugador.nombres.upper()
        jugador.save()
        crear_solicitud_validacion(
            "JUGADOR",
            f"Validar cambios de jugador: {jugador.nombres}",
            descripcion=f"El delegado actualizo datos de {jugador.nombres} en {jugador.equipo.nombre}.",
            user=request.user,
            equipo=jugador.equipo,
            jugador=jugador,
            datos={"equipo_id": jugador.equipo_id, "jugador_id": jugador.id, "accion": "EDITAR"},
        )
        registrar_actividad(
            request,
            "EDITAR_JUGADOR_DELEGADO",
            jugador,
            descripcion=f"El delegado actualizó a {jugador.nombres} en {jugador.equipo.nombre}.",
            datos={"equipo_id": jugador.equipo_id, "jugador_id": jugador.id, "foto_actualizada": bool(request.FILES.get("foto"))},
        )
        messages.success(request, "Jugador actualizado correctamente.")
        return redirect("delegado_equipo_editar", equipo_id=jugador.equipo_id)

    return render(request, "equipos/delegado_jugador_formulario.html", {
        "titulo": f"Editar jugador: {jugador.nombres}",
        "equipo": jugador.equipo,
        "form": form,
        "jugador": jugador,
    })


@login_required
@require_POST
def delegado_jugador_eliminar(request, jugador_id):
    jugador = get_object_or_404(
        Jugador.objects.select_related("equipo", "equipo__categoria").filter(equipo__in=equipos_editables_delegado_actual(request)),
        id=jugador_id,
    )
    equipo_id = jugador.equipo_id
    if not puede_editar_equipo_delegado(request.user, jugador.equipo):
        return HttpResponseForbidden("No tienes permiso para eliminar este jugador.")

    nombre = jugador.nombres
    equipo = jugador.equipo
    crear_solicitud_validacion(
        "JUGADOR",
        f"Validar eliminacion de jugador: {nombre}",
        descripcion=f"El delegado elimino a {nombre} del equipo {equipo.nombre}.",
        user=request.user,
        equipo=equipo,
        datos={"equipo_id": equipo.id, "jugador_id": jugador.id, "accion": "ELIMINAR", "jugador": nombre},
    )
    registrar_actividad(
        request,
        "ELIMINAR_JUGADOR_DELEGADO",
        jugador,
        descripcion=f"El delegado eliminó a {nombre} del equipo {equipo.nombre}.",
        datos={"equipo_id": equipo.id, "jugador_id": jugador.id, "jugador": nombre},
    )
    jugador.delete()
    messages.success(request, f"Jugador eliminado: {nombre}.")
    return redirect("delegado_equipo_editar", equipo_id=equipo_id)


@login_required
@user_passes_test(es_editor_torneo)
def crear_jugador_equipo(request, equipo_id):
    equipo = get_object_or_404(
        Equipo,
        id=equipo_id,
        responsable=request.user
    )

    if request.method == 'POST':
        nombres = request.POST.get('nombres')
        dorsal = request.POST.get('dorsal')
        cedula = request.POST.get('cedula')
        fecha_nacimiento = request.POST.get('fecha_nacimiento')
        foto = request.FILES.get('foto')

        Jugador.objects.create(
            equipo=equipo,
            nombres=nombres,
            dorsal=dorsal,
            cedula=cedula,
            fecha_nacimiento=fecha_nacimiento,
            foto=foto
        )

        return redirect('detalle_equipo', equipo_id=equipo.id)

    return render(request, 'equipos/crear_jugador.html', {
        'equipo': equipo
    })


@login_required
@user_passes_test(es_editor_torneo)
def gestion_probar_storage(request):
    def subir_prueba():
        nombre_archivo = f"pruebas/render-test-{datetime.now().strftime('%Y%m%d%H%M%S')}.txt"
        nombre = default_storage.save(nombre_archivo, ContentFile(b"ok"))
        return nombre, default_storage.url(nombre)

    try:
        executor = ThreadPoolExecutor(max_workers=1)
        future = executor.submit(subir_prueba)
        try:
            nombre, url = future.result(timeout=10)
        finally:
            executor.shutdown(wait=False, cancel_futures=True)
    except TimeoutError:
        return HttpResponse(
            "ERROR STORAGE\n\nTimeoutError: Supabase Storage no respondio en 10 segundos desde Render.",
            status=504,
            content_type="text/plain",
        )
    except Exception as exc:
        return HttpResponse(
            f"ERROR STORAGE\n\n{type(exc).__name__}: {exc}",
            status=500,
            content_type="text/plain",
        )

    return HttpResponse(
        f"STORAGE OK\n\nArchivo: {nombre}\nURL: {url}",
        content_type="text/plain",
    )


@login_required
@user_passes_test(es_editor_torneo)
def gestion_panel(request):
    torneo = torneo_actual(request)
    permisos = permisos_torneo_usuario(request.user, torneo)
    puede_editar = bool(permisos and permisos.puede_editar)
    puede_validar = bool(permisos and permisos.puede_validar)
    puede_programar = bool(permisos and permisos.puede_programar)
    puede_descargar_planillas = bool(permisos and getattr(permisos, "puede_descargar_planillas", False))
    organizadores = Organizador.objects.all() if tabla_disponible("torneos_organizador") else None
    categorias = Categoria.objects.all()
    equipos = Equipo.objects.all()
    jugadores = Jugador.objects.all()
    partidos = Partido.objects.all()
    documentos = Documento.objects.all()

    if torneo:
        categorias = categorias.filter(torneo=torneo)
        equipos = equipos.filter(categoria__torneo=torneo)
        jugadores = jugadores.filter(equipo__categoria__torneo=torneo)
        partidos = partidos.filter(categoria__torneo=torneo)
        documentos = documentos.filter(torneo=torneo)

    return render(request, "gestion/panel.html", {
        "torneo_seleccionado": torneo,
        "total_organizadores": organizadores.count() if organizadores is not None else 0,
        "total_categorias": categorias.count(),
        "total_equipos": equipos.count(),
        "total_jugadores": jugadores.count(),
        "total_partidos": partidos.count(),
        "total_documentos": documentos.count(),
        "puede_editar": puede_editar,
        "puede_validar": puede_validar,
        "puede_programar": puede_programar,
        "puede_descargar_planillas": puede_descargar_planillas,
        "puede_cargar_planillas_juego": puede_cargar_planillas_juego(request.user),
        "puede_gestionar_organizadores": puede_gestionar_organizadores(request.user),
    })


@login_required
def gestion_actividad(request):
    es_delegado = equipos_delegado_asignados(request.user).exists()
    if not request.user.is_superuser and not es_delegado:
        return HttpResponseForbidden("La auditoría está disponible únicamente para superusuarios y delegados asignados.")
    if request.GET.get("formato") == "csv" and not request.user.is_superuser:
        return HttpResponseForbidden("La descarga de auditoría está disponible únicamente para superusuarios.")
    if usuario_solo_descarga_planillas(request.user, torneo_actual(request)):
        return denegar_permiso_torneo()
    if not tabla_disponible("torneos_registroactividad"):
        messages.error(request, "La tabla de actividad todavia no esta creada. Ejecuta las migraciones.")
        return redirect("gestion_panel")

    torneo = torneo_actual(request)
    registros = RegistroActividad.objects.select_related("usuario", "torneo").order_by("-creado_en")
    if es_delegado and not request.user.is_superuser:
        torneos_delegado = Torneo.objects.filter(
            categorias__equipos__responsable=request.user,
        ).distinct()
        registros = registros.filter(
            Q(torneo__in=torneos_delegado)
            | Q(usuario=request.user, torneo__isnull=True)
        )
    elif torneo:
        registros = registros.filter(torneo=torneo)

    registros_disponibles = registros

    usuario_id = request.GET.get("usuario", "").strip()
    accion = request.GET.get("accion", "").strip()
    fecha_desde = request.GET.get("desde", "").strip()
    fecha_hasta = request.GET.get("hasta", "").strip()
    busqueda = request.GET.get("q", "").strip()

    if usuario_id:
        registros = registros.filter(usuario_id=usuario_id)

    if accion:
        registros = registros.filter(accion=accion)

    desde = parse_date(fecha_desde) if fecha_desde else None
    hasta = parse_date(fecha_hasta) if fecha_hasta else None
    if desde:
        registros = registros.filter(creado_en__date__gte=desde)
    if hasta:
        registros = registros.filter(creado_en__date__lte=hasta)
    if busqueda:
        registros = registros.filter(
            Q(usuario__username__icontains=busqueda)
            | Q(usuario__first_name__icontains=busqueda)
            | Q(usuario__last_name__icontains=busqueda)
            | Q(descripcion__icontains=busqueda)
            | Q(objeto_repr__icontains=busqueda)
        )

    if request.GET.get("formato") == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="auditoria_usuarios.csv"'
        response.write("\ufeff")
        writer = csv.writer(response)
        writer.writerow([
            "Fecha", "Usuario", "Tipo de usuario", "Acción", "Torneo",
            "Descripción", "Objeto", "Ruta", "IP", "Dispositivo",
        ])
        for registro in registros[:10000]:
            writer.writerow([
                timezone.localtime(registro.creado_en).strftime("%Y-%m-%d %H:%M:%S"),
                registro.usuario.username if registro.usuario else "Sistema",
                registro.datos.get("tipo_usuario", ""),
                registro.accion,
                registro.torneo.nombre if registro.torneo else "",
                registro.descripcion,
                registro.objeto_repr,
                registro.datos.get("ruta", ""),
                registro.ip or "",
                registro.user_agent,
            ])
        return response

    acciones = registros_disponibles.order_by("accion").values_list("accion", flat=True).distinct()
    if request.user.is_superuser:
        usuarios = User.objects.filter(actividad_admin__isnull=False).distinct().order_by("username")
    else:
        usuarios = User.objects.filter(actividad_admin__in=registros_disponibles).distinct().order_by("username")
    visitas_publicas = None
    if request.user.is_superuser:
        hoy = timezone.localdate()
        visitas_qs = VisitaPublicaDiaria.objects.all()
        if torneo:
            visitas_qs = visitas_qs.filter(torneo=torneo)
        canales = {
            item["canal"]: item["total"]
            for item in visitas_qs.filter(fecha__gte=hoy - timedelta(days=29))
            .values("canal")
            .annotate(total=Count("id"))
        }
        visitas_publicas = {
            "hoy": visitas_qs.filter(fecha=hoy).count(),
            "siete_dias": visitas_qs.filter(fecha__gte=hoy - timedelta(days=6)).count(),
            "treinta_dias": visitas_qs.filter(fecha__gte=hoy - timedelta(days=29)).count(),
            "canales": [
                {"nombre": "Aplicación", "total": canales.get("APK", 0)},
                {"nombre": "Navegador móvil", "total": canales.get("MOVIL", 0)},
                {"nombre": "Computador", "total": canales.get("ESCRITORIO", 0)},
            ],
        }

    return render(request, "gestion/actividad.html", {
        "registros": registros[:250],
        "torneo_seleccionado": torneo,
        "usuarios": usuarios,
        "acciones": acciones,
        "usuario_id": usuario_id,
        "accion": accion,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "busqueda": busqueda,
        "visitas_publicas": visitas_publicas,
        "puede_descargar_auditoria": request.user.is_superuser,
        "es_auditoria_delegado": es_delegado and not request.user.is_superuser,
        "volver_actividad_url": reverse("delegado_mis_equipos") if es_delegado and not request.user.is_superuser else reverse("gestion_panel"),
    })


@login_required
@user_passes_test(es_editor_torneo)
def gestion_validaciones(request):
    if not puede_gestionar_torneo(request, torneo_actual(request), "validar"):
        return denegar_permiso_torneo()
    if not tabla_disponible("torneos_solicitudvalidacion"):
        messages.error(request, "La tabla de validaciones todavia no esta creada. Espera que Render termine de aplicar las migraciones.")
        return redirect("gestion_panel")
    estado = (request.GET.get("estado") or "PENDIENTE").strip().upper()
    if estado not in {"PENDIENTE", "VALIDADO", "RECHAZADO", ""}:
        estado = "PENDIENTE"
    solicitudes = solicitudes_validacion_para_usuario(request.user, estado or None)
    return render(request, "gestion/validaciones.html", {
        "solicitudes": solicitudes[:300],
        "estado": estado,
        "estados": SolicitudValidacion.ESTADOS,
    })


@login_required
@user_passes_test(es_editor_torneo)
@require_POST
def gestion_validacion_resolver(request, solicitud_id):
    if not tabla_disponible("torneos_solicitudvalidacion"):
        messages.error(request, "La tabla de validaciones todavia no esta creada.")
        return redirect("gestion_panel")
    solicitud = get_object_or_404(SolicitudValidacion.objects.select_related("torneo", "partido"), id=solicitud_id)
    if not puede_gestionar_torneo(request, solicitud.torneo, "validar"):
        return denegar_permiso_torneo()

    accion = request.POST.get("accion")
    if accion == "validar":
        solicitud.estado = "VALIDADO"
        if solicitud.tipo == "ESTADISTICAS" and solicitud.partido_id:
            _validar_estadisticas_partido(solicitud.partido, request.user)
    elif accion == "rechazar":
        solicitud.estado = "RECHAZADO"
    else:
        messages.error(request, "Accion no valida.")
        return redirect("gestion_validaciones")

    solicitud.resuelto_por = request.user
    solicitud.resuelto_en = timezone.now()
    solicitud.save(update_fields=["estado", "resuelto_por", "resuelto_en"])
    registrar_actividad(
        request,
        f"VALIDACION_{solicitud.estado}",
        solicitud,
        torneo=solicitud.torneo,
        descripcion=f"{solicitud.get_estado_display()}: {solicitud.titulo}.",
        datos={"solicitud_id": solicitud.id, "tipo": solicitud.tipo},
    )
    messages.success(request, f"Solicitud {solicitud.get_estado_display().lower()}: {solicitud.titulo}.")
    return redirect(request.POST.get("next") or "gestion_validaciones")


@login_required
@user_passes_test(puede_gestionar_organizadores)
def gestion_organizadores(request):
    if not tabla_disponible("torneos_organizador"):
        messages.error(request, "La tabla de organizadores todavia no esta creada. Espera que Render termine de aplicar las migraciones.")
        return redirect("gestion_panel")

    organizadores = Organizador.objects.order_by("nombre")

    return render(request, "gestion/organizadores.html", {
        "organizadores": organizadores,
    })


@login_required
@user_passes_test(puede_gestionar_organizadores)
def gestion_organizador_nuevo(request):
    form = OrganizadorForm(request.POST or None, request.FILES or None)

    if request.method == "POST" and form.is_valid():
        organizador = form.save()
        registrar_actividad(request, "CREAR", organizador, descripcion=f"Creo organizador {organizador.nombre}.")
        messages.success(request, "Organizador creado correctamente.")
        return redirect("gestion_organizadores")

    return render(request, "gestion/formulario.html", {
        "titulo": "Nuevo organizador",
        "form": form,
        "volver_url": "gestion_organizadores",
    })


@login_required
@user_passes_test(puede_gestionar_organizadores)
def gestion_organizador_editar(request, organizador_id):
    organizador = get_object_or_404(Organizador, id=organizador_id)
    form = OrganizadorForm(request.POST or None, request.FILES or None, instance=organizador)

    if request.method == "POST" and form.is_valid():
        organizador = form.save()
        registrar_actividad(request, "EDITAR", organizador, descripcion=f"Actualizo organizador {organizador.nombre}.")
        messages.success(request, "Organizador actualizado correctamente.")
        return redirect("gestion_organizadores")

    return render(request, "gestion/formulario.html", {
        "titulo": f"Editar organizador: {organizador.nombre}",
        "form": form,
        "volver_url": "gestion_organizadores",
    })


@login_required
@user_passes_test(puede_gestionar_organizadores)
def gestion_organizador_admins(request, organizador_id):
    organizador = get_object_or_404(Organizador, id=organizador_id)
    asignaciones = AdminOrganizador.objects.select_related("usuario").filter(organizador=organizador)
    form = AdminOrganizadorForm()
    crear_form = CrearAdminOrganizadorForm()

    if request.method == "POST":
        accion = request.POST.get("accion")
        if accion == "crear_admin":
            crear_form = CrearAdminOrganizadorForm(request.POST)
            if crear_form.is_valid():
                usuario = crear_form.save_user()
                asignacion = AdminOrganizador.objects.create(
                    organizador=organizador,
                    usuario=usuario,
                    puede_editar=crear_form.cleaned_data["puede_editar"],
                    puede_validar=crear_form.cleaned_data["puede_validar"],
                    puede_programar=crear_form.cleaned_data["puede_programar"],
                    puede_descargar_planillas=crear_form.cleaned_data["puede_descargar_planillas"],
                    activo=crear_form.cleaned_data["activo"],
                )
                registrar_actividad(
                    request,
                    "CREAR_ADMIN_ORGANIZADOR",
                    torneo=None,
                    descripcion=f"Creo admin {usuario.username} y lo asigno al organizador {organizador.nombre}.",
                    datos={"organizador": organizador.nombre, "usuario": usuario.username},
                )
                messages.success(request, f"Admin creado y asignado: {usuario.username}.")
                return redirect("gestion_organizador_admins", organizador_id=organizador.id)
        else:
            form = AdminOrganizadorForm(request.POST)
            if form.is_valid():
                asignacion = form.save(commit=False)
                asignacion.organizador = organizador
                existente = AdminOrganizador.objects.filter(organizador=organizador, usuario=asignacion.usuario).first()

                if existente:
                    existente.puede_editar = asignacion.puede_editar
                    existente.puede_validar = asignacion.puede_validar
                    existente.puede_programar = asignacion.puede_programar
                    existente.puede_descargar_planillas = asignacion.puede_descargar_planillas
                    existente.activo = asignacion.activo
                    existente.save()
                    asignacion = existente
                    mensaje = "Admin actualizado correctamente."
                else:
                    asignacion.save()
                    mensaje = "Admin asignado correctamente."

                registrar_actividad(
                    request,
                    "ASIGNAR_ADMIN_ORGANIZADOR",
                    torneo=None,
                    descripcion=f"Asigno admin {asignacion.usuario.username} al organizador {organizador.nombre}.",
                    datos={
                        "organizador": organizador.nombre,
                        "usuario": asignacion.usuario.username,
                        "puede_editar": asignacion.puede_editar,
                        "puede_validar": asignacion.puede_validar,
                        "puede_programar": asignacion.puede_programar,
                        "puede_descargar_planillas": asignacion.puede_descargar_planillas,
                        "activo": asignacion.activo,
                    },
                )
                messages.success(request, mensaje)
                return redirect("gestion_organizador_admins", organizador_id=organizador.id)

    return render(request, "gestion/organizador_admins.html", {
        "organizador": organizador,
        "asignaciones": asignaciones,
        "form": form,
        "crear_form": crear_form,
        "torneos": organizador.torneos.order_by("-fecha_inicio", "nombre"),
    })


@login_required
@user_passes_test(puede_gestionar_organizadores)
@require_POST
def gestion_organizador_admin_eliminar(request, asignacion_id):
    asignacion = get_object_or_404(AdminOrganizador.objects.select_related("organizador", "usuario"), id=asignacion_id)
    organizador = asignacion.organizador
    usuario = asignacion.usuario.username
    registrar_actividad(
        request,
        "QUITAR_ADMIN_ORGANIZADOR",
        torneo=None,
        descripcion=f"Quito admin {usuario} del organizador {organizador.nombre}.",
        datos={"organizador": organizador.nombre, "usuario": usuario},
    )
    asignacion.delete()
    messages.success(request, f"Admin retirado: {usuario}.")
    return redirect("gestion_organizador_admins", organizador_id=organizador.id)


@login_required
@user_passes_test(es_editor_torneo)
def gestion_torneos(request):
    if usuario_solo_descarga_planillas(request.user, torneo_actual(request)):
        return denegar_permiso_torneo()
    torneos = torneos_para_usuario(request)

    return render(request, "gestion/torneos.html", {
        "torneos": torneos,
        "torneo_seleccionado": torneo_actual(request),
    })


@login_required
@user_passes_test(es_superadmin)
def gestion_torneo_nuevo(request):
    form = TorneoForm(request.POST or None, request.FILES or None)

    if request.method == "POST" and form.is_valid():
        torneo = form.save(commit=False)
        aplicar_imagenes_torneo_cloudinary(torneo, request.FILES)
        torneo.save()
        request.session["torneo_id"] = torneo.id
        registrar_actividad(request, "CREAR", torneo, descripcion=f"Creo torneo {torneo.nombre}.")
        messages.success(request, "Torneo creado correctamente.")
        return redirect("gestion_torneos")

    return render(request, "gestion/formulario.html", {
        "titulo": "Nuevo torneo",
        "form": form,
        "volver_url": "gestion_torneos",
    })


@login_required
@user_passes_test(es_editor_torneo)
def gestion_torneo_editar(request, torneo_id):
    torneo = get_object_or_404(torneos_para_usuario(request), id=torneo_id)
    if not puede_gestionar_torneo(request, torneo, "editar"):
        return denegar_permiso_torneo()
    form = TorneoForm(request.POST or None, request.FILES or None, instance=torneo)

    if request.method == "POST" and form.is_valid():
        torneo = form.save(commit=False)
        aplicar_imagenes_torneo_cloudinary(torneo, request.FILES)
        torneo.save()
        request.session["torneo_id"] = torneo.id
        registrar_actividad(request, "EDITAR", torneo, descripcion=f"Actualizo torneo {torneo.nombre}.")
        messages.success(request, "Torneo actualizado correctamente.")
        return redirect("gestion_torneos")

    return render(request, "gestion/formulario.html", {
        "titulo": f"Editar torneo: {torneo.nombre}",
        "form": form,
        "volver_url": "gestion_torneos",
    })


@login_required
@user_passes_test(es_superadmin)
def gestion_torneo_admins(request, torneo_id):
    torneo = get_object_or_404(Torneo, id=torneo_id)
    asignaciones = AdminTorneo.objects.select_related("usuario").filter(torneo=torneo)
    form = AdminTorneoForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        asignacion = form.save(commit=False)
        asignacion.torneo = torneo
        existente = AdminTorneo.objects.filter(torneo=torneo, usuario=asignacion.usuario).first()

        if existente:
            existente.puede_editar = asignacion.puede_editar
            existente.puede_validar = asignacion.puede_validar
            existente.puede_programar = asignacion.puede_programar
            existente.puede_descargar_planillas = asignacion.puede_descargar_planillas
            existente.activo = asignacion.activo
            existente.save()
            asignacion = existente
            mensaje = "Admin actualizado correctamente."
        else:
            asignacion.save()
            mensaje = "Admin asignado correctamente."

        registrar_actividad(
            request,
            "ASIGNAR_ADMIN",
            torneo,
            descripcion=f"Asigno admin {asignacion.usuario.username} al torneo {torneo.nombre}.",
            datos={
                "usuario": asignacion.usuario.username,
                "puede_editar": asignacion.puede_editar,
                "puede_validar": asignacion.puede_validar,
                "puede_programar": asignacion.puede_programar,
                "puede_descargar_planillas": asignacion.puede_descargar_planillas,
                "activo": asignacion.activo,
            },
        )
        messages.success(request, mensaje)
        return redirect("gestion_torneo_admins", torneo_id=torneo.id)

    return render(request, "gestion/torneo_admins.html", {
        "torneo": torneo,
        "asignaciones": asignaciones,
        "form": form,
    })


@login_required
@user_passes_test(es_superadmin)
@require_POST
def gestion_torneo_admin_eliminar(request, asignacion_id):
    asignacion = get_object_or_404(AdminTorneo.objects.select_related("torneo", "usuario"), id=asignacion_id)
    torneo = asignacion.torneo
    usuario = asignacion.usuario.username
    registrar_actividad(request, "QUITAR_ADMIN", torneo, descripcion=f"Quito admin {usuario} del torneo {torneo.nombre}.")
    asignacion.delete()
    messages.success(request, f"Admin retirado: {usuario}.")
    return redirect("gestion_torneo_admins", torneo_id=torneo.id)


@login_required
@user_passes_test(es_editor_torneo)
@require_POST
def gestion_torneo_activar(request, torneo_id):
    torneo = get_object_or_404(torneos_para_usuario(request), id=torneo_id)
    request.session["torneo_id"] = torneo.id
    messages.success(request, f"Ahora estás gestionando: {torneo.nombre}.")
    return redirect("gestion_torneos")


@login_required
@user_passes_test(es_superadmin)
@require_POST
def gestion_torneo_eliminar(request, torneo_id):
    torneo = get_object_or_404(torneos_para_usuario(request), id=torneo_id)
    nombre = torneo.nombre
    if request.session.get("torneo_id") == torneo.id:
        request.session.pop("torneo_id", None)
    registrar_actividad(request, "ELIMINAR", torneo, descripcion=f"Elimino torneo {nombre}.")
    torneo.delete()
    messages.success(request, f"Torneo eliminado: {nombre}.")
    return redirect("gestion_torneos")


@login_required
@user_passes_test(es_editor_torneo)
def gestion_categorias(request):
    torneo = torneo_actual(request)
    if usuario_solo_descarga_planillas(request.user, torneo):
        return denegar_permiso_torneo()
    categorias = Categoria.objects.select_related("torneo").prefetch_related("reglas_edad").order_by("nombre")
    if torneo:
        categorias = categorias.filter(torneo=torneo)

    return render(request, "gestion/categorias.html", {
        "categorias": categorias,
        "torneo_seleccionado": torneo,
    })


@login_required
@user_passes_test(es_editor_torneo)
def gestion_categoria_nueva(request):
    torneo = torneo_actual(request)
    if not puede_gestionar_torneo(request, torneo, "editar"):
        return denegar_permiso_torneo()
    categoria = Categoria(torneo=torneo)
    reglas_iniciales = [
        {"etiqueta": "+40", "edad_minima": 40, "edad_maxima": 44, "minimo_titulares": 0, "maximo_titulares": 4, "orden": 1, "activa": True},
        {"etiqueta": "+45", "edad_minima": 45, "edad_maxima": 49, "minimo_titulares": 4, "orden": 2, "activa": True},
        {"etiqueta": "+50", "edad_minima": 50, "minimo_titulares": 3, "orden": 3, "activa": True},
    ]
    form = CategoriaForm(request.POST or None, instance=categoria)
    reglas_formset = ReglaEdadCategoriaFormSet(
        None,
        instance=categoria,
        prefix="reglas",
        initial=reglas_iniciales,
    )

    if request.method == "POST" and form.is_valid():
        categoria = form.save(commit=False)
        categoria.torneo = torneo
        categoria.save()
        reglas_formset = ReglaEdadCategoriaFormSet(request.POST, instance=categoria, prefix="reglas")
        if not reglas_formset.is_valid():
            categoria.delete()
            return render(request, "gestion/categoria_formulario.html", {
                "titulo": "Nueva categorÃ­a",
                "form": form,
                "reglas_formset": reglas_formset,
                "volver_url": "gestion_categorias",
            })
        reglas_formset.save()
        registrar_actividad(request, "CREAR", categoria, descripcion=f"Creo categoria {categoria.nombre}.")
        messages.success(request, "Categoría creada correctamente.")
        return redirect("gestion_categorias")

    return render(request, "gestion/categoria_formulario.html", {
        "titulo": "Nueva categoría",
        "form": form,
        "reglas_formset": reglas_formset,
        "volver_url": "gestion_categorias",
    })


@login_required
@user_passes_test(es_editor_torneo)
def gestion_categoria_editar(request, categoria_id):
    torneo = torneo_actual(request)
    if not puede_gestionar_torneo(request, torneo, "editar"):
        return denegar_permiso_torneo()
    categorias = Categoria.objects.select_related("torneo")
    if torneo:
        categorias = categorias.filter(torneo=torneo)
    categoria = get_object_or_404(categorias, id=categoria_id)
    form = CategoriaForm(request.POST or None, instance=categoria)
    reglas_formset = ReglaEdadCategoriaFormSet(
        request.POST or None,
        instance=categoria,
        prefix="reglas",
    )

    if request.method == "POST" and form.is_valid() and reglas_formset.is_valid():
        categoria = form.save(commit=False)
        if torneo:
            categoria.torneo = torneo
        categoria.save()
        reglas_formset.instance = categoria
        reglas_formset.save()
        registrar_actividad(request, "EDITAR", categoria, descripcion=f"Actualizo categoria {categoria.nombre}.")
        messages.success(request, "Categoría actualizada correctamente.")
        return redirect("gestion_categorias")

    return render(request, "gestion/categoria_formulario.html", {
        "titulo": f"Editar categoría: {categoria.nombre}",
        "form": form,
        "reglas_formset": reglas_formset,
        "volver_url": "gestion_categorias",
    })


@login_required
@user_passes_test(es_editor_torneo)
@require_POST
def gestion_categoria_eliminar(request, categoria_id):
    torneo = torneo_actual(request)
    if not puede_gestionar_torneo(request, torneo, "editar"):
        return denegar_permiso_torneo()
    categorias = Categoria.objects.select_related("torneo")
    if torneo:
        categorias = categorias.filter(torneo=torneo)
    categoria = get_object_or_404(categorias, id=categoria_id)
    nombre = categoria.nombre
    registrar_actividad(request, "ELIMINAR", categoria, descripcion=f"Elimino categoria {nombre}.")
    categoria.delete()
    messages.success(request, f"Categoria eliminada: {nombre}.")
    return redirect("gestion_categorias")


@login_required
@user_passes_test(es_editor_torneo)
def gestion_documentos(request):
    torneo = torneo_actual(request)
    if usuario_solo_descarga_planillas(request.user, torneo):
        return denegar_permiso_torneo()
    documentos = Documento.objects.order_by("tipo", "-creado_en", "titulo")
    if torneo:
        documentos = documentos.filter(torneo=torneo)
    else:
        documentos = documentos.none()
    tipo = request.GET.get("tipo", "").strip()

    if tipo:
        documentos = documentos.filter(tipo=tipo)

    return render(request, "gestion/documentos.html", {
        "documentos": documentos,
        "tipo": tipo,
        "tipos": Documento.TIPOS,
    })


@login_required
@user_passes_test(es_editor_torneo)
def gestion_documento_nuevo(request):
    torneo = torneo_actual(request)
    if not puede_gestionar_torneo(request, torneo, "editar"):
        return denegar_permiso_torneo()
    if request.method == "POST" and request.POST.get("tipo") == "PLANILLA_JUEGO":
        messages.info(request, "Las planillas de juego se cargan desde el formulario especializado.")
        return redirect("gestion_planilla_juego_nueva")
    form = DocumentoForm(request.POST or None, request.FILES or None, initial={"torneo": torneo}, torneo=torneo)

    if request.method == "POST" and form.is_valid():
        documento = form.save(commit=False)
        if not documento.torneo:
            documento.torneo = torneo
        documento.archivo = subir_documento_torneo(
            form.cleaned_data["archivo_subido"],
            documento.tipo,
        )
        documento.save()
        registrar_actividad(request, "CREAR", documento, descripcion=f"Creo documento {documento.titulo}.")
        messages.success(request, "Documento creado correctamente.")
        return redirect("gestion_documentos")

    return render(request, "gestion/formulario.html", {
        "titulo": "Nuevo documento",
        "form": form,
        "volver_url": "gestion_documentos",
    })


@login_required
@user_passes_test(es_editor_torneo)
def gestion_documento_editar(request, documento_id):
    torneo = torneo_actual(request)
    documentos = Documento.objects.all()
    if torneo:
        documentos = documentos.filter(torneo=torneo)
    else:
        documentos = documentos.none()
    documento = get_object_or_404(documentos, id=documento_id)
    if not puede_gestionar_torneo(request, documento.torneo or torneo, "editar"):
        return denegar_permiso_torneo()
    form = DocumentoForm(request.POST or None, request.FILES or None, instance=documento, torneo=torneo)

    if request.method == "POST" and form.is_valid():
        documento = form.save(commit=False)
        archivo_subido = form.cleaned_data.get("archivo_subido")

        if archivo_subido:
            documento.archivo = subir_documento_torneo(archivo_subido, documento.tipo)

        documento.save()
        registrar_actividad(request, "EDITAR", documento, descripcion=f"Actualizo documento {documento.titulo}.")
        messages.success(request, "Documento actualizado correctamente.")
        return redirect("gestion_documentos")

    return render(request, "gestion/formulario.html", {
        "titulo": f"Editar documento: {documento.titulo}",
        "form": form,
        "volver_url": "gestion_documentos",
    })


def _planillas_juego_para_usuario(user, torneo=None):
    documentos = Documento.objects.select_related(
        "torneo",
        "categoria",
        "partido",
        "equipo_local",
        "equipo_visitante",
        "cargado_por",
    ).filter(tipo="PLANILLA_JUEGO")

    if torneo:
        documentos = documentos.filter(torneo=torneo)

    orden = ("categoria__nombre", "numero_fecha", "fecha_partido", "hora_partido", "equipo_local__nombre", "equipo_visitante__nombre", "-creado_en")

    if es_editor_torneo(user):
        return documentos.order_by(*orden)

    return documentos.filter(
        Q(cargado_por=user) | Q(partido__planilleros=user)
    ).distinct().order_by(*orden)


def _partidos_planillas_para_usuario(user, torneo=None):
    partidos = Partido.objects.select_related(
        "categoria",
        "categoria__torneo",
        "equipo_local",
        "equipo_visitante",
    ).order_by("categoria__nombre", "numero_fecha", "fecha", "hora", "equipo_local__nombre")

    if torneo:
        partidos = partidos.filter(categoria__torneo=torneo)

    if es_editor_torneo(user):
        return partidos

    return partidos.filter(planilleros=user)


def _agrupar_planillas_juego(documentos):
    categorias = {}

    for documento in documentos:
        categoria_id = documento.categoria_id or 0
        if categoria_id not in categorias:
            categorias[categoria_id] = SimpleNamespace(
                nombre=documento.categoria.nombre if documento.categoria else "Sin categoria",
                fechas={},
            )
        categoria = categorias[categoria_id]

        fecha_nombre = documento.numero_fecha or "Sin fecha fixture"
        if fecha_nombre not in categoria.fechas:
            categoria.fechas[fecha_nombre] = SimpleNamespace(
                nombre=fecha_nombre,
                partidos={},
            )
        fecha = categoria.fechas[fecha_nombre]

        partido_key = documento.partido_id or f"manual-{documento.equipo_local_id}-{documento.equipo_visitante_id}-{documento.fecha_partido}"
        if partido_key not in fecha.partidos:
            fecha.partidos[partido_key] = SimpleNamespace(
                partido=documento.partido,
                equipo_local=documento.equipo_local,
                equipo_visitante=documento.equipo_visitante,
                fecha_partido=documento.fecha_partido,
                hora_partido=documento.hora_partido,
                documentos=[],
            )
        fecha.partidos[partido_key].documentos.append(documento)

    return [
        SimpleNamespace(
            nombre=categoria.nombre,
            fechas=[
                SimpleNamespace(nombre=fecha.nombre, partidos=list(fecha.partidos.values()))
                for fecha in categoria.fechas.values()
            ],
        )
        for categoria in categorias.values()
    ]


def _agrupar_partidos_planillas(partidos, documentos):
    documentos_por_partido = defaultdict(list)
    for documento in documentos:
        documentos_por_partido[documento.partido_id].append(documento)

    categorias = {}
    for partido in partidos:
        categoria_id = partido.categoria_id or 0
        if categoria_id not in categorias:
            categorias[categoria_id] = SimpleNamespace(
                nombre=partido.categoria.nombre if partido.categoria else "Sin categoria",
                fechas={},
            )
        categoria = categorias[categoria_id]

        fecha_nombre = partido.numero_fecha or "Sin fecha fixture"
        if fecha_nombre not in categoria.fechas:
            categoria.fechas[fecha_nombre] = SimpleNamespace(nombre=fecha_nombre, partidos={})
        fecha = categoria.fechas[fecha_nombre]

        fecha.partidos[partido.id] = SimpleNamespace(
            partido=partido,
            equipo_local=partido.equipo_local,
            equipo_visitante=partido.equipo_visitante,
            fecha_partido=partido.fecha,
            hora_partido=partido.hora,
            documentos=documentos_por_partido.get(partido.id, []),
        )

    return [
        SimpleNamespace(
            nombre=categoria.nombre,
            fechas=[
                SimpleNamespace(nombre=fecha.nombre, partidos=list(fecha.partidos.values()))
                for fecha in categoria.fechas.values()
            ],
        )
        for categoria in categorias.values()
    ]


@login_required
@user_passes_test(puede_cargar_planillas_juego)
def gestion_planillas_juego(request):
    torneo = torneo_actual(request) if es_editor_torneo(request.user) else None
    documentos_base = _planillas_juego_para_usuario(request.user, torneo)
    partidos_base = _partidos_planillas_para_usuario(request.user, torneo)
    categoria_id = request.GET.get("categoria", "").strip()
    numero_fecha = request.GET.get("fecha", "").strip()
    partido_id = request.GET.get("partido", "").strip()
    partidos_resultado = partidos_base

    if categoria_id:
        partidos_resultado = partidos_resultado.filter(categoria_id=categoria_id)
    if numero_fecha:
        partidos_resultado = partidos_resultado.filter(numero_fecha=numero_fecha)
    if partido_id:
        partidos_resultado = partidos_resultado.filter(id=partido_id)

    documentos = documentos_base.filter(partido__in=partidos_resultado)
    categorias = Categoria.objects.filter(partido__in=partidos_base).distinct().order_by("nombre")
    fechas = partidos_base.exclude(numero_fecha__isnull=True).exclude(numero_fecha="").order_by("numero_fecha").values_list("numero_fecha", flat=True).distinct()
    partidos = partidos_base
    if categoria_id:
        partidos = partidos.filter(categoria_id=categoria_id)
    if numero_fecha:
        partidos = partidos.filter(numero_fecha=numero_fecha)
    partidos = partidos.distinct().order_by("categoria__nombre", "numero_fecha", "fecha", "hora", "equipo_local__nombre")

    return render(request, "gestion/planillas_juego.html", {
        "grupos_planillas": _agrupar_partidos_planillas(partidos_resultado[:500], documentos),
        "categorias": categorias,
        "fechas": fechas,
        "partidos": partidos,
        "categoria_id": categoria_id,
        "numero_fecha": numero_fecha,
        "partido_id": partido_id,
        "torneo_seleccionado": torneo,
        "es_editor": es_editor_torneo(request.user),
        "volver_panel_url": reverse("planillero_mis_partidos") if es_planillero_asignado(request.user) else reverse("panel"),
        "volver_panel_text": "Mis partidos" if es_planillero_asignado(request.user) else "Volver al panel",
    })


@login_required
@user_passes_test(puede_cargar_planillas_juego)
def gestion_planilla_juego_nueva(request):
    torneo = torneo_actual(request) if es_editor_torneo(request.user) else None
    initial = {}
    if request.method == "GET" and request.GET.get("partido"):
        initial["partido"] = request.GET.get("partido")
    form = PlanillaJuegoUploadForm(
        request.POST or None,
        request.FILES or None,
        initial=initial,
        user=request.user,
        torneo=torneo,
    )
    partidos_formulario = [
        {
            "id": partido.id,
            "categoria": partido.categoria_id,
            "categoria_nombre": partido.categoria.nombre,
            "numero_fecha": partido.numero_fecha or "",
            "equipo_local": partido.equipo_local_id,
            "equipo_local_nombre": partido.equipo_local.nombre,
            "equipo_visitante": partido.equipo_visitante_id,
            "equipo_visitante_nombre": partido.equipo_visitante.nombre,
            "fecha": partido.fecha.isoformat(),
            "hora": partido.hora.strftime("%H:%M"),
            "label": (
                f"{partido.categoria.nombre} - {partido.numero_fecha or 'Sin fecha'} - "
                f"{partido.equipo_local.nombre} vs {partido.equipo_visitante.nombre} - "
                f"{partido.fecha.strftime('%d/%m/%Y')} {partido.hora.strftime('%H:%M')}"
            ),
        }
        for partido in form.fields["partido"].queryset
    ]
    equipos_formulario = [
        {
            "id": equipo.id,
            "categoria": equipo.categoria_id,
            "nombre": equipo.nombre,
        }
        for equipo in form.fields["equipo_local"].queryset
    ]

    if request.method == "POST" and form.is_valid():
        categoria = form.cleaned_data["categoria"]
        equipo_local = form.cleaned_data["equipo_local"]
        equipo_visitante = form.cleaned_data["equipo_visitante"]
        numero_fecha = form.cleaned_data.get("numero_fecha") or (form.partido.numero_fecha if form.partido else "")
        fecha_partido = form.cleaned_data["fecha_partido"]
        hora_partido = form.cleaned_data["hora_partido"]
        archivos = form.cleaned_data["imagenes"]
        creados = 0

        for archivo in archivos:
            titulo = f"Planilla {categoria.nombre} - {numero_fecha or 'Sin fecha'} - {equipo_local.nombre} vs {equipo_visitante.nombre}"
            documento = Documento.objects.create(
                tipo="PLANILLA_JUEGO",
                torneo=categoria.torneo,
                categoria=categoria,
                partido=form.partido,
                equipo_local=equipo_local,
                equipo_visitante=equipo_visitante,
                titulo=titulo,
                descripcion=f"Cargada por {request.user.get_username()}",
                archivo=subir_documento_torneo(archivo, "PLANILLA_JUEGO"),
                numero_fecha=numero_fecha,
                fecha_partido=fecha_partido,
                hora_partido=hora_partido,
                cargado_por=request.user,
                activo=True,
            )
            registrar_actividad(
                request,
                "CARGAR_PLANILLA_JUEGO",
                documento,
                torneo=categoria.torneo,
                descripcion=f"Cargo planilla de juego de {equipo_local.nombre} vs {equipo_visitante.nombre}.",
                datos={
                    "categoria": categoria.nombre,
                    "numero_fecha": numero_fecha,
                    "fecha_partido": fecha_partido.isoformat(),
                    "hora_partido": hora_partido.strftime("%H:%M"),
                    "partido_id": form.partido.id if form.partido else None,
                },
            )
            creados += 1

        messages.success(request, f"Planilla cargada correctamente: {creados} archivo(s).")
        query = urlencode({
            "categoria": categoria.id,
            "fecha": numero_fecha,
            "partido": form.partido.id if form.partido else "",
        })
        return redirect(f"{reverse('gestion_planillas_juego')}?{query}")

    return render(request, "gestion/planilla_juego_form.html", {
        "titulo": "Cargar planilla de juego",
        "form": form,
        "partidos_formulario": partidos_formulario,
        "equipos_formulario": equipos_formulario,
        "torneo_seleccionado": torneo,
        "es_editor": es_editor_torneo(request.user),
        "volver_panel_url": reverse("planillero_mis_partidos") if es_planillero_asignado(request.user) else reverse("panel"),
        "volver_panel_text": "Mis partidos" if es_planillero_asignado(request.user) else "Volver al panel",
    })


@login_required
@user_passes_test(es_planillero_asignado)
def planillero_mis_partidos(request):
    torneo, torneos = torneo_actual_planillero(request)
    partidos = request.user.partidos_planillero.select_related(
        "categoria",
        "categoria__torneo",
        "equipo_local",
        "equipo_visitante",
    ).prefetch_related(
        Prefetch(
            "documentos_planilla",
            queryset=Documento.objects.filter(tipo="PLANILLA_JUEGO"),
            to_attr="planillas_juego_cargadas",
        )
    )

    if torneo:
        partidos = partidos.filter(categoria__torneo=torneo)

    estado = request.GET.get("estado", "").strip()
    if estado:
        partidos = partidos.filter(estado=estado)

    partidos_con_planilla = Documento.objects.filter(
        tipo="PLANILLA_JUEGO",
        partido_id__isnull=False,
    ).values("partido_id")
    partidos = partidos.exclude(
        estado="FINALIZADO",
        id__in=partidos_con_planilla,
    ).order_by("estado", "fecha", "hora", "categoria__nombre", "equipo_local__nombre")

    items = []
    for partido in partidos:
        items.append(SimpleNamespace(
            partido=partido,
            puede_editar=puede_diligenciar_partido(request.user, partido),
            planillas_count=len(getattr(partido, "planillas_juego_cargadas", [])),
        ))

    return render(request, "gestion/planillero_mis_partidos.html", {
        "items": items,
        "torneo": torneo,
        "torneos": torneos,
        "estado": estado,
        "estados": Partido.ESTADOS,
        "volver_panel_url": reverse("planillero_mis_partidos"),
        "volver_panel_text": "Mis partidos",
    })


def generar_fixture_grupo(equipos):
    equipos = list(equipos)

    if len(equipos) % 2:
        equipos.append(None)

    total = len(equipos)
    rondas = total - 1
    mitad = total // 2
    calendario = []

    for numero_fecha in range(1, rondas + 1):
        partidos_fecha = []

        for i in range(mitad):
            local = equipos[i]
            visitante = equipos[total - 1 - i]

            if local and visitante:
                if numero_fecha % 2 == 0:
                    local, visitante = visitante, local

                partidos_fecha.append((local, visitante))

        calendario.append(partidos_fecha)
        equipos = [equipos[0]] + [equipos[-1]] + equipos[1:-1]

    return calendario


def distribuir_equipos_en_grupos(equipos, cabezas, cantidad_grupos):
    grupos = {chr(65 + i): [] for i in range(cantidad_grupos)}
    usados = set()

    for indice, cabeza in enumerate(cabezas):
        if cabeza and cabeza.id not in usados:
            grupos[chr(65 + indice)].append(cabeza)
            usados.add(cabeza.id)

    restantes = [equipo for equipo in equipos if equipo.id not in usados]

    for equipo in restantes:
        grupo_destino = min(grupos, key=lambda nombre: len(grupos[nombre]))
        grupos[grupo_destino].append(equipo)

    return grupos


def armar_grupos_desde_formulario(equipos, cabezas, request_post, cantidad_grupos):
    grupos = {chr(65 + i): [] for i in range(cantidad_grupos)}
    equipos_por_id = {str(equipo.id): equipo for equipo in equipos}
    usados = set()

    for indice, cabeza in enumerate(cabezas):
        if cabeza and cabeza.id not in usados:
            grupos[chr(65 + indice)].append(cabeza)
            usados.add(cabeza.id)

    for indice in range(cantidad_grupos):
        grupo_nombre = chr(65 + indice)

        for equipo_id in request_post.getlist(f"equipos_grupo_{indice}"):
            equipo = equipos_por_id.get(equipo_id)

            if equipo and equipo.id not in usados:
                grupos[grupo_nombre].append(equipo)
                usados.add(equipo.id)

    sin_asignar = [equipo for equipo in equipos if equipo.id not in usados]

    return grupos, sin_asignar


FRANJAS_PROGRAMACION_FIXTURE = [
    ("SAB_16", "Sabado 4:00 pm", 5, time(16, 0)),
    ("SAB_18", "Sabado 6:00 pm", 5, time(18, 0)),
    ("DOM_08", "Domingo 8:00 am", 6, time(8, 0)),
    ("DOM_10", "Domingo 10:00 am", 6, time(10, 0)),
    ("DOM_14", "Domingo 2:00 pm", 6, time(14, 0)),
    ("DOM_16", "Domingo 4:00 pm", 6, time(16, 0)),
]


def fecha_desde_texto(valor):
    try:
        return datetime.strptime((valor or "").strip(), "%Y-%m-%d").date()
    except ValueError:
        return None


def canchas_desde_texto(valor):
    canchas = []
    for linea in (valor or "").splitlines():
        cancha = linea.strip()
        if cancha and cancha not in canchas:
            canchas.append(cancha)
    return canchas


def siguiente_dia_semana(fecha_base, dia_semana):
    diferencia = (dia_semana - fecha_base.weekday()) % 7
    return fecha_base + timedelta(days=diferencia)


def cupos_programacion_fixture(fecha_inicio, canchas, franjas, cantidad_partidos):
    if not fecha_inicio or not canchas or not franjas:
        return []

    cupos = []
    sabado_base = siguiente_dia_semana(fecha_inicio, 5)
    semanas = 0

    while len(cupos) < cantidad_partidos and semanas < 30:
        inicio_semana = sabado_base + timedelta(days=semanas * 7)
        for codigo, etiqueta, dia_semana, hora in franjas:
            fecha_cupo = siguiente_dia_semana(inicio_semana, dia_semana)
            for cancha in canchas:
                cupos.append({
                    "fecha": fecha_cupo,
                    "hora": hora,
                    "cancha": cancha,
                    "franja": codigo,
                    "etiqueta": etiqueta,
                })
        semanas += 1

    return cupos[:cantidad_partidos]


def elegir_cupo_balanceado(local, visitante, cupos, usados, conteos, equipos_ids, cancha_obligatoria):
    mejor_indice = None
    mejor_puntaje = None

    for indice, cupo in enumerate(cupos):
        if indice in usados:
            continue

        local_id = local.id
        visitante_id = visitante.id
        cancha = cupo["cancha"]
        franja = cupo["franja"]
        es_obligatoria = cancha_obligatoria and cancha.lower() == cancha_obligatoria.lower()

        conteo_cancha_local = conteos["cancha"].get(local_id, 0)
        conteo_cancha_visitante = conteos["cancha"].get(visitante_id, 0)
        conteo_franja_local = conteos["franjas"].setdefault(local_id, {}).get(franja, 0)
        conteo_franja_visitante = conteos["franjas"].setdefault(visitante_id, {}).get(franja, 0)
        conteo_fecha_local = conteos["fechas"].setdefault(local_id, {}).get(cupo["fecha"], 0)
        conteo_fecha_visitante = conteos["fechas"].setdefault(visitante_id, {}).get(cupo["fecha"], 0)

        puntaje = 0
        if es_obligatoria:
            puntaje += (conteo_cancha_local + conteo_cancha_visitante) * 40
        else:
            pendientes = sum(1 for equipo_id in equipos_ids if conteos["cancha"].get(equipo_id, 0) == 0)
            if pendientes:
                if conteo_cancha_local == 0:
                    puntaje += 12
                if conteo_cancha_visitante == 0:
                    puntaje += 12

        puntaje += (conteo_franja_local + conteo_franja_visitante) * 12
        puntaje += (conteo_fecha_local + conteo_fecha_visitante) * 25

        if mejor_puntaje is None or puntaje < mejor_puntaje:
            mejor_puntaje = puntaje
            mejor_indice = indice

    return mejor_indice


def resumen_equidad_programacion(conteos, equipos_por_id, franjas):
    resumen = []
    for equipo_id, equipo in sorted(equipos_por_id.items(), key=lambda item: item[1].nombre):
        resumen.append({
            "equipo": equipo,
            "cancha_obligatoria": conteos["cancha"].get(equipo_id, 0),
            "franjas": [
                (etiqueta, conteos["franjas"].setdefault(equipo_id, {}).get(codigo, 0))
                for codigo, etiqueta, _, _ in franjas
            ],
        })
    return resumen


@login_required
@user_passes_test(es_editor_torneo)
def gestion_generar_fixture(request):
    torneo = torneo_actual(request)
    if not puede_gestionar_torneo(request, torneo, "programar"):
        return denegar_permiso_torneo()
    categorias = Categoria.objects.order_by("nombre")
    if torneo:
        categorias = categorias.filter(torneo=torneo)
    categoria = None
    equipos = Equipo.objects.none()
    cantidad_grupos = 2
    tipo_fixture = request.GET.get("tipo_fixture") or request.POST.get("tipo_fixture") or "GRUPOS"
    if tipo_fixture not in {"GRUPOS", "MATA_MATA_IDA_VUELTA"}:
        tipo_fixture = "GRUPOS"
    grupos_generados = None
    resumen_programacion = None
    advertencias_programacion = []

    categoria_id = request.GET.get("categoria") or request.POST.get("categoria")

    if categoria_id:
        categoria = categorias.filter(id=categoria_id).first()

    if categoria:
        equipos = Equipo.objects.filter(categoria=categoria, activo=True).order_by("nombre")

    cantidad_grupos_valor = request.GET.get("grupos") or request.POST.get("grupos")

    if cantidad_grupos_valor:
        try:
            cantidad_grupos = max(1, min(8, int(cantidad_grupos_valor)))
        except ValueError:
            cantidad_grupos = 2

    letras_grupos = [chr(65 + i) for i in range(cantidad_grupos)]
    parejas_mata_mata = list(range((equipos.count() if categoria else 0) // 2))

    if request.method == "POST" and categoria:
        reemplazar = request.POST.get("reemplazar") == "on"
        generar_programacion = request.POST.get("generar_programacion") == "on"
        fecha_inicio_programacion = fecha_desde_texto(request.POST.get("fecha_inicio_programacion"))
        canchas_programacion = canchas_desde_texto(request.POST.get("canchas_programacion")) or ["Principal", "Porvenir"]
        cancha_obligatoria = (request.POST.get("cancha_obligatoria") or "Porvenir").strip()
        franjas_seleccionadas = request.POST.getlist("franjas_programacion")
        franjas_programacion = [
            franja for franja in FRANJAS_PROGRAMACION_FIXTURE
            if not franjas_seleccionadas or franja[0] in franjas_seleccionadas
        ]
        existentes = Partido.objects.filter(categoria=categoria, fase="GRUPOS").exists()

        if existentes and not reemplazar:
            messages.error(request, "Esta categoría ya tiene partidos de grupos. Marca reemplazar fixture para generarlo de nuevo.")
            return redirect(f"{request.path}?categoria={categoria.id}&grupos={cantidad_grupos}&tipo_fixture={tipo_fixture}")

        partidos_a_crear = []

        if tipo_fixture == "MATA_MATA_IDA_VUELTA":
            equipos_mata_mata = list(equipos)

            if len(equipos_mata_mata) < 8:
                messages.error(request, "El torneo mata-mata necesita minimo 8 equipos activos para poder formar los cuartos.")
                return redirect(f"{request.path}?categoria={categoria.id}&grupos={cantidad_grupos}&tipo_fixture={tipo_fixture}")

            if len(equipos_mata_mata) % 2 != 0:
                messages.error(request, "El torneo mata-mata por parejas ida y vuelta necesita un numero par de equipos.")
                return redirect(f"{request.path}?categoria={categoria.id}&grupos={cantidad_grupos}&tipo_fixture={tipo_fixture}")

            hay_sorteo_manual, parejas_manuales, error_sorteo_manual = parejas_mata_mata_desde_formulario(equipos_mata_mata, request.POST)
            if error_sorteo_manual:
                messages.error(request, error_sorteo_manual)
                return redirect(f"{request.path}?categoria={categoria.id}&grupos={cantidad_grupos}&tipo_fixture={tipo_fixture}")

            if hay_sorteo_manual:
                partidos_a_crear = crear_partidos_mata_mata_desde_parejas(parejas_manuales)
            else:
                partidos_a_crear = crear_partidos_mata_mata_ida_vuelta(categoria, equipos_mata_mata)

            grupos_generados = {}
            for grupo_nombre, _, local, visitante in partidos_a_crear:
                grupos_generados.setdefault(grupo_nombre, [])
                if local not in grupos_generados[grupo_nombre]:
                    grupos_generados[grupo_nombre].append(local)
                if visitante not in grupos_generados[grupo_nombre]:
                    grupos_generados[grupo_nombre].append(visitante)
        else:
            cabezas = []

            for indice in range(cantidad_grupos):
                cabeza_id = request.POST.get(f"cabeza_{indice}")
                cabeza = equipos.filter(id=cabeza_id).first() if cabeza_id else None
                cabezas.append(cabeza)

            grupos_generados, sin_asignar = armar_grupos_desde_formulario(equipos, cabezas, request.POST, cantidad_grupos)

            if sin_asignar:
                nombres_sin_asignar = ", ".join(equipo.nombre for equipo in sin_asignar)
                messages.error(request, f"Faltan equipos por asignar a un grupo: {nombres_sin_asignar}.")
                return redirect(f"{request.path}?categoria={categoria.id}&grupos={cantidad_grupos}&tipo_fixture={tipo_fixture}")

            grupos_vacios = [nombre for nombre, equipos_grupo in grupos_generados.items() if len(equipos_grupo) < 2]

            if grupos_vacios:
                messages.error(request, f"Cada grupo debe tener al menos 2 equipos. Revisa: {', '.join(grupos_vacios)}.")
                return redirect(f"{request.path}?categoria={categoria.id}&grupos={cantidad_grupos}&tipo_fixture={tipo_fixture}")

            for grupo_nombre, equipos_grupo in grupos_generados.items():
                calendario = generar_fixture_grupo(equipos_grupo)

                for indice_fecha, partidos_fecha in enumerate(calendario, start=1):
                    for local, visitante in partidos_fecha:
                        partidos_a_crear.append((grupo_nombre, indice_fecha, local, visitante))

        if reemplazar:
            Partido.objects.filter(categoria=categoria, fase="GRUPOS").delete()

        creados = 0

        cupos = []
        conteos = {"cancha": {}, "franjas": {}, "fechas": {}}
        usados = set()
        equipos_por_id = {equipo.id: equipo for equipo in equipos}
        if generar_programacion:
            if not fecha_inicio_programacion:
                messages.error(request, "Para generar programacion automatica debes indicar la fecha de inicio.")
                return redirect(f"{request.path}?categoria={categoria.id}&grupos={cantidad_grupos}&tipo_fixture={tipo_fixture}")

            cupos = cupos_programacion_fixture(
                fecha_inicio_programacion,
                canchas_programacion,
                franjas_programacion,
                len(partidos_a_crear),
            )

            if len(cupos) < len(partidos_a_crear):
                advertencias_programacion.append(
                    "No hay suficientes cupos configurados; algunos partidos quedaran sin fecha/hora/cancha balanceada."
                )

        for grupo_nombre, indice_fecha, local, visitante in partidos_a_crear:
            defaults = {
                "fecha": date.today(),
                "hora": time(0, 0),
                "estado": "PROGRAMADO",
                "cancha": "",
                "estado_programacion": "MANUAL",
            }

            if generar_programacion and cupos:
                indice_cupo = elegir_cupo_balanceado(
                    local,
                    visitante,
                    cupos,
                    usados,
                    conteos,
                    equipos_por_id.keys(),
                    cancha_obligatoria,
                )

                if indice_cupo is not None:
                    usados.add(indice_cupo)
                    cupo = cupos[indice_cupo]
                    defaults.update({
                        "fecha": cupo["fecha"],
                        "hora": cupo["hora"],
                        "cancha": cupo["cancha"],
                        "estado_programacion": "SUGERIDA",
                    })

                    for equipo_id in [local.id, visitante.id]:
                        conteos["franjas"].setdefault(equipo_id, {})
                        conteos["fechas"].setdefault(equipo_id, {})
                        conteos["franjas"][equipo_id][cupo["franja"]] = conteos["franjas"][equipo_id].get(cupo["franja"], 0) + 1
                        conteos["fechas"][equipo_id][cupo["fecha"]] = conteos["fechas"][equipo_id].get(cupo["fecha"], 0) + 1
                        if cancha_obligatoria and cupo["cancha"].lower() == cancha_obligatoria.lower():
                            conteos["cancha"][equipo_id] = conteos["cancha"].get(equipo_id, 0) + 1

            _, creado = Partido.objects.get_or_create(
                categoria=categoria,
                fase="GRUPOS",
                grupo=grupo_nombre,
                numero_fecha=str(indice_fecha),
                equipo_local=local,
                equipo_visitante=visitante,
                defaults=defaults,
            )

            if creado:
                creados += 1

        if generar_programacion:
            resumen_programacion = resumen_equidad_programacion(conteos, equipos_por_id, franjas_programacion)
            sin_cancha_obligatoria = [
                item["equipo"].nombre
                for item in resumen_programacion
                if item["cancha_obligatoria"] == 0
            ]
            if sin_cancha_obligatoria:
                advertencias_programacion.append(
                    f"Equipos pendientes por jugar en {cancha_obligatoria}: {', '.join(sin_cancha_obligatoria)}."
                )

            if advertencias_programacion:
                for advertencia in advertencias_programacion:
                    messages.warning(request, advertencia)
            else:
                messages.success(request, "Programacion automatica balanceada sin alertas de equidad.")

        messages.success(request, f"Fixture generado para {categoria.nombre}. Partidos creados: {creados}.")
        registrar_actividad(
            request,
            "GENERAR_FIXTURE",
            categoria,
            descripcion=f"Genero fixture para {categoria.nombre}. Partidos creados: {creados}.",
            datos={
                "partidos_creados": creados,
                "grupos": cantidad_grupos,
                "programacion_automatica": generar_programacion,
                "cancha_obligatoria": cancha_obligatoria if generar_programacion else "",
                "tipo_fixture": tipo_fixture,
            },
        )

    return render(request, "gestion/generar_fixture.html", {
        "categorias": categorias,
        "categoria": categoria,
        "equipos": equipos,
        "cantidad_grupos": cantidad_grupos,
        "tipo_fixture": tipo_fixture,
        "letras_grupos": letras_grupos,
        "parejas_mata_mata": parejas_mata_mata,
        "grupos_generados": grupos_generados,
        "franjas_programacion": FRANJAS_PROGRAMACION_FIXTURE,
        "resumen_programacion": resumen_programacion,
    })


@login_required
@user_passes_test(es_editor_torneo)
def gestion_equipos(request):
    torneo = torneo_actual(request)
    if usuario_solo_descarga_planillas(request.user, torneo):
        return denegar_permiso_torneo()
    categorias = Categoria.objects.order_by("nombre")
    equipos = equipos_gestion_filtrados(torneo, request.GET.get("q", ""), request.GET.get("categoria", ""))
    if torneo:
        categorias = categorias.filter(torneo=torneo)
    q = request.GET.get("q", "").strip()
    categoria_id = request.GET.get("categoria", "").strip()

    return render(request, "gestion/equipos.html", {
        "equipos": equipos,
        "categorias": categorias,
        "q": q,
        "categoria_id": categoria_id,
    })


def equipos_gestion_filtrados(torneo, q="", categoria_id=""):
    equipos = Equipo.objects.select_related("categoria").order_by("categoria__nombre", "nombre")
    if torneo:
        equipos = equipos.filter(categoria__torneo=torneo)
    q = (q or "").strip()
    categoria_id = (categoria_id or "").strip()
    if q:
        equipos = equipos.filter(nombre__icontains=q)
    if categoria_id:
        equipos = equipos.filter(categoria_id=categoria_id)
    return equipos


def equipos_gestionables_para_usuario(request):
    equipos = Equipo.objects.select_related("categoria", "categoria__torneo").order_by("categoria__nombre", "nombre")
    if request.user.is_superuser:
        return equipos
    return equipos.filter(categoria__torneo__in=torneos_para_usuario(request)).distinct()


def username_delegado_equipo(equipo, usuario_actual=None):
    nombre_equipo = (slugify(equipo.nombre) or f"equipo-{equipo.id}").replace("-", "")
    base = f"admin-{nombre_equipo}"[:140].strip("-") or f"admin-equipo-{equipo.id}"
    username = base
    contador = 2
    existentes = User.objects.all()
    if usuario_actual and usuario_actual.pk:
        existentes = existentes.exclude(pk=usuario_actual.pk)
    while existentes.filter(username__iexact=username).exists():
        sufijo = f"-{contador}"
        username = f"{base[:150 - len(sufijo)]}{sufijo}"
        contador += 1
    return username


@login_required
@user_passes_test(es_editor_torneo)
@require_POST
def gestion_equipos_acceso_delegado_masivo(request):
    torneo = torneo_actual(request)
    if not puede_gestionar_torneo(request, torneo, "editar"):
        return denegar_permiso_torneo()

    acceso_hasta_texto = (request.POST.get("acceso_delegado_hasta") or "").strip()
    q = request.POST.get("q", "")
    categoria_id = request.POST.get("categoria", "")

    if not acceso_hasta_texto:
        messages.error(request, "Selecciona la fecha/hora de vencimiento.")
        return redirect(f"{reverse('gestion_equipos')}?q={quote(q)}&categoria={quote(categoria_id)}")

    acceso_hasta = parse_datetime(acceso_hasta_texto)
    if not acceso_hasta:
        messages.error(request, "La fecha/hora de vencimiento no es valida.")
        return redirect(f"{reverse('gestion_equipos')}?q={quote(q)}&categoria={quote(categoria_id)}")
    if timezone.is_naive(acceso_hasta):
        acceso_hasta = timezone.make_aware(acceso_hasta, timezone.get_current_timezone())

    equipos = equipos_gestion_filtrados(torneo, q, categoria_id).filter(responsable__isnull=False)
    cantidad = equipos.update(acceso_delegado_hasta=acceso_hasta)

    registrar_actividad(
        request,
        "ACCESO_DELEGADO_MASIVO",
        torneo=torneo,
        descripcion=f"Actualizo vencimiento de acceso delegado para {cantidad} equipo(s) hasta {timezone.localtime(acceso_hasta).strftime('%d/%m/%Y %H:%M')}.",
        datos={
            "equipos": cantidad,
            "categoria_id": categoria_id,
            "q": q,
            "acceso_delegado_hasta": acceso_hasta.isoformat(),
        },
    )
    messages.success(request, f"Fecha de acceso delegado actualizada para {cantidad} equipo(s) con responsable asignado.")
    return redirect(f"{reverse('gestion_equipos')}?q={quote(q)}&categoria={quote(categoria_id)}")


@login_required
@user_passes_test(es_editor_torneo)
@require_POST
def gestion_equipos_crear_delegados_masivo(request):
    torneo = torneo_actual(request)
    if not puede_gestionar_torneo(request, torneo, "editar"):
        return denegar_permiso_torneo()

    password_temporal = request.POST.get("password_temporal") or ""
    acceso_hasta_texto = (request.POST.get("acceso_delegado_hasta") or "").strip()
    q = request.POST.get("q", "")
    categoria_id = request.POST.get("categoria", "")

    if len(password_temporal) < 8:
        messages.error(request, "La contrasena temporal debe tener minimo 8 caracteres.")
        return redirect(f"{reverse('gestion_equipos')}?q={quote(q)}&categoria={quote(categoria_id)}")

    acceso_hasta = None
    if acceso_hasta_texto:
        acceso_hasta = parse_datetime(acceso_hasta_texto)
        if not acceso_hasta:
            messages.error(request, "La fecha/hora de vencimiento no es valida.")
            return redirect(f"{reverse('gestion_equipos')}?q={quote(q)}&categoria={quote(categoria_id)}")
        if timezone.is_naive(acceso_hasta):
            acceso_hasta = timezone.make_aware(acceso_hasta, timezone.get_current_timezone())

    equipos = equipos_gestion_filtrados(torneo, q, categoria_id).filter(responsable__isnull=True)
    creados = []
    for equipo in equipos:
        username = username_delegado_equipo(equipo)
        usuario = User.objects.create_user(
            username=username,
            password=password_temporal,
            first_name=equipo.nombre[:150],
        )
        equipo.responsable = usuario
        if acceso_hasta:
            equipo.acceso_delegado_hasta = acceso_hasta
        equipo.save(update_fields=["responsable", "acceso_delegado_hasta"])
        creados.append(username)

    registrar_actividad(
        request,
        "CREAR_DELEGADOS_MASIVO",
        torneo=torneo,
        descripcion=f"Creo usuarios delegados para {len(creados)} equipo(s).",
        datos={
            "usuarios": creados,
            "categoria_id": categoria_id,
            "q": q,
            "acceso_delegado_hasta": acceso_hasta.isoformat() if acceso_hasta else "",
        },
    )
    if creados:
        messages.success(request, f"Usuarios delegados creados: {len(creados)}. Contrasena temporal aplicada.")
        messages.info(request, "Usuarios: " + ", ".join(creados[:20]))
        if len(creados) > 20:
            messages.info(request, f"Y {len(creados) - 20} usuario(s) mas.")
    else:
        messages.warning(request, "No se crearon usuarios: los equipos visibles ya tienen delegado responsable.")
    return redirect(f"{reverse('gestion_equipos')}?q={quote(q)}&categoria={quote(categoria_id)}")


@login_required
@user_passes_test(es_editor_torneo)
@require_POST
def gestion_equipos_renombrar_delegados_masivo(request):
    torneo = torneo_actual(request)
    if not puede_gestionar_torneo(request, torneo, "editar"):
        return denegar_permiso_torneo()

    q = request.POST.get("q", "")
    categoria_id = request.POST.get("categoria", "")
    equipos = equipos_gestion_filtrados(torneo, q, categoria_id).filter(responsable__isnull=False).select_related("responsable")
    cambios = []

    for equipo in equipos:
        usuario = equipo.responsable
        username_nuevo = username_delegado_equipo(equipo, usuario_actual=usuario)
        if usuario.username == username_nuevo:
            continue
        username_anterior = usuario.username
        usuario.username = username_nuevo
        usuario.save(update_fields=["username"])
        cambios.append({"equipo": equipo.nombre, "antes": username_anterior, "despues": username_nuevo})

    registrar_actividad(
        request,
        "RENOMBRAR_DELEGADOS_MASIVO",
        torneo=torneo,
        descripcion=f"Renombro usuarios delegados para {len(cambios)} equipo(s).",
        datos={
            "cambios": cambios,
            "categoria_id": categoria_id,
            "q": q,
        },
    )
    if cambios:
        messages.success(request, f"Usuarios delegados renombrados: {len(cambios)}.")
        resumen = ", ".join(f"{item['antes']} -> {item['despues']}" for item in cambios[:10])
        messages.info(request, resumen)
        if len(cambios) > 10:
            messages.info(request, f"Y {len(cambios) - 10} cambio(s) mas.")
    else:
        messages.warning(request, "No hubo usuarios para renombrar en los equipos visibles.")
    return redirect(f"{reverse('gestion_equipos')}?q={quote(q)}&categoria={quote(categoria_id)}")


@login_required
@user_passes_test(es_editor_torneo)
@require_POST
def gestion_equipos_permisos_delegados_masivo(request):
    torneo = torneo_actual(request)
    if not puede_gestionar_torneo(request, torneo, "editar"):
        return denegar_permiso_torneo()

    q = request.POST.get("q", "")
    categoria_id = request.POST.get("categoria", "")
    opciones_validas = {"MANTENER", "HABILITAR", "DESHABILITAR"}
    editar = (request.POST.get("permiso_editar") or "MANTENER").upper()
    fotos = (request.POST.get("permiso_fotos") or "MANTENER").upper()
    if editar not in opciones_validas or fotos not in opciones_validas:
        messages.error(request, "Selecciona opciones válidas para los permisos.")
        return redirect(f"{reverse('gestion_equipos')}?q={quote(q)}&categoria={quote(categoria_id)}")
    if editar == "MANTENER" and fotos == "MANTENER":
        messages.warning(request, "No seleccionaste ningún cambio de permisos.")
        return redirect(f"{reverse('gestion_equipos')}?q={quote(q)}&categoria={quote(categoria_id)}")

    equipos = equipos_gestion_filtrados(torneo, q, categoria_id).filter(responsable__isnull=False)
    cambios = {}
    if editar != "MANTENER":
        cambios["delegado_puede_editar_equipo"] = editar == "HABILITAR"
    if fotos != "MANTENER":
        cambios["delegado_puede_cargar_fotos_jugadores"] = fotos == "HABILITAR"

    cantidad = equipos.update(**cambios)
    registrar_actividad(
        request,
        "PERMISOS_DELEGADOS_MASIVO",
        torneo=torneo,
        descripcion=f"Actualizó permisos de delegados para {cantidad} equipo(s).",
        datos={
            "equipos": cantidad,
            "categoria_id": categoria_id,
            "q": q,
            "permiso_editar": editar,
            "permiso_fotos": fotos,
        },
    )
    messages.success(request, f"Permisos actualizados para {cantidad} equipo(s) con delegado asignado.")
    return redirect(f"{reverse('gestion_equipos')}?q={quote(q)}&categoria={quote(categoria_id)}")


@login_required
@user_passes_test(es_editor_torneo)
def gestion_equipo_nuevo(request):
    torneo = torneo_actual(request)
    if not puede_gestionar_torneo(request, torneo, "editar"):
        return denegar_permiso_torneo()
    form = EquipoForm(request.POST or None, request.FILES or None, torneo=torneo)

    if request.method == "POST" and form.is_valid():
        equipo = form.save(commit=False)
        aplicar_imagen_cloudinary(
            equipo,
            "escudo",
            request.POST.get("imagen_cloudinary"),
            request.FILES.get("escudo"),
        )
        equipo.save()
        form.save_m2m()
        registrar_actividad(request, "CREAR", equipo, descripcion=f"Creo equipo {equipo.nombre}.")
        messages.success(request, "Equipo creado correctamente.")
        return redirect("gestion_equipo_editar", equipo_id=equipo.id)

    return render(request, "gestion/formulario.html", {
        "titulo": "Nuevo equipo",
        "form": form,
        "volver_url": "gestion_equipos",
        "cloudinary_images": listar_imagenes_cloudinary(),
        "cloudinary_label": "Seleccionar escudo existente de Cloudinary",
    })


@login_required
@user_passes_test(es_editor_torneo)
def gestion_equipo_editar(request, equipo_id):
    equipo = get_object_or_404(equipos_gestionables_para_usuario(request), id=equipo_id)
    torneo_equipo = equipo.categoria.torneo if equipo.categoria_id else None
    if not puede_gestionar_torneo(request, torneo_equipo, "editar"):
        return denegar_permiso_torneo()
    form = EquipoForm(request.POST or None, request.FILES or None, instance=equipo, torneo=torneo_equipo)
    jugadores = equipo.jugadores.order_by("dorsal", "nombres")

    if request.method == "POST" and form.is_valid():
        equipo = form.save(commit=False)
        aplicar_imagen_cloudinary(
            equipo,
            "escudo",
            request.POST.get("imagen_cloudinary"),
            request.FILES.get("escudo"),
        )
        equipo.save()
        form.save_m2m()
        registrar_actividad(request, "EDITAR", equipo, descripcion=f"Actualizo equipo {equipo.nombre}.")
        messages.success(request, "Equipo actualizado correctamente.")
        return redirect("gestion_equipo_editar", equipo_id=equipo.id)

    return render(request, "gestion/equipo_formulario.html", {
        "titulo": f"Editar equipo: {equipo.nombre}",
        "form": form,
        "equipo": equipo,
        "jugadores": jugadores,
        "estados_jugador": Jugador.ESTADOS,
        "volver_url": "gestion_equipos",
        "cloudinary_images": listar_imagenes_cloudinary(),
        "cloudinary_label": "Seleccionar escudo existente de Cloudinary",
    })


@login_required
@user_passes_test(es_editor_torneo)
def gestion_equipo_reinscribir(request, equipo_id):
    equipo = get_object_or_404(Equipo.objects.select_related("categoria", "categoria__torneo"), id=equipo_id)
    if not puede_gestionar_torneo(request, equipo.categoria.torneo if equipo.categoria_id else None, "editar"):
        return denegar_permiso_torneo()

    jugadores = list(equipo.jugadores.order_by("estado", "dorsal", "nombres"))
    form = EquipoReinscripcionForm(
        request.POST or None,
        user=request.user,
        equipo_origen=equipo,
    )

    if request.method == "POST" and form.is_valid():
        categoria_destino = form.cleaned_data["categoria_destino"]
        if not puede_gestionar_torneo(request, categoria_destino.torneo, "editar"):
            return denegar_permiso_torneo()

        seleccionados = set(request.POST.getlist("jugadores"))
        incluir_retirados = form.cleaned_data["copiar_jugadores_retirados"]
        jugadores_a_copiar = [
            jugador for jugador in jugadores
            if str(jugador.id) in seleccionados and (incluir_retirados or jugador.estado != "RETIRADO")
        ]

        with transaction.atomic():
            nuevo_equipo = Equipo.objects.create(
                nombre=equipo.nombre,
                categoria=categoria_destino,
                responsable=equipo.responsable if form.cleaned_data["conservar_delegado"] else None,
                acceso_delegado_hasta=equipo.acceso_delegado_hasta if form.cleaned_data["conservar_acceso_delegado"] else None,
                delegado=equipo.delegado,
                telefono=equipo.telefono,
                director_tecnico=equipo.director_tecnico,
                telefono_dt=equipo.telefono_dt,
                asistente_tecnico=equipo.asistente_tecnico,
                telefono_at=equipo.telefono_at,
                escudo=equipo.escudo,
                activo=True,
            )
            copiados = 0
            omitidos = []
            for jugador in jugadores_a_copiar:
                nuevo_jugador = Jugador(
                    equipo=nuevo_equipo,
                    dorsal=jugador.dorsal,
                    nombres=jugador.nombres,
                    cedula=jugador.cedula,
                    fecha_nacimiento=jugador.fecha_nacimiento,
                    telefono=jugador.telefono,
                    estado="ACTIVO" if jugador.estado == "RETIRADO" else jugador.estado,
                    foto=jugador.foto,
                    es_foraneo=jugador.es_foraneo,
                )
                try:
                    nuevo_jugador.save()
                    copiados += 1
                except Exception as exc:
                    omitidos.append(f"{jugador.nombres}: {exc}")

        registrar_actividad(
            request,
            "REINSCRIBIR_EQUIPO",
            nuevo_equipo,
            torneo=categoria_destino.torneo,
            descripcion=f"Reinscribio {equipo.nombre} desde {equipo.categoria.nombre} hacia {categoria_destino.nombre}. Jugadores copiados: {copiados}.",
            datos={
                "equipo_origen_id": equipo.id,
                "equipo_nuevo_id": nuevo_equipo.id,
                "categoria_destino_id": categoria_destino.id,
                "jugadores_copiados": copiados,
                "omitidos": omitidos[:20],
            },
        )
        messages.success(request, f"Equipo reinscrito en {categoria_destino.nombre}. Jugadores copiados: {copiados}.")
        for error in omitidos[:5]:
            messages.warning(request, f"Omitido: {error}")
        if len(omitidos) > 5:
            messages.warning(request, f"Hay {len(omitidos) - 5} jugador(es) omitidos adicionales.")
        return redirect("gestion_equipo_editar", equipo_id=nuevo_equipo.id)

    return render(request, "gestion/equipo_reinscribir.html", {
        "titulo": f"Reinscribir equipo: {equipo.nombre}",
        "equipo": equipo,
        "form": form,
        "jugadores": jugadores,
        "volver_url": "gestion_equipos",
    })


@login_required
@user_passes_test(es_editor_torneo)
@require_POST
def gestion_equipo_jugadores_guardar(request, equipo_id):
    equipo = get_object_or_404(equipos_gestionables_para_usuario(request), id=equipo_id)
    torneo_equipo = equipo.categoria.torneo if equipo.categoria_id else None
    if not puede_gestionar_torneo(request, torneo_equipo, "editar"):
        return denegar_permiso_torneo()

    jugadores = list(equipo.jugadores.all())
    errores = []
    actualizados = 0

    for jugador in jugadores:
        prefijo = f"jugador_{jugador.id}_"
        nombres = (request.POST.get(prefijo + "nombres") or "").strip()
        cedula = (request.POST.get(prefijo + "cedula") or "").strip()
        fecha_nacimiento = (request.POST.get(prefijo + "fecha_nacimiento") or "").strip()
        estado = request.POST.get(prefijo + "estado") or "ACTIVO"

        if not nombres or not cedula or not fecha_nacimiento:
            errores.append(f"{jugador.nombres}: nombre, cedula y fecha son obligatorios.")
            continue

        jugador.dorsal = request.POST.get(prefijo + "dorsal") or None
        jugador.nombres = nombres.upper()
        jugador.cedula = cedula
        jugador.fecha_nacimiento = fecha_nacimiento
        jugador.estado = estado
        jugador.es_foraneo = request.POST.get(prefijo + "es_foraneo") == "on"

        try:
            jugador.save()
            actualizados += 1
        except Exception as exc:
            errores.append(f"{nombres}: {exc}")

    nuevo_nombre = (request.POST.get("nuevo_nombres") or "").strip()
    nuevo_cedula = (request.POST.get("nuevo_cedula") or "").strip()
    nuevo_fecha = (request.POST.get("nuevo_fecha_nacimiento") or "").strip()
    if nuevo_nombre or nuevo_cedula or nuevo_fecha:
        if not nuevo_nombre or not nuevo_cedula or not nuevo_fecha:
            errores.append("Para agregar jugador nuevo debes llenar nombre, cedula y fecha de nacimiento.")
        else:
            nuevo = Jugador(
                equipo=equipo,
                dorsal=request.POST.get("nuevo_dorsal") or None,
                nombres=nuevo_nombre.upper(),
                cedula=nuevo_cedula,
                fecha_nacimiento=nuevo_fecha,
                estado=request.POST.get("nuevo_estado") or "ACTIVO",
                es_foraneo=request.POST.get("nuevo_es_foraneo") == "on",
            )
            try:
                nuevo.save()
                registrar_actividad(request, "CREAR", nuevo, descripcion=f"Agrego jugador {nuevo.nombres} al equipo {equipo.nombre}.")
                messages.success(request, f"Jugador agregado: {nuevo.nombres}.")
            except Exception as exc:
                errores.append(f"No se pudo agregar {nuevo_nombre}: {exc}")

    if actualizados:
        registrar_actividad(
            request,
            "EDITAR_PLANTILLA",
            equipo,
            descripcion=f"Actualizo plantilla de {equipo.nombre}. Jugadores actualizados: {actualizados}.",
            datos={"actualizados": actualizados},
        )
        messages.success(request, f"Plantilla actualizada: {actualizados} jugador(es).")
    for error in errores[:8]:
        messages.error(request, error)
    if len(errores) > 8:
        messages.error(request, f"Hay {len(errores) - 8} errores adicionales.")

    return redirect("gestion_equipo_editar", equipo_id=equipo.id)


@login_required
@user_passes_test(es_editor_torneo)
@require_POST
def gestion_equipo_eliminar(request, equipo_id):
    equipo = get_object_or_404(equipos_gestionables_para_usuario(request), id=equipo_id)
    torneo_equipo = equipo.categoria.torneo if equipo.categoria_id else None
    if not puede_gestionar_torneo(request, torneo_equipo, "editar"):
        return denegar_permiso_torneo()
    nombre = equipo.nombre
    registrar_actividad(request, "ELIMINAR", equipo, descripcion=f"Elimino equipo {nombre}.")
    equipo.delete()
    messages.success(request, f"Equipo eliminado: {nombre}.")
    return redirect("gestion_equipos")


@login_required
@user_passes_test(es_editor_torneo)
def gestion_jugadores(request):
    torneo = torneo_actual(request)
    if usuario_solo_descarga_planillas(request.user, torneo):
        return denegar_permiso_torneo()
    categorias = Categoria.objects.order_by("nombre")
    equipos = Equipo.objects.select_related("categoria").order_by("categoria__nombre", "nombre")
    jugadores = Jugador.objects.select_related("equipo", "equipo__categoria").order_by(
        "equipo__categoria__nombre",
        "equipo__nombre",
        "dorsal",
        "nombres",
    )
    if torneo:
        categorias = categorias.filter(torneo=torneo)
        equipos = equipos.filter(categoria__torneo=torneo)
        jugadores = jugadores.filter(equipo__categoria__torneo=torneo)
    q = request.GET.get("q", "").strip()
    categoria_id = request.GET.get("categoria", "").strip()
    equipo_id = request.GET.get("equipo", "").strip()

    if q:
        jugadores = jugadores.filter(Q(nombres__icontains=q) | Q(cedula__icontains=q))

    if categoria_id:
        jugadores = jugadores.filter(equipo__categoria_id=categoria_id)

    if equipo_id:
        jugadores = jugadores.filter(equipo_id=equipo_id)

    return render(request, "gestion/jugadores.html", {
        "jugadores": jugadores,
        "categorias": categorias,
        "equipos": equipos,
        "q": q,
        "categoria_id": categoria_id,
        "equipo_id": equipo_id,
    })


@login_required
@user_passes_test(es_editor_torneo)
def gestion_jugador_nuevo(request):
    torneo = torneo_actual(request)
    if not puede_gestionar_torneo(request, torneo, "editar"):
        return denegar_permiso_torneo()
    form = JugadorForm(request.POST or None, request.FILES or None, torneo=torneo)

    if request.method == "POST" and form.is_valid():
        jugador = form.save(commit=False)
        aplicar_imagen_cloudinary(
            jugador,
            "foto",
            request.POST.get("imagen_cloudinary"),
            request.FILES.get("foto"),
        )
        jugador.save()
        form.save_m2m()
        registrar_actividad(request, "CREAR", jugador, descripcion=f"Creo jugador {jugador.nombres}.")
        messages.success(request, "Jugador creado correctamente.")
        return redirect("gestion_jugador_editar", jugador_id=jugador.id)

    return render(request, "gestion/formulario.html", {
        "titulo": "Nuevo jugador",
        "form": form,
        "volver_url": "gestion_jugadores",
        "cloudinary_images": listar_imagenes_cloudinary(),
        "cloudinary_label": "Seleccionar foto existente de Cloudinary",
    })


@login_required
@user_passes_test(es_editor_torneo)
def gestion_jugador_editar(request, jugador_id):
    torneo = torneo_actual(request)
    if not puede_gestionar_torneo(request, torneo, "editar"):
        return denegar_permiso_torneo()
    jugadores = Jugador.objects.select_related("equipo", "equipo__categoria")
    if torneo:
        jugadores = jugadores.filter(equipo__categoria__torneo=torneo)
    jugador = get_object_or_404(jugadores, id=jugador_id)
    form = JugadorForm(request.POST or None, request.FILES or None, instance=jugador, torneo=torneo)

    if request.method == "POST" and form.is_valid():
        jugador = form.save(commit=False)
        aplicar_imagen_cloudinary(
            jugador,
            "foto",
            request.POST.get("imagen_cloudinary"),
            request.FILES.get("foto"),
        )
        jugador.save()
        form.save_m2m()
        registrar_actividad(request, "EDITAR", jugador, descripcion=f"Actualizo jugador {jugador.nombres}.")
        messages.success(request, "Jugador actualizado correctamente.")
        return redirect("gestion_jugadores")

    return render(request, "gestion/formulario.html", {
        "titulo": f"Editar jugador: {jugador.nombres}",
        "form": form,
        "volver_url": "gestion_jugadores",
        "cloudinary_images": listar_imagenes_cloudinary(),
        "cloudinary_label": "Seleccionar foto existente de Cloudinary",
    })


@login_required
@user_passes_test(es_editor_torneo)
@require_POST
def gestion_jugador_eliminar(request, jugador_id):
    torneo = torneo_actual(request)
    if not puede_gestionar_torneo(request, torneo, "editar"):
        return denegar_permiso_torneo()
    jugadores = Jugador.objects.select_related("equipo", "equipo__categoria")
    if torneo:
        jugadores = jugadores.filter(equipo__categoria__torneo=torneo)
    jugador = get_object_or_404(jugadores, id=jugador_id)
    nombre = jugador.nombres
    registrar_actividad(request, "ELIMINAR", jugador, descripcion=f"Elimino jugador {nombre}.")
    jugador.delete()
    messages.success(request, f"Jugador eliminado: {nombre}.")
    return redirect("gestion_jugadores")


@login_required
@user_passes_test(es_editor_torneo)
def gestion_importar_planilla(request):
    torneo = torneo_actual(request)
    if not puede_gestionar_torneo(request, torneo, "editar"):
        return denegar_permiso_torneo()

    if request.method == "POST":
        archivo = request.FILES.get("archivo_excel")

        if not archivo:
            messages.error(request, "Selecciona un archivo Excel.")
            return redirect("gestion_importar_planilla")

        try:
            workbook = load_workbook(archivo, data_only=True)
            hoja = obtener_hoja_planilla_excel(workbook)

            categoria_nombre = limpiar_texto_excel(hoja["D3"].value)
            equipo_nombre = limpiar_texto_excel(hoja["I3"].value)
            delegado = limpiar_texto_excel(hoja["D4"].value)
            telefono_delegado = limpiar_cedula_excel(hoja["I4"].value)
            director_tecnico = limpiar_texto_excel(hoja["C39"].value)
            telefono_dt = limpiar_cedula_excel(hoja["G39"].value)
            asistente_tecnico = limpiar_texto_excel(hoja["C40"].value)
            telefono_at = limpiar_cedula_excel(hoja["G40"].value)

            if not categoria_nombre:
                messages.error(request, "No se encontró la categoría en la celda D3.")
                return redirect("gestion_importar_planilla")

            if not equipo_nombre:
                messages.error(request, "No se encontró el equipo en la celda I3.")
                return redirect("gestion_importar_planilla")

            categorias = Categoria.objects.filter(nombre__iexact=categoria_nombre)
            if torneo:
                categorias = categorias.filter(torneo=torneo)
            categoria = categorias.first()

            if not categoria:
                messages.error(request, f"No existe la categoría: {categoria_nombre}. Créala primero.")
                return redirect("gestion_importar_planilla")

            equipo, _ = Equipo.objects.get_or_create(
                nombre=equipo_nombre.upper(),
                categoria=categoria,
                defaults={"activo": True},
            )

            equipo.delegado = delegado.upper() if delegado else equipo.delegado
            equipo.telefono = telefono_delegado or equipo.telefono
            equipo.director_tecnico = director_tecnico.upper() if director_tecnico else equipo.director_tecnico
            equipo.telefono_dt = telefono_dt or equipo.telefono_dt
            equipo.asistente_tecnico = asistente_tecnico.upper() if asistente_tecnico else equipo.asistente_tecnico
            equipo.telefono_at = telefono_at or equipo.telefono_at
            equipo.activo = True
            equipo.save()

            creados = 0
            actualizados = 0
            omitidos = 0
            eliminados = 0
            errores = []
            cedulas_importadas = set()

            for fila in range(8, 38):
                nombre = limpiar_texto_excel(hoja[f"C{fila}"].value)
                dorsal = limpiar_entero_excel(hoja[f"D{fila}"].value)
                dia = hoja[f"E{fila}"].value
                mes = hoja[f"F{fila}"].value
                anio = hoja[f"G{fila}"].value
                cedula = limpiar_cedula_excel(hoja[f"H{fila}"].value)

                if not nombre and not cedula:
                    continue

                if not nombre:
                    omitidos += 1
                    errores.append(f"Fila {fila}: falta el nombre del jugador.")
                    continue

                if not cedula:
                    omitidos += 1
                    errores.append(f"Fila {fila}: falta la cédula de {nombre}.")
                    continue

                fecha_nacimiento = construir_fecha_excel(dia, mes, anio)

                if not fecha_nacimiento:
                    omitidos += 1
                    errores.append(f"Fila {fila}: fecha de nacimiento inválida para {nombre}.")
                    continue

                jugador_misma_categoria = Jugador.objects.filter(
                    cedula=cedula,
                    equipo__categoria=categoria,
                ).exclude(equipo=equipo).select_related("equipo").first()
                if jugador_misma_categoria:
                    omitidos += 1
                    errores.append(
                        f"Fila {fila}: {nombre} ya esta inscrito en {jugador_misma_categoria.equipo.nombre} "
                        f"para {categoria.nombre}."
                    )
                    continue

                _, creado = Jugador.objects.update_or_create(
                    equipo=equipo,
                    cedula=cedula,
                    defaults={
                        "dorsal": dorsal,
                        "nombres": nombre.upper(),
                        "fecha_nacimiento": fecha_nacimiento,
                        "estado": "ACTIVO",
                    },
                )
                cedulas_importadas.add(cedula)

                if creado:
                    creados += 1
                else:
                    actualizados += 1

            if cedulas_importadas:
                eliminados, _ = Jugador.objects.filter(equipo=equipo).exclude(
                    cedula__in=cedulas_importadas
                ).delete()

            messages.success(
                request,
                f"Planilla importada: {equipo.nombre} / {categoria.nombre}. Nuevos: {creados}. Actualizados: {actualizados}. Eliminados: {eliminados}. Omitidos: {omitidos}.",
            )
            registrar_actividad(
                request,
                "IMPORTAR_PLANILLA",
                equipo,
                descripcion=f"Importo planilla de {equipo.nombre}. Nuevos: {creados}. Actualizados: {actualizados}. Eliminados: {eliminados}. Omitidos: {omitidos}.",
                datos={"creados": creados, "actualizados": actualizados, "eliminados": eliminados, "omitidos": omitidos},
            )

            for error in errores[:12]:
                messages.warning(request, error)

            if len(errores) > 12:
                messages.warning(request, f"Hay {len(errores) - 12} advertencias adicionales.")

            return redirect("gestion_jugadores")

        except Exception as exc:
            messages.error(request, f"No se pudo importar la planilla: {exc}")
            return redirect("gestion_importar_planilla")

    return render(request, "gestion/importar_planilla.html")


@login_required
@user_passes_test(es_editor_torneo)
def gestion_partidos(request):
    torneo = torneo_actual(request)
    permisos = permisos_torneo_usuario(request.user, torneo)
    puede_editar = bool(permisos and permisos.puede_editar)
    puede_programar = bool(permisos and permisos.puede_programar)
    puede_validar = bool(permisos and permisos.puede_validar)
    puede_descargar_planillas = bool(permisos and getattr(permisos, "puede_descargar_planillas", False))
    categorias = Categoria.objects.order_by("nombre")
    if torneo:
        categorias = categorias.filter(torneo=torneo)

    partidos = Partido.objects.select_related("categoria", "equipo_local", "equipo_visitante").order_by(
        "fecha",
        "hora",
        "categoria__nombre",
        "grupo",
        "fase",
    )
    if torneo:
        partidos = partidos.filter(categoria__torneo=torneo)

    q = request.GET.get("q", "").strip()
    categoria_id = request.GET.get("categoria", "").strip()
    estado = request.GET.get("estado", "").strip()

    if q:
        partidos = partidos.filter(
            Q(equipo_local__nombre__icontains=q) |
            Q(equipo_visitante__nombre__icontains=q) |
            Q(cancha__icontains=q)
        )

    if categoria_id:
        partidos = partidos.filter(categoria_id=categoria_id)

    if estado:
        partidos = partidos.filter(estado=estado)

    return render(request, "gestion/partidos.html", {
        "partidos": partidos,
        "categorias": categorias,
        "estados": Partido.ESTADOS,
        "q": q,
        "categoria_id": categoria_id,
        "estado": estado,
        "puede_editar": puede_editar,
        "puede_programar": puede_programar,
        "puede_validar": puede_validar,
        "puede_descargar_planillas": puede_descargar_planillas,
    })


@login_required
@user_passes_test(es_editor_torneo)
def gestion_partido_nuevo(request):
    torneo = torneo_actual(request)
    if not puede_gestionar_torneo(request, torneo, "programar"):
        return denegar_permiso_torneo()
    form = PartidoProgramacionForm(request.POST or None, torneo=torneo)

    if request.method == "POST" and form.is_valid():
        partido = form.save()
        from django.utils import timezone

        if partido.estado == "EN_JUEGO" and not partido.inicio_en_vivo:
            partido.inicio_en_vivo = timezone.now()
            partido.save()
        registrar_actividad(request, "CREAR", partido, descripcion=f"Creo partido {partido.equipo_local} vs {partido.equipo_visitante}.")
        messages.success(request, "Partido creado correctamente.")
        return redirect("gestion_partido_editar", partido_id=partido.id)
        
    return render(request, "gestion/formulario.html", {
        "titulo": "Nuevo partido",
        "form": form,
        "volver_url": "gestion_partidos",
    })


@login_required
@user_passes_test(es_editor_torneo)
def gestion_partido_editar(request, partido_id):
    torneo = torneo_actual(request)
    if not puede_gestionar_torneo(request, torneo, "programar"):
        return denegar_permiso_torneo()
    volver_url = url_retorno_gestion(request, "gestion_partidos")
    partidos = Partido.objects.select_related("categoria", "equipo_local", "equipo_visitante")
    if torneo:
        partidos = partidos.filter(categoria__torneo=torneo)
    partido = get_object_or_404(partidos, id=partido_id)
    form = PartidoProgramacionForm(request.POST or None, instance=partido, torneo=torneo)

    if request.method == "POST" and form.is_valid():
        partido = form.save()
        if partido.estado_programacion == "SUGERIDA":
            partido.estado_programacion = "OFICIAL"
            partido.save(update_fields=["estado_programacion"])
        registrar_actividad(request, "PROGRAMAR", partido, descripcion=f"Programo partido {partido.equipo_local} vs {partido.equipo_visitante}.")
        messages.success(request, "Programacion actualizada correctamente.")
        return redirect(volver_url)

    return render(request, "gestion/formulario.html", {
        "titulo": f"Programar partido: {partido.equipo_local} vs {partido.equipo_visitante}",
        "form": form,
        "volver_url": "gestion_partidos",
        "volver_href": volver_url,
    })


@login_required
@user_passes_test(es_editor_torneo)
@require_POST
def gestion_partido_confirmar_programacion(request, partido_id):
    torneo = torneo_actual(request)
    if not puede_gestionar_torneo(request, torneo, "programar"):
        return denegar_permiso_torneo()
    partidos = Partido.objects.select_related("categoria", "equipo_local", "equipo_visitante")
    if torneo:
        partidos = partidos.filter(categoria__torneo=torneo)
    partido = get_object_or_404(partidos, id=partido_id)
    partido.estado_programacion = "OFICIAL"
    partido.save(update_fields=["estado_programacion"])
    registrar_actividad(
        request,
        "CONFIRMAR_PROGRAMACION",
        partido,
        descripcion=f"Confirmo programacion de {partido.equipo_local} vs {partido.equipo_visitante}.",
    )
    messages.success(request, f"Programacion oficial: {partido.equipo_local} vs {partido.equipo_visitante}.")
    return redirect(request.POST.get("next") or "gestion_partidos")


@login_required
@user_passes_test(es_editor_torneo)
@require_POST
def gestion_partido_eliminar(request, partido_id):
    torneo = torneo_actual(request)
    if not puede_gestionar_torneo(request, torneo, "programar"):
        return denegar_permiso_torneo()
    partidos = Partido.objects.select_related("categoria", "equipo_local", "equipo_visitante")
    if torneo:
        partidos = partidos.filter(categoria__torneo=torneo)
    partido = get_object_or_404(partidos, id=partido_id)
    descripcion = f"Elimino partido {partido.equipo_local} vs {partido.equipo_visitante}."
    datos = {
        "partido_id": partido.id,
        "categoria": partido.categoria.nombre if partido.categoria_id else "",
        "equipo_local": partido.equipo_local.nombre if partido.equipo_local_id else "",
        "equipo_visitante": partido.equipo_visitante.nombre if partido.equipo_visitante_id else "",
        "fecha": partido.fecha.isoformat() if partido.fecha else "",
        "hora": partido.hora.isoformat() if partido.hora else "",
        "estado": partido.estado,
        "numero_fecha": partido.numero_fecha or "",
        "grupo": partido.grupo or "",
        "fase": partido.fase,
    }
    registrar_actividad(request, "ELIMINAR", partido, descripcion=descripcion, datos=datos)
    partido.delete()
    messages.success(request, "Partido eliminado correctamente.")
    return redirect(request.POST.get("next") or "gestion_partidos")


@login_required
@user_passes_test(es_editor_torneo)
@require_POST
def gestion_partido_validar_estadisticas(request, partido_id):
    torneo = torneo_actual(request)
    if not puede_gestionar_torneo(request, torneo, "validar"):
        return denegar_permiso_torneo()
    partidos = Partido.objects.select_related("categoria", "equipo_local", "equipo_visitante")
    if torneo:
        partidos = partidos.filter(categoria__torneo=torneo)
    partido = get_object_or_404(partidos, id=partido_id)
    _validar_estadisticas_partido(partido, request.user)
    registrar_actividad(request, "VALIDAR_ESTADISTICAS", partido, descripcion=f"Valido estadisticas de {partido.equipo_local} vs {partido.equipo_visitante}.")
    messages.success(request, f"Estadisticas validadas: {partido.equipo_local} vs {partido.equipo_visitante}.")
    return redirect(request.POST.get("next") or "gestion_partidos")


@login_required
@user_passes_test(es_editor_torneo)
def gestion_importar_partidos(request):
    torneo = torneo_actual(request)
    if not puede_gestionar_torneo(request, torneo, "programar"):
        return denegar_permiso_torneo()

    if request.method == "POST":
        archivo = request.FILES.get("archivo_excel")

        if not archivo:
            messages.error(request, "Selecciona un archivo Excel.")
            return redirect("gestion_importar_partidos")

        try:
            workbook = load_workbook(archivo, data_only=True)
            hoja = workbook.active
            encabezados = {}

            for indice, celda in enumerate(hoja[1]):
                if celda.value:
                    encabezados[normalizar_encabezado_excel(celda.value)] = indice

            creados = 0
            actualizados = 0
            omitidos = 0
            errores = []
            planilleros_asignados = 0
            tiene_columna_planillero = encabezado_existe(
                encabezados,
                "planillero",
                "planilleros",
                "planillero_asignado",
                "planilleros_asignados",
                "usuario_planillero",
                "usuarios_planilleros",
            )

            for numero_fila, row in enumerate(hoja.iter_rows(min_row=2), start=2):
                categoria_nombre = limpiar_texto_excel(valor_por_encabezado(row, encabezados, "categoria", "categoría"))
                local_nombre = limpiar_texto_excel(valor_por_encabezado(row, encabezados, "equipo_local", "local", "equipo local"))
                visitante_nombre = limpiar_texto_excel(valor_por_encabezado(row, encabezados, "equipo_visitante", "visitante", "equipo visitante"))

                if not categoria_nombre and not local_nombre and not visitante_nombre:
                    continue

                if not categoria_nombre or not local_nombre or not visitante_nombre:
                    omitidos += 1
                    errores.append(f"Fila {numero_fila}: falta categoría, local o visitante.")
                    continue

                categorias = Categoria.objects.filter(nombre__iexact=categoria_nombre)
                if torneo:
                    categorias = categorias.filter(torneo=torneo)
                categoria = categorias.first()

                if not categoria:
                    omitidos += 1
                    errores.append(f"Fila {numero_fila}: no existe la categoría {categoria_nombre}.")
                    continue

                local = Equipo.objects.filter(nombre__iexact=local_nombre, categoria=categoria).first()
                visitante = Equipo.objects.filter(nombre__iexact=visitante_nombre, categoria=categoria).first()

                if not local:
                    omitidos += 1
                    errores.append(f"Fila {numero_fila}: no existe el equipo local {local_nombre} en {categoria.nombre}.")
                    continue

                if not visitante:
                    omitidos += 1
                    errores.append(f"Fila {numero_fila}: no existe el equipo visitante {visitante_nombre} en {categoria.nombre}.")
                    continue

                numero_fecha = limpiar_texto_excel(valor_por_encabezado(row, encabezados, "numero_fecha", "fecha fixture", "jornada")) or "1"
                grupo = limpiar_texto_excel(valor_por_encabezado(row, encabezados, "grupo")) or "SIN GRUPO"
                fase = limpiar_texto_excel(valor_por_encabezado(row, encabezados, "fase")) or "GRUPOS"
                cancha = limpiar_texto_excel(valor_por_encabezado(row, encabezados, "cancha"))
                estado = limpiar_texto_excel(valor_por_encabezado(row, encabezados, "estado")) or "PROGRAMADO"
                fecha_partido = construir_fecha_partido_excel(valor_por_encabezado(row, encabezados, "fecha", "fecha_partido", "dia", "día", "fecha calendario"))
                hora_partido = construir_hora_partido_excel(valor_por_encabezado(row, encabezados, "hora"))
                goles_local = limpiar_entero_excel(valor_por_encabezado(row, encabezados, "goles_local", "gl")) or 0
                goles_visitante = limpiar_entero_excel(valor_por_encabezado(row, encabezados, "goles_visitante", "gv")) or 0
                planilleros_excel = valor_por_encabezado(
                    row,
                    encabezados,
                    "planillero",
                    "planilleros",
                    "planillero_asignado",
                    "planilleros_asignados",
                    "usuario_planillero",
                    "usuarios_planilleros",
                )

                if fase not in dict(Partido.FASES):
                    fase = fase.upper().replace(" ", "_")

                if fase not in dict(Partido.FASES):
                    fase = "GRUPOS"

                if estado not in dict(Partido.ESTADOS):
                    estado = estado.upper().replace(" ", "_")

                if estado not in dict(Partido.ESTADOS):
                    estado = "PROGRAMADO"

                partido, creado = Partido.objects.update_or_create(
                    categoria=categoria,
                    fase=fase,
                    numero_fecha=numero_fecha,
                    equipo_local=local,
                    equipo_visitante=visitante,
                    defaults={
                        "fecha": fecha_partido or date.today(),
                        "hora": hora_partido or time(0, 0),
                        "estado": estado,
                        "estado_programacion": "MANUAL",
                        "grupo": grupo,
                        "cancha": cancha,
                        "goles_local": goles_local,
                        "goles_visitante": goles_visitante,
                    },
                )

                if creado:
                    creados += 1
                else:
                    actualizados += 1

                if tiene_columna_planillero and limpiar_texto_excel(planilleros_excel):
                    planilleros, planilleros_no_encontrados = buscar_planilleros_excel(planilleros_excel)
                    if planilleros:
                        partido.planilleros.set(planilleros)
                        planilleros_asignados += len(planilleros)
                    if planilleros_no_encontrados:
                        errores.append(
                            f"Fila {numero_fila}: planillero(s) no encontrado(s): {', '.join(planilleros_no_encontrados)}."
                        )

            messages.success(
                request,
                (
                    f"Partidos importados. Nuevos: {creados}. Actualizados: {actualizados}. "
                    f"Omitidos: {omitidos}. Planilleros asignados: {planilleros_asignados}."
                ),
            )
            registrar_actividad(
                request,
                "IMPORTAR_PARTIDOS",
                torneo=torneo,
                descripcion=(
                    f"Importo partidos. Nuevos: {creados}. Actualizados: {actualizados}. "
                    f"Omitidos: {omitidos}. Planilleros asignados: {planilleros_asignados}."
                ),
                datos={
                    "creados": creados,
                    "actualizados": actualizados,
                    "omitidos": omitidos,
                    "planilleros_asignados": planilleros_asignados,
                },
            )

            for error in errores[:12]:
                messages.warning(request, error)

            if len(errores) > 12:
                messages.warning(request, f"Hay {len(errores) - 12} advertencias adicionales.")

            return redirect("gestion_partidos")

        except Exception as exc:
            messages.error(request, f"No se pudo importar el archivo: {exc}")
            return redirect("gestion_importar_partidos")

    return render(request, "gestion/importar_partidos.html")
def partido_live(request, partido_id):
    partido = get_object_or_404(
        Partido.objects.select_related(
            "categoria",
            "categoria__torneo",
            "equipo_local",
            "equipo_visitante"
        ),
        id=partido_id
    )
    volver_url = request.GET.get("volver", "").strip()
    if volver_url and not url_has_allowed_host_and_scheme(
        volver_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        volver_url = ""
    if not volver_url:
        volver_url = f"{reverse('panel')}?torneo={partido.categoria.torneo_id}"

    goles = Gol.objects.filter(partido=partido).select_related("jugador", "equipo").order_by("minuto", "creado_en", "id")
    tarjetas = Tarjeta.objects.filter(partido=partido).select_related("jugador", "equipo").order_by("minuto", "creado_en", "id")
    alineaciones = AlineacionPartido.objects.filter(partido=partido).select_related("jugador", "equipo").order_by("equipo__nombre", "rol", "jugador__nombres")
    sustituciones = SustitucionPartido.objects.filter(partido=partido).select_related("equipo", "jugador_sale", "jugador_entra").order_by("minuto", "creado_en", "id")
    sustituciones_local = [cambio for cambio in sustituciones if cambio.equipo_id == partido.equipo_local_id]
    sustituciones_visitante = [cambio for cambio in sustituciones if cambio.equipo_id == partido.equipo_visitante_id]
    for cambio in sustituciones:
        cambio.jugador_entra_corto = nombre_corto_jugador(cambio.jugador_entra)
        cambio.jugador_sale_corto = nombre_corto_jugador(cambio.jugador_sale)
        cambio.jugador_entra_edad = etiqueta_edad_jugador(cambio.jugador_entra, partido.categoria, partido.fecha)
        cambio.jugador_sale_edad = etiqueta_edad_jugador(cambio.jugador_sale, partido.categoria, partido.fecha)
        cambio.jugador_entra_foto = foto_jugador_url(cambio.jugador_entra)
        cambio.jugador_sale_foto = foto_jugador_url(cambio.jugador_sale)
        cambio.jugador_entra_iniciales = iniciales_jugador(cambio.jugador_entra)
        cambio.jugador_sale_iniciales = iniciales_jugador(cambio.jugador_sale)

    eventos_por_jugador = defaultdict(dict)

    def agregar_evento_jugador(jugador_id, tipo, titulo, cantidad=1):
        if not jugador_id:
            return
        eventos = eventos_por_jugador[jugador_id]
        if tipo in eventos:
            eventos[tipo].cantidad += cantidad
        else:
            eventos[tipo] = SimpleNamespace(
                tipo=tipo,
                titulo=titulo,
                cantidad=cantidad,
            )

    for gol in goles:
        if not gol.jugador_id:
            continue
        cantidad = max(gol.cantidad or 1, 1)
        es_autogol = bool(
            getattr(gol, "autogol", False)
            or getattr(gol, "es_autogol", False)
            or getattr(gol, "tipo", "") == "AUTOGOL"
        )
        es_penal = bool(getattr(gol, "es_penal", False))
        if es_autogol:
            tipo_gol = "autogol"
            titulo_gol = "Autogol"
        elif es_penal:
            tipo_gol = "penal"
            titulo_gol = "Gol de penal"
        else:
            tipo_gol = "gol"
            titulo_gol = "Gol"
        agregar_evento_jugador(gol.jugador_id, tipo_gol, titulo_gol, cantidad)

    for tarjeta in tarjetas:
        if not tarjeta.jugador_id:
            continue
        agregar_evento_jugador(
            tarjeta.jugador_id,
            "roja" if tarjeta.tipo == "ROJA" else "amarilla",
            tarjeta.get_tipo_display(),
        )

    for sustitucion in sustituciones:
        agregar_evento_jugador(sustitucion.jugador_sale_id, "sale", "Sustituido")
        agregar_evento_jugador(sustitucion.jugador_entra_id, "entra", "Ingreso")

    orden_ingreso_suplente = {}
    for indice, sustitucion in enumerate(sustituciones, start=1):
        orden_ingreso_suplente.setdefault(sustitucion.jugador_entra_id, indice)

    alineaciones_local = []
    alineaciones_visitante = []
    suplentes_local = []
    suplentes_visitante = []
    no_disponibles_local = []
    no_disponibles_visitante = []
    for alineacion in alineaciones:
        jugador = alineacion.jugador
        item = SimpleNamespace(
            jugador=jugador,
            nombre=jugador.nombres,
            nombre_corto=nombre_corto_jugador(jugador),
            dorsal=jugador.dorsal,
            rol=alineacion.rol,
            posicion=alineacion.posicion_cancha,
            foto=foto_jugador_url(jugador),
            iniciales=iniciales_jugador(jugador),
            etiqueta_edad=etiqueta_edad_jugador(jugador, partido.categoria, partido.fecha),
            eventos=list(eventos_por_jugador.get(jugador.id, {}).values()),
            orden_ingreso=orden_ingreso_suplente.get(jugador.id, 9999),
        )
        if alineacion.equipo_id == partido.equipo_local_id:
            if alineacion.rol == "SUPLENTE":
                suplentes_local.append(item)
            elif alineacion.rol == "NO_DISPONIBLE":
                no_disponibles_local.append(item)
            else:
                alineaciones_local.append(item)
        elif alineacion.equipo_id == partido.equipo_visitante_id:
            if alineacion.rol == "SUPLENTE":
                suplentes_visitante.append(item)
            elif alineacion.rol == "NO_DISPONIBLE":
                no_disponibles_visitante.append(item)
            else:
                alineaciones_visitante.append(item)

    alineaciones_local = _ordenar_titulares_cancha(alineaciones_local)
    alineaciones_visitante = _ordenar_titulares_cancha(alineaciones_visitante)
    suplentes_local = sorted(suplentes_local, key=lambda item: (item.orden_ingreso, item.nombre_corto))
    suplentes_visitante = sorted(suplentes_visitante, key=lambda item: (item.orden_ingreso, item.nombre_corto))

    eventos_live = []
    orden = 0
    for gol in goles:
        orden += 1
        if gol.es_autogol:
            tipo_evento = "autogol"
            detalle_gol = "Autogol"
        elif gol.es_penal:
            tipo_evento = "gol"
            detalle_gol = "Gol de penal"
        else:
            tipo_evento = "gol"
            detalle_gol = f"{gol.cantidad} gol(es)" if gol.cantidad > 1 else "Gol"
        eventos_live.append(SimpleNamespace(
            tipo=tipo_evento,
            icono="\u26bd",
            minuto=gol.minuto,
            equipo_id=gol.equipo_id,
            texto=nombre_resumen_jugador(gol.jugador),
            detalle=detalle_gol,
            creado_en=gol.creado_en,
            orden=gol.id,
        ))

    for tarjeta in tarjetas:
        orden += 1
        eventos_live.append(SimpleNamespace(
            tipo="tarjeta",
            icono="🟥" if tarjeta.tipo == "ROJA" else "🟨",
            minuto=tarjeta.minuto,
            equipo_id=tarjeta.equipo_id,
            texto=nombre_resumen_jugador(tarjeta.jugador),
            detalle=tarjeta.get_tipo_display(),
            creado_en=tarjeta.creado_en,
            orden=tarjeta.id,
        ))

    for sustitucion in sustituciones:
        orden += 1
        eventos_live.append(SimpleNamespace(
            tipo="sustitucion",
            icono="🔁",
            minuto=sustitucion.minuto,
            equipo_id=sustitucion.equipo_id,
            texto=nombre_resumen_jugador(sustitucion.jugador_entra),
            detalle=f"Sale {nombre_resumen_jugador(sustitucion.jugador_sale)}",
            jugador_entra=nombre_resumen_jugador(sustitucion.jugador_entra),
            jugador_sale=nombre_resumen_jugador(sustitucion.jugador_sale),
            creado_en=sustitucion.creado_en,
            orden=sustitucion.id,
        ))

    eventos_live = sorted(
        eventos_live,
        key=_clave_orden_evento_resumen,
        reverse=True,
    )
    eventos_live_grupos = _agrupar_eventos_resumen_live(eventos_live)

    return render(request, "partido_live.html", {
        "partido": partido,
        "escudo_local": escudo_url(partido.equipo_local),
        "escudo_visitante": escudo_url(partido.equipo_visitante),
        "fecha_inicio_live": partido.fecha.strftime("%Y-%m-%d") if partido.fecha else "",
        "hora_inicio_live": partido.hora.strftime("%H:%M") if partido.hora else "",
        "goles": goles,
        "tarjetas": tarjetas,
        "sustituciones": sustituciones,
        "sustituciones_local": sustituciones_local,
        "sustituciones_visitante": sustituciones_visitante,
        "alineaciones_local": alineaciones_local,
        "alineaciones_visitante": alineaciones_visitante,
        "suplentes_local": suplentes_local,
        "suplentes_visitante": suplentes_visitante,
        "no_disponibles_local": no_disponibles_local,
        "no_disponibles_visitante": no_disponibles_visitante,
        "eventos_live": eventos_live,
        "eventos_live_grupos": eventos_live_grupos,
        "segundos_vivos": segundos_vivos_partido(partido),
        "volver_url": volver_url,
        "delegado_alineacion_url": url_alineacion_delegado_si_aplica(request.user, partido),
        "puede_diligenciar_partido": puede_diligenciar_partido(request.user, partido),
    })
def _pausar_cronometro(partido):
    if partido.inicio_en_vivo:
        diferencia = timezone.now() - partido.inicio_en_vivo
        partido.segundos_acumulados += int(diferencia.total_seconds())

    partido.inicio_en_vivo = None
    partido.cronometro_pausado = True
    partido.save()


@login_required
def cronometro_primer_tiempo(request, partido_id):
    partido = get_object_or_404(Partido, id=partido_id)
    if not puede_diligenciar_partido(request.user, partido):
        return denegar_partido_no_autorizado()
    partido.estado = "EN_JUEGO"
    partido.periodo_en_vivo = "PT"
    partido.cronometro_pausado = False

    if not partido.inicio_en_vivo:
        partido.inicio_en_vivo = timezone.now()

    partido.save()
    return redirect("editor_partido_movil", partido_id=partido.id)


@login_required
def cronometro_entretiempo(request, partido_id):
    partido = get_object_or_404(Partido, id=partido_id)
    if not puede_diligenciar_partido(request.user, partido):
        return denegar_partido_no_autorizado()
    _pausar_cronometro(partido)
    partido.periodo_en_vivo = "ET"
    partido.save()
    return redirect("editor_partido_movil", partido_id=partido.id)


@login_required
def cronometro_segundo_tiempo(request, partido_id):
    partido = get_object_or_404(Partido, id=partido_id)
    if not puede_diligenciar_partido(request.user, partido):
        return denegar_partido_no_autorizado()
    partido.estado = "EN_JUEGO"
    partido.periodo_en_vivo = "ST"
    partido.cronometro_pausado = False
    partido.inicio_en_vivo = timezone.now()
    partido.save()
    actualizar_incidencia_regla_edad(partido, partido.equipo_local, request=request, permitir_crear=True)
    actualizar_incidencia_regla_edad(partido, partido.equipo_visitante, request=request, permitir_crear=True)
    return redirect("editor_partido_movil", partido_id=partido.id)


@login_required
def cronometro_pausar(request, partido_id):
    partido = get_object_or_404(Partido, id=partido_id)
    if not puede_diligenciar_partido(request.user, partido):
        return denegar_partido_no_autorizado()
    _pausar_cronometro(partido)
    return redirect("editor_partido_movil", partido_id=partido.id)


@login_required
def cronometro_reanudar(request, partido_id):
    partido = get_object_or_404(Partido, id=partido_id)
    if not puede_diligenciar_partido(request.user, partido):
        return denegar_partido_no_autorizado()
    partido.estado = "EN_JUEGO"
    partido.cronometro_pausado = False
    partido.inicio_en_vivo = timezone.now()
    partido.save()
    return redirect("editor_partido_movil", partido_id=partido.id)


@login_required
def cronometro_suspender(request, partido_id):
    partido = get_object_or_404(Partido, id=partido_id)
    if not puede_diligenciar_partido(request.user, partido):
        return denegar_partido_no_autorizado()
    _pausar_cronometro(partido)
    partido.estado = "SUSPENDIDO"
    partido.save()
    return redirect("editor_partido_movil", partido_id=partido.id)


@login_required
def cronometro_finalizar(request, partido_id):
    partido = get_object_or_404(Partido, id=partido_id)
    if not puede_diligenciar_partido(request.user, partido):
        return denegar_partido_no_autorizado()
    _pausar_cronometro(partido)
    partido.estado = "FINALIZADO"
    partido.periodo_en_vivo = "FIN"
    partido.save()
    if not es_editor_torneo(request.user):
        return redirect("partido_live", partido_id=partido.id)
    return redirect("editor_partido_movil", partido_id=partido.id)


def partidos_para_planillas(torneo, categoria_id=None):
    partidos = Partido.objects.select_related(
        "categoria",
        "categoria__torneo",
        "equipo_local",
        "equipo_visitante",
    ).prefetch_related(
        "equipo_local__jugadores",
        "equipo_visitante__jugadores",
    ).filter(
        estado="PROGRAMADO",
        estado_programacion__in=["MANUAL", "OFICIAL"],
    ).exclude(
        cancha__isnull=True,
    ).exclude(
        cancha__exact="",
    ).order_by("fecha", "hora", "categoria__nombre", "fase", "numero_fecha", "equipo_local__nombre")
    if torneo:
        partidos = partidos.filter(categoria__torneo=torneo)
    if categoria_id:
        partidos = partidos.filter(categoria_id=categoria_id)
    return partidos


def respuesta_pdf_planilla(partido):
    contenido = generar_planilla_juego_pdf(partido)
    response = HttpResponse(contenido, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo_planilla(partido)}"'
    return response


def respuesta_archivo_descarga_app(request, contenido, nombre_archivo, content_type, volver_url, archivo_url):
    data_url = f"data:{content_type};base64,{base64.b64encode(contenido).decode('ascii')}"
    return render(request, "descargas/archivo_descarga.html", {
        "data_url": data_url,
        "nombre_archivo": nombre_archivo,
        "content_type": content_type,
        "volver_url": volver_url,
        "archivo_url": archivo_url,
    })


@login_required
@user_passes_test(es_editor_torneo)
def descargar_planilla_juego_partido(request, partido_id):
    partido = get_object_or_404(
        partidos_para_planillas(None),
        id=partido_id,
    )
    if not puede_gestionar_torneo(request, partido.categoria.torneo if partido.categoria_id else None, "descargar_planillas"):
        return denegar_permiso_torneo()
    if request.GET.get("app") == "1":
        return respuesta_archivo_descarga_app(
            request,
            generar_planilla_juego_pdf(partido),
            nombre_archivo_planilla(partido),
            "application/pdf",
            request.GET.get("volver") or reverse("gestion_partidos"),
            request.build_absolute_uri(reverse("descargar_planilla_juego_partido", args=[partido.id])),
        )
    return respuesta_pdf_planilla(partido)


def generar_zip_planillas(partidos):
    buffer = BytesIO()
    usados = set()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archivo_zip:
        for partido in partidos:
            nombre = nombre_archivo_planilla(partido)
            if nombre in usados:
                base, ext = nombre.rsplit(".", 1)
                nombre = f"{base}-{partido.id}.{ext}"
            usados.add(nombre)
            archivo_zip.writestr(nombre, generar_planilla_juego_pdf(partido))
    buffer.seek(0)
    return buffer.getvalue()


def respuesta_zip_planillas(partidos, nombre_zip):
    response = HttpResponse(generar_zip_planillas(partidos), content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="{nombre_zip}"'
    return response


@login_required
@user_passes_test(es_editor_torneo)
def descargar_planillas_juego_categoria(request, categoria_id):
    torneo = torneo_actual(request)
    categoria = get_object_or_404(Categoria, id=categoria_id)
    if not puede_gestionar_torneo(request, categoria.torneo, "descargar_planillas"):
        return HttpResponseForbidden("No tienes permiso para esta categoria.")
    partidos = list(partidos_para_planillas(torneo, categoria_id=categoria.id))
    if not partidos:
        messages.warning(request, "No hay partidos para generar planillas en esta categoria.")
        return redirect("gestion_partidos")
    nombre_zip = f"PLANILLAS_{limpiar_nombre(categoria.nombre)}.zip"
    if request.GET.get("app") == "1":
        return respuesta_archivo_descarga_app(
            request,
            generar_zip_planillas(partidos),
            nombre_zip,
            "application/zip",
            request.GET.get("volver") or reverse("gestion_partidos"),
            request.build_absolute_uri(reverse("descargar_planillas_juego_categoria", args=[categoria.id])),
        )
    return respuesta_zip_planillas(partidos, nombre_zip)


@login_required
@user_passes_test(es_editor_torneo)
def descargar_planillas_juego_torneo(request):
    torneo = torneo_actual(request)
    if not puede_gestionar_torneo(request, torneo, "descargar_planillas"):
        return denegar_permiso_torneo()
    partidos = list(partidos_para_planillas(torneo))
    if not partidos:
        messages.warning(request, "No hay partidos para generar planillas.")
        return redirect("gestion_partidos")
    nombre_base = limpiar_nombre(torneo.nombre) if torneo else "TORNEO"
    nombre_zip = f"PLANILLAS_{nombre_base}.zip"
    if request.GET.get("app") == "1":
        return respuesta_archivo_descarga_app(
            request,
            generar_zip_planillas(partidos),
            nombre_zip,
            "application/zip",
            request.GET.get("volver") or reverse("gestion_partidos"),
            request.build_absolute_uri(reverse("descargar_planillas_juego_torneo")),
        )
    return respuesta_zip_planillas(partidos, nombre_zip)
